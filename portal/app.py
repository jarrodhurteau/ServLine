# portal/app.py 
from flask import (
    Flask, jsonify, render_template, abort, request, redirect, url_for,
    session, send_from_directory, flash, make_response, send_file, g     # ← added g (Day 84)
)

# --- Standard libs & typing ---
import sqlite3
from pathlib import Path
from functools import wraps
import uuid
import os
import threading
import json
import shutil
from datetime import datetime
from typing import Optional, Iterable, Tuple, List, Dict, Any
import hashlib  # <-- added for cat_hue filter
import time     # <-- NEW: for gentle polling after upload

# --- Paths ---
ROOT = Path(__file__).resolve().parents[1]

# Make project root importable so we can import storage.*
import sys
if str(ROOT) not in sys.path:
    sys.path.append(str(ROOT))

from storage import import_jobs as import_jobs_store  # <-- NEW: structured import helpers
# segment_document import removed — facade provides layout data; no need for duplicate call


# --- Forward decls for type checkers (real implementations appear later) ---
def _ocr_image_to_text(img_path: Path) -> str: ...
def _pdf_to_text(pdf_path: Path) -> str: ...


# stdlib for exports
import io
import csv
import re  # <-- for OCR parsing
import statistics  # <-- for export metrics (Day 81)

# NEW: optional Excel export dependency
openpyxl = None  # type: ignore[assignment]
try:
    import openpyxl as _openpyxl  # type: ignore[import]
    openpyxl = _openpyxl  # type: ignore[assignment]
except Exception:
    openpyxl = None  # type: ignore[assignment]


# safer filename + big-file error handling
from werkzeug.utils import secure_filename
from werkzeug.exceptions import RequestEntityTooLarge
from werkzeug.security import check_password_hash

# OCR health imports
import pytesseract

# OCR health (single source of truth)
ocr_health_lib = None  # type: ignore[assignment]

try:
    from storage.ocr_facade import health as _ocr_health_lib  # type: ignore
    ocr_health_lib = _ocr_health_lib
except Exception as e:
    _ocr_health_import_error = repr(e)

    def _fallback_ocr_health_lib():
        return {
            "engine": "error",
            "error": f"ocr_facade health import failed: {_ocr_health_import_error}",
        }

    ocr_health_lib = _fallback_ocr_health_lib




# ✅ Try to import the contract validator (with safe fallback if file not added yet)
try:
    from portal.contracts import validate_draft_payload  # type: ignore
except Exception:
    def validate_draft_payload(_payload):  # type: ignore
        # No-op validator so app still runs if contracts.py isn't present yet.
        return True, ""


# ✅ Make sure the OCR worker is imported at app startup
#    (this triggers the version banner inside ocr_worker.py)
from portal import ocr_worker
print("[App] Imported portal.ocr_worker")  # optional confirmation from app.py


# ------------------------
# App & Config
# ------------------------
app = Flask(__name__)
log = app.logger

# --- Config (dev) ---
app.config["SECRET_KEY"] = "dev-secret-change-me"          # replace later with env var
app.config["MAX_CONTENT_LENGTH"] = 20 * 1024 * 1024        # ~20 MB
# Dev QoL: auto-reload templates and disable static caching when iterating on UI
app.config["TEMPLATES_AUTO_RELOAD"] = True
app.config["SEND_FILE_MAX_AGE_DEFAULT"] = 0

DEV_USERNAME = "admin"
DEV_PASSWORD = "letmein"

# --- Debug OCR routes ---
from portal.routes_debug_preocr import debug_preocr
app.register_blueprint(debug_preocr)

# --- Paths ---
DB_PATH = ROOT / "storage" / "servline.db"

# --- Load .env if available (so TESSERACT_CMD / POPPLER_PATH work even without PATH) ---
try:
    from dotenv import load_dotenv  # optional; ok if not installed
    load_dotenv(ROOT / ".env")
except Exception:
    pass

# --- OCR paths from env ---
TESSERACT_CMD = os.getenv("TESSERACT_CMD")
POPPLER_PATH = os.getenv("POPPLER_PATH") or None
# Allow tuning Tesseract without code changes
TESSERACT_LANG = os.getenv("TESSERACT_LANG") or "eng"
TESSERACT_CONFIG = os.getenv("TESSERACT_CONFIG") or "--oem 1 --psm 3"

# Day 20 — Canonical taxonomy seed (editable)
TAXONOMY_SEED = [
    "pizzas", "calzones", "salads", "wings", "appetizers", "burgers", "sandwiches", "subs",
    "pasta", "steaks", "seafood", "tacos", "burritos", "sides", "desserts", "beverages",
    "breakfast", "lunch specials", "dinner specials", "kids menu"
]

# Windows-friendly Tesseract discovery:
# 1) Respect explicit env var if it exists
# 2) Else use PATH
# 3) Else try common install locations
if TESSERACT_CMD and Path(TESSERACT_CMD).exists():
    pytesseract.pytesseract.tesseract_cmd = TESSERACT_CMD
else:
    _which = shutil.which("tesseract") or shutil.which("tesseract.exe")
    if _which:
        pytesseract.pytesseract.tesseract_cmd = _which
    else:
        _common = [
            r"C:\Program Files\Tesseract-OCR\tesseract.exe",
            r"C:\Program Files (x86)\Tesseract-OCR\tesseract.exe",
        ]
        for p in _common:
            if Path(p).exists():
                pytesseract.pytesseract.tesseract_cmd = p
                break

# storage layer for drafts (DB-first, Day 12+)
try:
    from storage import drafts as drafts_store
except Exception:
    drafts_store = None  # guarded below

# storage layer for menus (multi-menu & versioning, Phase 10 Day 87+)
menus_store = None  # type: ignore[assignment]
try:
    from storage import menus as menus_store
except Exception:
    menus_store = None  # guarded below

# storage layer for user accounts (Phase 13 Day 126+)
users_store = None  # type: ignore[assignment]
try:
    from storage import users as users_store
    users_store._ensure_users_schema()
    users_store._ensure_restaurant_columns()  # Day 128: cuisine_type, website, updated_at
    users_store._ensure_tier_column()  # Day 131: account_tier
except Exception:
    users_store = None  # guarded below

# Price comparison intelligence (Phase 13 Day 134+)
price_intel = None  # type: ignore[assignment]
try:
    from storage import price_intel
    print("[APP] Loaded price_intel OK")
except Exception:
    price_intel = None

# Price intelligence — Claude Call 4 (Phase 13 Day 135+)
ai_price_intel = None  # type: ignore[assignment]
try:
    from storage import ai_price_intel
    print("[APP] Loaded ai_price_intel OK")
except Exception:
    ai_price_intel = None

# Day 136: ensure pipeline_stage column exists on import_jobs
# Day 139: use sqlite3 directly since db_connect isn't defined yet at import time
def _ensure_import_jobs_columns():
    """Add missing columns to import_jobs (idempotent)."""
    try:
        import sqlite3 as _sq
        from pathlib import Path as _P
        _db_path = _P(__file__).resolve().parents[1] / "storage" / "servline.db"
        if not _db_path.exists():
            return
        conn = _sq.connect(str(_db_path))
        try:
            cols = {r[1] for r in conn.execute("PRAGMA table_info(import_jobs)").fetchall()}
            if "pipeline_stage" not in cols:
                conn.execute("ALTER TABLE import_jobs ADD COLUMN pipeline_stage TEXT")
                conn.commit()
                print("[APP] Added pipeline_stage column to import_jobs")
            # Day 141: store extra page filenames (JSON array) for multi-page menus
            if "extra_filenames" not in cols:
                conn.execute("ALTER TABLE import_jobs ADD COLUMN extra_filenames TEXT")
                conn.commit()
                print("[APP] Added extra_filenames column to import_jobs")
        finally:
            conn.close()
    except Exception as _e:
        print(f"[APP] import_jobs column backfill: {_e}")

_ensure_import_jobs_columns()

# OCR engine (Day-21 revamp / One Brain façade)
try:
    from storage.ocr_facade import build_structured_menu
    extract_items_from_path = build_structured_menu
    print("[APP] Loaded OCR facade OK")
except Exception as e:
    print("[APP] OCR facade failed:", e)

    extract_items_from_path = None
    _ocr_facade_error = repr(e)

# AI OCR Heuristics (Day 20) — removed from pipeline in Day 100.5
# analyze_ocr_text import removed: heuristic fallback no longer used in pipeline or routes

# uploads (kept out of git via .gitignore)
UPLOAD_FOLDER = ROOT / "uploads"
UPLOAD_FOLDER.mkdir(parents=True, exist_ok=True)
TRASH_FOLDER = UPLOAD_FOLDER / ".trash"
TRASH_FOLDER.mkdir(parents=True, exist_ok=True)

# drafts + raw artifacts
DRAFTS_FOLDER = ROOT / "storage" / "drafts"
DRAFTS_FOLDER.mkdir(parents=True, exist_ok=True)

# Primary raw folder (newer layout) and legacy raw folder support
RAW_FOLDER = DRAFTS_FOLDER / "raw"
RAW_FOLDER.mkdir(parents=True, exist_ok=True)
LEGACY_RAW_FOLDER = ROOT / "storage" / "raw"
LEGACY_RAW_FOLDER.mkdir(parents=True, exist_ok=True)

# trash bins for artifacts
TRASH_DRAFTS = DRAFTS_FOLDER / ".trash"
TRASH_DRAFTS.mkdir(parents=True, exist_ok=True)
TRASH_RAW = RAW_FOLDER / ".trash"
TRASH_RAW.mkdir(parents=True, exist_ok=True)
LEGACY_TRASH_RAW = LEGACY_RAW_FOLDER / ".trash"
LEGACY_TRASH_RAW.mkdir(parents=True, exist_ok=True)

# Allowed upload types
ALLOWED_EXTENSIONS = {"jpg", "jpeg", "png", "pdf"}
def allowed_file(filename: str) -> bool:
    return "." in filename and filename.rsplit(".", 1)[1].lower() in ALLOWED_EXTENSIONS

def _now_iso() -> str:
    return datetime.utcnow().isoformat(timespec="seconds") + "Z"

# ------------------------------------------------------------
# Day 84: In-memory rate limiter for REST API
# ------------------------------------------------------------
from collections import deque

_rate_limit_lock = threading.Lock()
_rate_limit_windows: Dict[int, deque] = {}


def _check_rate_limit(key_record: Dict[str, Any]) -> Tuple[bool, Dict[str, str]]:
    """Sliding-window rate limiter (per minute). Returns (allowed, headers)."""
    key_id = key_record["id"]
    limit = key_record.get("rate_limit_rpm", 60)
    now = time.time()
    window_start = now - 60.0

    with _rate_limit_lock:
        if key_id not in _rate_limit_windows:
            _rate_limit_windows[key_id] = deque()
        window = _rate_limit_windows[key_id]

        # Evict expired entries
        while window and window[0] <= window_start:
            window.popleft()

        remaining = max(0, limit - len(window))
        reset_at = int(now + 60)

        headers = {
            "X-RateLimit-Limit": str(limit),
            "X-RateLimit-Remaining": str(remaining),
            "X-RateLimit-Reset": str(reset_at),
        }

        if len(window) >= limit:
            headers["Retry-After"] = "60"
            return False, headers

        window.append(now)
        headers["X-RateLimit-Remaining"] = str(max(0, limit - len(window)))
        return True, headers


def api_key_required(view_func):
    """Decorator: authenticate via X-API-Key or Authorization: Bearer header."""
    @wraps(view_func)
    def wrapper(*args, **kwargs):
        _require_drafts_storage()

        # Extract key from header
        raw_key = request.headers.get("X-API-Key", "").strip()
        if not raw_key:
            auth = request.headers.get("Authorization", "")
            if auth.lower().startswith("bearer "):
                raw_key = auth[7:].strip()

        if not raw_key:
            return jsonify({"ok": False, "error": "Missing API key"}), 401

        key_record = drafts_store.validate_api_key(raw_key)
        if key_record is None:
            return jsonify({"ok": False, "error": "Invalid API key"}), 401

        if not key_record.get("active"):
            return jsonify({"ok": False, "error": "API key is revoked"}), 403

        # Rate limit check
        allowed, rl_headers = _check_rate_limit(key_record)
        if not allowed:
            resp = jsonify({"ok": False, "error": "Rate limit exceeded"})
            resp.status_code = 429
            for k, v in rl_headers.items():
                resp.headers[k] = v
            return resp

        g.api_key = key_record

        response = view_func(*args, **kwargs)
        if isinstance(response, tuple):
            resp_obj = make_response(*response)
        else:
            resp_obj = response
        for k, v in rl_headers.items():
            resp_obj.headers[k] = v
        return resp_obj

    return wrapper


# ------------------------
# Safe render helper (prevents template-caused 500 loops)
# ------------------------
def _safe_render(template_name: str, **ctx):
    # Try to print the exact template file being used
    try:
        if app and app.jinja_loader:
            try:
                # get_source returns (source, filename, uptodate)
                _src, _filename, _ = app.jinja_loader.get_source(app.jinja_env, template_name)
                print(f"[TEMPLATE DEBUG] -> {template_name} from {_filename}")
            except Exception as e:
                print(f"[TEMPLATE DEBUG] (could not resolve path for {template_name}): {e}")
    except Exception:
        # If printing fails for any reason, we still render normally
        pass

    try:
        return render_template(template_name, **ctx)
    except Exception:
        # Minimal inline fallback to expose the real traceback
        import html, traceback
        tb = html.escape(traceback.format_exc())
        body = f"<h1>{html.escape(template_name)} missing or failed</h1><pre>{tb}</pre>"
        return body, 200, {"Content-Type": "text/html; charset=utf-8"}

# ------------------------
# Template globals + filters
# ------------------------
def _cat_hue(value: Optional[str]) -> int:
    """
    Deterministic hue (0–359) from a category string.
    Usage in templates: style="--hue: {{ category|cat_hue }};"
    """
    s = (value or "").strip().lower()
    if not s:
        return 210  # default-ish blue
    h = int(hashlib.md5(s.encode("utf-8")).hexdigest()[:8], 16) % 360
    return h

@app.template_filter("cat_hue")
def jinja_cat_hue(value: Optional[str]) -> int:
    return _cat_hue(value)

# ------------------------
# Confidence → CSS class mapping
# ------------------------
@app.template_filter("confidence_class")
def confidence_class(conf) -> str:
    """
    Map numeric confidence (0–100) to a CSS class.
    """
    try:
        c = int(conf or 0)
    except (TypeError, ValueError):
        c = 0

    if c >= 80:
        return "conf-high"
    elif c >= 50:
        return "conf-med"
    elif c > 0:
        return "conf-low"
    else:
        return "conf-unknown"


def score_item_quality(item: Dict[str, Any]) -> Tuple[int, bool]:
    """
    Compute a 0–100 quality score for a cleaned draft item and whether it's low-confidence.

    Inputs (post-AI-cleanup fields, if present):
      - item["confidence"]: OCR/parse confidence (0–100)
      - item["price_cents"]: int cents (0 if missing/invalid)
      - item["category"]: string (may be 'Uncategorized')
      - item["name"]: cleaned name
      - item["description"]: cleaned description

    Returns:
      (quality_score, is_low_confidence)
    """
    name = (item.get("name") or "").strip()
    desc = (item.get("description") or "").strip()
    category = (item.get("category") or "").strip() or "Uncategorized"

    # Confidence as integer baseline
    conf_raw = item.get("confidence")
    try:
        conf = int(conf_raw) if conf_raw is not None else 0
    except Exception:
        conf = 0

    # Price in cents
    price_raw = item.get("price_cents")
    try:
        price_cents = int(price_raw) if price_raw is not None else 0
    except Exception:
        price_cents = 0

    # Start from OCR confidence (already 0–100)
    score = conf

    # --- Price validity ---
    if price_cents <= 0:
        # Missing/zero price is a strong negative
        score -= 15
    else:
        # Very low price (likely sides) is a small nudge
        if price_cents < 300:      # <$3
            score -= 3
        # Very high price (probably parse issue)
        if price_cents > 6000:     # >$60
            score -= 8

    # --- Category quality ---
    if category.lower() in {"uncategorized", "other", "misc"}:
        score -= 10
    else:
        score += 3

    # --- Name length sanity ---
    nlen = len(name)
    if nlen == 0:
        score -= 25
    elif nlen < 6:
        score -= 10
    elif nlen > 120:
        score -= 15
    elif nlen > 80:
        score -= 10

    # --- Junk-symbol density (proxy for OCR/cleanup difficulty) ---
    if nlen:
        clean_chars = sum(
            1 for c in name
            if (c.isalnum() or c.isspace() or c in "&()/+'-.$")
        )
        junk_ratio = 1.0 - (clean_chars / max(nlen, 1))
        if junk_ratio > 0.45:
            score -= 25
        elif junk_ratio > 0.30:
            score -= 15
        elif junk_ratio > 0.20:
            score -= 8
    else:
        junk_ratio = 0.0  # not used directly, kept for clarity

    # --- Description sanity (tiny nudge) ---
    if not desc and nlen > 40:
        score -= 3

    # Clamp to [0, 100]
    if score < 0:
        score = 0
    if score > 100:
        score = 100

    low_conf = score < 65
    return score, low_conf


@app.context_processor
def inject_globals():
    """Provide `now`, `show_admin`, `is_customer`, `account_tier`, and
    `dev_tools` (env-gated opt-in for in-editor testing helpers) to templates."""
    raw = session.get("user")
    u = raw if isinstance(raw, dict) else {}
    role = u.get("role")
    dev_tools = os.environ.get("DEV_TOOLS_ENABLED", "").strip().lower() in ("1", "true", "yes")
    return {
        "now": datetime.utcnow().strftime("%Y-%m-%d %H:%M UTC"),
        "show_admin": role == "admin" or (bool(u) and role is None),
        "is_customer": role == "customer",
        "account_tier": u.get("account_tier"),
        "dev_tools": dev_tools,
    }

# ------------------------
# DB helpers (align with schema.sql: status only)
# ------------------------
def db_connect():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA foreign_keys = ON;")
    return conn

def create_import_job(filename: str, restaurant_id: Optional[int] = None) -> int:
    with db_connect() as conn:
        cur = conn.cursor()
        cur.execute(
            """
            INSERT INTO import_jobs (restaurant_id, filename, status, created_at, updated_at)
            VALUES (?, ?, 'pending', datetime('now'), datetime('now'))
            """,
            (restaurant_id, filename),
        )
        job_id = cur.lastrowid
        conn.commit()
        return int(job_id)

def update_import_job(job_id: int, **fields):
    if not fields:
        return
    sets = ", ".join(f"{k}=?" for k in fields.keys())
    values = list(fields.values())
    sets = f"{sets}, updated_at=datetime('now')"
    with db_connect() as conn:
        conn.execute(f"UPDATE import_jobs SET {sets} WHERE id=?", (*values, job_id))
        conn.commit()

def get_import_job(job_id: int):
    with db_connect() as conn:
        return conn.execute("SELECT * FROM import_jobs WHERE id=?", (job_id,)).fetchone()

def list_import_jobs(limit: int = 100, include_deleted: bool = False):
    with db_connect() as conn:
        if include_deleted:
            sql = """
                SELECT * FROM import_jobs
                ORDER BY datetime(created_at) DESC
                LIMIT ?
            """
            args = (limit,)
        else:
            sql = """
                SELECT * FROM import_jobs
                WHERE COALESCE(status,'') != 'deleted'
                ORDER BY datetime(created_at) DESC
                LIMIT ?
            """
            args = (limit,)
        return conn.execute(sql, args).fetchall()

def _jobs_for_upload_filename(upload_name: str):
    with db_connect() as conn:
        return conn.execute(
            "SELECT id, draft_path FROM import_jobs WHERE filename=?",
            (upload_name, ),
        ).fetchall()

# ------------------------
# Auth helper
# ------------------------
def login_required(view_func):
    @wraps(view_func)
    def wrapper(*args, **kwargs):
        if not session.get("user"):
            return redirect(url_for("login", next=request.path))
        return view_func(*args, **kwargs)
    return wrapper

# ------------------------
# Role/restaurant helpers (ADMIN vs CUSTOMER scoping) — Day 126-127
# ------------------------
def _is_admin() -> bool:
    """True if the logged-in user has the admin role."""
    return (session.get("user") or {}).get("role") == "admin"


def _is_customer() -> bool:
    """True if the logged-in user has the customer role."""
    return (session.get("user") or {}).get("role") == "customer"


def require_restaurant_access(view_func):
    """Decorator: ensures a customer user owns the restaurant in the route.

    Expects `rest_id` in the route kwargs.  Admins pass through.
    Customers must have a user_restaurants link (checked via users_store).
    """
    @wraps(view_func)
    def wrapper(*args, **kwargs):
        u = session.get("user") or {}
        role = u.get("role")
        # Admins and legacy sessions (no role field) pass through
        if role == "admin" or role is None:
            return view_func(*args, **kwargs)
        # Customer: verify ownership via user_restaurants link
        rest_id = kwargs.get("rest_id")
        user_id = u.get("user_id")
        if rest_id and user_id and users_store:
            if users_store.user_owns_restaurant(user_id, rest_id):
                return view_func(*args, **kwargs)
        abort(403, description="You do not have access to this restaurant.")
    return wrapper
def _resolve_restaurant_id_for_action(job_row: sqlite3.Row) -> Optional[int]:
    """
    Determine the restaurant context for an action:
      - customer side: session['user']['restaurant_id']
      - admin side: use job_row.restaurant_id (must be chosen)
    """
    u = (session.get("user") or {})
    if u.get("role") == "customer" and u.get("restaurant_id"):
        try:
            return int(u["restaurant_id"])
        except Exception:
            return None
    rid = job_row["restaurant_id"]
    try:
        return int(rid) if rid is not None else None
    except Exception:
        return None

def _resolve_restaurant_id_from_request() -> Optional[int]:
    """
    Used at upload time:
      - Explicit form['restaurant_id'] takes priority (e.g. "Add Menu" from restaurant page).
      - Else customer: fall back to session restaurant_id.
      - Else (admin): try form['restaurant_id'] if provided.
    """
    # Explicit restaurant_id from form (Day 133: "Add Menu" flow)
    rid = request.form.get("restaurant_id")
    if rid and str(rid).strip():
        try:
            return int(rid)
        except Exception:
            pass
    # Customer fallback: session restaurant
    u = (session.get("user") or {})
    if u.get("role") == "customer" and u.get("restaurant_id"):
        try:
            return int(u["restaurant_id"])
        except Exception:
            return None
    return None

def _find_or_create_menu_for_restaurant(conn: sqlite3.Connection, restaurant_id: int) -> int:
    """Return an existing active menu for the restaurant, or create a new one."""
    cur = conn.cursor()
    row = cur.execute(
        "SELECT id FROM menus WHERE restaurant_id=? AND active=1 ORDER BY id LIMIT 1",
        (int(restaurant_id),)
    ).fetchone()
    if row:
        return int(row["id"] if isinstance(row, sqlite3.Row) else row[0])
    name = f"Imported {datetime.utcnow().date()}"
    cur.execute(
        "INSERT INTO menus (restaurant_id, name, active) VALUES (?, ?, 1)",
        (int(restaurant_id), name)
    )
    return int(cur.lastrowid)

# ------------------------
# Small helpers for Day 13 flow
# ------------------------
def _require_drafts_storage():
    if drafts_store is None:
        abort(500, description="Drafts storage layer not available. Ensure storage/drafts.py exists and is importable.")

def _abs_from_rel(rel_path: Optional[str]) -> Optional[Path]:
    if not rel_path:
        return None
    return (ROOT / rel_path).resolve()

def _price_to_cents(v) -> int:
    """Accept floats/strings like 12.5, '12.50', '$12.5', '12 50' and return cents."""
    if v is None:
        return 0
    try:
        if isinstance(v, (int, float)):
            return int(round(float(v) * 100))
        s = str(v)
        s = s.replace("$", "").replace(",", " ").strip()
        # handle '12 50' -> '12.50'
        parts = [p for p in s.split() if p]
        if len(parts) == 2 and parts[1].isdigit() and len(parts[1]) in (1, 2):
            s = parts[0] + "." + parts[1].rjust(2, "0")
        return int(round(float(s) * 100))
    except Exception:
        return 0
def _find_or_create_menu_for_job(conn: sqlite3.Connection, job_row: sqlite3.Row) -> int:
    """(Legacy) Pick an existing active menu for the job's restaurant, else create one."""
    rest_id = job_row["restaurant_id"]
    cur = conn.cursor()
    if rest_id:
        m = cur.execute(
            "SELECT id FROM menus WHERE restaurant_id=? AND active=1 ORDER  BY id LIMIT 1",
            (rest_id,),
        ).fetchone()
        if m:
            return int(m["id"])
    # create a new menu
    name = f"Imported {datetime.utcnow().date()}"
    cur.execute(
        "INSERT INTO menus (restaurant_id, name, active) VALUES (?, ?, 1)",
        (rest_id, name),
    )
    return int(cur.lastrowid)

def _find_draft_for_job(job_id: int) -> Optional[int]:
    """
    Read-only lookup: return an existing DB-backed draft_id linked via source_job_id.
    Does NOT create drafts from preview/legacy JSON.
    """
    _require_drafts_storage()

    if hasattr(drafts_store, "find_draft_by_source_job"):
        existing = drafts_store.find_draft_by_source_job(job_id)
        if existing and (existing.get("id") or existing.get("draft_id")):
            try:
                return int(existing.get("id") or existing.get("draft_id"))
            except Exception:
                return None

    return None


def _get_or_create_draft_for_job(job_id: int, allow_create: bool = False) -> Optional[int]:
    """
    Return a draft_id for this import job.

    If allow_create=False (default): read-only behavior, returns existing draft_id or None.
    If allow_create=True: will create a draft from the best available JSON source.

    Priority (when allow_create=True):
      1. Reuse existing DB-backed draft linked via source_job_id.
      2. For OCR/AI jobs: use ai_preview_path / preview_path JSON if present.
      3. For structured JSON jobs: use embedded payload_json when present.
      4. Fallback: legacy JSON at draft_path.
    """
    _require_drafts_storage()
    row = get_import_job(job_id)
    if not row:
        return None

    # -----------------------------
    # 1) Existing DB-backed draft?
    # -----------------------------
    draft_id = _find_draft_for_job(job_id)
    if draft_id:
        # Sync restaurant_id if the job has one and draft is missing it
        try:
            keys = set(row.keys()) if hasattr(row, "keys") else set()
            if "restaurant_id" in keys and row["restaurant_id"]:
                existing = drafts_store.get_draft(draft_id) if hasattr(drafts_store, "get_draft") else {}
                if isinstance(existing, dict) and not existing.get("restaurant_id"):
                    drafts_store.save_draft_metadata(
                        draft_id,
                        restaurant_id=int(row["restaurant_id"]),
                    )
        except Exception:
            pass
        return draft_id

    # If we are not allowed to create drafts, stop here (GET-safe default).
    if not allow_create:
        return None

    # We’ll build draft_json from the best available source.
    draft_json: Optional[Dict[str, Any]] = None
    keys = set(row.keys()) if hasattr(row, "keys") else set()

    # -----------------------------
    # 2) Prefer file-based AI preview JSON (OCR jobs)
    #    Try ai_preview_path → preview_path → draft_path
    # -----------------------------
    abs_path: Optional[Path] = None
    used_source_label: str = ""
    for key in ("ai_preview_path", "preview_path", "draft_path"):
        raw_path = row[key] if key in keys else None
        if raw_path:
            candidate = _abs_from_rel(raw_path)
            if candidate and candidate.exists():
                abs_path = candidate
                used_source_label = key
                break

    if abs_path and abs_path.exists():
        try:
            with open(abs_path, "r", encoding="utf-8") as f:
                loaded = json.load(f)
            if isinstance(loaded, dict):
                draft_json = loaded
        except Exception:
            draft_json = None

    # -----------------------------
    # 3) Fallback: embedded payload_json (structured JSON jobs)
    # -----------------------------
    if draft_json is None and "payload_json" in keys:
        payload_raw = row["payload_json"]
        if payload_raw:
            try:
                candidate = json.loads(payload_raw)
                if isinstance(candidate, dict):
                    draft_json = candidate
                    used_source_label = "payload_json"
            except Exception:
                draft_json = None

    if draft_json is None:
        return None

    # -----------------------------
    # 4) Create new draft from chosen JSON
    # -----------------------------
    draft_id = None

    create_fn = getattr(drafts_store, "create_draft_from_import", None)
    if callable(create_fn):
        try:
            created = create_fn(draft_json, import_job_id=job_id)
        except Exception:
            created = None

        if isinstance(created, dict):
            new_id_raw = created.get("id") or created.get("draft_id")
            if new_id_raw:
                try:
                    draft_id = int(new_id_raw)
                except Exception:
                    draft_id = None

    # If create_draft_from_import didn’t produce a usable draft id, fall back to structured-items path.
    if draft_id is None:
        create_structured = getattr(drafts_store, "create_draft_from_structured_items", None)
        if callable(create_structured):
            flat = _draft_items_from_draft_json(draft_json)

            structured_items: List[Dict[str, Any]] = []
            for it in flat:
                name = (it.get("name") or "").strip()
                if not name:
                    continue
                desc = (it.get("description") or "").strip()
                category = (it.get("category") or "Uncategorized").strip() or "Uncategorized"
                confidence = it.get("confidence")
                position = it.get("position")

                price_val = it.get("price")
                price_cents = 0
                if price_val is not None:
                    try:
                        price_cents = int(round(float(price_val) * 100.0))
                    except Exception:
                        price_cents = 0

                structured_items.append(
                    {
                        "name": name,
                        "description": desc,
                        "category": category,
                        "price_cents": int(price_cents),
                        "confidence": confidence,
                        "position": position,
                    }
                )

            restaurant_id = None
            if "restaurant_id" in keys and row["restaurant_id"]:
                try:
                    restaurant_id = int(row["restaurant_id"])
                except Exception:
                    restaurant_id = None

            title = f"Imported OCR {datetime.utcnow().date()}"
            try:
                src = draft_json.get("source") if isinstance(draft_json, dict) else None
                if isinstance(src, dict) and src.get("file"):
                    title = f"{src.get('file')}"
            except Exception:
                title = f"Imported OCR {datetime.utcnow().date()}"

            if structured_items:
                try:
                    created2 = create_structured(
                        title=title,
                        restaurant_id=restaurant_id,
                        items=structured_items,
                        source_type="ocr_legacy_categories",
                        source_job_id=int(job_id),
                        source_meta={
                            "import_job_id": int(job_id),
                            "source": used_source_label or "unknown",
                            "draft_json_kind": "categories",
                        },
                    )
                except Exception:
                    created2 = None

                if isinstance(created2, dict):
                    new_id_raw2 = created2.get("id") or created2.get("draft_id")
                    if new_id_raw2:
                        try:
                            draft_id = int(new_id_raw2)
                        except Exception:
                            draft_id = None

    if draft_id is None:
        return None

    # -----------------------------
    # 5) Sync restaurant_id
    # -----------------------------
    if "restaurant_id" in keys and row["restaurant_id"]:
        try:
            drafts_store.save_draft_metadata(
                draft_id,
                restaurant_id=int(row["restaurant_id"]),
            )
        except Exception:
            pass

    # Best-effort: link back to import_jobs.draft_id if the column exists.
    try:
        update_import_job(job_id, draft_id=int(draft_id))
    except Exception:
        pass

    return int(draft_id)



# ---------- MISSING HELPERS (wired to /imports actions) ----------
def _dedupe_exists(conn: sqlite3.Connection, menu_id: int, name: str, price_cents: int) -> bool:
    cur = conn.cursor()
    row = cur.execute(
        """
        SELECT 1 FROM menu_items
        WHERE menu_id=? AND lower(trim(name))=lower(trim(?)) AND price_cents=?
        LIMIT 1
        """,
        (menu_id, name, price_cents),
    ).fetchone()
    return bool(row)


def approve_draft_to_menu(job_id: int) -> Tuple[int, int]:
    """
    Commit draft rows for job -> menu_items with simple dedupe.
    Requires a restaurant context:
      - customer side: session['user']['restaurant_id']
      - admin side: job_row.restaurant_id (must be chosen)
    Returns (menu_id, inserted_count).
    """
    _require_drafts_storage()
    job = get_import_job(job_id)
    if not job:
        abort(404, description="Job not found")

    # Determine restaurant context
    restaurant_id = _resolve_restaurant_id_for_action(job)
    if not restaurant_id:
        abort(400, description="No restaurant selected. Choose a restaurant for this import before approving.")

    # Explicit user action (POST) -> allow creating draft from preview/legacy JSON if needed
    draft_id = _get_or_create_draft_for_job(job_id, allow_create=True)
    if not draft_id:
        abort(400, description="No draft available to approve")

    items = drafts_store.get_draft_items(draft_id) or []

    with db_connect() as conn:
        menu_id = _find_or_create_menu_for_restaurant(conn, int(restaurant_id))
        cur = conn.cursor()

        inserted = 0
        for it in items:
            name = (it.get("name") or "").strip()
            if not name:
                continue

            price_cents = it.get("price_cents")
            if price_cents is None:
                price_cents = _price_to_cents(it.get("price") or it.get("price_text"))

            desc = (it.get("description") or "").strip()

            if not _dedupe_exists(conn, menu_id, name, int(price_cents)):
                cur.execute(
                    "INSERT INTO menu_items (menu_id, name, description, price_cents, is_available) VALUES (?, ?, ?, ?, 1)",
                    (menu_id, name, desc, int(price_cents)),
                )
                inserted += 1

        conn.commit()

    # Mark job approved and sync restaurant linkage
    try:
        update_import_job(job_id, status="approved", restaurant_id=int(restaurant_id))
    except Exception:
        pass
    try:
        drafts_store.save_draft_metadata(draft_id, restaurant_id=int(restaurant_id))
    except Exception:
        pass

    return menu_id, inserted


def discard_draft_for_job(job_id: int) -> int:
    """
    Delete all draft items (keep the draft shell so editor can be used later).
    Returns deleted count.
    """
    _require_drafts_storage()

    # Explicit user action (POST) -> allow creating draft from preview/legacy JSON if needed
    draft_id = _get_or_create_draft_for_job(job_id, allow_create=True)
    if not draft_id:
        return 0

    items = drafts_store.get_draft_items(draft_id) or []
    ids = []
    for it in items:
        if it.get("id") is not None:
            ids.append(it.get("id"))

    deleted = 0
    if ids:
        deleted = drafts_store.delete_draft_items(draft_id, ids)

    try:
        update_import_job(job_id, status="discarded")
    except Exception:
        pass

    return int(deleted)
# ---------- /MISSING HELPERS ----------


# ------------------------
# Health / DB / OCR Health
# ------------------------
@app.get("/ocr/health")
def ocr_health_route():
    # Prefer explicit cmd if set; otherwise look up via PATH
    explicit_cmd = getattr(pytesseract.pytesseract, "tesseract_cmd", "") or ""
    which_cmd = shutil.which("tesseract") or shutil.which("tesseract.exe") or ""
    display_cmd = explicit_cmd or which_cmd
    tess_path_exists = (bool(explicit_cmd) and Path(explicit_cmd).exists()) or bool(which_cmd)

    try:
        tess_version = str(pytesseract.get_tesseract_version())
    except Exception:
        tess_version = None

    # App (Flask) context libs — likely False on 3.13
    from importlib.util import find_spec as _find_spec
    have_pandas_app = _find_spec("pandas") is not None
    have_sklearn_app = _find_spec("sklearn") is not None

    # Probe the OCR worker's interpreter (.venv311) to get true Column Mode status
    worker_probe = {
        "active": False,
        "env_ok": False,
        "pandas": False,
        "scikit_learn": False,
        "python": None,
        "path": None,
        "error": None,
    }
    try:
        worker_py = str(ROOT / ".venv311" / "Scripts" / "python.exe")
        if Path(worker_py).exists():
            worker_probe["path"] = worker_py
            import subprocess
            code = (
                "import json,sys\n"
                "res={'python':sys.executable}\n"
                "try:\n"
                " import pandas; res['pandas']=pandas.__version__\n"
                "except Exception:\n"
                " res['pandas']=False\n"
                "try:\n"
                " import sklearn; res['scikit_learn']=sklearn.__version__\n"
                "except Exception:\n"
                " res['scikit_learn']=False\n"
                "print(json.dumps(res))\n"
            )
            out = subprocess.check_output([worker_py, "-c", code], text=True)
            data = json.loads(out.strip())
            worker_probe.update(data)
            worker_probe["env_ok"] = bool(data.get("pandas")) and bool(data.get("scikit_learn"))
            worker_probe["active"] = bool(worker_probe["env_ok"])
    except Exception as e:
        worker_probe["error"] = str(e)

    column_mode = "active" if worker_probe.get("env_ok") else "fallback"

    # Surface OCR worker version string so you can confirm live-reload worked
    worker_version = getattr(ocr_worker, "OCR_WORKER_VERSION", None)

    return jsonify({
        "tesseract": {
            "cmd": display_cmd,
            "found_on_disk": bool(tess_path_exists),
            "version": tess_version
        },
        "poppler": {
            "poppler_path_env": os.getenv("POPPLER_PATH") or "",
            "poppler_bin_present": bool((os.getenv("POPPLER_PATH") or "") and Path(os.getenv("POPPLER_PATH")).exists())
        },
        # Show both app and worker context (useful for debugging)
        "columns": {
            "mode": column_mode,
            "pandas": worker_probe.get("pandas", False),
            "scikit_learn": worker_probe.get("scikit_learn", False),
            "app_context": {"pandas": have_pandas_app, "scikit_learn": have_sklearn_app},
            "worker_python": worker_probe.get("python"),
            "worker_path": worker_probe.get("path"),
            "probe_error": worker_probe.get("error"),
        },
        "ocr_worker_version": worker_version,
        "ocr_lib_health": (ocr_health_lib() if ocr_health_lib else None),

    })

@app.get("/db/health")
def db_health():
    try:
        with db_connect() as conn:
            cur = conn.cursor()
            def count(table):
                cur.execute(f"SELECT COUNT(*) AS c FROM {table}")
                return cur.fetchone()["c"]
            data = {
                "db_path": str(DB_PATH),
                "restaurants": count("restaurants"),
                "menus": count("menus"),
                "menu_items": count("menu_items"),
                "import_jobs": count("import_jobs"),
            }
        return jsonify({"status": "ok", "data": data})
    except Exception as e:
        return jsonify({"status": "error", "error": str(e)}), 500

# ------------------------
# JSON API (existing)
# ------------------------
@app.get("/api/restaurants")
def get_restaurants():
    with db_connect() as conn:
        rows = conn.execute("SELECT * FROM restaurants WHERE active=1").fetchall()
        return jsonify([dict(r) for r in rows])

@app.get("/api/restaurants/<int:rest_id>/menus")
def get_menus(rest_id):
    with db_connect() as conn:
        rows = conn.execute(
            "SELECT * FROM menus WHERE restaurant_id=? AND active=1", (rest_id,),
        ).fetchall()
        if not rows:
            abort(404, description="No menus found for that restaurant")
        return jsonify([dict(r) for r in rows])

@app.get("/api/menus/<int:menu_id>/items")
def get_menu_items(menu_id):
    with db_connect() as conn:
        rows = conn.execute(
            "SELECT * FROM menu_items WHERE menu_id=? AND is_available=1", (menu_id,),
        ).fetchall()
        if not rows:
            abort(404, description="No items found for that menu")
        return jsonify([dict(r) for r in rows])

# =========================================================
#  NEW: Manual rotate preview (working copy) helpers/routes
# =========================================================

def _work_image_path(job_id: int) -> Path:
    """JPEG working copy (user-rotatable) tied to job id."""
    return RAW_FOLDER / f"job_{job_id}_work.jpg"

def _ensure_work_image(job_id: int, src_path: Path) -> Optional[Path]:
    """
    Ensure a JPEG preview exists for the job:
    - Images: convert/copy to RGB JPEG
    - PDFs: rasterize first page
    
    IMPORTANT: Apply orientation normalization so preview matches OCR input.
    """
    try:
        p = _work_image_path(job_id)
        if p.exists():
            return p

        # Import orientation normalizer from One Brain pipeline
        from storage.ocr_utils import normalize_orientation

        suffix = src_path.suffix.lower()
        if suffix in (".png", ".jpg", ".jpeg"):
            from PIL import Image
            with Image.open(src_path) as im:
                if im.mode != "RGB":
                    im = im.convert("RGB")
                # Apply orientation correction
                im, deg = normalize_orientation(im)
                if deg != 0:
                    print(f"[Preview] job={job_id} orientation corrected by {deg}° for preview")
                p.parent.mkdir(parents=True, exist_ok=True)
                im.save(p, "JPEG", quality=92, optimize=True)
            return p

        if suffix == ".pdf":
            try:
                from pdf2image import convert_from_path
            except Exception:
                return None
            pages = convert_from_path(str(src_path), dpi=200, poppler_path=POPPLER_PATH)
            if not pages:
                return None
            im = pages[0]
            if im.mode != "RGB":
                im = im.convert("RGB")
            # Apply orientation correction
            im, deg = normalize_orientation(im)
            if deg != 0:
                print(f"[Preview] job={job_id} orientation corrected by {deg}° for preview (PDF)")
            p.parent.mkdir(parents=True, exist_ok=True)
            im.save(p, "JPEG", quality=92, optimize=True)
            return p
    except Exception:
        return None
    return None

def _get_work_image_if_any(job_id: int) -> Optional[Path]:
    p = _work_image_path(job_id)
    return p if p.exists() else None

def _path_for_ocr(job_id: int, original: Path) -> Tuple[Path, str]:
    """
    Return (path, type_tag). If the user has rotated a working copy,
    use it and tag as 'image'. Otherwise return the original path.
    """
    p = _get_work_image_if_any(job_id)
    if p and p.exists():
        return p, "image"
    return original, ("image" if original.suffix.lower() in (".jpg", ".jpeg", ".png") else "pdf")

def _score_ocr_text_for_orientation(text: str) -> int:
    """
    Heuristic score: prefer outputs that look like real menu text.
    Higher is better.
    """
    t = (text or "").strip()
    if not t:
        return 0

    total = len(t)

    letters = 0
    digits = 0
    spaces = 0
    other = 0
    for ch in t:
        if ch.isalpha():
            letters += 1
        elif ch.isdigit():
            digits += 1
        elif ch.isspace():
            spaces += 1
        else:
            other += 1

    # Ratio of "good" characters
    good = letters + digits + spaces
    good_ratio = good / max(total, 1)

    # Junk penalty (punctuation/symbol soup tends to spike on wrong rotations)
    junk_ratio = other / max(total, 1)

    # Word count (rough)
    words = [w for w in t.split() if w.strip()]
    word_count = len(words)

    # Simple "menu-ish" signal: presence of prices ($ or digit+dot patterns)
    has_dollar = "$" in t
    has_decimal = False
    for i in range(0, len(t) - 2):
        if t[i].isdigit() and t[i + 1] == "." and t[i + 2].isdigit():
            has_decimal = True
            break

    score = 0

    # Length helps, but cap it
    score += min(total, 2500)

    # Favor readable text, punish junk
    score += int(good_ratio * 2000)
    score -= int(junk_ratio * 1500)

    # Word count helps (cap)
    score += min(word_count, 300) * 10

    # Tiny boosts if it looks like menu pricing
    if has_dollar:
        score += 250
    if has_decimal:
        score += 250

    return int(max(score, 0))


def _auto_rotate_work_image_if_needed(job_id: int, original_image: Path) -> Optional[Path]:
    """
    Ensure a working preview exists, then attempt auto-rotation (0/90/180/270)
    by OCR-scoring each orientation. If a better orientation is found, rotate
    the working JPEG in-place.

    Returns the work image Path (rotated if needed) or None on failure.
    """
    try:
        wp = _ensure_work_image(job_id, original_image) or _get_work_image_if_any(job_id)
        if not wp or not wp.exists():
            return None

        from PIL import Image

        # Load the working JPEG once
        with Image.open(wp) as im0:
            if im0.mode != "RGB":
                im0 = im0.convert("RGB")

            candidates = [
                ("0", 0, im0),
                ("90", 90, im0.rotate(90, expand=True)),
                ("180", 180, im0.rotate(180, expand=True)),
                ("270", 270, im0.rotate(270, expand=True)),
            ]

            best_tag = "0"
            best_angle = 0
            best_score = -1

            for tag, angle, im in candidates:
                try:
                    # OCR directly from PIL image
                    txt = pytesseract.image_to_string(im)
                except Exception:
                    txt = ""
                s = _score_ocr_text_for_orientation(txt)
                if s > best_score:
                    best_score = s
                    best_tag = tag
                    best_angle = angle

            # If best is not 0, rotate the work image in-place
            if best_angle != 0:
                best_im = im0.rotate(best_angle, expand=True)
                if best_im.mode != "RGB":
                    best_im = best_im.convert("RGB")
                best_im.save(wp, "JPEG", quality=92, optimize=True)

        return wp
    except Exception:
        return None


@app.get("/imports/<int:job_id>/preview.jpg")
@login_required
def imports_preview_image(job_id: int):
    """
    Serve/create the rotatable working preview JPEG for this job.
    """
    row = get_import_job(job_id)
    if not row:
        abort(404)
    src = (UPLOAD_FOLDER / (row["filename"] or "")).resolve()
    if not src.exists():
        abort(404)
    p = _ensure_work_image(job_id, src) or _get_work_image_if_any(job_id)
    if not p or not p.exists():
        # As last resort, stream the original if it's already an image
        if src.suffix.lower() in (".jpg", ".jpeg", ".png"):
            return send_file(str(src), mimetype="image/jpeg")
        abort(404)
    return send_file(str(p), mimetype="image/jpeg")


def _get_import_pages(job_id: int) -> list:
    """Day 141: return all image file paths for a multi-page upload.

    Returns a list of (page_index, Path) tuples, page_index is 0-based.
    Page 0 is the primary file; pages 1+ come from extra_filenames.
    For single-page PDFs, also rasterizes subsequent PDF pages on demand.
    """
    row = get_import_job(job_id)
    if not row:
        return []

    pages: list = []
    primary_name = row["filename"] or ""
    primary = (UPLOAD_FOLDER / primary_name).resolve()
    if primary.exists():
        pages.append(primary)

    # Add any extra pages from multi-file upload
    try:
        extra_json = row["extra_filenames"] if "extra_filenames" in row.keys() else None
    except Exception:
        extra_json = None
    if extra_json:
        import json as _json
        try:
            for name in _json.loads(extra_json):
                path = (UPLOAD_FOLDER / name).resolve()
                if path.exists():
                    pages.append(path)
        except Exception:
            pass

    return pages


@app.get("/imports/<int:job_id>/pages.json")
@login_required
def imports_pages_info(job_id: int):
    """Day 141: return JSON metadata about how many pages this import has."""
    pages = _get_import_pages(job_id)
    return jsonify({"ok": True, "job_id": job_id, "page_count": len(pages)})


@app.get("/imports/<int:job_id>/preview/<int:page>.jpg")
@login_required
def imports_preview_image_page(job_id: int, page: int):
    """Day 141: serve a specific page (0-indexed) for multi-page menu uploads.

    Page 0 reuses the existing _ensure_work_image path (the rotatable primary).
    Pages 1+ are served as-is from the extra_filenames list, rasterized if PDF.
    """
    row = get_import_job(job_id)
    if not row:
        abort(404)

    pages = _get_import_pages(job_id)
    if not pages or page < 0 or page >= len(pages):
        abort(404)

    # Page 0 = the primary working image (rotatable, already handled)
    if page == 0:
        src = pages[0]
        p = _ensure_work_image(job_id, src) or _get_work_image_if_any(job_id)
        if not p or not p.exists():
            if src.suffix.lower() in (".jpg", ".jpeg", ".png"):
                return send_file(str(src), mimetype="image/jpeg")
            abort(404)
        return send_file(str(p), mimetype="image/jpeg")

    # Extra pages: serve image files directly, rasterize PDFs on the fly
    src = pages[page]
    if src.suffix.lower() in (".jpg", ".jpeg", ".png"):
        return send_file(str(src), mimetype="image/jpeg")
    if src.suffix.lower() == ".pdf":
        try:
            from pdf2image import convert_from_path
            from io import BytesIO
            imgs = convert_from_path(str(src), dpi=200, poppler_path=POPPLER_PATH)
            if not imgs:
                abort(404)
            buf = BytesIO()
            imgs[0].save(buf, format="JPEG", quality=85)
            buf.seek(0)
            return send_file(buf, mimetype="image/jpeg")
        except Exception:
            abort(500)
    abort(404)

@app.route("/imports/<int:job_id>/rotate", methods=["POST", "GET"])
@login_required
def imports_rotate_image(job_id: int):
    """
    Rotate working preview: dir=left|right OR angle=±90/180.
    - GET/POST both supported.
    - If form/redirect requested, flash+redirect back to import view.
    - Else returns JSON.
    """
    row = get_import_job(job_id)
    if not row:
        abort(404)
    src = (UPLOAD_FOLDER / (row["filename"] or "")).resolve()
    if not src.exists():
        abort(404)

    # Ensure working copy exists first
    wp = _ensure_work_image(job_id, src) or _get_work_image_if_any(job_id)
    if not wp or not wp.exists():
        return jsonify({"ok": False, "error": "preview not available"}), 400

    # Parse direction/angle
    direction = (request.values.get("dir") or "").strip().lower()
    angle_param = request.values.get("angle")
    angle = 0
    if angle_param:
        try:
            angle = int(angle_param)
        except Exception:
            angle = 0
    elif direction in ("left", "counterclockwise", "ccw"):
        angle = 90
    elif direction in ("right", "clockwise", "cw"):
        angle = -90
    else:
        angle = 90  # default: left

    from PIL import Image
    try:
        with Image.open(wp) as im:
            im = im.rotate(angle, expand=True)  # PIL is CCW-positive
            if im.mode != "RGB":
                im = im.convert("RGB")
            im.save(wp, "JPEG", quality=92, optimize=True)
    except Exception as e:
        wants_redirect = request.args.get("redirect") == "1" or request.method == "POST"
        if wants_redirect:
            flash(f"Rotate failed: {e}", "error")
            return redirect(url_for("imports_detail", job_id=job_id))
        return jsonify({"ok": False, "error": str(e)}), 500

    wants_redirect = request.args.get("redirect") == "1" or request.method == "POST"
    if wants_redirect:
        flash("Image rotated.", "success")
        return redirect(url_for("imports_detail", job_id=job_id))
    return jsonify({"ok": True, "angle": angle})


# ------------------------
# Column mapping / preview (Phase 6 pt.9)
# ------------------------
@app.get("/imports/<int:job_id>/mapping")
@login_required
def imports_mapping(job_id: int):
    """
    Read-only column mapping preview for structured imports (CSV/XLSX).

    We rebuild the preview directly from the original source file using the
    One Brain parsers in storage.import_jobs:

      - parse_structured_csv(path)
      - parse_structured_xlsx(path)

    The template expects:
      mapping = {
        "header_map":   { original_header -> canonical_field },
        "sample_rows":  [ {original_header: value, ...}, ... ],
        "column_names": [ "Header 1", "Header 2", ... ],
        "is_structured": bool,
        "file_ext": "csv" | "xlsx",
        "source_type": "structured_csv" | "structured_xlsx" | ...
      }
    """
    row = get_import_job(job_id)
    if not row:
        abort(404, description="Import job not found")

    # sqlite3.Row supports .keys() and item access
    col_names = set(row.keys()) if hasattr(row, "keys") else set()

    filename = (row["filename"] if "filename" in col_names else "") or ""
    ext = filename.rsplit(".", 1)[-1].lower() if "." in filename else ""

    src_type_raw = row["source_type"] if "source_type" in col_names else ""
    src_type = (src_type_raw or "").lower()

    is_structured_ext = ext in ("csv", "xlsx")
    is_structured_type = src_type.startswith("structured_")
    is_structured = is_structured_ext or is_structured_type

    # For now we only support CSV/XLSX in the mapping preview.
    if not is_structured or ext not in ("csv", "xlsx"):
        flash("Column mapping is currently only available for structured CSV/XLSX imports.", "error")
        return redirect(url_for("imports_detail", job_id=job_id))

    source_path_str = row["source_path"] if "source_path" in col_names else ""
    if not source_path_str:
        flash("This structured import does not have a source_path recorded.", "error")
        return redirect(url_for("imports_detail", job_id=job_id))

    src_path = Path(source_path_str)
    if not src_path.exists():
        flash("The original structured import file could not be found on disk.", "error")
        return redirect(url_for("imports_detail", job_id=job_id))

    header_map_canon_to_header: Dict[str, str] = {}
    raw_rows: List[Dict[str, Any]] = []

    try:
        # Use the One Brain helpers to parse the file and recover raw rows
        if ext == "csv" or src_type == "structured_csv":
            _, _, _, header_map_canon_to_header, raw_rows = import_jobs_store.parse_structured_csv(src_path)
        elif ext == "xlsx" or src_type == "structured_xlsx":
            _, _, _, header_map_canon_to_header, raw_rows = import_jobs_store.parse_structured_xlsx(src_path)
        else:
            header_map_canon_to_header = {}
            raw_rows = []
    except Exception as exc:
        app.logger.exception("Failed to build column mapping preview for import job %s", job_id)
        flash(f"Could not build column mapping preview: {exc}", "error")
        return redirect(url_for("imports_detail", job_id=job_id))

    # raw_rows are the original tabular rows keyed by the file's headers.
    column_names: List[str] = []
    sample_rows: List[Dict[str, Any]] = []

    if raw_rows and isinstance(raw_rows, list) and isinstance(raw_rows[0], dict):
        first_row = raw_rows[0]
        column_names = list(first_row.keys())
        sample_rows = raw_rows[:5]

    # Invert canonical -> header mapping into header -> canonical for the UI
    header_map_original_to_canonical: Dict[str, str] = {}
    for canonical, original in (header_map_canon_to_header or {}).items():
        if original:
            header_map_original_to_canonical[original] = canonical

    mapping_ctx = {
        "header_map": header_map_original_to_canonical,
        "sample_rows": sample_rows,
        "column_names": column_names,
        "is_structured": True,
        "file_ext": ext,
        "source_type": src_type,
    }

    return _safe_render(
        "import_mapping.html",
        job=row,
        mapping=mapping_ctx,
    )



# ------------------------
# Import flow: Upload -> Job -> Worker -> Draft JSON (OCR)
# ------------------------
def _save_draft_json(job_id: int, draft: dict) -> str:
    draft_name = f"draft_{job_id}_{uuid.uuid4().hex[:8]}.json"
    abs_path = DRAFTS_FOLDER / draft_name
    with open(abs_path, "w", encoding="utf-8") as f:
        json.dump(draft, f, indent=2)
    return str(abs_path.relative_to(ROOT)).replace("\\", "/")

# --- OCR helpers: image/PDF → text, then text → draft -----------------
def _ocr_via_google_vision(img_path: Path) -> str:
    """OCR using Google Cloud Vision API (DOCUMENT_TEXT_DETECTION).

    Returns full text extracted from the image. Falls back to empty string
    on any error. Requires GOOGLE_CLOUD_API_KEY in environment.
    """
    try:
        import requests as _requests
        import base64 as _b64
        key = os.environ.get("GOOGLE_CLOUD_API_KEY", "").strip()
        if not key:
            return ""
        with open(str(img_path), "rb") as f:
            img_data = _b64.b64encode(f.read()).decode()
        resp = _requests.post(
            f"https://vision.googleapis.com/v1/images:annotate?key={key}",
            json={
                "requests": [{
                    "image": {"content": img_data},
                    "features": [{"type": "DOCUMENT_TEXT_DETECTION"}]
                }]
            },
            timeout=30,
        )
        data = resp.json()
        if "responses" in data and data["responses"]:
            r = data["responses"][0]
            if "error" not in r:
                text = r.get("fullTextAnnotation", {}).get("text", "")
                if text:
                    print(f"[OCR] Google Vision: {len(text)} chars from {img_path.name}")
                    # Day 140: save OCR text for debugging price verification
                    try:
                        _ocr_log = Path(__file__).resolve().parent.parent / "storage" / "logs" / "vision_ocr_latest.txt"
                        _ocr_log.write_text(text, encoding="utf-8")
                    except Exception:
                        pass
                    return text
    except Exception as e:
        print(f"[OCR] Google Vision failed: {e}")
    return ""


def _ocr_image_to_text(img_path: Path) -> str:
    # Day 140: Try Google Vision first, fall back to Tesseract
    vision_text = _ocr_via_google_vision(img_path)
    if vision_text:
        return vision_text
    print(f"[OCR] Falling back to Tesseract for {img_path.name}")
    try:
        from PIL import Image, ImageOps, ImageFilter
        img = Image.open(str(img_path)).convert("L")
        img = ImageOps.autocontrast(img)
        img = img.filter(ImageFilter.SHARPEN)
        return pytesseract.image_to_string(
            img,
            lang=TESSERACT_LANG,
            config=TESSERACT_CONFIG
        ) or ""
    except Exception:
        return ""

def _pdf_to_text(pdf_path: Path) -> str:
    """
    Try to rasterize each PDF page with pdf2image (+poppler) and OCR it.
    Falls back gracefully if pdf2image/poppler are not available.
    """
    try:
        from pdf2image import convert_from_path
        from PIL import ImageOps, ImageFilter
    except Exception:
        return ""

    poppler_path = POPPLER_PATH
    try:
        pages = convert_from_path(str(pdf_path), dpi=300, poppler_path=poppler_path)
    except Exception:
        return ""

    buf = []
    for i, pg in enumerate(pages):
        try:
            # Save page to temp file for Vision API
            import tempfile
            with tempfile.NamedTemporaryFile(suffix=".png", delete=False) as tmp:
                pg.save(tmp.name, "PNG")
                txt = _ocr_via_google_vision(Path(tmp.name))
                if not txt:
                    img = pg.convert("L")
                    img = ImageOps.autocontrast(img)
                    img = img.filter(ImageFilter.SHARPEN)
                    txt = pytesseract.image_to_string(
                        img, lang=TESSERACT_LANG, config=TESSERACT_CONFIG
                    )
                try:
                    os.unlink(tmp.name)
                except OSError:
                    pass
            if txt:
                buf.append(txt)
        except Exception:
            continue
    return "\n".join(buf).strip()

_price_rx = re.compile(r"""
    (?P<name>.+?)                         # item name
    [\s\-\–\—·:]*                         # optional separators (dashes, middots, colon)
    (?P<price>                            # price at end
        [\$€£]?\s*
        \d{1,3}(?:[.,]\d{3})*             # thousands with . or ,
        (?:[.,]\d{1,2})?                  # optional decimals
        |\$?\d+(?:[.,]\d{1,2})?           # or simple number with decimals
    )\s*$
""", re.X)

def _text_to_draft(text: str, job_id: int, src_file: str, engine_label: str) -> dict:
    """
    Heuristic parser:
      - Default category until a new one detected
      - New category when 'Category: Foo' or an ALL-CAPS line
      - Item line if it ends in a price; otherwise, append to previous item's description
    """
    categories = []
    current_cat = {"name": "Uncategorized", "items": []}
    categories.append(current_cat)

    def new_cat(name: str):
        nonlocal current_cat
        name = (name or "Misc").strip()
        current_cat = {"name": name, "items": []}
        categories.append(current_cat)

    prev_item = None
    for raw in (text or "").splitlines():
        line = raw.strip()
        if not line:
            continue

        # Category cues
        if line.lower().startswith("category:"):
            new_cat(line.split(":", 1)[1].strip() or "Misc")
            prev_item = None
            continue
        if len(line) <= 40 and line.isupper() and any(c.isalpha() for c in line):
            new_cat(line.title())
            prev_item = None
            continue

        # Item with trailing price
        m = _price_rx.match(line)
        if m:
            name = m.group("name").strip(" -·:—–")
            price = (m.group("price") or "").strip()
            price_norm = price.replace("$", "").replace("€", "").replace("£", "").replace(" ", "")
            if ("," in price_norm) and ("." not in price_norm):
                price_norm = price_norm.replace(",", ".")
            try:
                p = float(re.sub(r"[^\d.]", "", price_norm))
            except Exception:
                p = 0.0
            current_cat["items"].append({
                "name": name,
                "description": "",
                "sizes": [{"name": "One Size", "price": round(p, 2)}] if p else [],
            })
            prev_item = current_cat["items"][-1]
            continue

        # Otherwise, treat as description line for the last item
        if prev_item:
            desc = (prev_item.get("description") or "").strip()
            prev_item["description"] = (desc + " " + line).strip()
        else:
            current_cat["items"].append({"name": line, "description": "", "sizes": []})
            prev_item = current_cat["items"][-1]

    # Drop empty leading category if it has no items
    categories = [c for c in categories if c.get("items")]

    return {
        "job_id": job_id,
        "source": {"type": "upload", "file": src_file, "ocr_engine": engine_label},
        "extracted_at": _now_iso(),
        "categories": categories or [{"name": "Uncategorized", "items": []}],
    }

# ----- Day 14: helper-backed draft builder (facade-aware) -----
def _build_draft_from_helper(job_id: int, saved_file_path: Path):
    """
    Use storage/ocr_facade.extract_menu_from_pdf to build a draft JSON.

    New façade shape (preferred):

      extract_items_from_path(path) ->

        • (categories_dict, debug_payload)  OR
        • categories_dict

      where categories_dict looks like:
        {
          "categories": [
            {
              "name": "Pizza",
              "items": [
                {
                  "name": "...",
                  "description": "...",
                  "sizes": [
                    {"label": "L", "price": 12.99},
                    {"label": "XL", "price_cents": 1899},
                    ...
                  ],
                  "confidence": 92,
                },
                ...
              ]
            },
            ...
          ],
          "extracted_at": "...Z",
          "source": { "type": "upload", "file": "...", "ocr_engine": "ocr_helper+tesseract" }
        }

    For backward compatibility, we still support the old shape:
      • dict[str, list[items]]  (category_name -> items)
    """
    if extract_items_from_path is None:
        raise RuntimeError("ocr_facade not available")

    debug_payload = None
    cats_raw = extract_items_from_path(str(saved_file_path)) or {}

    # Unpack (categories_dict, debug_payload) vs just categories_dict
    if isinstance(cats_raw, tuple) and len(cats_raw) == 2:
        cats_dict, debug_payload = cats_raw
    else:
        cats_dict, debug_payload = cats_raw, None

    categories: List[Dict[str, Any]] = []
    source_meta: Dict[str, Any] = {}
    extracted_at: Optional[str] = None

    # -------- New façade shape: {"categories": [...], "source": {...}, "extracted_at": "..."} --------
    if isinstance(cats_dict, dict) and "categories" in cats_dict:
        extracted_at = cats_dict.get("extracted_at")
        source_meta = (cats_dict.get("source") or {}) if isinstance(cats_dict.get("source"), dict) else {}

        for cat_obj in (cats_dict.get("categories") or []):
            if not isinstance(cat_obj, dict):
                continue
            cat_name = (cat_obj.get("name") or "Uncategorized").strip() or "Uncategorized"

            out_items: List[Dict[str, Any]] = []
            for it in (cat_obj.get("items") or []):
                if not isinstance(it, dict):
                    continue

                name = (it.get("name") or "").strip() or "Untitled"
                desc = (it.get("description") or "").strip()

                # Normalize sizes: support {"label"/"name", "price"} or {"price_cents"}
                sizes_out: List[Dict[str, Any]] = []
                for s in (it.get("sizes") or []):
                    if not isinstance(s, dict):
                        continue
                    label = (s.get("label") or s.get("name") or "").strip()

                    raw_price = s.get("price", None)
                    if raw_price is None and s.get("price_cents") is not None:
                        try:
                            raw_price = float(s.get("price_cents"))
                            raw_price = raw_price / 100.0
                        except Exception:
                            raw_price = 0.0
                    try:
                        pr = float(raw_price or 0.0)
                    except Exception:
                        pr = 0.0

                    if pr > 0:
                        sizes_out.append({"name": label, "price": round(pr, 2)})

                out_items.append(
                    {
                        "name": name,
                        "description": desc,
                        "sizes": sizes_out,
                        "category": cat_name,
                        "confidence": it.get("confidence"),
                    }
                )

            if out_items:
                categories.append({"name": cat_name, "items": out_items})

    # -------- Backward compat: old dict[category_name] -> [items...] shape --------
    elif isinstance(cats_dict, dict):
        for cat_name, items in (cats_dict or {}).items():
            out_items: List[Dict[str, Any]] = []
            for it in (items or []):
                if not isinstance(it, dict):
                    continue
                name = (it.get("name") or "").strip()
                desc = (it.get("description") or "").strip()
                price = it.get("price")

                if isinstance(price, str):
                    try:
                        price_val = float(price.replace("$", "").strip())
                    except Exception:
                        price_val = 0.0
                elif isinstance(price, (int, float)):
                    # historical behavior: ints were sometimes cents
                    price_val = float(price) / 100.0 if isinstance(price, int) and price >= 100 else float(price)
                else:
                    price_val = 0.0

                sizes = [{"name": "One Size", "price": round(float(price_val), 2)}] if price_val else []

                out_items.append(
                    {
                        "name": name or "Untitled",
                        "description": desc,
                        "sizes": sizes,
                        "category": cat_name,
                        "confidence": it.get("confidence"),
                        "raw": it.get("raw"),
                    }
                )
            if out_items:
                categories.append({"name": cat_name, "items": out_items})

    # -------- Fallback if everything came back empty --------
    if not categories:
        categories = [
            {
                "name": "Uncategorized",
                "items": [
                    {
                        "name": "No items recognized",
                        "description": "OCR returned no items.",
                        "sizes": [],
                    }
                ],
            }
        ]

    # Source + engine metadata (prefer façade's own source block if present)
    engine = (source_meta.get("ocr_engine") if isinstance(source_meta, dict) else None) or "ocr_helper+tesseract"
    source = {
        "type": (source_meta.get("type") if isinstance(source_meta, dict) else None) or "upload",
        "file": (source_meta.get("file") if isinstance(source_meta, dict) else None) or saved_file_path.name,
        "ocr_engine": engine,
    }

    draft_dict = {
        "job_id": job_id,
        "source": source,
        "extracted_at": extracted_at or _now_iso(),
        "categories": categories,
    }
    return draft_dict, debug_payload




def _draft_items_from_draft_json(draft: Dict[str, Any]) -> List[Dict[str, Any]]:
    """
    Convert saved draft JSON format:
      {"categories":[{"name": "...", "items":[{...}]}]}
    into a flat list of DB draft_items rows expected by drafts_store.upsert_draft_items.

    Each item may include a '_variants' key with structured variant data
    that upsert_draft_items() will insert into draft_item_variants.
    """
    out: List[Dict[str, Any]] = []

    categories = draft.get("categories") or []
    if not isinstance(categories, list):
        return out

    pos = 1
    for cat in categories:
        if not isinstance(cat, dict):
            continue
        cat_name = (cat.get("name") or "Uncategorized").strip() or "Uncategorized"

        items = cat.get("items") or []
        if not isinstance(items, list):
            continue

        for it in items:
            if not isinstance(it, dict):
                continue

            name = (it.get("name") or "").strip()
            desc = (it.get("description") or "").strip()

            # Build structured variants from sizes array
            sizes = it.get("sizes")
            variants: list = []
            if isinstance(sizes, list) and sizes:
                for vi, s in enumerate(sizes):
                    if not isinstance(s, dict):
                        continue
                    lbl = (s.get("name") or s.get("label") or "").strip()
                    try:
                        pr = float(s.get("price", 0))
                    except Exception:
                        pr = 0.0
                    pr_cents = int(round(pr * 100))
                    if lbl or pr_cents > 0:
                        variants.append({
                            "label": lbl or f"Size {vi + 1}",
                            "price_cents": pr_cents,
                            "kind": "size",
                            "position": vi,
                        })

            # Base price: first variant price or item-level price
            price_val = None
            if variants:
                price_val = variants[0]["price_cents"] / 100.0
            if price_val is None:
                raw_price = it.get("price")
                try:
                    price_val = float(raw_price) if raw_price is not None and str(raw_price).strip() != "" else None
                except Exception:
                    price_val = None

            if not name and not desc and price_val is None:
                continue

            row: Dict[str, Any] = {
                "name": name or "Untitled",
                "description": desc,
                "price": price_val,
                "category": cat_name,
                "position": pos,
            }
            if variants:
                row["_variants"] = variants
            out.append(row)
            pos += 1

    return out


def _build_draft_from_worker(job_id: int, saved_file_path: Path, worker_obj: Dict[str, Any]) -> Dict[str, Any]:
    """
    Convert the ocr_worker draft object into the on-disk draft JSON format expected by the rest of the app.

    Critical: preserve categories/items so /imports/raw/<job_id> is not empty and Draft Editor has items.
    """
    from datetime import datetime, timezone

    categories = worker_obj.get("categories")
    if not isinstance(categories, list):
        categories = []

    source = worker_obj.get("source")
    if not isinstance(source, dict):
        source = {}

    # Ensure source has the basics
    source.setdefault("type", "upload")
    source.setdefault("file", saved_file_path.name)
    source.setdefault("ocr_engine", "ocr_worker")

    extracted_at = datetime.now(timezone.utc).replace(microsecond=0).isoformat().replace("+00:00", "Z")

    draft: Dict[str, Any] = {
        "job_id": int(job_id),
        "restaurant_id": worker_obj.get("restaurant_id"),
        "currency": worker_obj.get("currency") or "USD",
        "categories": categories,
        "source": source,
        "extracted_at": extracted_at,
    }
    return draft


def run_ocr_and_make_draft(job_id: int, saved_file_path: Path, *, extra_pages: list = None):
    draft: Any = None
    debug_payload: Any = None
    engine = ""
    helper_error: Optional[str] = None
    all_pages = [saved_file_path] + (extra_pages or [])

    try:
        update_import_job(job_id, status="processing", pipeline_stage="extracting")

        # Always create a user-rotatable working preview up-front
        try:
            _ensure_work_image(job_id, saved_file_path)
        except Exception:
            pass

        # Prefer working image (respects later user rotation too if we rerun)
        src_for_ocr, type_tag = _path_for_ocr(job_id, saved_file_path)

        # 0) RETIRED: portal/ocr_worker is legacy and must not be the preferred path.
        # Images should flow through the One Brain pipeline via storage.ocr_facade/build_structured_menu.
        # If facade fails, we will fall back later.
        if type_tag == "image" and False:
            try:
                raw_text, worker_obj = ocr_worker.run_image_pipeline(
                    Path(src_for_ocr),
                    job_id=str(job_id),
                )
                draft = _build_draft_from_worker(
                    job_id,
                    saved_file_path,
                    worker_obj or {},
                )
                engine = "ocr_worker"

                if isinstance(worker_obj, dict):
                    dbg_obj = worker_obj.get("debug")
                    if isinstance(dbg_obj, dict):
                        debug_payload = dbg_obj

            except Exception as e:
                helper_error = f"ocr_worker_failed: {e}"
                draft = None

        # 1) Helper path (typically for PDFs), fallback if image path failed
        if draft is None and extract_items_from_path is not None:
            try:
                draft, debug_payload = _build_draft_from_helper(job_id, saved_file_path)
                if isinstance(draft, dict):
                    engine = (draft.get("source") or {}).get("ocr_engine") or "ocr_helper+tesseract"
                else:
                    engine = "ocr_helper+tesseract"

                # NOTE: segment_document was previously called here for layout_debug
                # but it re-runs the full OCR pipeline (~3min), doubling processing time.
                # The facade already provides text_blocks/preview_blocks which is sufficient.


            except Exception as e:
                helper_error = str(e)
                draft = None
                debug_payload = None

        # 2) Legacy fallback to direct Tesseract text OCR + heuristics
        if draft is None:
            text = ""
            if type_tag == "image":
                engine = "tesseract"
                text = _ocr_image_to_text(src_for_ocr)
            else:
                engine = "tesseract+pdf2image"
                text = _pdf_to_text(saved_file_path)

            if text:
                draft = _text_to_draft(
                    text,
                    job_id,
                    saved_file_path.name,
                    engine or "tesseract",
                )
                debug_payload = {
                    "notes": [f"fallback_engine={engine}", f"len_text={len(text)}"],
                    "items": [],
                    "lines": text.splitlines()[:500],
                }
            else:
                draft = {
                    "job_id": job_id,
                    "source": {
                        "type": "upload",
                        "file": saved_file_path.name,
                        "ocr_engine": engine or "unavailable",
                    },
                    "extracted_at": _now_iso(),
                    "categories": [
                        {
                            "name": "Uncategorized",
                            "items": [
                                {
                                    "name": "OCR not configured",
                                    "description": "Install Tesseract; for PDFs also install pdf2image + Poppler.",
                                    "sizes": [],
                                }
                            ],
                        }
                    ],
                }
                debug_payload = {"notes": ["no_text_extracted"]}

        # ✅ Normalize draft["source"] if it was produced as a JSON string (prevents weird UI/source rendering)
        if isinstance(draft, dict):
            src = draft.get("source")
            if isinstance(src, str):
                try:
                    draft["source"] = json.loads(src)
                except Exception:
                    draft["source"] = {"raw": src}

        rel_draft_path = _save_draft_json(job_id, draft)

        # Save draft_path to DB NOW so _get_or_create_draft_for_job can find it.
        # status="done" is set later, after items are in the DB.
        update_import_job(job_id, draft_path=rel_draft_path)

        try:
            raw_dump = f"(engine={engine})\n"
            if helper_error:
                raw_dump += f"[helper_error] {helper_error}\n"
            (RAW_FOLDER / f"{job_id}.txt").write_text(raw_dump, encoding="utf-8")
        except Exception:
            pass

        # =====================================================================
        # ITEM EXTRACTION — Claude API only (Day 100.5: heuristic/legacy removed)
        # Get clean OCR text, then extract via Claude API.
        # No API key = empty draft for manual input (free tier).
        # =====================================================================
        items = []
        extraction_strategy = "none"

        # Pipeline metrics tracker (Day 99) — records per-step timing & counts
        try:
            from storage.pipeline_metrics import (
                PipelineTracker, STEP_OCR_TEXT, STEP_CALL1_EXTRACT,
                STEP_CALL2_VISION, STEP_SEMANTIC, STEP_CALL3_RECONCILE,
            )
            tracker = PipelineTracker()
        except Exception:
            tracker = None

        # Get clean OCR text via simple Tesseract (same path as /ai/preview)
        # Day 139: handle multiple pages — OCR each and concatenate with page markers
        clean_ocr_text = ""
        try:
            if tracker:
                tracker.start_step(STEP_OCR_TEXT)
            if len(all_pages) > 1:
                # Multi-page: OCR each file, concatenate with page markers
                page_texts = []
                for pi, pg_path in enumerate(all_pages, start=1):
                    _pg_suffix = pg_path.suffix.lower()
                    pg_text = ""
                    if _pg_suffix == ".pdf":
                        pg_text = _pdf_to_text(pg_path)
                    elif _pg_suffix in (".png", ".jpg", ".jpeg"):
                        pg_text = _ocr_image_to_text(pg_path)
                    if pg_text.strip():
                        page_texts.append(f"--- PAGE {pi} ---\n{pg_text}")
                clean_ocr_text = "\n\n".join(page_texts)
                print(f"[Draft] Multi-page OCR: {len(all_pages)} pages, {len(clean_ocr_text)} chars total")
            else:
                _suffix = saved_file_path.suffix.lower()
                if _suffix == ".pdf":
                    clean_ocr_text = _pdf_to_text(saved_file_path)
                elif _suffix in (".png", ".jpg", ".jpeg"):
                    clean_ocr_text = _ocr_image_to_text(src_for_ocr)
                print(f"[Draft] Clean OCR text: {len(clean_ocr_text)} chars")
            if tracker:
                tracker.end_step(STEP_OCR_TEXT, chars=len(clean_ocr_text))
        except Exception as _ocr_err:
            print(f"[Draft] Clean OCR failed: {_ocr_err}")
            if tracker:
                tracker.fail_step(STEP_OCR_TEXT, str(_ocr_err))

        # Strategy 1: Day 139 Detect+Classify+Locate pipeline
        # New Call 1 returns classified elements with bounding boxes.
        # Code layer assembles structure (no AI for grouping).
        # Falls back to legacy extraction if detection fails.
        vision_result = None
        _thinking_active = False
        _detected_elements = None  # raw elements for storage
        _coord_data = None  # bounding box data for post-insert linking
        if clean_ocr_text and not items:
            try:
                if tracker:
                    tracker.start_step(STEP_CALL1_EXTRACT)
                from storage.ai_menu_extract import (
                    detect_menu_elements, elements_to_draft_rows,
                    extract_menu_items_via_claude, claude_items_to_draft_rows,
                    EXTENDED_THINKING, PIPELINE_MODE,
                    SKIP_CALL2, SKIP_CALL3,
                )
                _thinking_active = EXTENDED_THINKING
                print(f"[Draft] Pipeline mode: detect+classify+locate (Day 139)")

                # New pipeline: detect elements with bounding boxes
                # Day 139: pass extra pages for multi-file uploads
                _extra_paths = [str(p) for p in all_pages[1:]] if len(all_pages) > 1 else None
                _detected_elements = detect_menu_elements(
                    clean_ocr_text, image_path=str(saved_file_path),
                    extra_image_paths=_extra_paths,
                    use_thinking=_thinking_active,
                )

                if _detected_elements:
                    # Code assembly: elements -> structured items + coordinates
                    items, _coord_data = elements_to_draft_rows(_detected_elements, ocr_text=clean_ocr_text)
                    n_elements = len(_detected_elements)
                    if tracker:
                        tracker.end_step(STEP_CALL1_EXTRACT, items=len(items))
                    extraction_strategy = "detect+assemble"
                    print(f"[Draft] Detection: {n_elements} elements -> {len(items)} items assembled"
                          f" ({len(_coord_data or [])} with coordinates)")
                else:
                    # Fallback: legacy extraction
                    print("[Draft] Detection returned no elements, falling back to legacy extraction")
                    claude_items = extract_menu_items_via_claude(
                        clean_ocr_text, image_path=str(saved_file_path),
                        use_thinking=_thinking_active,
                    )
                    if claude_items:
                        items = claude_items_to_draft_rows(claude_items)
                        extraction_strategy = "claude_api+thinking" if _thinking_active else "claude_api"
                        if tracker:
                            tracker.end_step(STEP_CALL1_EXTRACT, items=len(claude_items))
                        print(f"[Draft] Fallback extraction: {len(claude_items)} items")
                    else:
                        if tracker:
                            tracker.end_step(STEP_CALL1_EXTRACT, items=0)
            except Exception as _claude_err:
                if tracker:
                    tracker.fail_step(STEP_CALL1_EXTRACT, str(_claude_err))
                print(f"[Draft] Detection/extraction failed: {_claude_err}")

        # (Day 100.5: Strategy 2 heuristic AI and Strategy 3 legacy JSON removed.
        #  No API key = empty draft for manual input.)

        # =====================================================================
        # CALL 2: PER-CATEGORY VISUAL DIFF — Day 140
        # Day 140: SKIP_CALL2 flag bypasses this step when Vision OCR
        # provides accurate text, saving ~60-90s per upload.
        # =====================================================================
        if items and not SKIP_CALL2:
            try:
                if tracker:
                    tracker.start_step(STEP_CALL2_VISION)
                from storage.ai_vision_verify import verify_category_visual

                _cat_groups: Dict[str, list] = {}
                for it in items:
                    cat = (it.get("category") or "Other").strip()
                    _cat_groups.setdefault(cat, []).append(it)

                _pos_to_item = {it.get("position", 0): it for it in items}
                total_corrections = 0
                total_added = 0

                for cat_name, cat_items in _cat_groups.items():
                    update_import_job(job_id, pipeline_stage=f"verifying:{cat_name}")

                    _items_for_api = []
                    for it in cat_items:
                        api_item = dict(it)
                        api_item["id"] = it.get("position", 0)
                        _items_for_api.append(api_item)

                    result = verify_category_visual(
                        str(saved_file_path), cat_name, _items_for_api, {},
                        model="claude-sonnet-4-5-20250929",
                        ocr_text=clean_ocr_text,
                    )

                    if result.get("error"):
                        print(f"[Call2] '{cat_name}': error: {result['error']}")
                        continue

                    _items_to_remove = set()
                    for corr in (result.get("corrections") or []):
                        if not isinstance(corr, dict):
                            continue
                        pos = corr.get("item_id")
                        target = _pos_to_item.get(pos)
                        if not target:
                            continue

                        # Call 2 can flag section headers for removal
                        if corr.get("remove"):
                            _items_to_remove.add(pos)
                            total_corrections += 1
                            continue

                        fixes = corr.get("fixes") or {}
                        for field, val in fixes.items():
                            if field == "price_cents":
                                try:
                                    target["price_cents"] = int(round(float(val)))
                                except (ValueError, TypeError):
                                    pass
                            elif field == "name":
                                target["name"] = str(val).strip()
                            elif field == "description":
                                target["description"] = (str(val).strip() if val else None)

                        vf = corr.get("variant_fixes")
                        if vf and isinstance(vf, list):
                            new_variants = []
                            for vi, v in enumerate(vf):
                                if not isinstance(v, dict):
                                    continue
                                lbl = (v.get("label") or "").strip()
                                vp = 0
                                try:
                                    vp = int(round(float(v.get("price_cents", 0))))
                                except (ValueError, TypeError):
                                    pass
                                if lbl or vp:
                                    new_variants.append({
                                        "label": lbl or f"Size {vi+1}",
                                        "price_cents": vp,
                                        "kind": v.get("kind", "size"),
                                        "position": vi,
                                    })
                            if new_variants:
                                target["_variants"] = new_variants
                                if target.get("price_cents", 0) == 0:
                                    target["price_cents"] = new_variants[0]["price_cents"]
                        total_corrections += 1

                    for mi in (result.get("missing_items") or []):
                        if not isinstance(mi, dict):
                            continue
                        name = (mi.get("name") or "").strip()
                        if not name:
                            continue
                        pc = 0
                        try:
                            pc = int(round(float(mi.get("price_cents", 0))))
                        except (ValueError, TypeError):
                            pass
                        next_pos = max((it.get("position", 0) for it in items), default=0) + 1
                        items.append({
                            "name": name,
                            "description": (mi.get("description") or "").strip() or None,
                            "price_cents": pc,
                            "category": cat_name,
                            "position": next_pos,
                            "confidence": 85,
                        })
                        _pos_to_item[next_pos] = items[-1]
                        total_added += 1

                    # Remove items flagged as section headers (not orderable)
                    if _items_to_remove:
                        items = [it for it in items if it.get("position", 0) not in _items_to_remove]
                        for rpos in _items_to_remove:
                            _pos_to_item.pop(rpos, None)
                        print(f"[Call2] '{cat_name}': removed {len(_items_to_remove)} non-items (section headers)")

                    n_c = len(result.get("corrections") or [])
                    n_m = len(result.get("missing_items") or [])
                    if n_c or n_m:
                        print(f"[Call2] '{cat_name}': {n_c} corrections, {n_m} missing")

                extraction_strategy = "detect+assemble+vision"
                if tracker:
                    tracker.end_step(STEP_CALL2_VISION,
                                     categories=len(_cat_groups),
                                     corrections=total_corrections,
                                     items_added=total_added)
                print(f"[Call2] Done: {len(_cat_groups)} categories, "
                      f"{total_corrections} corrections, {total_added} added")

                vision_result = {
                    "skipped": False, "confidence": 0.0,
                    "model": "claude-sonnet-4-5-20250929",
                    "changes_count": total_corrections,
                    "changes": [], "gap_warnings": [],
                    "notes": f"{total_corrections} corrections across {len(_cat_groups)} categories",
                }

            except Exception as _call2_err:
                if tracker:
                    tracker.fail_step(STEP_CALL2_VISION, str(_call2_err))
                print(f"[Call2] Failed: {_call2_err}")
        elif items and SKIP_CALL2:
            if tracker:
                tracker.skip_step(STEP_CALL2_VISION, "SKIP_CALL2=True")
            print("[Call2] Skipped (SKIP_CALL2=True, Day 140 pipeline optimization)")

        # =====================================================================
        # SEMANTIC PIPELINE — Phase 8 quality checks on Claude-extracted items
        # Runs cross-item consistency, confidence scoring, tiers, repair recs,
        # auto-repair, and generates quality report.
        # =====================================================================
        semantic_result = None
        if items and extraction_strategy in ("claude_api", "claude_api+vision", "detect+assemble", "detect+assemble+vision"):
            try:
                if tracker:
                    tracker.start_step(STEP_SEMANTIC)
                from storage.semantic_bridge import run_semantic_pipeline
                semantic_result = run_semantic_pipeline(items)
                n_repairs = semantic_result.get("repairs_applied", 0)
                grade = semantic_result.get("quality_grade", "?")
                mean_conf = semantic_result.get("mean_confidence", 0.0)
                if tracker:
                    tracker.end_step(STEP_SEMANTIC, items=len(items),
                                     quality_grade=grade, repairs=n_repairs,
                                     mean_confidence=mean_conf)
                print(f"[Draft] Semantic pipeline: grade={grade}, "
                      f"mean_confidence={mean_conf:.2f}, repairs={n_repairs}")
            except Exception as _sem_err:
                if tracker:
                    tracker.fail_step(STEP_SEMANTIC, str(_sem_err))
                print(f"[Draft] Semantic pipeline failed: {_sem_err}")

        # =====================================================================
        # CALL 3: TARGETED RECONCILIATION — Sprint 11.2 (Day 102)
        # Day 140: SKIP_CALL3 flag bypasses reconciliation when Vision OCR
        # provides accurate extraction, saving ~30-60s per upload.
        # =====================================================================
        reconcile_result = None
        update_import_job(job_id, pipeline_stage="reconciling")
        if items and not SKIP_CALL3 and semantic_result and semantic_result.get("items"):
            try:
                if tracker:
                    tracker.start_step(STEP_CALL3_RECONCILE)
                from storage.ai_reconcile import (
                    collect_flagged_items, reconcile_flagged_items,
                    merge_reconciled_items,
                )
                from storage.semantic_confidence import (
                    score_semantic_confidence as _rescore,
                    classify_confidence_tiers as _reclassify,
                )

                # Collect items flagged by semantic pipeline
                flagged = collect_flagged_items(semantic_result["items"])

                if flagged:
                    reconcile_result = reconcile_flagged_items(
                        str(saved_file_path), flagged
                    )
                    if (not reconcile_result.get("skipped")
                            and not reconcile_result.get("error")):
                        # Merge corrections back into the semantic-processed items
                        # (which are deep copies, not the draft_items list itself)
                        sem_items = semantic_result["items"]
                        sem_items, merge_changes = merge_reconciled_items(
                            sem_items, reconcile_result["items"]
                        )
                        reconcile_result["merge_changes"] = merge_changes

                        # Stamp Call 3 confidence for signal #6 in re-scoring (Day 106)
                        try:
                            from storage.semantic_confidence import stamp_claude_confidence as _stamp_c3
                            _c3_conf = reconcile_result.get("confidence", 0)
                            if _c3_conf:
                                _stamp_c3(sem_items, _c3_conf)
                        except Exception:
                            pass

                        # Re-score confidence after reconciliation corrections
                        _rescore(sem_items)
                        _reclassify(sem_items)

                        # Apply reconciliation field fixes back to draft items
                        # Draft items are matched by position (1:1 with sem_items)
                        for draft_it, sem_it in zip(items, sem_items):
                            for field in ("name", "category", "description"):
                                new_val = sem_it.get(field)
                                if new_val and new_val != draft_it.get(field):
                                    draft_it[field] = new_val
                            # Price: sem_items store price_cents directly
                            new_price = sem_it.get("price_cents")
                            if new_price and new_price != draft_it.get("price_cents"):
                                draft_it["price_cents"] = new_price
                            # Confidence: convert 0-100 back from 0-1 semantic scale
                            new_conf = sem_it.get("confidence")
                            if new_conf is not None:
                                if isinstance(new_conf, (int, float)) and new_conf <= 1.0:
                                    draft_it["confidence"] = int(round(new_conf * 100))
                                else:
                                    draft_it["confidence"] = int(round(new_conf))

                        confirmed = reconcile_result.get("items_confirmed", 0)
                        corrected = reconcile_result.get("items_corrected", 0)
                        not_found = reconcile_result.get("items_not_found", 0)
                        if tracker:
                            tracker.end_step(
                                STEP_CALL3_RECONCILE, items=len(flagged),
                                confirmed=confirmed, corrected=corrected,
                                not_found=not_found,
                                confidence=reconcile_result.get("confidence", 0),
                            )
                        print(f"[Draft] Call 3 (Reconciliation): {len(flagged)} flagged -> "
                              f"{confirmed} confirmed, {corrected} corrected, "
                              f"{not_found} not_found")
                    else:
                        skip = (reconcile_result.get("skip_reason")
                                or reconcile_result.get("error", "unknown"))
                        if tracker:
                            tracker.skip_step(STEP_CALL3_RECONCILE, skip)
                        print(f"[Draft] Call 3 skipped ({skip})")
                else:
                    reconcile_result = {"skipped": True, "skip_reason": "no_flagged_items"}
                    if tracker:
                        tracker.skip_step(STEP_CALL3_RECONCILE, "no_flagged_items")
                    print("[Draft] Call 3 skipped (no flagged items)")
            except Exception as _recon_err:
                if tracker:
                    tracker.fail_step(STEP_CALL3_RECONCILE, str(_recon_err))
                print(f"[Draft] Call 3 (Reconciliation) failed: {_recon_err}")
        elif items and SKIP_CALL3:
            if tracker:
                tracker.skip_step(STEP_CALL3_RECONCILE, "SKIP_CALL3=True")
            print("[Draft] Call 3 skipped (SKIP_CALL3=True, Day 140 pipeline optimization)")

        # =====================================================================
        # CONFIDENCE GATE — Sprint 11.3 (Day 106)
        # Evaluates binary pass/fail at the menu level.  Failed gates log a
        # rejection entry and set job status to "rejected" so the frontend
        # can surface the customer_message retry prompt instead of the editor.
        # =====================================================================
        gate_result = None
        if False and items and not _thinking_active:  # Day 140: gate disabled — never tested with new pipeline
            try:
                from storage.confidence_gate import evaluate_confidence_gate
                _gate_items = (
                    semantic_result["items"]
                    if semantic_result and semantic_result.get("items")
                    else items
                )
                _call2_conf = (
                    vision_result.get("confidence") or None
                    if vision_result and not vision_result.get("skipped")
                       and not vision_result.get("error")
                    else None
                )
                _call3_conf = (
                    reconcile_result.get("confidence") or None
                    if reconcile_result and not reconcile_result.get("skipped")
                       and not reconcile_result.get("error")
                    else None
                )
                gate_result = evaluate_confidence_gate(
                    _gate_items,
                    call2_confidence=_call2_conf,
                    call3_confidence=_call3_conf,
                    ocr_char_count=len(clean_ocr_text),
                )
                if gate_result.passed:
                    print(f"[Draft] Gate PASS: score={gate_result.score:.4f}")
                else:
                    print(f"[Draft] Gate FAIL: score={gate_result.score:.4f} [{gate_result.reason}]")
                    if drafts_store is not None and hasattr(drafts_store, "log_pipeline_rejection"):
                        try:
                            _draft_id_for_gate = None
                            try:
                                _draft_id_for_gate = _get_or_create_draft_for_job(
                                    job_id, allow_create=False
                                )
                            except Exception:
                                pass
                            drafts_store.log_pipeline_rejection(
                                restaurant_id=None,
                                draft_id=_draft_id_for_gate,
                                image_path=str(saved_file_path),
                                ocr_chars=len(clean_ocr_text),
                                item_count=len(items),
                                gate_score=gate_result.score,
                                gate_reason=gate_result.reason,
                                pipeline_signals=gate_result.signals,
                            )
                        except Exception as _rej_err:
                            print(f"[Draft] Rejection log failed: {_rej_err}")
            except Exception as _gate_err:
                print(f"[Draft] Confidence gate failed: {_gate_err}")

        price_intel_result = None  # populated after draft persistence (Call 4)

        update_import_job(job_id, pipeline_stage="finalizing")

        # ✅ CRITICAL (SUCCESS PATH): hydrate DB-backed draft items + save OCR debug payload
        # status="done" is set AFTER items are in the DB so auto-redirect
        # lands on a populated editor.
        try:
            if drafts_store is not None:
                draft_id = _get_or_create_draft_for_job(job_id, allow_create=True)

                if draft_id and hasattr(drafts_store, "upsert_draft_items"):
                    if items:
                        # Clear any heuristic items loaded by _get_or_create_draft_for_job
                        # before inserting Claude pipeline items.  Without this, garbage
                        # items from the Tesseract heuristic path persist in the draft.
                        try:
                            existing = drafts_store.get_draft_items(draft_id, include_variants=False)
                            if existing:
                                old_ids = [it["id"] for it in existing if "id" in it]
                                if old_ids:
                                    drafts_store.delete_draft_items(draft_id, old_ids)
                                    print(f"[Draft] Cleared {len(old_ids)} heuristic items before Claude upsert")
                        except Exception as _clear_err:
                            print(f"[Draft] Warning: could not clear old items: {_clear_err}")
                        upsert_result = drafts_store.upsert_draft_items(draft_id, items)

                        # Day 139: Store bounding box coordinates + source elements
                        if _detected_elements and _coord_data:
                            try:
                                import json as _json
                                # Save raw classified elements on the draft
                                drafts_store.save_source_elements(
                                    draft_id, _json.dumps(_detected_elements)
                                )
                                # Link coordinates to inserted item IDs by position
                                inserted_ids = upsert_result.get("inserted_ids", [])
                                if inserted_ids:
                                    # Build position→item_id map from the freshly inserted items
                                    fresh_items = drafts_store.get_draft_items(draft_id, include_variants=False) or []
                                    pos_to_id = {it["position"]: it["id"] for it in fresh_items if it.get("position")}
                                    coord_rows = []
                                    for cd in _coord_data:
                                        item_id = pos_to_id.get(cd["position"])
                                        if item_id:
                                            coord_rows.append({
                                                "item_id": item_id,
                                                "x_pct": cd["x_pct"],
                                                "y_pct": cd["y_pct"],
                                                "w_pct": cd["w_pct"],
                                                "h_pct": cd["h_pct"],
                                                "page": cd.get("page", 1),
                                                "element_type": cd.get("element_type", "item"),
                                            })
                                    if coord_rows:
                                        drafts_store.store_item_coordinates_bulk(coord_rows)
                                        print(f"[Draft] Stored {len(coord_rows)} bounding box coordinates")
                            except Exception as _coord_err:
                                print(f"[Draft] Warning: coordinate storage failed: {_coord_err}")

                        # Day 139.5: Store Call 2 gap warnings on draft
                        if vision_result and vision_result.get("gap_warnings"):
                            try:
                                import json as _json
                                drafts_store.save_gap_warnings(
                                    draft_id, _json.dumps(vision_result["gap_warnings"])
                                )
                                print(f"[Draft] Stored {len(vision_result['gap_warnings'])} gap warnings")
                            except Exception as _gw_err:
                                print(f"[Draft] Warning: gap warning storage failed: {_gw_err}")

                if draft_id and hasattr(drafts_store, "save_ocr_debug"):
                    payload = debug_payload if isinstance(debug_payload, dict) else {}
                    payload.setdefault("import_job_id", int(job_id))
                    payload.setdefault("pipeline", engine or "unknown")
                    payload.setdefault("bridge", "run_ocr_and_make_draft")
                    payload["extraction_strategy"] = extraction_strategy
                    payload["clean_ocr_chars"] = len(clean_ocr_text)
                    if vision_result is not None:
                        payload["vision_verification"] = {
                            "skipped": vision_result.get("skipped", False),
                            "skip_reason": vision_result.get("skip_reason"),
                            "error": vision_result.get("error"),
                            "confidence": vision_result.get("confidence", 0.0),
                            "model": vision_result.get("model"),
                            "changes_count": len(vision_result.get("changes", [])),
                            "changes": vision_result.get("changes", []),
                            "notes": vision_result.get("notes"),
                            "item_count_before": len(vision_result.get("items", [])),
                            "gap_warnings": vision_result.get("gap_warnings", []),
                        }
                    if semantic_result is not None:
                        payload["semantic_pipeline"] = {
                            "quality_grade": semantic_result.get("quality_grade"),
                            "mean_confidence": semantic_result.get("mean_confidence", 0.0),
                            "tier_counts": semantic_result.get("tier_counts", {}),
                            "repairs_applied": semantic_result.get("repairs_applied", 0),
                            "repair_results": semantic_result.get("repair_results", {}),
                            "items_metadata": semantic_result.get("items_metadata", []),
                        }
                    if reconcile_result is not None:
                        payload["targeted_reconciliation"] = {
                            "skipped": reconcile_result.get("skipped", False),
                            "skip_reason": reconcile_result.get("skip_reason"),
                            "error": reconcile_result.get("error"),
                            "confidence": reconcile_result.get("confidence", 0.0),
                            "model": reconcile_result.get("model"),
                            "items_confirmed": reconcile_result.get("items_confirmed", 0),
                            "items_corrected": reconcile_result.get("items_corrected", 0),
                            "items_not_found": reconcile_result.get("items_not_found", 0),
                            "changes": reconcile_result.get("changes", []),
                            "merge_changes": reconcile_result.get("merge_changes", []),
                            "notes": reconcile_result.get("notes"),
                        }
                    if gate_result is not None:
                        payload["confidence_gate"] = {
                            "passed": gate_result.passed,
                            "score": round(gate_result.score, 4),
                            "threshold": gate_result.threshold,
                            "signals": gate_result.signals,
                            "reason": gate_result.reason,
                        }
                    if tracker:
                        tracker.strategy = extraction_strategy
                        payload["pipeline_metrics"] = tracker.summary()
                    drafts_store.save_ocr_debug(draft_id, payload)
        except Exception as _draft_err:
            print(f"[Draft] ERROR creating draft items: {_draft_err}")
            import traceback; traceback.print_exc()

        # Day 141.7: Call 4 (Price Intelligence) moved OUT of the import
        # pipeline. It now runs on-demand after the restaurant-details
        # confirmation screen, gated by tier. Keeps upload fast and lets
        # users verify their own restaurant info before we scrape
        # competitor prices for comparison.

        # Mark done AFTER items are in DB so auto-redirect shows populated editor
        if gate_result is not None and not gate_result.passed:
            update_import_job(job_id, status="rejected", pipeline_stage="done", error=gate_result.customer_message)
        else:
            update_import_job(job_id, status="done", pipeline_stage="done")


    except Exception as e:
        import traceback as _tb
        _tb.print_exc()
        print(f"[Draft] run_ocr_and_make_draft CRASHED: {type(e).__name__}: {e}")
        try:
            update_import_job(job_id, status="failed", pipeline_stage="done", error=str(e))
        except Exception as _upd_err:
            print(f"[Draft] update_import_job also failed: {_upd_err}")
            try:
                update_import_job(job_id, status="failed", error=str(e))
            except Exception:
                pass


# Upload route (JSON API) — returns job id
@app.post("/api/menus/import")
@login_required
def import_menu():
    # Tier gate: free-tier users cannot upload images/PDFs via API
    _raw = session.get("user")
    _u = _raw if isinstance(_raw, dict) else {}
    if _u.get("role") != "admin" and _u.get("user_id") and _u.get("account_tier") != "premium":
        return jsonify({"error": "Photo/PDF upload requires the Premium Package."}), 403
    try:
        # Day 139: Accept multiple files (multi-page menu support)
        files = request.files.getlist("file")
        if not files or all(f.filename == "" for f in files):
            # Fallback: single file field
            if "file" in request.files:
                files = [request.files["file"]]
            else:
                return jsonify({"error": "No file field 'file' provided"}), 400

        # Validate and save all files
        saved_paths = []
        primary_name = None
        for file in files:
            if not file or file.filename == "":
                continue
            if not allowed_file(file.filename):
                return jsonify({"error": f"Unsupported file type: {file.filename}. Allowed: jpg, jpeg, png, pdf"}), 400
            base_name = secure_filename(file.filename) or "upload"
            tmp_name = f"{uuid.uuid4().hex[:8]}_{base_name}"
            save_path = UPLOAD_FOLDER / tmp_name
            file.save(str(save_path))
            saved_paths.append(save_path)
            if primary_name is None:
                primary_name = tmp_name

        if not saved_paths:
            return jsonify({"error": "No valid files uploaded"}), 400

        restaurant_id = _resolve_restaurant_id_from_request()
        job_id = create_import_job(filename=primary_name, restaurant_id=restaurant_id)

        # Day 141: persist extra page filenames so the wizard menu viewer can
        # cycle through all pages later.
        if len(saved_paths) > 1:
            import json as _json
            extra_names = [p.name for p in saved_paths[1:]]
            try:
                with db_connect() as _c:
                    _c.execute(
                        "UPDATE import_jobs SET extra_filenames=? WHERE id=?",
                        (_json.dumps(extra_names), job_id),
                    )
                    _c.commit()
            except Exception as _e:
                print(f"[APP] failed to save extra_filenames: {_e}")

        # Pass all file paths for multi-page processing
        if len(saved_paths) == 1:
            t = threading.Thread(target=run_ocr_and_make_draft, args=(job_id, saved_paths[0]), daemon=True)
        else:
            # Multi-file: pass extra_pages kwarg
            t = threading.Thread(
                target=run_ocr_and_make_draft,
                args=(job_id, saved_paths[0]),
                kwargs={"extra_pages": saved_paths[1:]},
                daemon=True,
            )
        t.start()

        return jsonify({
            "job_id": job_id, "status": "pending",
            "file": primary_name, "page_count": len(saved_paths),
            "restaurant_id": restaurant_id,
        }), 200

    except RequestEntityTooLarge:
        return jsonify({"error": "File too large. Try a smaller image or raise MAX_CONTENT_LENGTH."}), 413
    except Exception as e:
        return jsonify({"error": f"Server error while saving upload: {e}"}), 500

# Check job status
@app.get("/api/menus/import/<int:job_id>/status")
@login_required
def import_status(job_id):
    row = get_import_job(job_id)
    if not row:
        abort(404)
    data = dict(row)
    abs_draft = _abs_from_rel(row["draft_path"])
    data["draft_ready"] = bool(abs_draft and abs_draft.exists())
    return jsonify(data)

# ------------------------
# HTML Pages (Portal UI)
# ------------------------
@app.get("/restaurants")
@login_required
def restaurants_page():
    u = session.get("user") or {}
    # Customers: redirect to restaurant detail (which is now the restaurants management page)
    if _is_customer() and u.get("user_id") and users_store:
        links = users_store.get_user_restaurants(u["user_id"])
        rest_ids = [lnk["restaurant_id"] for lnk in links]
        if rest_ids:
            return redirect(url_for("restaurant_detail", rest_id=rest_ids[0]))
        # No restaurants yet — show the page with empty grid
        return redirect(url_for("restaurant_detail", rest_id=0))
    # Admins: show the old table view
    with db_connect() as conn:
        rows = conn.execute("SELECT * FROM restaurants WHERE active=1 ORDER BY id").fetchall()
    return _safe_render("restaurants.html", restaurants=rows)

@app.post("/restaurants")
@login_required
def create_restaurant():
    """Create a new restaurant. Auto-links to the logged-in customer user."""
    name = (request.form.get("name") or "").strip()
    redirect_to = request.form.get("_redirect") or None
    want_json = redirect_to == "__json__"
    if not name:
        if want_json:
            return jsonify({"error": "Restaurant name is required."}), 400
        flash("Restaurant name is required.", "error")
        return redirect(redirect_to or url_for("restaurants_page"))
    phone = (request.form.get("phone") or "").strip() or None
    address = (request.form.get("address") or "").strip() or None
    address_line2 = (request.form.get("address_line2") or "").strip() or None
    city = (request.form.get("city") or "").strip() or None
    state = (request.form.get("state") or "").strip().upper() or None
    zip_code = (request.form.get("zip_code") or "").strip() or None
    cuisine_type = (request.form.get("cuisine_type") or "").strip() or None
    try:
        with db_connect() as conn:
            cur = conn.execute(
                "INSERT INTO restaurants (name, phone, address, address_line2, city, state, zip_code, cuisine_type, active, created_at) VALUES (?, ?, ?, ?, ?, ?, ?, ?, 1, datetime('now'))",
                (name, phone, address, address_line2, city, state, zip_code, cuisine_type),
            )
            rest_id = cur.lastrowid
            conn.commit()

        # Auto-link customer user to the new restaurant (Day 127)
        u = session.get("user") or {}
        user_id = u.get("user_id")
        if user_id and users_store:
            try:
                users_store.link_user_restaurant(user_id, rest_id, role="owner")
            except Exception:
                pass  # link failed (e.g. already linked) — restaurant still created
            # Update session with first restaurant if none set
            if not u.get("restaurant_id"):
                session["user"] = {**u, "restaurant_id": rest_id}

        if want_json:
            return jsonify({"id": rest_id, "name": name})
        flash(f'Restaurant "{name}" created.', "success")
    except Exception as e:
        if want_json:
            return jsonify({"error": str(e)}), 500
        flash(f"Failed to create restaurant: {e}", "error")
    if want_json:
        return jsonify({"error": "Unknown error"}), 500
    return redirect(redirect_to or url_for("restaurants_page"))

@app.get("/restaurants/<int:rest_id>/menus")
@login_required
@require_restaurant_access
def menus_page(rest_id):
    with db_connect() as conn:
        rest = conn.execute("SELECT * FROM restaurants WHERE id=?", (rest_id,)).fetchone()
    if not rest:
        abort(404)
    # Use menus_store for richer data (version_count, menu_type, etc.)
    if menus_store:
        menu_list = menus_store.list_menus(rest_id)
    else:
        with db_connect() as conn:
            menu_list = [dict(r) for r in conn.execute(
                "SELECT * FROM menus WHERE restaurant_id=? AND active=1 ORDER BY id", (rest_id,),
            ).fetchall()]
    valid_types = sorted(menus_store.VALID_MENU_TYPES) if menus_store else []
    # Day 141.6: Also show drafts linked to this restaurant
    draft_list = []
    try:
        import json as _json
        with db_connect() as conn:
            draft_list = [dict(r) for r in conn.execute(
                "SELECT id, title, status, source, wizard_completed, updated_at FROM drafts WHERE restaurant_id=? ORDER BY updated_at DESC",
                (rest_id,),
            ).fetchall()]
        # Pre-fetch export history and wizard review counts for all drafts
        draft_ids = [d["id"] for d in draft_list]
        export_map = {}  # draft_id -> latest export format
        review_started = set()  # draft_ids that have wizard reviews
        if draft_ids:
            with db_connect() as conn:
                placeholders = ",".join("?" * len(draft_ids))
                # Latest export per draft
                for row in conn.execute(
                    f"SELECT draft_id, format, exported_at FROM draft_export_history WHERE draft_id IN ({placeholders}) ORDER BY exported_at DESC",
                    draft_ids,
                ).fetchall():
                    if row["draft_id"] not in export_map:
                        export_map[row["draft_id"]] = row["format"]
                # Which drafts have wizard reviews started
                for row in conn.execute(
                    f"SELECT DISTINCT draft_id FROM draft_category_reviews WHERE draft_id IN ({placeholders})",
                    draft_ids,
                ).fetchall():
                    review_started.add(row["draft_id"])

        # Pre-fetch item counts per draft
        item_count_map = {}
        if draft_ids:
            with db_connect() as conn:
                for row in conn.execute(
                    f"SELECT draft_id, COUNT(*) as cnt FROM draft_items WHERE draft_id IN ({placeholders}) GROUP BY draft_id",
                    draft_ids,
                ).fetchall():
                    item_count_map[row["draft_id"]] = row["cnt"]

        for d in draft_list:
            d["item_count"] = item_count_map.get(d["id"], 0)
            # Derive import method from source JSON
            src = {}
            try:
                src = _json.loads(d.get("source") or "{}")
            except Exception:
                pass
            st = src.get("source_type") or src.get("type") or ""
            if "csv" in st:
                d["import_method"] = "CSV"
            elif "xlsx" in st or "excel" in st:
                d["import_method"] = "Excel"
            elif "json" in st and "structured" in st:
                d["import_method"] = "JSON"
            elif st == "upload" or "ocr" in str(src.get("ocr_engine", "")):
                d["import_method"] = "Photo"
            else:
                d["import_method"] = "Manual"
            # Derive display status (priority order)
            did = d["id"]
            if did in export_map:
                fmt = export_map[did]
                fmt_names = {
                    "csv": "CSV", "csv_flat": "CSV", "csv_variants": "CSV",
                    "json": "JSON", "xlsx": "Excel", "xlsx_by_category": "Excel",
                    "square_csv": "Square", "toast_csv": "Toast",
                    "clover_csv": "Clover", "pos_json": "POS JSON",
                }
                export_name = fmt_names.get(fmt, fmt.replace("_", " ").title() if fmt else "File")
                d["display_status"] = f"Exported to {export_name}"
                d["status_class"] = "exported"
            elif d.get("status") == "saved":
                d["display_status"] = "Saved"
                d["status_class"] = "saved"
            elif d.get("wizard_completed"):
                d["display_status"] = "In Editor"
                d["status_class"] = "in-editor"
            elif did in review_started:
                d["display_status"] = "In Review"
                d["status_class"] = "in-review"
            else:
                d["display_status"] = "Draft"
                d["status_class"] = "draft"
            del d["source"]  # don't send raw JSON to template
    except Exception:
        pass
    return _safe_render("menus.html", restaurant=rest, menus=menu_list, drafts=draft_list, valid_types=valid_types)

@app.post("/restaurants/<int:rest_id>/menus")
@login_required
@require_restaurant_access
def create_menu_route(rest_id):
    """Create a new menu for a restaurant."""
    if not menus_store:
        abort(500, description="Menus storage not available.")
    name = (request.form.get("name") or "").strip()
    if not name:
        flash("Menu name is required.", "error")
        return redirect(url_for("menus_page", rest_id=rest_id))
    menu_type = (request.form.get("menu_type") or "").strip() or None
    description = (request.form.get("description") or "").strip() or None
    try:
        menu = menus_store.create_menu(rest_id, name, menu_type=menu_type, description=description)
        flash(f"Menu \"{name}\" created.", "success")
        # Day 128: redirect customers to next-step page instead of back to list
        if _is_customer():
            return redirect(url_for("menu_next_steps", rest_id=rest_id, menu_id=menu["id"]))
    except Exception as e:
        flash(f"Failed to create menu: {e}", "error")
    return redirect(url_for("menus_page", rest_id=rest_id))


@app.get("/restaurants/<int:rest_id>/menus/<int:menu_id>/next")
@login_required
@require_restaurant_access
def menu_next_steps(rest_id, menu_id):
    """Day 128: Post-create next-steps page — upload or start editing."""
    rest = None
    with db_connect() as conn:
        rest = conn.execute("SELECT * FROM restaurants WHERE id=? AND active=1", (rest_id,)).fetchone()
    if not rest:
        abort(404)
    menu = menus_store.get_menu(menu_id) if menus_store else None
    if not menu:
        abort(404)
    return _safe_render("menu_next_steps.html", restaurant=rest, menu=menu)


@app.post("/restaurants/<int:rest_id>/menus/<int:menu_id>/new-draft")
@login_required
@require_restaurant_access
def create_blank_draft(rest_id, menu_id):
    """Day 128: Create a blank draft linked to a restaurant + menu, redirect to editor."""
    _require_drafts_storage()
    menu = menus_store.get_menu(menu_id) if menus_store else None
    menu_name = menu["name"] if menu else "Menu"
    title = f"{menu_name} — Draft"
    draft_id = drafts_store._insert_draft(
        title=title,
        restaurant_id=rest_id,
        menu_id=menu_id,
        status="editing",
    )
    flash(f'Draft created for "{menu_name}".', "success")
    return redirect(f"/drafts/{draft_id}/edit")


@app.get("/menus/<int:menu_id>/edit")
@login_required
def menu_edit_redirect(menu_id):
    """Find or create an editing draft for this menu, then redirect to editor."""
    _require_drafts_storage()
    if not menus_store:
        abort(500, description="Menus storage not available.")
    menu = menus_store.get_menu(menu_id)
    if not menu:
        abort(404, description="Menu not found")
    rest_id = menu["restaurant_id"]
    # Look for an existing draft linked to this menu in editing status
    with db_connect() as conn:
        row = conn.execute(
            "SELECT id FROM drafts WHERE menu_id = ? AND status = 'editing' ORDER BY id DESC LIMIT 1",
            (menu_id,),
        ).fetchone()
    if row:
        return redirect(f"/drafts/{row['id']}/edit")
    # No existing draft — create one
    title = f"{menu['name']} — Draft"
    draft_id = drafts_store._insert_draft(
        title=title,
        restaurant_id=rest_id,
        menu_id=menu_id,
        status="editing",
    )
    flash(f'Draft created for "{menu["name"]}".', "success")
    return redirect(f"/drafts/{draft_id}/edit")


@app.post("/menus/<int:menu_id>/update")
@login_required
def update_menu_route(menu_id):
    """Update menu metadata (name, type, description)."""
    if not menus_store:
        abort(500, description="Menus storage not available.")
    menu = menus_store.get_menu(menu_id)
    if not menu:
        abort(404, description="Menu not found")
    name = (request.form.get("name") or "").strip() or None
    menu_type = (request.form.get("menu_type") or "").strip() or None
    description = (request.form.get("description") or "").strip() or None
    try:
        menus_store.update_menu(menu_id, name=name, menu_type=menu_type, description=description)
        flash("Menu updated.", "success")
    except Exception as e:
        flash(f"Failed to update menu: {e}", "error")
    return redirect(url_for("menus_page", rest_id=menu["restaurant_id"]))

@app.post("/menus/<int:menu_id>/delete")
@login_required
def delete_menu_route(menu_id):
    """Soft-delete a menu."""
    if not menus_store:
        abort(500, description="Menus storage not available.")
    menu = menus_store.get_menu(menu_id)
    if not menu:
        abort(404, description="Menu not found")
    rest_id = menu["restaurant_id"]
    try:
        menus_store.delete_menu(menu_id)
        flash(f"Menu \"{menu['name']}\" deleted.", "success")
    except Exception as e:
        flash(f"Failed to delete menu: {e}", "error")
    return redirect(url_for("menus_page", rest_id=rest_id))

@app.get("/menus/<int:menu_id>/items")
def items_page(menu_id):
    with db_connect() as conn:
        menu = conn.execute("SELECT * FROM menus WHERE id=?", (menu_id,)).fetchone()
        if not menu:
            abort(404)
        rest = conn.execute("SELECT * FROM restaurants WHERE id=?", (menu["restaurant_id"],)).fetchone()
        items = conn.execute(
            "SELECT * FROM menu_items WHERE menu_id=? AND is_available=1 ORDER BY id", (menu_id,),
        ).fetchall()
        return _safe_render("items.html", restaurant=rest, menu=menu, items=items)


# --- Day 88: Menu Detail & Version Views ---

@app.get("/menus/<int:menu_id>/detail")
@login_required
def menu_detail(menu_id):
    """Menu detail page showing version history (Day 88)."""
    if not menus_store:
        abort(500, description="Menus storage not available.")
    menu = menus_store.get_menu(menu_id)
    if not menu:
        abort(404)
    with db_connect() as conn:
        rest = conn.execute(
            "SELECT * FROM restaurants WHERE id=?", (menu["restaurant_id"],)
        ).fetchone()
    versions = menus_store.list_menu_versions(menu_id)
    current = menus_store.get_current_version(menu_id) if versions else None
    # Day 92: activity log + stats
    activities = menus_store.list_menu_activity(menu_id, limit=10)
    stats = menus_store.get_version_stats(menu_id)
    # Day 93: schedule summary
    schedule_summary = menus_store.get_menu_schedule_summary(menu)
    valid_seasons = sorted(menus_store.VALID_SEASONS)
    return _safe_render(
        "menu_detail.html",
        restaurant=rest,
        menu=menu,
        versions=versions,
        current_version=current,
        activities=activities,
        stats=stats,
        schedule_summary=schedule_summary,
        valid_seasons=valid_seasons,
    )


@app.get("/menus/versions/<int:version_id>")
@login_required
def menu_version_detail(version_id):
    """Version detail page with full item list (Day 88)."""
    if not menus_store:
        abort(500, description="Menus storage not available.")
    version = menus_store.get_menu_version(version_id, include_items=True)
    if not version:
        abort(404)
    menu = menus_store.get_menu(version["menu_id"])
    if not menu:
        abort(404)
    with db_connect() as conn:
        rest = conn.execute(
            "SELECT * FROM restaurants WHERE id=?", (menu["restaurant_id"],)
        ).fetchone()
    version_items = version.get("items", [])
    return _safe_render(
        "menu_version_detail.html",
        restaurant=rest,
        menu=menu,
        version=version,
        version_items=version_items,
    )


# --- Day 89: Version Comparison / Diff ---

@app.get("/menus/<int:menu_id>/compare")
@login_required
def menu_version_compare(menu_id):
    """Compare two menu versions side by side (Day 89)."""
    if not menus_store:
        abort(500, description="Menus storage not available.")
    menu = menus_store.get_menu(menu_id)
    if not menu:
        abort(404)
    a = request.args.get("a", type=int)
    b = request.args.get("b", type=int)
    if a is None or b is None:
        flash("Select two versions to compare.", "warning")
        return redirect(url_for("menu_detail", menu_id=menu_id))
    diff = menus_store.compare_menu_versions(a, b)
    if diff is None or diff["menu_id"] != menu_id:
        abort(404)
    with db_connect() as conn:
        rest = conn.execute(
            "SELECT * FROM restaurants WHERE id=?", (menu["restaurant_id"],)
        ).fetchone()
    versions = menus_store.list_menu_versions(menu_id)
    return _safe_render(
        "menu_version_compare.html",
        restaurant=rest,
        menu=menu,
        diff=diff,
        versions=versions,
        version_a_id=a,
        version_b_id=b,
    )


@app.post("/menus/versions/<int:version_id>/restore")
@login_required
def menu_version_restore(version_id):
    """Restore a menu version to a new draft (Day 90)."""
    if not menus_store:
        abort(500, description="Menus storage not available.")
    # Fetch version to get menu_id before restore
    _ver = menus_store.get_menu_version(version_id, include_items=False)
    if not _ver:
        abort(404)
    result = menus_store.restore_version_to_draft(version_id)
    if result is None:
        abort(404)
    # Day 92: record restore activity
    try:
        _actor = session.get("user", {}).get("email") or session.get("user", {}).get("name")
        menus_store.record_menu_activity(
            _ver["menu_id"], "version_restored",
            version_id=version_id,
            detail=f"Restored {result['version_label']} → draft #{result['draft_id']}",
            actor=_actor,
        )
    except Exception:
        pass
    flash(
        f"Created draft #{result['draft_id']} from {result['version_label']} "
        f"({result['item_count']} items, {result['variant_count']} variants).",
        "success",
    )
    return redirect(url_for("draft_editor", draft_id=result["draft_id"]))


# --- Day 91: Edit Version Metadata ---

@app.post("/menus/versions/<int:version_id>/edit")
@login_required
def menu_version_edit(version_id):
    """Edit version label and notes (Day 91)."""
    if not menus_store:
        abort(500, description="Menus storage not available.")
    version = menus_store.get_menu_version(version_id, include_items=False)
    if not version:
        abort(404)
    label = request.form.get("label")
    notes = request.form.get("notes")
    updated = menus_store.update_menu_version(
        version_id,
        label=label if label is not None else None,
        notes=notes if notes is not None else None,
    )
    if updated:
        # Day 92: record edit activity
        try:
            _actor = session.get("user", {}).get("email") or session.get("user", {}).get("name")
            menus_store.record_menu_activity(
                version["menu_id"], "version_edited",
                version_id=version_id,
                detail=f"Edited {label or version.get('label', '')}",
                actor=_actor,
            )
        except Exception:
            pass
        flash(f"Version {version.get('label', '')} updated.", "success")
    else:
        flash("No changes to save.", "info")
    return redirect(url_for("menu_detail", menu_id=version["menu_id"]))


# --- Day 92: Version Lifecycle — Pin, Delete & Activity ---

@app.post("/menus/versions/<int:version_id>/pin")
@login_required
def menu_version_pin(version_id):
    """Toggle pin/unpin on a version (Day 92)."""
    if not menus_store:
        abort(500, description="Menus storage not available.")
    version = menus_store.get_menu_version(version_id, include_items=False)
    if not version:
        abort(404)
    _actor = None
    try:
        _actor = session.get("user", {}).get("email") or session.get("user", {}).get("name")
    except Exception:
        pass
    if version.get("pinned"):
        menus_store.unpin_menu_version(version_id)
        menus_store.record_menu_activity(
            version["menu_id"], "version_unpinned",
            version_id=version_id,
            detail=f"Unpinned {version.get('label', '')}",
            actor=_actor,
        )
        flash(f"Unpinned {version.get('label', '')}.", "success")
    else:
        menus_store.pin_menu_version(version_id)
        menus_store.record_menu_activity(
            version["menu_id"], "version_pinned",
            version_id=version_id,
            detail=f"Pinned {version.get('label', '')}",
            actor=_actor,
        )
        flash(f"Pinned {version.get('label', '')}.", "success")
    return redirect(url_for("menu_detail", menu_id=version["menu_id"]))


@app.post("/menus/versions/<int:version_id>/delete")
@login_required
def menu_version_delete(version_id):
    """Delete a menu version (Day 92). Safety: no pinned, not sole version."""
    if not menus_store:
        abort(500, description="Menus storage not available.")
    version = menus_store.get_menu_version(version_id, include_items=False)
    if not version:
        abort(404)
    _actor = None
    try:
        _actor = session.get("user", {}).get("email") or session.get("user", {}).get("name")
    except Exception:
        pass
    menu_id = version["menu_id"]
    try:
        result = menus_store.delete_menu_version(version_id)
        if result is None:
            abort(404)
        menus_store.record_menu_activity(
            menu_id, "version_deleted",
            detail=f"Deleted {result.get('label', '')} (v{result.get('version_number', '?')})",
            actor=_actor,
        )
        flash(f"Deleted {result.get('label', '')}.", "success")
    except ValueError as e:
        flash(str(e), "error")
    return redirect(url_for("menu_detail", menu_id=menu_id))


@app.get("/menus/<int:menu_id>/activity")
@login_required
def menu_activity_feed(menu_id):
    """Menu activity log page (Day 92)."""
    if not menus_store:
        abort(500, description="Menus storage not available.")
    menu = menus_store.get_menu(menu_id)
    if not menu:
        abort(404)
    with db_connect() as conn:
        rest = conn.execute(
            "SELECT * FROM restaurants WHERE id=?", (menu["restaurant_id"],)
        ).fetchone()
    activities = menus_store.list_menu_activity(menu_id)
    return _safe_render(
        "menu_activity.html",
        restaurant=rest,
        menu=menu,
        activities=activities,
    )


# --- Day 93: Menu Scheduling ---

@app.post("/menus/<int:menu_id>/schedule")
@login_required
def menu_schedule(menu_id):
    """Set or clear schedule on a menu (Day 93)."""
    if not menus_store:
        abort(500, description="Menus storage not available.")
    menu = menus_store.get_menu(menu_id)
    if not menu:
        abort(404)

    if request.form.get("clear") == "1":
        menus_store.clear_menu_schedule(menu_id)
        try:
            _actor = session.get("user", {}).get("email") or session.get("user", {}).get("name")
            menus_store.record_menu_activity(
                menu_id, "schedule_updated",
                detail="Cleared schedule",
                actor=_actor,
            )
        except Exception:
            pass
        flash("Schedule cleared.", "success")
        return redirect(url_for("menu_detail", menu_id=menu_id))

    season = (request.form.get("season") or "").strip() or None
    effective_from = (request.form.get("effective_from") or "").strip() or None
    effective_to = (request.form.get("effective_to") or "").strip() or None
    active_days_list = request.form.getlist("active_days")
    active_days = ",".join(active_days_list) if active_days_list else None
    active_start_time = (request.form.get("active_start_time") or "").strip() or None
    active_end_time = (request.form.get("active_end_time") or "").strip() or None

    updated = menus_store.set_menu_schedule(
        menu_id,
        season=season,
        effective_from=effective_from,
        effective_to=effective_to,
        active_days=active_days,
        active_start_time=active_start_time,
        active_end_time=active_end_time,
    )

    if updated:
        try:
            _actor = session.get("user", {}).get("email") or session.get("user", {}).get("name")
            summary = menus_store.get_menu_schedule_summary(
                menus_store.get_menu(menu_id)
            )
            menus_store.record_menu_activity(
                menu_id, "schedule_updated",
                detail=f"Schedule set: {summary or 'custom'}",
                actor=_actor,
            )
        except Exception:
            pass
        flash("Schedule updated.", "success")
    else:
        flash("No changes to schedule.", "info")

    return redirect(url_for("menu_detail", menu_id=menu_id))


# --- Day 94: Active Menu Switching ---

@app.get("/restaurants/<int:rest_id>/active_menus")
@login_required
def active_menus_page(rest_id):
    """Active menu switching dashboard (Day 94)."""
    if not menus_store:
        abort(500, description="Menus storage not available.")
    with db_connect() as conn:
        rest = conn.execute("SELECT * FROM restaurants WHERE id=?", (rest_id,)).fetchone()
    if not rest:
        abort(404)

    now_date = request.args.get("date") or None
    now_time = request.args.get("time") or None

    summary = menus_store.get_active_menu_summary(
        rest_id,
        now_date=now_date,
        now_time=now_time,
    )
    return _safe_render(
        "active_menus.html",
        restaurant=rest,
        summary=summary,
        query_date=now_date,
        query_time=now_time,
    )


@app.get("/api/restaurants/<int:rest_id>/active_menus")
def api_active_menus(rest_id):
    """API endpoint: get currently active menus for a restaurant (Day 94)."""
    if not menus_store:
        return jsonify({"ok": False, "error": "Menus storage not available"}), 500

    now_date = request.args.get("date") or None
    now_time = request.args.get("time") or None
    now_day = request.args.get("day") or None

    active = menus_store.get_active_menus(
        rest_id,
        now_date=now_date,
        now_time=now_time,
        now_day=now_day,
    )

    primary = active[0] if active else None
    next_trans = menus_store.get_next_transition(
        rest_id,
        now_date=now_date,
        now_time=now_time,
    )

    return jsonify({
        "ok": True,
        "restaurant_id": rest_id,
        "active_count": len(active),
        "active_menus": [
            {
                "id": m["id"],
                "name": m.get("name"),
                "menu_type": m.get("menu_type"),
                "specificity_score": m.get("specificity_score", 0),
                "is_scheduled": m.get("is_scheduled", False),
            }
            for m in active
        ],
        "primary_menu": {
            "id": primary["id"],
            "name": primary.get("name"),
            "menu_type": primary.get("menu_type"),
        } if primary else None,
        "next_transition": {
            "time": next_trans["time"],
            "type": next_trans["type"],
            "menu_name": next_trans["menu"].get("name"),
            "label": next_trans["label"],
        } if next_trans else None,
    })


# -----------------------------------------------
# Day 95: Menu Health Dashboard
# -----------------------------------------------
@app.get("/restaurants/<int:rest_id>/menu_health")
@login_required
def menu_health_page(rest_id):
    """Menu health dashboard — conflicts, coverage, scores (Day 95)."""
    if not menus_store:
        abort(500, description="Menus storage not available.")
    with db_connect() as conn:
        rest = conn.execute("SELECT * FROM restaurants WHERE id=?", (rest_id,)).fetchone()
    if not rest:
        abort(404)

    summary = menus_store.get_phase10_summary(rest_id)
    return _safe_render(
        "menu_health.html",
        restaurant=rest,
        summary=summary,
    )


@app.get("/api/restaurants/<int:rest_id>/menu_health")
def api_menu_health(rest_id):
    """API endpoint: menu health scores, conflicts, coverage (Day 95)."""
    if not menus_store:
        return jsonify({"ok": False, "error": "Menus storage not available"}), 500

    summary = menus_store.get_phase10_summary(rest_id)
    return jsonify({
        "ok": True,
        "restaurant_id": rest_id,
        "total_menus": summary["total_menus"],
        "total_versions": summary["total_versions"],
        "total_items": summary["total_items"],
        "avg_health_score": summary["avg_health_score"],
        "grade": summary["grade"],
        "conflict_count": summary["conflict_count"],
        "conflicts": summary["conflicts"],
        "coverage_score": summary["coverage"]["coverage_score"],
        "coverage_gaps": summary["coverage"]["gaps"],
        "menu_health": [
            {
                "menu_id": h["menu_id"],
                "name": h["name"],
                "health_score": h["health_score"],
                "issues": h["issues"],
                "version_count": h["version_count"],
                "latest_item_count": h["latest_item_count"],
            }
            for h in summary["menu_health"]
        ],
    })


# ------------------------
# Day 6: Auth (Login / Logout) — upgraded Day 126 (user accounts)
# ------------------------
@app.get("/login")
def login():
    return _safe_render("login.html", error=None, next=request.args.get("next"))

@app.post("/login")
def login_post():
    email_or_username = (request.form.get("username") or request.form.get("email") or "").strip()
    password = (request.form.get("password") or "").strip()
    nxt = request.form.get("next") or ""

    # Legacy dev admin login (backward compat)
    if email_or_username == DEV_USERNAME and password == DEV_PASSWORD:
        session["user"] = {"username": email_or_username, "role": "admin"}
        flash("Welcome back!", "success")
        return redirect(nxt or url_for("core.index"))

    # Database user login (Phase 13)
    if users_store:
        user = users_store.verify_password(email_or_username, password)
        if user:
            restaurants = users_store.get_user_restaurants(user["id"])
            tier = users_store.get_user_tier(user["id"])
            session["user"] = {
                "user_id": user["id"],
                "username": user["display_name"] or user["email"],
                "email": user["email"],
                "role": "customer",
                "restaurant_id": restaurants[0]["restaurant_id"] if restaurants else None,
                "account_tier": tier,
            }
            flash("Welcome back!", "success")
            # If user hasn't chosen a tier yet, redirect to plan selection
            if not tier:
                return redirect(url_for("choose_plan"))
            return redirect(nxt or url_for("dashboard"))

    flash("Invalid credentials", "error")
    return redirect(url_for("login", next=request.form.get("next") or ""))

@app.get("/register")
def register():
    return _safe_render("register.html", error=None)

@app.post("/register")
def register_post():
    email = (request.form.get("email") or "").strip()
    password = (request.form.get("password") or "").strip()
    confirm = (request.form.get("confirm_password") or "").strip()
    display_name = (request.form.get("display_name") or "").strip() or None

    if password != confirm:
        flash("Passwords do not match", "error")
        return redirect(url_for("register"))

    if not users_store:
        flash("Registration is not available", "error")
        return redirect(url_for("register"))

    try:
        user = users_store.create_user(email, password, display_name=display_name)
    except ValueError as exc:
        flash(str(exc), "error")
        return redirect(url_for("register"))

    # Generate email verification token (Day 130)
    verification_token = users_store.generate_verification_token(user["id"])

    # Auto-login after registration
    session["user"] = {
        "user_id": user["id"],
        "username": user["display_name"] or user["email"],
        "email": user["email"],
        "role": "customer",
        "restaurant_id": None,
        "email_verified": False,
    }
    flash("Account created! Welcome to ServLine.", "success")
    flash(f"Verification link: /verify-email/{verification_token}", "success")
    return redirect(url_for("choose_plan"))

@app.post("/logout")
def logout():
    session.clear()
    flash("You have been logged out.", "info")
    return redirect(url_for("core.index"))


# ------------------------
# Day 131: Choose Your Plan
# ------------------------
@app.get("/choose-plan")
@login_required
def choose_plan():
    """Show the plan selection page (free vs premium).

    Always accessible so users can upgrade their plan.
    """
    u = session.get("user") or {}
    user_id = u.get("user_id")
    current_tier = None
    if user_id and users_store:
        current_tier = users_store.get_user_tier(user_id)
    return _safe_render("choose_plan.html", current_tier=current_tier)


@app.post("/choose-plan")
@login_required
def choose_plan_post():
    """Set the user's account tier and redirect to dashboard."""
    tier = (request.form.get("tier") or "").strip().lower()
    u = session.get("user") or {}
    user_id = u.get("user_id")

    if not user_id or not users_store:
        flash("Unable to set plan.", "error")
        return redirect(url_for("choose_plan"))

    if tier not in ("free", "premium"):
        flash("Please select a plan.", "error")
        return redirect(url_for("choose_plan"))

    try:
        users_store.set_user_tier(user_id, tier)
    except ValueError as exc:
        flash(str(exc), "error")
        return redirect(url_for("choose_plan"))

    # Update session with tier
    session["user"] = {**u, "account_tier": tier}

    if tier == "premium":
        flash("Premium Package activated! All features unlocked — upload your menu to get started.", "success")
        return redirect(url_for("import_upload", unlocked="1"))
    else:
        flash("Free plan activated! You can start building your menu with CSV, Excel, or JSON imports.", "success")
        return redirect(url_for("import_upload"))


def _require_tier_chosen(f):
    """Decorator: redirect to /choose-plan if user hasn't picked a tier yet."""
    from functools import wraps

    @wraps(f)
    def decorated(*args, **kwargs):
        raw = session.get("user")
        u = raw if isinstance(raw, dict) else {}
        user_id = u.get("user_id")
        # Admins, legacy sessions (no user_id), and non-dict sessions bypass
        if u.get("role") == "admin" or (raw and not user_id):
            return f(*args, **kwargs)
        # Check if tier is in session first (fast path)
        if u.get("account_tier"):
            return f(*args, **kwargs)
        # Check DB
        if user_id and users_store:
            tier = users_store.get_user_tier(user_id)
            if tier:
                session["user"] = {**u, "account_tier": tier}
                return f(*args, **kwargs)
        return redirect(url_for("choose_plan"))

    return decorated


def _require_premium(f):
    """Decorator: block access unless user has premium tier.
    Returns 403 with upgrade prompt for free-tier users."""
    from functools import wraps

    @wraps(f)
    def decorated(*args, **kwargs):
        raw = session.get("user")
        u = raw if isinstance(raw, dict) else {}
        user_id = u.get("user_id")
        # Admins, legacy sessions (no user_id), and non-dict sessions bypass
        if u.get("role") == "admin" or (raw and not user_id):
            return f(*args, **kwargs)
        tier = u.get("account_tier")
        if not tier:
            if user_id and users_store:
                tier = users_store.get_user_tier(user_id)
        if tier == "premium":
            return f(*args, **kwargs)
        flash("This feature requires the Premium Package.", "error")
        return redirect(url_for("dashboard"))

    return decorated


# ------------------------
# Day 130: Email Verification & Password Reset
# ------------------------
@app.get("/verify-email/<token>")
def verify_email(token):
    """Verify a user's email via token link."""
    if not users_store:
        flash("Email verification is not available.", "error")
        return redirect(url_for("core.index"))
    user = users_store.verify_email_token(token)
    if user:
        # Update session if the verified user is logged in
        u = session.get("user") or {}
        if u.get("user_id") == user["id"]:
            session["user"] = {**u, "email_verified": True}
        flash("Email verified successfully!", "success")
        return redirect(url_for("account_page") if session.get("user") else url_for("login"))
    flash("Invalid or expired verification link.", "error")
    return redirect(url_for("core.index"))


@app.post("/resend-verification")
@login_required
def resend_verification():
    """Generate a new email verification token for the logged-in user."""
    u = session.get("user") or {}
    user_id = u.get("user_id")
    if not user_id or not users_store:
        flash("Verification not available.", "error")
        return redirect(url_for("account_page"))
    user_data = users_store.get_user_by_id(user_id)
    if user_data and user_data["email_verified"]:
        flash("Your email is already verified.", "info")
        return redirect(url_for("account_page"))
    token = users_store.generate_verification_token(user_id)
    # In production, send this via email. For now, flash the link.
    flash(f"Verification link: /verify-email/{token}", "success")
    return redirect(url_for("account_page"))


@app.get("/forgot-password")
def forgot_password():
    return _safe_render("forgot_password.html")


@app.post("/forgot-password")
def forgot_password_post():
    email = (request.form.get("email") or "").strip()
    if not email:
        flash("Please enter your email address.", "error")
        return redirect(url_for("forgot_password"))
    if not users_store:
        flash("Password reset is not available.", "error")
        return redirect(url_for("forgot_password"))
    token = users_store.generate_reset_token(email)
    if token:
        # In production, send this via email. For now, flash the link.
        flash(f"Password reset link: /reset-password/{token}", "success")
    else:
        # Don't reveal whether the email exists — always show same message
        flash("If an account with that email exists, a reset link has been generated.", "info")
    return redirect(url_for("forgot_password"))


@app.get("/reset-password/<token>")
def reset_password(token):
    if not users_store:
        flash("Password reset is not available.", "error")
        return redirect(url_for("core.index"))
    user_id = users_store.validate_reset_token(token)
    if not user_id:
        flash("Invalid or expired reset link.", "error")
        return redirect(url_for("forgot_password"))
    return _safe_render("reset_password.html", token=token)


@app.post("/reset-password/<token>")
def reset_password_post(token):
    if not users_store:
        flash("Password reset is not available.", "error")
        return redirect(url_for("core.index"))
    new_pw = (request.form.get("new_password") or "").strip()
    confirm_pw = (request.form.get("confirm_password") or "").strip()
    if new_pw != confirm_pw:
        flash("Passwords do not match.", "error")
        return redirect(url_for("reset_password", token=token))
    try:
        ok = users_store.consume_reset_token(token, new_pw)
    except ValueError as exc:
        flash(str(exc), "error")
        return redirect(url_for("reset_password", token=token))
    if ok:
        flash("Password reset successfully. Please sign in.", "success")
        return redirect(url_for("login"))
    flash("Invalid or expired reset link.", "error")
    return redirect(url_for("forgot_password"))


# ------------------------
# Day 127: Customer Dashboard
# ------------------------
@app.get("/dashboard")
@login_required
@_require_tier_chosen
def dashboard():
    """Customer dashboard — shows 'My Restaurants' and quick actions."""
    u = session.get("user") or {}
    user_id = u.get("user_id")

    # Fetch restaurants the customer owns/manages
    my_restaurants = []
    if user_id and users_store:
        links = users_store.get_user_restaurants(user_id)
        if links:
            rest_ids = [lnk["restaurant_id"] for lnk in links]
            role_map = {lnk["restaurant_id"]: lnk["role"] for lnk in links}
            placeholders = ",".join("?" * len(rest_ids))
            with db_connect() as conn:
                rows = conn.execute(
                    f"SELECT * FROM restaurants WHERE id IN ({placeholders}) AND active=1 ORDER BY id",
                    rest_ids,
                ).fetchall()
            for r in rows:
                rd = dict(r)
                rd["role"] = role_map.get(r["id"], "owner")
                # Stats for each restaurant (Day 128: use get_restaurant_stats)
                if users_store:
                    try:
                        stats = users_store.get_restaurant_stats(r["id"])
                        rd.update(stats)
                    except Exception:
                        rd.setdefault("draft_count", 0)
                        rd.setdefault("menu_count", 0)
                        rd.setdefault("item_count", 0)
                else:
                    with db_connect() as conn:
                        rd["draft_count"] = conn.execute(
                            "SELECT COUNT(*) FROM drafts WHERE restaurant_id=?", (r["id"],)
                        ).fetchone()[0]
                        rd["menu_count"] = conn.execute(
                            "SELECT COUNT(*) FROM menus WHERE restaurant_id=? AND active=1", (r["id"],)
                        ).fetchone()[0]
                    rd["item_count"] = 0
                my_restaurants.append(rd)

    return _safe_render("dashboard.html", my_restaurants=my_restaurants, user=u)


@app.get("/account")
@login_required
def account_page():
    """Account settings page — profile, restaurants, future sections."""
    u = session.get("user") or {}
    user_id = u.get("user_id")
    user_data = None
    my_restaurants = []

    if user_id and users_store:
        user_data = users_store.get_user_by_id(user_id)
        links = users_store.get_user_restaurants(user_id)
        if links:
            rest_ids = [lnk["restaurant_id"] for lnk in links]
            role_map = {lnk["restaurant_id"]: lnk["role"] for lnk in links}
            placeholders = ",".join("?" * len(rest_ids))
            with db_connect() as conn:
                rows = conn.execute(
                    f"SELECT * FROM restaurants WHERE id IN ({placeholders}) AND active=1 ORDER BY id",
                    rest_ids,
                ).fetchall()
            for r in rows:
                rd = dict(r)
                rd["role"] = role_map.get(r["id"], "owner")
                rd["menus"] = []
                if menus_store:
                    rd["menus"] = menus_store.list_menus(r["id"])
                my_restaurants.append(rd)

    return _safe_render("account.html", user=u, user_data=user_data,
                        my_restaurants=my_restaurants)


@app.post("/account/update")
@login_required
def account_update():
    """Update account settings (display name)."""
    u = session.get("user") or {}
    user_id = u.get("user_id")
    if not user_id or not users_store:
        flash("Account update not available.", "error")
        return redirect(url_for("account_page"))

    display_name = (request.form.get("display_name") or "").strip() or None
    users_store.update_user(user_id, display_name=display_name)

    # Update session username
    session["user"] = {**u, "username": display_name or u.get("email", "")}

    flash("Account updated.", "success")
    return redirect(url_for("account_page"))


@app.post("/account/change-password")
@login_required
def account_change_password():
    """Change the logged-in user's password."""
    u = session.get("user") or {}
    user_id = u.get("user_id")
    if not user_id or not users_store:
        flash("Password change not available.", "error")
        return redirect(url_for("account_page"))

    current_pw = (request.form.get("current_password") or "").strip()
    new_pw = (request.form.get("new_password") or "").strip()
    confirm_pw = (request.form.get("confirm_password") or "").strip()

    # Verify current password
    user_data = users_store.get_user_by_id(user_id)
    if not user_data or not check_password_hash(user_data["password_hash"], current_pw):
        flash("Current password is incorrect.", "error")
        return redirect(url_for("account_page"))

    if new_pw != confirm_pw:
        flash("New passwords do not match.", "error")
        return redirect(url_for("account_page"))

    try:
        users_store.change_password(user_id, new_pw)
        flash("Password changed successfully.", "success")
    except ValueError as exc:
        flash(str(exc), "error")

    return redirect(url_for("account_page"))


@app.post("/account/delete")
@login_required
def account_delete():
    """Delete the logged-in user's account and all associated data."""
    u = session.get("user") or {}
    user_id = u.get("user_id")
    if user_id and users_store:
        users_store.delete_user(user_id)
    session.clear()
    flash("Your account has been deleted.", "info")
    return redirect(url_for("core.index"))


# ------------------------
# Day 128: Restaurant Detail, Edit & Multi-Restaurant Switching
# ------------------------
@app.get("/restaurants/<int:rest_id>/detail")
@login_required
def restaurant_detail(rest_id):
    """Customer-facing restaurants management page."""
    # rest_id=0 means "no restaurant yet, just show the add form"
    rest = None
    if rest_id > 0:
        # Verify access (non-zero rest_id)
        u_check = session.get("user") or {}
        role = u_check.get("role")
        if role not in ("admin", None):
            uid = u_check.get("user_id")
            if not (uid and users_store and users_store.user_owns_restaurant(uid, rest_id)):
                abort(403)
        if users_store:
            rest = users_store.get_restaurant(rest_id)
        if not rest:
            with db_connect() as conn:
                row = conn.execute(
                    "SELECT * FROM restaurants WHERE id = ? AND active = 1", (rest_id,)
                ).fetchone()
                rest = dict(row) if row else None
        if not rest:
            abort(404)
    else:
        # Dummy restaurant object for the template (empty state)
        rest = {"id": 0, "name": ""}

    stats = {"draft_count": 0, "menu_count": 0, "item_count": 0}
    if users_store:
        try:
            stats = users_store.get_restaurant_stats(rest_id)
        except Exception:
            pass

    # Fetch recent drafts for this restaurant
    recent_drafts = []
    try:
        with db_connect() as conn:
            recent_drafts = [dict(r) for r in conn.execute(
                "SELECT id, title, status, created_at FROM drafts WHERE restaurant_id = ? ORDER BY id DESC LIMIT 5",
                (rest_id,),
            ).fetchall()]
    except Exception:
        pass

    # Fetch menus
    menu_list = []
    if menus_store:
        try:
            menu_list = menus_store.list_menus(rest_id)
        except Exception:
            pass

    cuisine_types = sorted(users_store.VALID_CUISINE_TYPES) if users_store else []

    # Fetch all user's restaurants for the tiles panel
    all_restaurants = []
    u = session.get("user") or {}
    if _is_customer() and u.get("user_id") and users_store:
        links = users_store.get_user_restaurants(u["user_id"])
        r_ids = [lnk["restaurant_id"] for lnk in links]
        if r_ids:
            ph = ",".join("?" * len(r_ids))
            with db_connect() as conn:
                all_restaurants = [dict(r) for r in conn.execute(
                    f"SELECT * FROM restaurants WHERE id IN ({ph}) AND active=1 ORDER BY id", r_ids
                ).fetchall()]

    # Price intel: cached comparisons + market summary (Day 134)
    market_summary = {"has_data": False}
    comparisons = []
    if price_intel and rest_id > 0:
        try:
            comparisons = price_intel.get_cached_comparisons(rest_id)
            market_summary = price_intel.get_market_summary(rest_id)
        except Exception:
            pass

    return _safe_render("restaurant_detail.html",
                        restaurant=rest, stats=stats,
                        recent_drafts=recent_drafts,
                        menus=menu_list,
                        cuisine_types=cuisine_types,
                        all_restaurants=all_restaurants,
                        market_summary=market_summary,
                        comparisons=comparisons)


@app.post("/restaurants/<int:rest_id>/update")
@login_required
@require_restaurant_access
def update_restaurant(rest_id):
    """Update restaurant details (name, phone, address, cuisine_type, website, zip_code)."""
    if not users_store:
        flash("Restaurant updates not available.", "error")
        return redirect(url_for("restaurant_detail", rest_id=rest_id))

    name = (request.form.get("name") or "").strip()
    phone = (request.form.get("phone") or "").strip() or None
    address = (request.form.get("address") or "").strip() or None
    address_line2 = (request.form.get("address_line2") or "").strip() or None
    city = (request.form.get("city") or "").strip() or None
    state = (request.form.get("state") or "").strip() or None
    zip_code = (request.form.get("zip_code") or "").strip() or None
    cuisine_type = (request.form.get("cuisine_type") or "").strip() or None
    website = (request.form.get("website") or "").strip() or None

    # Detect AJAX/JSON callers (wizard inline editor) — respond with JSON
    # instead of redirecting so the caller can refresh the panel in place.
    wants_json = (
        request.headers.get("X-Requested-With") == "XMLHttpRequest"
        or "application/json" in (request.headers.get("Accept") or "")
    )

    try:
        users_store.update_restaurant(rest_id,
                                      name=name, phone=phone, address=address,
                                      address_line2=address_line2, city=city,
                                      state=state, zip_code=zip_code,
                                      cuisine_type=cuisine_type, website=website)
        if wants_json:
            return jsonify({
                "ok": True,
                "restaurant": {
                    "id": rest_id, "name": name, "phone": phone,
                    "address": address, "address_line2": address_line2,
                    "city": city, "state": state, "zip_code": zip_code,
                    "cuisine_type": cuisine_type, "website": website,
                },
            })
        flash("Restaurant updated.", "success")
    except ValueError as exc:
        if wants_json:
            return jsonify({"ok": False, "error": str(exc)}), 400
        flash(str(exc), "error")
    except Exception as e:
        if wants_json:
            return jsonify({"ok": False, "error": f"Update failed: {e}"}), 500
        flash(f"Update failed: {e}", "error")

    redirect_to = request.form.get("_redirect") or None
    return redirect(redirect_to or url_for("restaurant_detail", rest_id=rest_id))


@app.post("/restaurants/<int:rest_id>/delete")
@login_required
@require_restaurant_access
def delete_restaurant(rest_id):
    """Soft-delete a restaurant."""
    if not users_store:
        flash("Restaurant deletion not available.", "error")
        return redirect(url_for("dashboard"))

    try:
        users_store.delete_restaurant(rest_id)
        # Clear session restaurant_id if it was the deleted one
        u = session.get("user") or {}
        if u.get("restaurant_id") == rest_id:
            session["user"] = {**u, "restaurant_id": None}
        flash("Restaurant deleted.", "success")
    except Exception as e:
        flash(f"Delete failed: {e}", "error")

    return redirect(url_for("dashboard"))


# -------------------------------------------------------------------
# Price Comparison Intelligence (Day 134)
# -------------------------------------------------------------------
@app.post("/restaurants/<int:rest_id>/price_intel")
@login_required
@require_restaurant_access
def run_price_intel(rest_id):
    """Trigger a Google Places nearby search for price comparison."""
    if not price_intel:
        flash("Price comparison not available.", "error")
        return redirect(url_for("restaurant_detail", rest_id=rest_id))
    force = request.form.get("force_refresh") == "1"
    try:
        result = price_intel.search_nearby_restaurants(rest_id, force_refresh=force)
        if result.get("error"):
            flash(result["error"], "error")
        else:
            flash(f"Found {result['result_count']} comparable restaurants nearby.", "success")
    except RuntimeError as e:
        flash(str(e), "error")
    except Exception as e:
        flash(f"Price comparison failed: {e}", "error")
    return redirect(url_for("restaurant_detail", rest_id=rest_id))


@app.get("/api/restaurants/<int:rest_id>/price_intel")
@login_required
@require_restaurant_access
def api_price_intel(rest_id):
    """JSON endpoint: return cached price comparison data + market summary."""
    if not price_intel:
        return jsonify({"error": "Price comparison not available"}), 503
    try:
        comps = price_intel.get_cached_comparisons(rest_id)
        summary = price_intel.get_market_summary(rest_id)
        return jsonify({"comparisons": comps, "summary": summary})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


# Day 141.9: per-competitor Opus comparison endpoint deleted. The route
# computed expensive Opus+thinking comparisons that the editor UI never
# surfaced — burning ~$0.30-0.50 per click for data nobody saw. If a
# future product brings back side-by-side compare, it'll be designed
# UI-first and lazy-computed on demand, not background-preloaded.


# -------------------------------------------------------------------
# Price Intelligence — Claude Call 4 (Day 135)
# -------------------------------------------------------------------
@app.post("/drafts/<int:draft_id>/price_intelligence")
@login_required
def run_price_intelligence(draft_id):
    """Trigger Claude Call 4: price intelligence analysis on a draft."""
    if not ai_price_intel:
        flash("Price intelligence not available.", "error")
        return redirect(url_for("draft_editor", draft_id=draft_id))

    u = session.get("user") or {}
    rest_id = u.get("restaurant_id", 0)
    force = request.form.get("force_refresh") == "1"

    # New run starting — clear cached editor state so the next render
    # rebuilds against fresh pipeline output.
    _invalidate_editor_cache(draft_id, rest_id)

    try:
        result = ai_price_intel.analyze_menu_prices(
            draft_id, rest_id, force_refresh=force,
        )
        if result.get("error"):
            flash(result["error"], "error")
        else:
            assessed = result.get("items_assessed", 0)
            total = result.get("total_items", 0)
            flash(f"Price intelligence complete: {assessed}/{total} items assessed.", "success")
    except Exception as exc:
        print(f"[PriceIntel] error: {exc}")
        flash("Price intelligence failed. Please try again.", "error")

    return redirect(url_for("draft_editor", draft_id=draft_id))


@app.get("/api/drafts/<int:draft_id>/price_intelligence")
@login_required
def api_price_intelligence(draft_id):
    """JSON endpoint: return price intelligence results for a draft."""
    if not ai_price_intel:
        return jsonify({"error": "Price intelligence not available"}), 503
    try:
        result = ai_price_intel.get_price_intelligence(draft_id)
        if not result:
            return jsonify({"error": "No price intelligence data", "has_data": False})
        result["has_data"] = True
        return jsonify(result)
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.get("/api/drafts/<int:draft_id>/price_intelligence/<int:item_id>")
@login_required
def api_item_price_assessment(draft_id, item_id):
    """JSON endpoint: return price assessment for a single item."""
    if not ai_price_intel:
        return jsonify({"error": "Price intelligence not available"}), 503
    try:
        result = ai_price_intel.get_item_assessment(draft_id, item_id)
        if not result:
            return jsonify({"error": "No assessment for this item"})
        return jsonify(result)
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.post("/switch-restaurant")
@login_required
def switch_restaurant():
    """Switch the active restaurant in the session."""
    u = session.get("user") or {}
    user_id = u.get("user_id")
    new_rest_id = request.form.get("restaurant_id")

    if not new_rest_id or not user_id or not users_store:
        flash("Could not switch restaurant.", "error")
        return redirect(url_for("dashboard"))

    try:
        new_rest_id = int(new_rest_id)
    except (ValueError, TypeError):
        flash("Invalid restaurant.", "error")
        return redirect(url_for("dashboard"))

    # Verify ownership
    if not users_store.user_owns_restaurant(user_id, new_rest_id):
        flash("You do not have access to that restaurant.", "error")
        return redirect(url_for("dashboard"))

    session["user"] = {**u, "restaurant_id": new_rest_id}
    rest = users_store.get_restaurant(new_rest_id)
    name = rest["name"] if rest else f"#{new_rest_id}"
    flash(f'Switched to "{name}".', "success")
    return redirect(request.form.get("_redirect") or url_for("dashboard"))


# ------------------------
# Dev helper page: simple upload form
# ------------------------
@app.get("/dev/upload")
@login_required
def dev_upload_form():
    return """
    <!doctype html>
    <html>
      <head>
        <meta charset="utf-8">
        <title>Dev Upload Test</title>
        <style>
          :root { --bg:#0b1220; --panel:#111a2f; --ink:#e8eefc; --muted:#9fb0d1; --line:#1f2a44; --brand:#7aa2ff; --brandH:#5a86f7; }
          * { box-sizing: border-box; }
          body { font-family: system-ui, -apple-system, Segoe UI, Roboto, Helvetica, Arial, sans-serif; margin: 0; background: var(--bg); color: var(--ink); }
          .wrap { max-width: 640px; margin: 32px auto; padding: 0 16px; }
          .card { background: var(--panel); border: 1px solid var(--line); border-radius: 14px; padding: 18px; box-shadow: 0 8px 24px rgba(0,0,0,.25); }
          h2 { margin: 0 0 8px; }
          p { margin: 0 0 12px; color: var(--muted); }
          .btn {
            display: inline-block; padding: 8px 14px; border-radius: 12px;
            border: 1px solid var(--brand); background: var(--brand);
            color: #000; font-weight: 600; cursor: pointer; text-decoration: none;
            transition: background .2s, border-color .2s;
          }
          .btn:hover { background: var(--brandH); border-color: var(--brandH); }
          .row { display:flex; gap:8px; align-items:center; flex-wrap:wrap; }
          .mt-3{ margin-top: .75rem; } .mt-4{ margin-top: 1rem; }
          input[type="file"] { color: #000; background: #fff; border: 1px solid var(--line); border-radius: 10px; padding: 8px; }
          code { color: #b7cdfb; }
        </style>
      </head>
      <body>
        <div class="wrap">
          <div class="card">
            <h2>Dev Upload Test</h2>
            <p>Pick an image or PDF and submit to <code>/api/menus/import</code>.</p>
            <form class="mt-3" action="/api/menus/import" method="post" enctype="multipart/form-data">
              <input type="file" name="file" accept="image/*,.pdf" required />
              <div class="mt-3 row">
                <button type="submit" class="btn">Upload</button>
                <a href="/import" class="btn">Back to Import</a>
              </div>
            </form>
            <p class="mt-4">Max file size: 20&nbsp;MB. Allowed: PNG, JPG, PDF.</p>
          </div>
        </div>
      </body>
    </html>
    """


@app.get("/test_json_form")
@login_required
def test_json_form():
    """
    Dev-only helper to POST a structured JSON file into /import/json.
    """
    return """
    <!doctype html>
    <html>
      <head>
        <meta charset="utf-8">
        <title>Dev JSON Import Test</title>
        <style>
          :root { --bg:#0b1220; --panel:#111a2f; --ink:#e8eefc; --muted:#9fb0d1; --line:#1f2a44; --brand:#7aa2ff; --brandH:#5a86f7; }
          * { box-sizing: border-box; }
          body { font-family: system-ui, -apple-system, Segoe UI, Roboto, Helvetica, Arial, sans-serif; margin: 0; background: var(--bg); color: var(--ink); }
          .wrap { max-width: 640px; margin: 32px auto; padding: 0 16px; }
          .card { background: var(--panel); border: 1px solid var(--line); border-radius: 14px; padding: 18px; box-shadow: 0 8px 24px rgba(0,0,0,.25); }
          h2 { margin: 0 0 8px; }
          p { margin: 0 0 12px; color: var(--muted); }
          .btn {
            display: inline-block; padding: 8px 14px; border-radius: 12px;
            border: 1px solid var(--brand); background: var(--brand);
            color: #000; font-weight: 600; cursor: pointer; text-decoration: none;
            transition: background .2s, border-color .2s;
          }
          .btn:hover { background: var(--brandH); border-color: var(--brandH); }
          .row { display:flex; gap:8px; align-items:center; flex-wrap:wrap; }
          .mt-3{ margin-top: .75rem; } .mt-4{ margin-top: 1rem; }
          input[type="file"] { color: #000; background: #fff; border: 1px solid var(--line); border-radius: 10px; padding: 8px; }
          code { color: #b7cdfb; }
        </style>
      </head>
      <body>
        <div class="wrap">
          <div class="card">
            <h2>Dev JSON Import Test</h2>
            <p>Upload a structured JSON payload and send it to <code>/import/json</code>.</p>
            <form class="mt-3" action="/import/json" method="post" enctype="multipart/form-data">
              <input type="file" name="json_file" accept="application/json" required />
              <div class="mt-3 row">
                <button type="submit" class="btn">Upload JSON</button>
                <a href="/import" class="btn">Back to Import</a>
              </div>
            </form>
            <p class="mt-4">Expected: JSON that passes the structured item contract.</p>
          </div>
        </div>
      </body>
    </html>
    """


# ------------------------
# Blank draft (manual menu entry)
# ------------------------
@app.post("/drafts/new-blank")
@login_required
def create_blank_draft_manual():
    """Create an empty draft and redirect to the editor for manual entry."""
    restaurant_id = _resolve_restaurant_id_from_request()
    title = f"New Menu — {datetime.utcnow().strftime('%b %d, %Y')}"
    create_fn = getattr(drafts_store, "create_draft_from_structured_items", None)
    if not callable(create_fn):
        flash("Draft creation not available.", "error")
        return redirect(url_for("import_upload"))
    draft = create_fn(title=title, restaurant_id=restaurant_id, items=[])
    draft_id = draft.get("id") or draft.get("draft_id")
    if draft_id:
        return redirect(f"/drafts/{draft_id}/edit")
    flash("Failed to create draft.", "error")
    return redirect(url_for("import_upload"))


# ------------------------
# **NEW** Import landing page + HTML POST handler
# ------------------------
@app.route("/import", methods=["GET", "POST"], strict_slashes=False)
@login_required
@_require_tier_chosen
def import_upload():
    """
    Handles uploaded menu files (images or PDFs) and launches the OCR import job.

    GET  -> render the import upload page.
    POST -> save the file, launch OCR job, then redirect to Import Preview.
    """
    # Handle landing-page GET so /import from navbar doesn't 405
    if request.method == "GET":
        _raw_g = session.get("user")
        u_g = _raw_g if isinstance(_raw_g, dict) else {}
        tier = u_g.get("account_tier")
        # Day 133: fetch restaurant profile for cuisine/zip prompt
        rest_profile = None
        rest_id = u_g.get("restaurant_id")
        if rest_id and users_store:
            try:
                rest_profile = users_store.get_restaurant(int(rest_id))
            except Exception:
                pass
        # Day 133: "Add Menu" banner when coming from restaurant page
        for_rest_name = request.args.get("rest_name", "")
        for_rest_addr = request.args.get("rest_addr", "")
        for_rest_id = request.args.get("for_restaurant", "")
        return _safe_render("import.html", account_tier=tier, rest_profile=rest_profile,
                            for_rest_name=for_rest_name, for_rest_addr=for_rest_addr,
                            for_rest_id=for_rest_id)

    # POST: actual upload handler — OCR image upload requires premium tier
    _raw = session.get("user")
    u = _raw if isinstance(_raw, dict) else {}
    tier = u.get("account_tier")
    if u.get("role") != "admin" and u.get("user_id") and tier != "premium":
        flash("Photo/PDF upload requires the Premium Package.", "error")
        return redirect(url_for("import_upload"))

    try:
        file = request.files.get("file")
        if not file or file.filename == "":
            flash("Please choose a file to upload.", "error")
            return redirect(url_for("import_upload"))

        if not allowed_file(file.filename):
            flash("Unsupported file type. Allowed: JPG, JPEG, PNG, PDF.", "error")
            return redirect(url_for("import_upload"))

        base_name = secure_filename(file.filename) or "upload"
        tmp_name = f"{uuid.uuid4().hex[:8]}_{base_name}"
        save_path = UPLOAD_FOLDER / tmp_name
        file.save(str(save_path))

        restaurant_id = _resolve_restaurant_id_from_request()
        job_id = create_import_job(filename=tmp_name, restaurant_id=restaurant_id)

        # Run OCR asynchronously
        threading.Thread(
            target=run_ocr_and_make_draft, args=(job_id, save_path), daemon=True
        ).start()

        # Flash success with optional restaurant info
        if restaurant_id:
            flash(
                f"Import started for {base_name} (job #{job_id}) — linked to restaurant #{restaurant_id}.",
                "success",
            )
        else:
            flash(
                f"Import started for {base_name} (job #{job_id}). "
                "Tip: assign a restaurant on the import page before approving.",
                "success",
            )

        # ✅ Redirect straight to Import Preview instead of waiting for Draft Editor
        flash("Import complete — review preview below and rotate if needed.", "success")
        return redirect(url_for("imports_view", job_id=job_id))

    except RequestEntityTooLarge:
        flash("File too large. Try a smaller file or raise MAX_CONTENT_LENGTH.", "error")

    except Exception as e:
        flash(f"Server error while saving upload: {e}", "error")

    # On error, bounce back to the same import page
    return redirect(url_for("import_upload"))



# ------------------------
# NEW: Structured CSV import (Phase 6 pt.1)
# ------------------------
@app.post("/import/csv")
@login_required
def import_csv():
    """
    Structured ingestion for CSV menus (bypasses OCR).

    Expected form fields:
      - csv_file: uploaded .csv file
      - restaurant_id: optional (uses session restaurant for customers)

    Flow:
      - Save CSV into uploads/
      - Call storage.import_jobs.create_csv_import_job_from_file(...)
      - Create a DB-backed draft via drafts_store.create_draft_from_structured_items
      - Redirect to Draft Editor (if draft exists) or the import detail page.
    """
    _require_drafts_storage()

    # Ensure storage/import_jobs helpers are present
    if not hasattr(import_jobs_store, "create_csv_import_job_from_file"):
        flash("CSV import helpers are not available yet (storage.import_jobs.create_csv_import_job_from_file missing).", "error")
        return redirect(url_for("import_upload"))

    try:
        # Accept either 'csv_file' (preferred) or fallback to 'file'
        file = request.files.get("csv_file") or request.files.get("file")
        if not file or file.filename == "":
            flash("Please choose a CSV file to upload.", "error")
            return redirect(url_for("import_upload"))

        base_name = secure_filename(file.filename) or "upload.csv"
        if not base_name.lower().endswith(".csv"):
            flash("Structured CSV import currently only accepts .csv files.", "error")
            return redirect(url_for("import_upload"))

        tmp_name = f"{uuid.uuid4().hex[:8]}_{base_name}"
        save_path = UPLOAD_FOLDER / tmp_name
        file.save(str(save_path))

        restaurant_id = _resolve_restaurant_id_from_request()

        # Let the storage layer parse + validate the CSV
        result = import_jobs_store.create_csv_import_job_from_file(
            save_path,
            restaurant_id=restaurant_id,
        )

        job_id = int(result.get("job_id"))
        items = result.get("items") or []
        summary = result.get("summary") or {}
        errors = result.get("errors") or []

        # Build a draft from the structured items using the shared drafts storage
        title = summary.get("title") or f"Imported CSV {datetime.utcnow().date()}"
        create_structured = getattr(
            drafts_store, "create_draft_from_structured_items", None
        )
        draft_id = None
        if callable(create_structured) and items:
            draft = create_structured(
                title=title,
                restaurant_id=restaurant_id,
                items=items,
                source_type="structured_csv",
                # 🔑 link this draft back to import_jobs.id
                source_job_id=job_id,
                source_meta={
                    "filename": base_name,
                    "row_count": summary.get("row_count"),
                    "valid_rows": summary.get("valid_rows"),
                    "invalid_rows": summary.get("invalid_rows"),
                    "job_id": job_id,
                },
            )
            draft_id = int(draft.get("id") or draft.get("draft_id"))



        # Flash a concise summary
        row_count = summary.get("row_count", len(items))
        valid_rows = summary.get("valid_rows", len(items))
        invalid_rows = summary.get("invalid_rows", len(errors))
        msg = f"CSV import created job #{job_id}: {valid_rows} item(s) imported"
        if row_count is not None:
            msg += f" out of {row_count} row(s)"
        if invalid_rows:
            msg += f" ({invalid_rows} row(s) skipped)."
            level = "warning"
        else:
            msg += "."
            level = "success"
        flash(msg, level)

        if errors:
            # Keep the message short; detailed surfacing can be added in the template later.
            flash("Some CSV rows could not be imported. Check your CSV headers and formats.", "warning")

        # Day 141.6: All imports go through the wizard for review
        if draft_id:
            return redirect(url_for("draft_wizard", draft_id=draft_id))

        # Fallback: show the structured job detail
        return redirect(url_for("imports_detail", job_id=job_id))

    except RequestEntityTooLarge:
        flash("CSV file too large. Try a smaller file or raise MAX_CONTENT_LENGTH.", "error")
        return redirect(url_for("import_upload"))
    except Exception as e:
        flash(f"CSV import failed: {e}", "error")
        return redirect(url_for("import_upload"))




# ------------------------
# NEW: Structured XLSX import (Phase 6 pt.3)
# ------------------------
@app.post("/import/xlsx")
@login_required
def import_xlsx():
    """
    Structured ingestion for Excel menus (XLSX, bypasses OCR).

    Expected form fields:
      - xlsx_file: uploaded .xlsx file (preferred)
      - file:      fallback field name
      - restaurant_id: optional (uses session restaurant for customers)

    Flow:
      - Save XLSX into uploads/
      - Call storage.import_jobs.create_xlsx_import_job_from_file(...)
      - Create a DB-backed draft via drafts_store.create_draft_from_structured_items
      - Redirect to Draft Editor (if draft exists) or the import detail page.
    """
    _require_drafts_storage()

    # Ensure storage/import_jobs helpers are present
    if not hasattr(import_jobs_store, "create_xlsx_import_job_from_file"):
        flash(
            "XLSX import helpers are not available yet (storage.import_jobs.create_xlsx_import_job_from_file missing).",
            "error",
        )
        return redirect(url_for("import_upload"))

    try:
        # Accept either 'xlsx_file' (preferred) or fallback to 'file'
        file = request.files.get("xlsx_file") or request.files.get("file")
        if not file or file.filename == "":
            flash("Please choose an XLSX file to upload.", "error")
            return redirect(url_for("import_upload"))

        base_name = secure_filename(file.filename) or "upload.xlsx"
        if not base_name.lower().endswith(".xlsx"):
            flash("Structured Excel import currently only accepts .xlsx files.", "error")
            return redirect(url_for("import_upload"))

        tmp_name = f"{uuid.uuid4().hex[:8]}_{base_name}"
        save_path = UPLOAD_FOLDER / tmp_name
        file.save(str(save_path))

        restaurant_id = _resolve_restaurant_id_from_request()

        # Let the storage layer parse + validate the XLSX
        result = import_jobs_store.create_xlsx_import_job_from_file(
            save_path,
            restaurant_id=restaurant_id,
        )

        job_id = int(result.get("job_id"))
        items = result.get("items") or []
        summary = result.get("summary") or {}
        errors = result.get("errors") or []

        # Build a draft from the structured items using the shared drafts storage
        title = summary.get("title") or f"Imported XLSX {datetime.utcnow().date()}"
        create_structured = getattr(
            drafts_store, "create_draft_from_structured_items", None
        )
        draft_id = None
        if callable(create_structured) and items:
            draft = create_structured(
                title=title,
                restaurant_id=restaurant_id,
                items=items,
                source_type="structured_xlsx",
                # 🔗 link draft → import job using the canonical kwarg
                source_job_id=job_id,
                source_meta={
                    "filename": base_name,
                    "row_count": summary.get("row_count"),
                    "valid_rows": summary.get("valid_rows"),
                    "invalid_rows": summary.get("invalid_rows"),
                    "job_id": job_id,
                },
            )
            draft_id = int(draft.get("id") or draft.get("draft_id"))

        # Flash a concise summary
        row_count = summary.get("row_count", len(items))
        valid_rows = summary.get("valid_rows", len(items))
        invalid_rows = summary.get("invalid_rows", len(errors))
        msg = f"XLSX import created job #{job_id}: {valid_rows} item(s) imported"
        if row_count is not None:
            msg += f" out of {row_count} row(s)"
        if invalid_rows:
            msg += f" ({invalid_rows} row(s) skipped)."
            level = "warning"
        else:
            msg += "."
            level = "success"
        flash(msg, level)

        if errors:
            flash(
                "Some XLSX rows could not be imported. Check your column headers and formats.",
                "warning",
            )

        # Day 141.6: All imports go through the wizard for review
        if draft_id:
            return redirect(url_for("draft_wizard", draft_id=draft_id))

        # Fallback: show the structured job detail
        return redirect(url_for("imports_detail", job_id=job_id))

    except RequestEntityTooLarge:
        flash("XLSX file too large. Try a smaller file or raise MAX_CONTENT_LENGTH.", "error")
        return redirect(url_for("import_upload"))
    except Exception as e:
        flash(f"XLSX import failed: {e}", "error")
        return redirect(url_for("import_upload"))



# ------------------------
# NEW: Structured JSON import (Phase 6 pt.7)
# ------------------------
@app.post("/import/json")
@login_required
def import_json():
    """
    Structured ingestion for JSON menus (bypasses OCR).

    Expected form fields:
      - json_file: uploaded .json file (preferred)
      - file:      fallback field name
      - restaurant_id: optional (uses session restaurant for customers)

    Flow:
      - Save JSON into uploads/
      - Call storage.import_jobs.create_json_import_job_from_file(...)
      - Create a DB-backed draft via drafts_store.create_draft_from_structured_items
      - Redirect to Draft Editor (if draft exists) or the import detail page.
    """
    _require_drafts_storage()

    # Ensure storage/import_jobs helpers are present
    if not hasattr(import_jobs_store, "create_json_import_job_from_file"):
        flash(
            "JSON import helpers are not available yet (storage.import_jobs.create_json_import_job_from_file missing).",
            "error",
        )
        return redirect(url_for("import_upload"))

    try:
        # Accept either 'json_file' (preferred) or fallback to 'file'
        file = request.files.get("json_file") or request.files.get("file")
        if not file or file.filename == "":
            flash("Please choose a JSON file to upload.", "error")
            return redirect(url_for("import_upload"))

        base_name = secure_filename(file.filename) or "upload.json"
        if not base_name.lower().endswith(".json"):
            flash("Structured JSON import currently only accepts .json files.", "error")
            return redirect(url_for("import_upload"))

        tmp_name = f"{uuid.uuid4().hex[:8]}_{base_name}"
        save_path = UPLOAD_FOLDER / tmp_name
        file.save(str(save_path))

        restaurant_id = _resolve_restaurant_id_from_request()

        # Let the storage layer parse + validate the JSON
        result = import_jobs_store.create_json_import_job_from_file(
            save_path,
            restaurant_id=restaurant_id,
        )

        job_id = int(result.get("job_id"))
        items = result.get("items") or []
        summary = result.get("summary") or {}
        errors = result.get("errors") or []

        # Build a draft from the structured items using the shared drafts storage
        title = summary.get("title") or f"Imported JSON {datetime.utcnow().date()}"
        create_structured = getattr(drafts_store, "create_draft_from_structured_items", None)
        draft_id = None
        if callable(create_structured) and items:
            draft = create_structured(
                title=title,
                restaurant_id=restaurant_id,
                items=items,
                source_type="structured_json",
                # 🔑 link this draft back to import_jobs.id (matches CSV/XLSX behavior)
                source_job_id=job_id,
                source_meta={
                    "filename": base_name,
                    "row_count": summary.get("row_count"),
                    "valid_rows": summary.get("valid_rows"),
                    "invalid_rows": summary.get("invalid_rows"),
                    "job_id": job_id,
                },
            )

            raw_id = (draft.get("id") or draft.get("draft_id") or 0)
            draft_id = int(raw_id) if raw_id else None


        # Flash a concise summary
        row_count = summary.get("row_count", len(items))
        valid_rows = summary.get("valid_rows", len(items))
        invalid_rows = summary.get("invalid_rows", len(errors))
        msg = f"JSON import created job #{job_id}: {valid_rows} item(s) imported"
        if row_count is not None:
            msg += f" out of {row_count} row(s)"
        if invalid_rows:
            msg += f" ({invalid_rows} row(s) skipped)."
            level = "warning"
        else:
            msg += "."
            level = "success"
        flash(msg, level)

        if errors:
            flash("Some JSON rows could not be imported. Check your JSON structure and field names.", "warning")

        # Day 141.6: All imports go through the wizard for review
        if draft_id:
            return redirect(url_for("draft_wizard", draft_id=draft_id))

        # Fallback: show the structured job detail
        return redirect(url_for("imports_detail", job_id=job_id))

    except RequestEntityTooLarge:
        flash("JSON file too large. Try a smaller file or raise MAX_CONTENT_LENGTH.", "error")
        return redirect(url_for("import_upload"))
    except Exception as e:
        flash(f"JSON import failed: {e}", "error")
        return redirect(url_for("import_upload"))



# ------------------------
# Imports pages
# ------------------------
@app.get("/imports")
@login_required
def imports():
    jobs = list_import_jobs()
    return _safe_render("imports.html", jobs=jobs)

@app.get("/imports/<int:job_id>")
@login_required
def imports_detail(job_id):
    """Job detail page with actions."""
    row = get_import_job(job_id)
    if not row:
        abort(404)
    draft = None
    if row["draft_path"]:
        abs_path = _abs_from_rel(row["draft_path"])
        if abs_path and abs_path.exists():
            with open(abs_path, "r", encoding="utf-8") as f:
                draft = json.load(f)
    with db_connect() as conn:
        restaurants = conn.execute(
            "SELECT id, name FROM restaurants WHERE active=1 ORDER BY name"
        ).fetchall()

    # NEW: surface preview + rotate endpoints
    preview_url = url_for("imports_preview_image", job_id=job_id)
    rotate_url = url_for("imports_rotate_image", job_id=job_id)

    return _safe_render("import_view.html", job=row, draft=draft, restaurants=restaurants,
                        preview_img_url=preview_url, rotate_action_url=rotate_url)


# Day 136: Pipeline progress preview — cycles through stages without uploading
@app.get("/debug/pipeline-progress")
@login_required
def debug_pipeline_progress():
    """Preview the 5-stage pipeline progress screen. Cycles through stages automatically."""
    stage = request.args.get("stage", "extracting")
    return _safe_render("pipeline_preview.html", stage=stage)


# === NEW ===
# Visual OCR Blocks Debugger page (renders debug_blocks.html)
@app.get("/debug/blocks/<int:job_id>")
@login_required
def debug_blocks_page(job_id: int):
    """
    Render the overlay debugger for a given import job.
    Template expects:
      - preview_img_url: image to overlay boxes on
      - blocks_json_url: JSON feed with 'preview_blocks' / 'text_blocks'
      - rotate_action_url: rotate handler so users can fix orientation
      - back_url: link back to the import detail page
    """
    row = get_import_job(job_id)
    if not row:
        abort(404)

    # Best-effort: ensure a preview exists
    try:
        src = (UPLOAD_FOLDER / (row["filename"] or "")).resolve()
        if src.exists():
            _ensure_work_image(job_id, src)
    except Exception:
        pass

    return _safe_render(
        "debug_blocks.html",
        job=row,
        preview_img_url=url_for("imports_preview_image", job_id=job_id),
        blocks_json_url=url_for("imports_blocks", job_id=job_id),
        rotate_action_url=url_for("imports_rotate_image", job_id=job_id),
        back_url=url_for("imports_detail", job_id=job_id),
    )


@app.get("/imports/raw/<int:job_id>")
@login_required
def imports_raw(job_id: int):
    """
    DB-first "raw" view for an import job.
    Returns the draft payload synthesized from the drafts storage layer
    (the same source the Draft Editor uses), NOT the legacy draft_path file.
    """
    row = get_import_job(job_id)
    if not row:
        abort(404)

    _require_drafts_storage()
    draft_id = _get_or_create_draft_for_job(job_id)
    if not draft_id:
        return jsonify({"ok": False, "job_id": job_id, "message": "No draft available for this job yet"}), 200

    draft_meta = drafts_store.get_draft(draft_id) if hasattr(drafts_store, "get_draft") else {"id": draft_id}
    items = drafts_store.get_draft_items(draft_id) if hasattr(drafts_store, "get_draft_items") else []

    categories_map: Dict[str, List[Dict[str, Any]]] = {}
    for it in (items or []):
        cat = (it.get("category") or "Uncategorized").strip() or "Uncategorized"
        categories_map.setdefault(cat, []).append(
            {
                "id": it.get("id"),
                "name": it.get("name") or "",
                "description": it.get("description") or "",
                "price_cents": it.get("price_cents") or 0,
                "price": (float(it.get("price_cents") or 0) / 100.0) if it.get("price_cents") else 0.0,
                "confidence": it.get("confidence"),
                "position": it.get("position"),
            }
        )

    categories = [{"name": k, "items": v} for k, v in categories_map.items()]

    payload = {
        "job_id": int(job_id),
        "draft_id": int(draft_id),
        "source": (draft_meta.get("source") if isinstance(draft_meta, dict) else None) or {"type": "db"},
        "extracted_at": (draft_meta.get("created_at") if isinstance(draft_meta, dict) else None) or _now_iso(),
        "categories": categories,
    }
    return jsonify(payload)


# ---- NEW: Segmentation preview bridges (JSON) ----
def _load_debug_for_draft(draft_id: int) -> Dict[str, Any]:
    """Helper to fetch OCR debug payload saved by worker/helper."""
    _require_drafts_storage()
    load_fn = getattr(drafts_store, "load_ocr_debug", None)
    if not load_fn:
        return {}
    dbg = load_fn(draft_id) or {}
    if not isinstance(dbg, dict):
        return {}
    return dbg


def _load_layout_debug_for_draft(draft_id: int) -> Dict[str, Any]:
    """
    Phase 7 — layout/geometry debug payload.
    Expected keys (optional, experimental):
      - blocks
      - proto_sections
      - block_labels
      - geometry_stats
    """
    dbg = _load_debug_for_draft(draft_id)
    layout = dbg.get("layout_debug") or {}
    return layout if isinstance(layout, dict) else {}


@app.get("/drafts/<int:draft_id>/blocks")
@login_required
def drafts_blocks(draft_id: int):
    """
    Returns segmentation overlays for the draft editor:
    {
      preview_blocks: [ {bbox:[x1,y1,x2,y2], block_type, merged_text, lines:[...]}, ... ],
      text_blocks:    [ raw text-blocks if available ],
    }
    """
    dbg = _load_debug_for_draft(draft_id)

    preview_blocks = dbg.get("preview_blocks") or []
    if not isinstance(preview_blocks, list):
        preview_blocks = []

    text_blocks = dbg.get("text_blocks") or dbg.get("blocks") or []
    if not isinstance(text_blocks, list):
        text_blocks = []

    # Defensive: only allow dict blocks (prevents jsonify from choking on weird objects)
    preview_blocks = [b for b in preview_blocks if isinstance(b, dict)]
    text_blocks = [b for b in text_blocks if isinstance(b, dict)]

    return jsonify({
        "ok": True,
        "draft_id": draft_id,
        "preview_blocks": preview_blocks,
        "text_blocks": text_blocks,
    })


@app.get("/imports/<int:job_id>/blocks")
@login_required
def imports_blocks(job_id: int):
    """
    Convenience bridge: look up draft for import job and delegate to /drafts/<id>/blocks.
    """
    _require_drafts_storage()
    @app.get("/imports/<int:job_id>/blocks")
    @login_required
    def imports_blocks(job_id: int):
        """
        Read-only bridge: serve blocks only if a DB draft already exists.
        Must NOT create drafts or trigger OCR.
        """
        _require_drafts_storage()

        row = get_import_job(job_id)
        if not row:
            abort(404)

        data = dict(row)
        draft_id = data.get("draft_id") or data.get("draftId") or data.get("draft")
        try:
            draft_id = int(draft_id) if draft_id is not None else None
        except Exception:
            draft_id = None

        if not draft_id:
            return jsonify({"ok": False, "error": "Draft not ready"}), 404

        return drafts_blocks(draft_id)



# Bridge to Draft Editor (DB-first)
@app.get("/imports/<int:job_id>/draft")
@login_required
def imports_draft(job_id: int):
    """
    Bridge to Draft Editor.

    User intent is explicit here ("Open Draft Editor"), so it is OK to:
      - ensure a DB-backed draft exists for this import job, and
      - link it back onto import_jobs.draft_id when possible.

    This does NOT trigger OCR. It only ensures the editor has a draft.
    """
    row = get_import_job(job_id)
    if not row:
        abort(404)

    data = dict(row)

    # Prefer DB-backed draft id if present on the job row.
    draft_id = data.get("draft_id") or data.get("draftId") or data.get("draft")
    try:
        draft_id = int(draft_id) if draft_id is not None else None
    except Exception:
        draft_id = None

    if draft_id:
        # Day 137: Route to wizard for first-time review, editor if already completed
        try:
            d = drafts_store.get_draft(int(draft_id))
            if d and not d.get("wizard_completed"):
                return redirect(url_for("draft_wizard", draft_id=int(draft_id)))
        except Exception:
            pass
        return redirect(url_for("draft_editor", draft_id=draft_id))

    # NEW: On-demand DB draft creation/linking (no OCR)
    try:
        draft_id = _ensure_draft_for_job(job_id, row=row)
    except Exception:
        draft_id = None

    if draft_id:
        # Day 137: Route to wizard for first-time review
        try:
            d = drafts_store.get_draft(int(draft_id))
            if d and not d.get("wizard_completed"):
                return redirect(url_for("draft_wizard", draft_id=int(draft_id)))
        except Exception:
            pass
        return redirect(url_for("draft_editor", draft_id=int(draft_id)))

    # Fallback: legacy file draft (still read-only)
    abs_draft = _abs_from_rel(data.get("draft_path")) if data.get("draft_path") else None
    if abs_draft and abs_draft.exists():
        flash("Legacy draft file is ready, but no DB draft id is linked yet.", "info")
        return redirect(url_for("imports_detail", job_id=job_id))

    flash("Draft not ready yet for the editor. Try Clone Draft, or re-run Finalize.", "error")
    return redirect(url_for("imports_detail", job_id=job_id))




@app.post("/imports/<int:job_id>/set_restaurant")
@login_required
def imports_set_restaurant(job_id: int):
    rid = request.form.get("restaurant_id")
    if not rid:
        flash("Please choose a restaurant.", "error")
        return redirect(url_for("imports_detail", job_id=job_id))
    try:
        restaurant_id = int(rid)
    except Exception:
        flash("Invalid restaurant id.", "error")
        return redirect(url_for("imports_detail", job_id=job_id))

    try:
        update_import_job(job_id, restaurant_id=restaurant_id)
        try:
            _require_drafts_storage()
            draft_id = _get_or_create_draft_for_job(job_id)
            if draft_id:
                drafts_store.save_draft_metadata(draft_id, restaurant_id=restaurant_id)
        except Exception:
            pass
        flash("Linked import to restaurant.", "success")
    except Exception as e:
        flash(f"Failed to link restaurant: {e}", "error")
    return redirect(url_for("imports_detail", job_id=job_id))

@app.post("/imports/<int:job_id>/approve")
@login_required
def imports_approve(job_id: int):
    try:
        menu_id, inserted = approve_draft_to_menu(job_id)
        flash(f"Approved: inserted {inserted} item(s) into menu #{menu_id}.", "success")
        return redirect(url_for("items_page", menu_id=menu_id))
    except Exception as e:
        flash(f"Approve failed: {e}", "error")
        return redirect(url_for("imports_detail", job_id=job_id))

@app.post("/imports/<int:job_id>/discard")
@login_required
def imports_discard(job_id: int):
    try:
        deleted = discard_draft_for_job(job_id)
        flash(f"Discarded draft items ({deleted} removed).", "success")
    except Exception as e:
        flash(f"Discard failed: {e}", "error")
    return redirect(url_for("imports_detail", job_id=job_id))

@app.post("/imports/<int:job_id>/clone")
@login_required
def imports_clone(job_id: int):
    try:
        _require_drafts_storage()
        draft_id = _get_or_create_draft_for_job(job_id)
        if not draft_id:
            flash("No draft available to clone for this import.", "error")
            return redirect(url_for("imports_detail", job_id=job_id))
        if hasattr(drafts_store, "clone_draft"):
            clone = drafts_store.clone_draft(draft_id)
            new_id = int(clone.get("id") or clone.get("draft_id"))
            flash(f"Cloned draft #{draft_id} → #{new_id}.", "success")
            return redirect(url_for("draft_editor", draft_id=new_id))
        else:
            flash("Clone operation is not supported by the drafts storage layer.", "error")
            return redirect(url_for("draft_editor", draft_id=draft_id))
    except Exception as e:
        flash(f"Clone failed: {e}", "error")
        return redirect(url_for("imports_detail", job_id=job_id))

@app.get("/imports/view/<int:job_id>")
@login_required
def imports_view(job_id):
    # Pylance-safe redirect instead of direct function reference
    return redirect(url_for("imports_detail", job_id=job_id))

@app.route("/imports/cleanup", methods=["GET", "POST"])
@login_required
def imports_cleanup():
    with db_connect() as conn:
        rows = conn.execute("SELECT id, filename, COALESCE(status,'') AS st FROM import_jobs").fetchall()
        to_delete = [
            r["id"]
            for r in rows
            if not (UPLOAD_FOLDER / r["filename"]).exists() and r["st"] != "deleted"
        ]
        if to_delete:
            conn.executemany(
                "UPDATE import_jobs SET status='deleted', updated_at=datetime('now') WHERE id=?",
                [(jid,) for jid in to_delete],
            )
            conn.commit()
    flash("Imports cleanup completed.", "success")
    return redirect(url_for("imports"))

@app.post("/imports/<int:job_id>/delete")
@login_required
def imports_delete_job(job_id):
    with db_connect() as conn:
        conn.execute(
            "UPDATE import_jobs SET status='deleted', updated_at=datetime('now') WHERE id=?",
            (job_id,),
        )
        conn.commit()
    flash(f"Job #{job_id} moved to deleted.", "success")
    return redirect(url_for("imports"))


# ------------------------------------------------------------
# Phase 6 pt.1 — Structured CSV import → Drafts
# ------------------------------------------------------------

def _canonical_field_for_header(header: str) -> Optional[str]:
    """
    Map a CSV header to a canonical field name using storage.import_jobs
    HEADER_ALIASES / CANONICAL_FIELDS if available.

    Returns one of:
      - "name", "description", "category", "subcategory",
        "price", "price_cents", "size", "sku"
      - or None if we don't recognize the header.
    """
    h = (header or "").strip().lower()
    if not h:
        return None

    aliases = getattr(import_jobs_store, "HEADER_ALIASES", {}) or {}
    for field, names in aliases.items():
        try:
            if h in {n.lower() for n in names}:
                return field
        except Exception:
            continue

    canon_fields = set(getattr(import_jobs_store, "CANONICAL_FIELDS", []) or [])
    if h in canon_fields:
        return h

    return None


def _price_to_cents_loose(value: Any) -> int:
    """
    Loose parser for price columns in CSV (dollars → cents).

    Accepts things like:
      - "12.99"
      - "$12.99"
      - " 12 "
    Returns 0 on failure and clamps negatives to 0.
    """
    if value is None:
        return 0
    try:
        txt = str(value).strip().replace("$", "")
        cents = int(round(float(txt) * 100))
    except Exception:
        return 0
    return cents if cents > 0 else 0


def _csv_to_structured_items(file_storage) -> List[Dict[str, Any]]:
    """
    Read an uploaded CSV file and normalize rows into structured menu items
    suitable for drafts.create_draft_from_structured_items(...).

    Supported canonical fields (columns can use any alias defined in
    storage.import_jobs.HEADER_ALIASES):

      name (required)
      description
      category
      subcategory
      price          (dollars; we convert via _price_to_cents_loose)
      price_cents    (integer cents)
      size
      sku
    """
    raw = file_storage.read()
    # Basic charset handling; utf-8-sig for common BOM'd exports, else latin-1 fallback
    try:
        text = raw.decode("utf-8-sig")
    except Exception:
        text = raw.decode("latin-1", errors="ignore")

    f = io.StringIO(text)
    reader = csv.DictReader(f)

    if not reader.fieldnames:
        return []

    # Build header → canonical mapping once
    header_map: Dict[str, Optional[str]] = {}
    for col in reader.fieldnames:
        header_map[col] = _canonical_field_for_header(col)

    items: List[Dict[str, Any]] = []

    for row in reader:
        if not isinstance(row, dict):
            continue

        normalized: Dict[str, Any] = {}

        for orig_col, value in row.items():
            canon = header_map.get(orig_col)
            if not canon:
                continue
            normalized[canon] = value

        name = (normalized.get("name") or "").strip()
        if not name:
            # No name → skip this row entirely
            continue

        description = (normalized.get("description") or "").strip()

        subcat = (normalized.get("subcategory") or "").strip() or None
        cat = (normalized.get("category") or "").strip() or None
        category = subcat or cat

        # Price handling: prefer explicit cents column, else parse dollars
        price_cents: int = 0
        if normalized.get("price_cents") not in (None, ""):
            try:
                price_cents = int(str(normalized["price_cents"]).strip())
            except Exception:
                price_cents = 0
        elif normalized.get("price") not in (None, ""):
            price_cents = _price_to_cents_loose(normalized.get("price"))

        # Never allow negative
        if price_cents < 0:
            price_cents = 0

        item: Dict[str, Any] = {
            "name": name,
            "description": description,
            "category": category,
            "subcategory": subcat,
            "price_cents": price_cents,
        }

        # Optional extras we might care about later (ignored by drafts for now)
        size = (normalized.get("size") or "").strip()
        if size:
            item["size_name"] = size

        sku = (normalized.get("sku") or "").strip()
        if sku:
            item["sku"] = sku

        items.append(item)

    return items


@app.post("/api/drafts/import_structured")
@login_required
def import_structured_draft():
    """
    Phase 6 pt.2 — Structured CSV import route (One Brain-backed).

    Flow:
      1) Accept CSV upload via multipart/form-data.
      2) Save CSV to disk (under UPLOAD_FOLDER).
      3) Use storage.import_jobs.create_csv_import_job_from_file(...) to:
         - parse + validate rows via One Brain contracts
         - create an import_jobs row (source_type=structured_csv)
      4) Create a DB-backed draft via storage.drafts.create_draft_from_structured_items.
      5) Return JSON with job_id, draft_id, summary, and redirect_url.

    Request (multipart/form-data):
      - file: CSV file
      - restaurant_id: optional (used to pre-link the draft)
      - title: optional draft title

    Response (JSON on success):
      {
        "ok": true,
        "job_id": 42,
        "draft_id": 123,
        "summary": {...},
        "errors": [...],
        "redirect_url": "/drafts/123/edit"
      }
    """
    try:
        file = request.files.get("file")
        if not file or file.filename == "":
            return jsonify({"ok": False, "error": "No file uploaded."}), 400

        if not file.filename.lower().endswith(".csv"):
            return jsonify(
                {
                    "ok": False,
                    "error": "Only CSV structured imports are supported right now.",
                }
            ), 400

        # Ensure drafts storage is available and supports structured creation
        _require_drafts_storage()
        create_fn = getattr(drafts_store, "create_draft_from_structured_items", None)
        if not callable(create_fn):
            return jsonify(
                {
                    "ok": False,
                    "error": "Structured draft creation is not available in this environment.",
                }
            ), 500

        # Resolve restaurant + title
        restaurant_id = _resolve_restaurant_id_from_request()
        title = (
            (request.form.get("title") or "").strip()
            or f"Structured Import {datetime.utcnow().date()}"
        )

        # --- Phase 6 pt.2: save CSV to disk and create a One Brain import_job ---
        safe_name = secure_filename(file.filename) or "structured.csv"
        unique_name = f"{uuid.uuid4().hex[:8]}_{safe_name}"
        csv_path = (UPLOAD_FOLDER / unique_name).resolve()
        csv_path.parent.mkdir(parents=True, exist_ok=True)
        file.save(str(csv_path))

        # Use the One Brain helper to parse + create import_jobs row
        job_result = import_jobs_store.create_csv_import_job_from_file(
            csv_path, restaurant_id=restaurant_id
        )


        job_id = int(job_result.get("job_id") or 0)
        items = job_result.get("items") or []
        errors = job_result.get("errors") or []
        summary = job_result.get("summary") or {}
        job_summary = job_result.get("job_summary") or {}

        if job_id <= 0:
            return jsonify(
                {"ok": False, "error": "Failed to create structured import job."}
            ), 500

        if not items:
            # We DID create a job row, but there were no valid items.
            return jsonify(
                {
                    "ok": False,
                    "job_id": job_id,
                    "error": "CSV parsed but produced no valid structured items.",
                    "summary": summary,
                    "errors": errors,
                }
            ), 400

        # --- Create the DB-backed draft from structured items ---
        source_meta = {
            "filename": file.filename,
            "csv_path": str(csv_path),
            "job_id": job_id,
            "summary": job_summary,
        }

        draft = create_fn(
            title=title,
            restaurant_id=restaurant_id,
            items=items,
            source_type="structured_csv",
            # 🔑 link draft back to this import job
            source_job_id=job_id,
            source_meta=source_meta,
        )
        draft_id = int(draft.get("id") or draft.get("draft_id") or 0)

        if not draft_id:
            return jsonify(
                {
                    "ok": False,
                    "job_id": job_id,
                    "error": "Structured draft creation did not return a draft id.",
                }
            ), 500

        return jsonify(
            {
                "ok": True,
                "job_id": job_id,
                "draft_id": draft_id,
                "summary": summary,
                "errors": errors,
                "redirect_url": url_for("draft_wizard", draft_id=draft_id),
            }
        ), 200

    except Exception as e:
        # Log for dev; keep JSON response stable for the UI
        app.logger.exception("Structured import failed")
        return jsonify({"ok": False, "error": f"Structured import failed: {e}"}), 500



# ------------------------
# Serving uploads (secure; block .trash)
# ------------------------
@app.get("/uploads/<path:filename>")
@login_required
def serve_upload(filename):
    requested = (UPLOAD_FOLDER / filename).resolve()

    # Strong containment check: requested must be inside UPLOAD_FOLDER
    try:
        requested.relative_to(UPLOAD_FOLDER.resolve())
    except Exception:
        abort(403)

    # Block anything inside .trash
    try:
        requested.relative_to(TRASH_FOLDER.resolve())
        abort(403)
    except Exception:
        pass

    return send_from_directory(str(UPLOAD_FOLDER), str(requested.relative_to(UPLOAD_FOLDER.resolve())), as_attachment=False)



# ------------------------
# Upload Management (Recycle Bin) + Artifact cleanup
# ------------------------
def _safe_in_uploads(path: Path) -> bool:
    try:
        path.resolve().relative_to(UPLOAD_FOLDER.resolve())
        return True
    except Exception:
        return False


def _is_direct_child_file(p: Path) -> bool:
    return p.parent.resolve() == UPLOAD_FOLDER.resolve() and p.is_file()

def _list_uploads_files():
    files = []
    for p in UPLOAD_FOLDER.iterdir():
        if p.name == ".trash":
            continue
        if p.is_file():
            stat = p.stat()
            files.append({
                "name": p.name,
                "size": stat.st_size,
                "modified": datetime.fromtimestamp(stat.st_mtime),
            })
    files.sort(key=lambda x: x["modified"], reverse=True)
    return files

def _list_trash_files():
    results = []
    try:
        trash_root = TRASH_FOLDER.resolve()
        if not trash_root.exists():
            return results

        for root, _, filenames in os.walk(trash_root, topdown=True):
            try:
                root_path = Path(root).resolve()
            except Exception:
                continue

            for fname in filenames:
                try:
                    p = (root_path / fname).resolve()
                    if not str(p).startswith(str(trash_root)):
                        continue
                    try:
                        st = p.stat()
                        mtime = datetime.fromtimestamp(st.st_mtime)
                        size = st.st_size
                    except FileNotFoundError:
                        continue
                    except Exception:
                        continue

                    try:
                        rel = p.relative_to(trash_root)
                    except Exception:
                        rel = p.name

                    results.append({
                        "trash_path": str(rel).replace("\\", "/"),
                        "name": p.name,
                        "size": size,
                        "modified": mtime,
                        "modified_iso": mtime.isoformat(timespec="seconds"),
                    })
                except Exception:
                    continue
    except Exception:
        return []

    results.sort(key=lambda x: x.get("modified", datetime.min), reverse=True)
    return results

def _batch_trash_dir() -> Path:
    return TRASH_FOLDER / datetime.utcnow().strftime("%Y%m%d_%H%M%S")

def _mark_jobs_for_upload(upload_name: str, new_status: str):
    with db_connect() as conn:
        conn.execute(
            "UPDATE import_jobs SET status=?, updated_at=datetime('now') WHERE filename=?",
            (new_status, upload_name),
        )
        conn.commit()

def _iter_draft_json_files():
    for p in DRAFTS_FOLDER.iterdir():
        if p.name in (".trash", "raw"):
            continue
        if p.is_file() and p.suffix.lower() == ".json":
            yield p

def _trash_draft_file(p: Path, ts: str):
    dest_dir = TRASH_DRAFTS / ts
    dest_dir.mkdir(parents=True, exist_ok=True)
    try:
        shutil.move(str(p), str(dest_dir / p.name))
    except Exception:
        pass

def _trash_job_artifacts_for_upload(upload_name: str):
    ts = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
    u = upload_name.lower()
    for p in _iter_draft_json_files():
        try:
            data = json.loads(p.read_text(encoding="utf-8"))
            src_file = (((data or {}).get("source") or {}).get("file") or "").strip()
            if Path(src_file).name.lower() == u:
                _trash_draft_file(p, ts)
        except Exception:
            continue
    def _sweep_raw_dir(src_root: Path, trash_root: Path):
        if not src_root.exists():
            return
        dest_dir = trash_root / ts
        dest_dir.mkdir(parents=True, exist_ok=True)
        for p in list(src_root.iterdir()):
            if p.name == ".trash":
                continue
            try:
                name_lower = p.name.lower()
                if u in name_lower:
                    shutil.move(str(p), str(dest_dir / p.name))
            except Exception:
                continue
    _sweep_raw_dir(RAW_FOLDER, TRASH_RAW)
    _sweep_raw_dir(LEGACY_RAW_FOLDER, LEGACY_TRASH_RAW)

def _sweep_all_raw_to_trash() -> int:
    ts = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
    total = 0
    for src_root, trash_root in [(RAW_FOLDER, TRASH_RAW), (LEGACY_RAW_FOLDER, LEGACY_TRASH_RAW)]:
        if not src_root.exists():
            continue
        dest_dir = trash_root / ts
        dest_dir.mkdir(parents=True, exist_ok=True)
        for p in list(src_root.iterdir()):
            if p.name == ".trash":
                continue
            try:
                shutil.move(str(p), str(dest_dir / p.name))
                total += 1
            except Exception:
                continue
    return total

def _sweep_all_drafts_to_trash() -> int:
    ts = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
    total = 0
    for p in list(_iter_draft_json_files()):
        try:
            _trash_draft_file(p, ts)
            total += 1
        except Exception:
            continue
    return total

def _move_to_trash(names: Iterable[str]) -> Iterable[str]:
    """Move given upload filenames into /uploads/.trash/<batch>/ and mark related jobs deleted."""
    batch_dir = _batch_trash_dir()
    batch_dir.mkdir(parents=True, exist_ok=True)

    moved: List[str] = []
    for raw in names:
        name = secure_filename(Path(raw).name)
        if not name:
            continue
        src = (UPLOAD_FOLDER / name).resolve()
        if not _safe_in_uploads(src) or not _is_direct_child_file(src) or not src.exists():
            continue
        dest = batch_dir / name
        try:
            shutil.move(str(src), str(dest))
            moved.append(name)
            _mark_jobs_for_upload(name, "deleted")
            _trash_job_artifacts_for_upload(name)
        except Exception:
            continue
    return moved

def _restore_from_trash(trash_paths: Iterable[str]) -> Iterable[Tuple[str, str]]:
    """Restore files from /uploads/.trash by relative trash paths and mark jobs restored."""
    restored: List[Tuple[str, str]] = []
    for rel in trash_paths:
        rel_path = (TRASH_FOLDER / rel).resolve()
        if not str(rel_path).startswith(str(TRASH_FOLDER.resolve())):
            continue
        if not rel_path.exists() or not rel_path.is_file():
            continue

        original_name = rel_path.name
        dest = UPLOAD_FOLDER / original_name
        if dest.exists():
            base = dest.stem
            ext = dest.suffix
            idx = 1
            while True:
                candidate = UPLOAD_FOLDER / f"{base} (restored {idx}){ext}"
                if not candidate.exists():
                    dest = candidate
                    break
                idx += 1
        try:
            shutil.move(str(rel_path), str(dest))
            _mark_jobs_for_upload(original_name, "restored")
            restored.append((original_name, dest.name))
        except Exception:
            continue
    return restored

def _empty_dir_tree(path: Path) -> int:
    deleted = 0
    if not path.exists():
        return 0
    for child in list(path.iterdir()):
        try:
            if child.is_file():
                child.unlink()
                deleted += 1
            elif child.is_dir():
                count = 0
                for _, _, files in os.walk(child):
                    count += len(files)
                shutil.rmtree(child, ignore_errors=True)
                deleted += count
        except Exception:
            continue
    return deleted

def _sweep_artifacts() -> dict:
    moved = []
    def _move(pattern: str):
        for p in ROOT.glob(f"storage/{pattern}"):
            if p.is_file():
                dest = DRAFTS_FOLDER / p.name
                try:
                    shutil.move(str(p), str(dest))
                    moved.append(str(dest.relative_to(ROOT)).replace("\\", "/"))
                except Exception:
                    pass
    _move("*.jsonl")
    _move("*.tmp")
    _move("*.raw.json")
    _move("*.ocr.txt")
    def _sweep_raw_dir(src_root: Path, trash_root: Path):
        if not src_root.exists():
            return
        ts = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
        dest_dir = trash_root / ts
        dest_dir.mkdir(parents=True, exist_ok=True)
        for p in list(src_root.iterdir()):
            if p.name == ".trash":
                continue
            try:
                shutil.move(str(p), str(dest_dir / p.name))
                moved.append(str((dest_dir / p.name).relative_to(ROOT)).replace("\\", "/"))
            except Exception:
                pass
    _sweep_raw_dir(RAW_FOLDER, TRASH_RAW)
    _sweep_raw_dir(LEGACY_RAW_FOLDER, LEGACY_TRASH_RAW)
    return {"moved": moved, "count": len(moved)}

@app.get("/uploads")
@login_required
def uploads_page():
    files = _list_uploads_files()
    return _safe_render("uploads.html", files=files)

@app.post("/uploads/delete")
@login_required
def uploads_delete():
    names = request.form.getlist("names")
    if not names:
        abort(400, description="No files selected")
    moved = list(_move_to_trash(names))
    flash(f"Moved {len(moved)} file(s) to Recycle Bin.", "success")
    return redirect(url_for("uploads_page"))


@app.get("/uploads/trash")
@login_required
def uploads_trash_page():
    err_note = ""
    trashed = []
    try:
        trashed = _list_trash_files()
        if not isinstance(trashed, list):
            trashed = []
    except Exception as e:
        err_note = f"Note: failed to enumerate recycle bin ({e.__class__.__name__}). Showing empty list."
        trashed = []
    return _safe_render("uploads_trash.html", trashed=trashed, err_note=err_note)

@app.post("/uploads/restore")
@login_required
def uploads_restore():
    paths = request.form.getlist("trash_paths")
    if not paths:
        abort(400, description="No trash items selected")
    restored = list(_restore_from_trash(paths))
    flash(f"Restored {len(restored)} file(s).", "success")
    return redirect(url_for("uploads_trash_page"))

@app.post("/uploads/empty_trash")
@login_required
def uploads_empty_trash():
    total = 0
    for p in (TRASH_FOLDER, TRASH_DRAFTS, TRASH_RAW, LEGACY_TRASH_RAW):
        total += _empty_dir_tree(p)
    flash(f"Permanently removed {total} file(s) from trash.", "success")
    return redirect(url_for("uploads_trash_page"))

@app.post("/uploads/clean_raw")
@login_required
def uploads_clean_raw():
    count = _sweep_all_raw_to_trash()
    flash(f"Moved {count} raw artifact file(s) to trash.", "success")
    return redirect(url_for("uploads_trash_page"))

@app.post("/uploads/clean_drafts")
@login_required
def uploads_clean_drafts():
    count = _sweep_all_drafts_to_trash()
    flash(f"Moved {count} draft file(s) to trash.", "success")
    return redirect(url_for("uploads_trash_page"))

@app.post("/admin/artifacts/sweep")
@login_required
def artifacts_sweep():
    report = _sweep_artifacts()
    return jsonify({"status": "ok", **report})

# ------------------------
# Draft Review (legacy JSON-file flow) & Publish (kept)
# ------------------------
def _load_draft_json_by_job(job_id: int):
    row = get_import_job(job_id)
    if not row or not row["draft_path"]:
        abort(404, description="Draft not found")
    abs_path = _abs_from_rel(row["draft_path"])
    if not abs_path or not abs_path.exists():
        abort(404, description="Draft file missing on disk")
    with open(abs_path, "r", encoding="utf-8") as f:
        return json.load(f)

def _is_image(name: str) -> bool:
    n = (name or "").lower()
    return n.endswith(".png") or n.endswith(".jpg") or n.endswith(".jpeg")

@app.get("/drafts/<int:job_id>")
@login_required
def draft_review_page(job_id: int):
    draft = _load_draft_json_by_job(job_id)
    with db_connect() as conn:
        restaurants = conn.execute(
            "SELECT id, name FROM restaurants WHERE active=1 ORDER BY name"
        ).fetchall()
    src_file = (draft.get("source", {}) or {}).get("file")
    preview_url = url_for("serve_upload", filename=src_file) if src_file and _is_image(src_file) else None
    return _safe_render("draft_review.html", draft=draft, restaurants=restaurants, preview_url=preview_url)

@app.post("/drafts/<int:job_id>/publish")
@login_required
def publish_draft(job_id: int):
    draft = _load_draft_json_by_job(job_id)
    restaurant_id = request.form.get("restaurant_id")
    menu_name = (request.form.get("menu_name") or "").strip() or f"Imported {datetime.utcnow().date()}"

    if not restaurant_id:
        abort(400, description="restaurant_id is required")

    with db_connect() as conn:
        cur = conn.cursor()
        cur.execute(
            "INSERT INTO menus (restaurant_id, name, active) VALUES (?, ?, 1)",
            (int(restaurant_id), menu_name),
        )
        menu_id = cur.lastrowid

        for cat in draft.get("categories", []):
            for item in (cat.get("items") or []):
                base_name = (item.get("name") or "").strip() or "Untitled"
                desc = (item.get("description") or "").strip()
                sizes = item.get("sizes") or []

                if sizes:
                    for s in sizes:
                        size_name = (s.get("name") or "").strip()
                        price_val = s.get("price", 0)
                        try:
                            price_cents = int(round(float(price_val) * 100))
                        except Exception:
                            price_cents = 0
                        display_name = f"{base_name} ({size_name})" if size_name else base_name
                        cur.execute(
                            "INSERT INTO menu_items (menu_id, name, description, price_cents, is_available) VALUES (?, ?, ?, ?, 1)",
                            (menu_id, display_name, desc, price_cents),
                        )
                else:
                    price_val = item.get("price", 0)
                    try:
                        price_cents = int(round(float(price_val) * 100))
                    except Exception:
                        price_cents = 0
                    cur.execute(
                        "INSERT INTO menu_items (menu_id, name, description, price_cents, is_available) VALUES (?, ?, ?, ?, 1)",
                        (menu_id, base_name, desc, price_cents),
                    )
        conn.commit()

    try:
        update_import_job(job_id, status="published")
    except Exception:
        pass

    flash(f"Published draft #{job_id} to menu #{menu_id}.", "success")
    return redirect(url_for("items_page", menu_id=menu_id))

# ======================================================================
# Drafts (DB-first): List + Editor + Save + Submit
# ======================================================================

QUALITY_LOW_THRESHOLD = 65  # items below this are considered "low confidence"

def _compute_item_quality(it: dict) -> tuple[int, bool]:
    """
    Compute a 0–100 quality score for a draft item based on:
      - OCR confidence (0–100)
      - Valid price vs missing/zero
      - Category presence/quality
      - Name length sanity
      - Junk-symbol density in the name

    Returns (score, is_low_confidence).
    """
    name = (it.get("name") or "").strip()
    desc = (it.get("description") or "").strip()
    cat = (it.get("category") or "").strip()
    price_cents = it.get("price_cents")
    try:
        conf = int(it.get("confidence")) if it.get("confidence") is not None else None
    except Exception:
        conf = None

    # ---------- Base score ----------
    score = 100

    # --- Confidence component ---
    if conf is None:
        score -= 10
    else:
        if conf < 30:
            score -= 35
        elif conf < 50:
            score -= 25
        elif conf < 70:
            score -= 15
        elif conf < 85:
            score -= 5
        # 85+ → no penalty

    # --- Price component ---
    if price_cents is None:
        # Try to infer from loose fields if present
        from_price = 0
        for key in ("price", "price_text"):
            if it.get(key) is not None:
                try:
                    from_price = int(round(float(str(it[key]).replace("$", "").strip()) * 100))
                    break
                except Exception:
                    continue
        price_cents = from_price

    if not price_cents or price_cents <= 0:
        score -= 20

    # --- Category component ---
    if not cat:
        score -= 15
    else:
        cl = cat.lower()
        if cl in ("uncategorized", "misc", "other"):
            score -= 8

    # --- Name length sanity ---
    nlen = len(name)
    if nlen == 0:
        score -= 40
    elif nlen < 3:
        score -= 25
    elif nlen < 8:
        score -= 5
    elif nlen > 120:
        score -= 30
    elif nlen > 80:
        score -= 15

    # --- Junk-symbol density in name ---
    if name:
        bad_chars = 0
        for ch in name:
            if not (ch.isalnum() or ch.isspace() or ch in "$.,&()/+'-"):
                bad_chars += 1
        junk_ratio = bad_chars / max(len(name), 1)
        if junk_ratio > 0.40:
            score -= 25
        elif junk_ratio > 0.25:
            score -= 15
        elif junk_ratio > 0.15:
            score -= 5

    # Clamp
    if score < 0:
        score = 0
    if score > 100:
        score = 100

    is_low = score < QUALITY_LOW_THRESHOLD
    return int(score), bool(is_low)


@app.get("/drafts")
@login_required
def drafts_list():
    """List drafts (optionally filter by status or restaurant)."""
    _require_drafts_storage()
    status = request.args.get("status") or None
    try:
        restaurant_id = int(request.args.get("restaurant_id")) if request.args.get("restaurant_id") else None
    except Exception:
        restaurant_id = None

    drafts = drafts_store.list_drafts(status=status, restaurant_id=restaurant_id, limit=200, offset=0)
    return _safe_render("drafts.html", drafts=drafts, status=status, restaurant_id=restaurant_id)

def _compute_editor_stats(items: list) -> dict:
    """Day 122: Compute summary stats for the editor stats bar."""
    total = len(items)
    cats = set()
    with_mg = 0
    prices = []
    for it in items:
        cat = (it.get("category") or "").strip()
        if cat:
            cats.add(cat)
        if it.get("modifier_groups"):
            with_mg += 1
        pc = it.get("price_cents")
        if pc and int(pc) > 0:
            prices.append(int(pc))
    mg_pct = round(with_mg / total * 100) if total else 0
    return {
        "item_count": total,
        "category_count": len(cats),
        "mg_coverage_pct": mg_pct,
        "price_min": min(prices) if prices else 0,
        "price_max": max(prices) if prices else 0,
    }


# Restaurant-name sanity check. Gemini's Google-Search-grounded pricing
# path pulls snippets from article titles and list pages, which show up as
# `price_sources[].restaurant = "Best Pizza in Cape May Court House | My
# Pizza Heaven"`. When we hand those to Google Places `findplacefromtext`
# we get matches to random pizzerias hundreds of miles away.
#
# The filter is intentionally conservative — real restaurant names CAN be
# long and CAN start with "The", so we only flag the unmistakable article
# markers. Prefer false negatives (junk slips through, caught by the 25km
# distance filter downstream) over false positives (dropping a legit place).
_BAD_NAME_MARKERS = (
    " | ",             # "Name | Website" / "Restaurant | City" title separator
    " · ",             # middle-dot variant
    "...",             # truncated snippet
    " near me",
    " near you",
    " near here",
    "best pizza in ",
    "best pizza of ",
    "best pizzas in ",
    "best restaurants in ",
    "best restaurants of ",
    "top restaurants in ",
    "places to eat in ",
    "where to eat in ",
)
_BAD_NAME_PREFIXES = (
    "best ", "top ", "the best ", "the top ",
    "10 ", "5 ", "7 ", "12 ", "15 ", "20 ", "25 ",
)


def _looks_like_article_title(name: str) -> bool:
    """Return True when `name` is probably a Google Search article title or
    list header rather than a real restaurant name."""
    if not name:
        return True
    n = name.strip()
    if len(n) > 80:
        return True
    low = n.lower()
    if any(m in low for m in _BAD_NAME_MARKERS):
        return True
    if any(low.startswith(p) for p in _BAD_NAME_PREFIXES):
        return True
    return False


def _build_editor_competitors(draft, price_intel, restaurant):
    """Build the full competitor list shown on the editor map panel.

    Combines Google Places cache (from Day 134's nearby search) with
    Gemini-sourced restaurants scraped out of `price_sources`, enriching
    unseen names via up to 30 parallel Places `findplacefromtext` calls.

    Expensive on cold builds: the Places enrichment burst can cost ~5-10s
    even parallelized. Callers should wrap this in the _EDITOR_COMPETITORS_CACHE
    so pure reloads skip all of it. The cache is explicitly invalidated at
    every run-start via _invalidate_editor_cache.
    """
    competitors = []
    if not draft.get("restaurant_id"):
        return competitors

    # Determine if Gemini produced real cites for this draft. The 20
    # Places-nearby list is Haiku reference data — we only want it in
    # the editor sidebar when Haiku was actually used. When Gemini
    # delivered real cites, the sidebar should show JUST the Gemini-
    # sourced restaurants (the ones our pricing actually came from).
    _gemini_real_count = 0
    try:
        if price_intel and price_intel.get("assessments"):
            for a in price_intel["assessments"]:
                ps = a.get("price_sources")
                if isinstance(ps, str):
                    try: ps = json.loads(ps)
                    except: ps = []
                if not isinstance(ps, list): continue
                for src in ps:
                    if not isinstance(src, dict): continue
                    if any(s.get("restaurant") for s in (src.get("sources") or []) if isinstance(s, dict)):
                        _gemini_real_count += 1
                        break
                    if any(
                        any(s.get("restaurant") for s in (sz.get("sources") or []) if isinstance(s, dict))
                        for sz in (src.get("sizes") or {}).values() if isinstance(sz, dict)
                    ):
                        _gemini_real_count += 1
                        break
    except Exception:
        pass
    _show_places_nearby = _gemini_real_count == 0

    # Always fetch Places nearby — we need their lat/lng to compute the
    # restaurant's geographic center for distance-filtering Gemini-cited
    # restaurants. But we only ADD them to the competitors list when
    # Haiku mode is active (no real Gemini cites).
    _places_nearby = []
    try:
        from storage.price_intel import get_cached_comparisons
        _places_nearby = get_cached_comparisons(draft["restaurant_id"])
    except Exception:
        pass
    if _show_places_nearby:
        competitors = _places_nearby

    import tempfile as _tf2
    _gpc = os.path.join(_tf2.gettempdir(), f"menuflow_places_cache_{draft['restaurant_id']}.json")

    # Pre-compute the lookup context that both backfill passes (shell +
    # website) need. Cheaper to compute once than re-derive twice.
    _rest_for_bf = restaurant or {}
    _bf_city = _rest_for_bf.get("city", "") or ""
    _bf_state = _rest_for_bf.get("state", "") or ""
    _bf_api_key = os.environ.get("GOOGLE_PLACES_API_KEY", "")
    # Compute geographic center from Places nearby (always available, even
    # when we're not adding them to the competitors list). Used for
    # distance-filtering Gemini-cited entries.
    _bf_our_lat, _bf_our_lng = None, None
    _geo_source = competitors if competitors else _places_nearby
    if _geo_source:
        _lats = [c["latitude"] for c in _geo_source if c.get("latitude")]
        _lngs = [c["longitude"] for c in _geo_source if c.get("longitude")]
        if _lats and _lngs:
            _bf_our_lat = sum(_lats) / len(_lats)
            _bf_our_lng = sum(_lngs) / len(_lngs)

    try:
        with open(_gpc, "r") as _gcf:
            _cached_places = json.loads(_gcf.read())
        # Backfill missing place_id metadata for "shell" entries — restaurants
        # Gemini cited but the parallel-burst Places lookup either timed out,
        # exceeded the per-render LIVE_LOOKUP_CAP, or got an empty candidates
        # list. Symptom: sidebar entry shows only the name, no rating/$/
        # address. Self-heal by retrying findplacefromtext.
        _shells = [
            (k, _cd) for k, _cd in _cached_places.items()
            if _cd and _cd.get("place_name") and not _cd.get("place_id")
        ]
        if _shells and _bf_api_key:
            from concurrent.futures import ThreadPoolExecutor, as_completed
            import requests as _bf_rq

            def _shell_lookup(k, cd):
                try:
                    _name = cd.get("place_name") or ""
                    _params = {
                        "key": _bf_api_key,
                        "input": f"{_name} restaurant {_bf_city} {_bf_state}",
                        "inputtype": "textquery",
                        "fields": "place_id,name,formatted_address,geometry,rating,user_ratings_total,price_level",
                    }
                    if _bf_our_lat and _bf_our_lng:
                        _params["locationbias"] = f"circle:16000@{_bf_our_lat},{_bf_our_lng}"
                    _r = _bf_rq.get(
                        "https://maps.googleapis.com/maps/api/place/findplacefromtext/json",
                        params=_params, timeout=8,
                    )
                    if _r.status_code != 200:
                        return k, None
                    cands = _r.json().get("candidates", [])
                    if not cands:
                        return k, None
                    _c = cands[0]
                    _loc = _c.get("geometry", {}).get("location", {})
                    if _bf_our_lat and _bf_our_lng and _loc.get("lat") and _loc.get("lng"):
                        import math as _math
                        _dl = _math.radians(_loc["lat"] - _bf_our_lat)
                        _dn = _math.radians(_loc["lng"] - _bf_our_lng)
                        _a = _math.sin(_dl/2)**2 + _math.cos(_math.radians(_bf_our_lat)) * _math.cos(_math.radians(_loc["lat"])) * _math.sin(_dn/2)**2
                        _dist = 6371 * 2 * _math.atan2(_math.sqrt(_a), _math.sqrt(1-_a))
                        if _dist > 25:
                            return k, None
                    _pl = _c.get("price_level")
                    return k, {
                        "place_name": _c.get("name", _name),
                        "place_address": _c.get("formatted_address", ""),
                        "rating": _c.get("rating"),
                        "price_label": {1:"$",2:"$$",3:"$$$",4:"$$$$"}.get(_pl),
                        "price_level": _pl,
                        "user_ratings": _c.get("user_ratings_total"),
                        "latitude": _loc.get("lat"),
                        "longitude": _loc.get("lng"),
                        "cuisine_match": None,
                        "place_id": _c.get("place_id"),
                        "website_url": None,
                    }
                except Exception:
                    return k, None

            SHELL_BACKFILL_CAP = 30
            shells_filled = 0
            with ThreadPoolExecutor(max_workers=10) as _spool:
                _futs = [_spool.submit(_shell_lookup, k, cd)
                         for k, cd in _shells[:SHELL_BACKFILL_CAP]]
                for _fut in as_completed(_futs):
                    try:
                        _k, _new = _fut.result()
                    except Exception:
                        continue
                    if _new and _k in _cached_places:
                        _cached_places[_k] = _new
                        shells_filled += 1
            if shells_filled:
                log.info(
                    "Editor cache: backfilled %d/%d shell entries with Places metadata",
                    shells_filled, len(_shells),
                )

        # Backfill missing website_url for cached entries that have a
        # place_id but no website. Cause: the original parallel build
        # burst sometimes loses the get_place_details follow-up call to
        # rate limits, leaving website_url=None on entries whose Place
        # Details actually exist. The editor's "click competitor → load
        # website iframe" flow then shows "no website available" for
        # restaurants that DO have websites. Self-heal on each render.
        _missing = [
            (k, _cd) for k, _cd in _cached_places.items()
            if _cd and _cd.get("place_id") and not _cd.get("website_url")
        ]
        if _missing:
            from concurrent.futures import ThreadPoolExecutor, as_completed
            from storage.price_intel import get_place_details

            def _backfill_one(k, cd):
                try:
                    details = get_place_details(cd["place_id"])
                    if details and details.get("website"):
                        return k, details["website"]
                except Exception:
                    pass
                return k, None

            BACKFILL_CAP = 30
            backfilled = 0
            with ThreadPoolExecutor(max_workers=10) as _pool:
                futs = [_pool.submit(_backfill_one, k, cd)
                        for k, cd in _missing[:BACKFILL_CAP]]
                for fut in as_completed(futs):
                    try:
                        k, ws = fut.result()
                    except Exception:
                        continue
                    if ws and k in _cached_places and _cached_places[k]:
                        _cached_places[k]["website_url"] = ws
                        backfilled += 1
            if backfilled:
                # Persist the enriched cache back to disk so subsequent
                # loads don't have to re-do these calls.
                try:
                    with open(_gpc, "w") as _gcf:
                        _gcf.write(json.dumps(_cached_places))
                except Exception:
                    pass
                log.info(
                    "Editor cache: backfilled website_url for %d/%d entries",
                    backfilled, len(_missing),
                )
        _existing_lower = {c.get("place_name", "").lower() for c in competitors}
        for _cn, _cd in _cached_places.items():
            if not _cd:
                continue
            # Skip article-title / list-page junk that may be in caches
            # written before the filter existed.
            if _looks_like_article_title(_cd.get("place_name", "")):
                continue
            if _cd.get("place_name", "").lower() not in _existing_lower:
                competitors.append(_cd)
                _existing_lower.add(_cd.get("place_name", "").lower())
    except Exception:
        pass
    # Always populate gemini_restaurants from price_intel — the loop below
    # dedupes against places_cache so already-resolved entries are skipped.
    # The previous `not os.path.exists(_gpc)` gate created a coverage gap:
    # the first build resolved ~30 restaurants but Gemini cites often
    # reference 50+ unique names, and subsequent renders never attempted
    # the missing lookups because the cache file existed. User sees the
    # symptom as "only 20 results in Similar Restaurants" — the count
    # plateaus at whatever the first build managed to resolve.
    gemini_restaurants = set()
    if price_intel and price_intel.get("assessments"):
        import re as _re
        _norm_seen = {}
        for a in price_intel["assessments"]:
            ps = a.get("price_sources")
            if isinstance(ps, str):
                try: ps = json.loads(ps)
                except: ps = []
            if not isinstance(ps, list): continue
            for src in ps:
                if not isinstance(src, dict): continue
                for s in src.get("sources", []):
                    if isinstance(s, dict) and s.get("restaurant"):
                        raw = s["restaurant"]
                        n = _re.sub(r'\s*\(.*?\)\s*', '', raw).strip()
                        if _looks_like_article_title(n):
                            continue
                        if n.lower() not in _norm_seen:
                            _norm_seen[n.lower()] = n
                            gemini_restaurants.add(n)
                for sz in (src.get("sizes") or {}).values():
                    if isinstance(sz, dict):
                        for s in sz.get("sources", []):
                            if isinstance(s, dict) and s.get("restaurant"):
                                raw = s["restaurant"]
                                n = _re.sub(r'\s*\(.*?\)\s*', '', raw).strip()
                                if _looks_like_article_title(n):
                                    continue
                                if n.lower() not in _norm_seen:
                                    _norm_seen[n.lower()] = n
                                    gemini_restaurants.add(n)
    if gemini_restaurants:
        rest_info = restaurant or {}
        rest_city = rest_info.get("city", "")
        rest_state = rest_info.get("state", "")
        rest_zip = rest_info.get("zip_code", "")

        def _normalize_restaurant_name(name):
            import re
            n = name.strip()
            n = re.sub(r'\s*\(.*?\)\s*', '', n)
            for suffix in [f" {rest_city}", f" {rest_city}, {rest_state}", f" {rest_city} {rest_state}",
                           f", {rest_city}", f" {rest_state}", f" MA", f" {rest_zip}"]:
                if suffix and n.lower().endswith(suffix.lower()):
                    n = n[:len(n) - len(suffix)].strip()
            return n.strip()
        # Reuse the geo center we already computed (works whether or not
        # Places nearby is included in `competitors`).
        our_lat, our_lng = _bf_our_lat, _bf_our_lng

        existing_names = {c.get("place_name", "").lower() for c in competitors}
        added_normalized = set()
        for n in existing_names:
            added_normalized.add(_normalize_restaurant_name(n).lower())

        import requests as _rq
        api_key = os.environ.get("GOOGLE_PLACES_API_KEY", "")
        import tempfile
        cache_path = os.path.join(tempfile.gettempdir(), f"menuflow_places_cache_{draft.get('restaurant_id', 0)}.json")
        places_cache = {}
        try:
            with open(cache_path, "r") as cf:
                places_cache = json.loads(cf.read())
        except Exception:
            pass

        cache_dirty = False
        to_lookup: list = []
        for rname in sorted(gemini_restaurants):
            norm = _normalize_restaurant_name(rname).lower()
            if not norm or norm in added_normalized or rname.lower() in existing_names:
                continue
            added_normalized.add(norm)

            if norm in places_cache:
                c_data = places_cache[norm]
                if c_data:
                    # Dedupe by Places-resolved name. Multiple Gemini-cited
                    # variants ("141 Main Street Restaurant & Catering" +
                    # "141 Main Street Roka") often resolve to the same
                    # actual business in Places API ("141 Main Street").
                    # Without this check, the same restaurant got appended
                    # twice — visible to the user as "27 results" with the
                    # map only plotting 20 unique pins.
                    pn_lower = (c_data.get("place_name") or "").lower()
                    if pn_lower and pn_lower not in existing_names:
                        competitors.append(c_data)
                        existing_names.add(pn_lower)
                continue

            if not api_key:
                comp_entry = {
                    "place_name": rname,
                    "place_address": f"{rest_city}, {rest_state}",
                    "rating": None, "price_label": None, "price_level": None,
                    "user_ratings": None, "latitude": None, "longitude": None,
                    "cuisine_match": None, "place_id": None, "website_url": None,
                }
                competitors.append(comp_entry)
                places_cache[norm] = comp_entry
                cache_dirty = True
                continue

            to_lookup.append((rname, norm))

        LIVE_LOOKUP_CAP = 60  # was 30 — Gemini routinely cites 50+ unique
                              # restaurants on a 158-item menu. Cap of 30
                              # was capping the editor sidebar artificially.
        for rname, norm in to_lookup[LIVE_LOOKUP_CAP:]:
            comp_entry = {
                "place_name": rname,
                "place_address": f"{rest_city}, {rest_state}",
                "rating": None, "price_label": None, "price_level": None,
                "user_ratings": None, "latitude": None, "longitude": None,
                "cuisine_match": None, "place_id": None, "website_url": None,
            }
            competitors.append(comp_entry)
            places_cache[norm] = comp_entry
            cache_dirty = True

        def _lookup_place(rname: str, norm: str):
            try:
                params = {
                    "key": api_key,
                    "input": f"{rname} restaurant {rest_city} {rest_state}",
                    "inputtype": "textquery",
                    "fields": "place_id,name,formatted_address,geometry,rating,user_ratings_total,price_level",
                }
                if our_lat and our_lng:
                    params["locationbias"] = f"circle:16000@{our_lat},{our_lng}"
                resp = _rq.get("https://maps.googleapis.com/maps/api/place/findplacefromtext/json",
                               params=params, timeout=5)
                if resp.status_code != 200:
                    return norm, None, None
                candidates = resp.json().get("candidates", [])
                if not candidates:
                    return norm, None, None
                c = candidates[0]
                loc = c.get("geometry", {}).get("location", {})
                if our_lat and our_lng and loc.get("lat") and loc.get("lng"):
                    import math
                    dlat = math.radians(loc["lat"] - our_lat)
                    dlng = math.radians(loc["lng"] - our_lng)
                    a2 = math.sin(dlat/2)**2 + math.cos(math.radians(our_lat)) * math.cos(math.radians(loc["lat"])) * math.sin(dlng/2)**2
                    dist_km = 6371 * 2 * math.atan2(math.sqrt(a2), math.sqrt(1-a2))
                    if dist_km > 25:
                        return norm, None, None
                pid = c.get("place_id")
                website_url = None
                if pid:
                    try:
                        from storage.price_intel import get_place_details
                        details = get_place_details(pid)
                        if details:
                            website_url = details.get("website")
                    except Exception:
                        pass
                price_level = c.get("price_level")
                comp_entry = {
                    "place_name": c.get("name", rname),
                    "place_address": c.get("formatted_address", ""),
                    "rating": c.get("rating"),
                    "price_label": {1:"$",2:"$$",3:"$$$",4:"$$$$"}.get(price_level),
                    "price_level": price_level,
                    "user_ratings": c.get("user_ratings_total"),
                    "latitude": loc.get("lat"),
                    "longitude": loc.get("lng"),
                    "cuisine_match": None,
                    "place_id": pid,
                    "website_url": website_url,
                }
                db_args = (
                    draft.get("restaurant_id") or 0, pid, c.get("name", rname),
                    c.get("formatted_address", ""), price_level,
                    {1:"$",2:"$$",3:"$$$",4:"$$$$"}.get(price_level),
                    c.get("rating"), c.get("user_ratings_total"),
                    loc.get("lat"), loc.get("lng"),
                )
                return norm, comp_entry, db_args
            except Exception:
                return norm, None, None

        batch = to_lookup[:LIVE_LOOKUP_CAP]
        if batch:
            from concurrent.futures import ThreadPoolExecutor, as_completed
            db_writes = []
            with ThreadPoolExecutor(max_workers=10) as pool:
                futs = [pool.submit(_lookup_place, rname, norm) for rname, norm in batch]
                for fut in as_completed(futs):
                    try:
                        norm, comp_entry, db_args = fut.result()
                    except Exception:
                        continue
                    if comp_entry:
                        competitors.append(comp_entry)
                    places_cache[norm] = comp_entry
                    cache_dirty = True
                    if db_args:
                        db_writes.append(db_args)
            if db_writes:
                try:
                    with db_connect() as conn2:
                        conn2.executemany(
                            """INSERT OR IGNORE INTO price_comparison_results
                               (restaurant_id, place_id, place_name, place_address, price_level, price_label, rating, user_ratings, latitude, longitude)
                               VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
                            db_writes,
                        )
                        conn2.commit()
                except Exception:
                    pass

        if cache_dirty:
            try:
                with open(cache_path, "w") as cf:
                    cf.write(json.dumps(places_cache))
            except Exception:
                pass

        # After the live build, do a single backfill sweep on whatever
        # shells / no-website entries remain in the just-built cache.
        # Reasons they exist after a fresh build:
        #   - Overflow beyond LIVE_LOOKUP_CAP=30 was deliberately written
        #     as shells. Backfill those now (capped separately) so the user
        #     doesn't have to reload to see metadata.
        #   - get_place_details may have lost the website_url race during
        #     the parallel-burst.
        # Without this, the first render after a rerun shows bare names
        # for many restaurants and the in-memory editor cache then freezes
        # that state until the next rerun.
        try:
            _post_build_backfill(places_cache, cache_path,
                                  _bf_api_key, _bf_city, _bf_state,
                                  _bf_our_lat, _bf_our_lng)
        except Exception as _bbe:
            log.warning("Post-build cache backfill failed: %s", _bbe)
        # Re-load the freshly-saved cache into competitors so the caller
        # gets the enriched entries (the for-loop above only added the
        # original cached shells/None entries).
        try:
            with open(cache_path, "r") as _gcf2:
                _enriched = json.loads(_gcf2.read())
            _have = {c.get("place_name", "").lower() for c in competitors}
            for _k, _cd in _enriched.items():
                if not _cd: continue
                if _looks_like_article_title(_cd.get("place_name", "")): continue
                _name_lc = _cd.get("place_name", "").lower()
                # Replace shell entries that were already in competitors with
                # the enriched version (match by name).
                for i, _existing in enumerate(competitors):
                    if _existing.get("place_name", "").lower() == _name_lc:
                        competitors[i] = _cd
                        _name_lc = None
                        break
                if _name_lc and _name_lc not in _have:
                    competitors.append(_cd)
        except Exception:
            pass

    return competitors


def _post_build_backfill(places_cache: dict, cache_path: str,
                         api_key: str, city: str, state: str,
                         our_lat, our_lng) -> None:
    """Shared backfill sweep for shells (no place_id) + no-website entries.
    Modifies `places_cache` in place and persists if anything changed."""
    if not api_key:
        return
    shells = [
        (k, cd) for k, cd in places_cache.items()
        if cd and cd.get("place_name") and not cd.get("place_id")
    ]
    no_urls = [
        (k, cd) for k, cd in places_cache.items()
        if cd and cd.get("place_id") and not cd.get("website_url")
    ]
    if not shells and not no_urls:
        return

    from concurrent.futures import ThreadPoolExecutor, as_completed
    from storage.price_intel import get_place_details
    import requests as _bf_rq2
    import math as _math2

    def _shell_lookup(k, cd):
        try:
            _name = cd.get("place_name") or ""
            _params = {
                "key": api_key,
                "input": f"{_name} restaurant {city} {state}",
                "inputtype": "textquery",
                "fields": "place_id,name,formatted_address,geometry,rating,user_ratings_total,price_level",
            }
            if our_lat and our_lng:
                _params["locationbias"] = f"circle:16000@{our_lat},{our_lng}"
            _r = _bf_rq2.get(
                "https://maps.googleapis.com/maps/api/place/findplacefromtext/json",
                params=_params, timeout=8,
            )
            if _r.status_code != 200:
                return k, None
            cands = _r.json().get("candidates", [])
            if not cands:
                return k, None
            _c = cands[0]
            _loc = _c.get("geometry", {}).get("location", {})
            if our_lat and our_lng and _loc.get("lat") and _loc.get("lng"):
                _dl = _math2.radians(_loc["lat"] - our_lat)
                _dn = _math2.radians(_loc["lng"] - our_lng)
                _a = _math2.sin(_dl/2)**2 + _math2.cos(_math2.radians(our_lat)) * _math2.cos(_math2.radians(_loc["lat"])) * _math2.sin(_dn/2)**2
                if 6371 * 2 * _math2.atan2(_math2.sqrt(_a), _math2.sqrt(1-_a)) > 25:
                    return k, None
            _pl = _c.get("price_level")
            return k, {
                "place_name": _c.get("name", _name),
                "place_address": _c.get("formatted_address", ""),
                "rating": _c.get("rating"),
                "price_label": {1:"$",2:"$$",3:"$$$",4:"$$$$"}.get(_pl),
                "price_level": _pl,
                "user_ratings": _c.get("user_ratings_total"),
                "latitude": _loc.get("lat"),
                "longitude": _loc.get("lng"),
                "cuisine_match": None,
                "place_id": _c.get("place_id"),
                "website_url": None,
            }
        except Exception:
            return k, None

    def _website_lookup(k, cd):
        try:
            details = get_place_details(cd["place_id"])
            if details and details.get("website"):
                return k, details["website"]
        except Exception:
            pass
        return k, None

    BACKFILL_CAP = 60
    shells_filled = urls_filled = 0
    with ThreadPoolExecutor(max_workers=10) as pool:
        s_futs = [pool.submit(_shell_lookup, k, cd) for k, cd in shells[:BACKFILL_CAP]]
        u_futs = [pool.submit(_website_lookup, k, cd) for k, cd in no_urls[:BACKFILL_CAP]]
        for fut in as_completed(s_futs):
            try:
                k, new = fut.result()
            except Exception:
                continue
            if new and k in places_cache:
                places_cache[k] = new
                shells_filled += 1
        for fut in as_completed(u_futs):
            try:
                k, ws = fut.result()
            except Exception:
                continue
            if ws and k in places_cache and places_cache[k]:
                places_cache[k]["website_url"] = ws
                urls_filled += 1

    if shells_filled or urls_filled:
        try:
            with open(cache_path, "w") as cf:
                cf.write(json.dumps(places_cache))
        except Exception:
            pass
        log.info(
            "Editor cache backfill: shells %d/%d, websites %d/%d",
            shells_filled, len(shells), urls_filled, len(no_urls),
        )


@app.get("/drafts/<int:draft_id>/edit")
@login_required
def draft_editor(draft_id: int):
    """Render the Draft Editor UI (Day 141.5: wizard-style card layout)."""
    _require_drafts_storage()
    draft = drafts_store.get_draft(draft_id)
    if not draft:
        abort(404, description=f"Draft {draft_id} not found")

    items = drafts_store.get_draft_items(draft_id, include_modifier_groups=True) or []

    # Position backfill (same as wizard route)
    needs_backfill = [it for it in items if it.get("position") is None]
    if needs_backfill:
        max_pos = max((it.get("position") or 0) for it in items) if items else 0
        try:
            with db_connect() as conn:
                for it in needs_backfill:
                    max_pos += 1
                    conn.execute(
                        "UPDATE draft_items SET position=?, updated_at=? WHERE id=? AND draft_id=?",
                        (max_pos, _now_iso(), it["id"], draft_id),
                    )
                    it["position"] = max_pos
                conn.commit()
        except Exception:
            pass

    # Compute per-item quality
    for it in items:
        try:
            score, is_low = _compute_item_quality(it)
        except Exception:
            score, is_low = 70, False
        it["quality"] = score
        it["low_confidence"] = bool(is_low)

    # Initialize wizard category tracking (reuse wizard infra for sidebar)
    drafts_store.init_wizard_categories(draft_id)
    progress = drafts_store.get_wizard_progress(draft_id)

    # Group items by category + subcategory
    flat_groups = {}
    subcat_groups = {}
    for it in items:
        cat = (it.get("category") or "Uncategorized").strip()
        subcat = (it.get("subcategory") or "").strip()
        flat_groups.setdefault(cat, []).append(it)
        subcat_groups.setdefault(cat, {}).setdefault(subcat, []).append(it)

    # Build category list (main categories only, in progress order)
    category_list = []
    _seen_cats = set()
    for c in progress["categories"]:
        if not c.get("subcategory") and c["name"] not in _seen_cats:
            _seen_cats.add(c["name"])
            category_list.append(c["name"])

    # Apply saved category order
    try:
        saved_order = drafts_store.get_category_order(draft_id)
        if saved_order:
            order_map = {name: i for i, name in enumerate(saved_order)}
            category_list.sort(key=lambda c: order_map.get(c, 9999))
            progress["categories"].sort(key=lambda c: order_map.get(c["name"], 9999))
    except Exception:
        pass

    # Determine current category (default to first)
    requested_cat = request.args.get("category", "").strip()
    requested_subcat = request.args.get("subcategory", "").strip() or None
    # None = "All Items" view; only set a category if explicitly requested
    current_category = None
    if requested_cat and requested_cat in category_list:
        current_category = requested_cat

    # Price intelligence
    price_intel = None
    try:
        from storage.ai_price_intel import get_price_intelligence
        price_intel = get_price_intelligence(draft_id)
    except Exception:
        pass

    price_map = {}
    # Detect "Gemini fully failed" state for the banner: every assessed item
    # has price_sources that contain only `market_estimate` source type with
    # no real restaurant citations. Surfaces a customer-facing warning when
    # we couldn't reach Google's grounded-search API at all and Haiku had to
    # fill everything.
    gemini_total_assessed = 0
    gemini_with_real_sources = 0
    if price_intel and price_intel.get("assessments"):
        for a in price_intel["assessments"]:
            if a.get("item_id"):
                # Day 141.8: parse price_sources JSON for template rendering
                ps = a.get("price_sources")
                if isinstance(ps, str):
                    try:
                        a["price_sources"] = json.loads(ps)
                    except (json.JSONDecodeError, TypeError):
                        a["price_sources"] = []
                price_map[a["item_id"]] = a
                # Count items with real Gemini sources (not just Haiku estimates)
                ps_parsed = a.get("price_sources") or []
                if a.get("regional_avg") is not None:
                    gemini_total_assessed += 1
                    has_real = False
                    for src in ps_parsed if isinstance(ps_parsed, list) else []:
                        if not isinstance(src, dict):
                            continue
                        for s in src.get("sources", []) or []:
                            if isinstance(s, dict) and s.get("restaurant"):
                                has_real = True
                                break
                        if has_real:
                            break
                        for sz in (src.get("sizes") or {}).values():
                            if isinstance(sz, dict):
                                for s in sz.get("sources", []) or []:
                                    if isinstance(s, dict) and s.get("restaurant"):
                                        has_real = True
                                        break
                            if has_real:
                                break
                        if has_real:
                            break
                    if has_real:
                        gemini_with_real_sources += 1
    # Banner condition: less than 60% of the batches in the most recent
    # pricing run completed successfully. This measures whether Gemini's
    # API was actually reachable, not how many items survived downstream
    # quote-validation. Read from gemini_call_log (batch-level only, ts
    # since the current run's stub was created).
    _MARKET_DATA_DOWN_BATCH_THRESHOLD = 0.60
    market_data_down = False
    if price_intel and price_intel.get("created_at"):
        try:
            with db_connect() as _conn:
                _stats = _conn.execute(
                    "SELECT outcome, COUNT(*) c FROM gemini_call_log "
                    "WHERE draft_id=? AND ts >= ? AND batch_size > 1 "
                    "GROUP BY outcome",
                    (draft_id, price_intel["created_at"]),
                ).fetchall()
            _ok = sum(r["c"] for r in _stats if r["outcome"] == "ok")
            _total = sum(r["c"] for r in _stats)
            if _total > 0:
                market_data_down = (_ok / _total) < _MARKET_DATA_DOWN_BATCH_THRESHOLD
        except Exception:
            # If we can't read the log, fall back to no banner — better
            # than a false alarm.
            market_data_down = False

    # Restaurant info
    restaurant = None
    if draft.get("restaurant_id"):
        try:
            with db_connect() as conn:
                restaurant = conn.execute(
                    "SELECT id, name, cuisine_type, zip_code FROM restaurants WHERE id = ?",
                    (draft["restaurant_id"],),
                ).fetchone()
                if restaurant:
                    restaurant = dict(restaurant)
        except Exception:
            pass

    # Editor competitors — cached per-draft. _invalidate_editor_cache(draft_id)
    # wipes the entry at every new run so fresh pipeline output is picked up.
    with _EDITOR_COMPETITORS_LOCK:
        competitors = _EDITOR_COMPETITORS_CACHE.get(draft_id)
    if competitors is None:
        competitors = _build_editor_competitors(draft, price_intel, restaurant)
        with _EDITOR_COMPETITORS_LOCK:
            _EDITOR_COMPETITORS_CACHE[draft_id] = competitors

    # Day 141.6: User restaurants + their drafts for the menu picker overlay
    user_restaurants = []
    u = session.get("user") or {}
    uid = u.get("user_id")
    if uid and users_store:
        try:
            user_restaurants = users_store.get_user_restaurants(uid)
            for ur in user_restaurants:
                full = users_store.get_restaurant(ur["restaurant_id"])
                if full:
                    ur["address"] = full.get("address") or ""
                    ur["phone"] = full.get("phone") or ""
                    ur["city"] = full.get("city") or ""
                    ur["state"] = full.get("state") or ""
                # Get drafts for this restaurant
                try:
                    with db_connect() as conn:
                        ur["drafts"] = [
                            dict(r)
                            for r in conn.execute(
                                "SELECT id, title, updated_at FROM drafts WHERE restaurant_id = ? ORDER BY updated_at DESC",
                                (ur["restaurant_id"],),
                            ).fetchall()
                        ]
                except Exception:
                    ur["drafts"] = []
        except Exception:
            pass

    # Google Maps API key (for JS embed)
    google_maps_key = os.environ.get("GOOGLE_PLACES_API_KEY", "").strip() or None

    # Source job for original menu viewer
    source_job_id = draft.get("source_job_id")

    # Day 141.6: Detect structured imports (CSV/XLSX/JSON) to show file content
    # instead of trying to render as image
    source_is_structured = False
    source_file_path = None
    try:
        import json as _json
        src = _json.loads(draft.get("source") or "{}")
        if src.get("source_type", "").startswith("structured_"):
            source_is_structured = True
            source_file_path = src.get("meta", {}).get("csv_path") or src.get("meta", {}).get("file_path")
    except Exception:
        pass

    # Bounding box coordinates
    item_coordinates = {}
    try:
        if drafts_store and hasattr(drafts_store, "get_draft_coordinates"):
            raw_coords = drafts_store.get_draft_coordinates(draft_id)
            for c in raw_coords:
                item_coordinates[c["item_id"]] = {
                    "x_pct": c["x_pct"], "y_pct": c["y_pct"],
                    "w_pct": c["w_pct"], "h_pct": c["h_pct"],
                    "page": c["page"],
                }
    except Exception:
        pass

    # Build display_groups: the items to render in the main area
    display_groups = {}
    if not current_category:
        # "All Items" view — group by category (and subcategory)
        for cat in category_list:
            for sc, sc_items in subcat_groups.get(cat, {}).items():
                key = f"{cat} > {sc}" if sc else cat
                display_groups[key] = sc_items
    elif requested_subcat:
        display_groups[requested_subcat] = subcat_groups.get(current_category, {}).get(requested_subcat, [])
    else:
        display_groups[""] = subcat_groups.get(current_category, {}).get("", [])

    return _safe_render(
        "draft_editor.html",
        draft=draft,
        items=items,
        flat_groups=flat_groups,
        subcat_groups=subcat_groups,
        display_groups=display_groups,
        progress=progress,
        category_list=category_list,
        current_category=current_category,
        current_subcategory=requested_subcat,
        restaurant=restaurant,
        price_map=price_map,
        price_intel=price_intel,
        competitors=(lambda comps: sorted(
            list({c.get("place_id") or re.sub(r'[^a-z0-9]', '', (c.get("place_name","")).lower()): c
                  for c in comps}.values()),
            key=lambda c: (c.get("place_name") or "").lower()
        ))(competitors),
        google_maps_key=google_maps_key,
        source_job_id=source_job_id,
        source_is_structured=source_is_structured,
        item_coordinates=item_coordinates,
        user_restaurants=user_restaurants,
        market_data_down=market_data_down,
    )


# ===== Day 137: Guided Onboarding Wizard =====

@app.get("/drafts/<int:draft_id>/wizard")
@login_required
def draft_wizard(draft_id: int):
    """Guided category-by-category review wizard."""
    _require_drafts_storage()
    draft = drafts_store.get_draft(draft_id)
    if not draft:
        abort(404, description=f"Draft {draft_id} not found")

    items = drafts_store.get_draft_items(draft_id, include_modifier_groups=True) or []

    # --- Position backfill: auto-assign positions to items with position=None ---
    needs_backfill = [it for it in items if it.get("position") is None]
    if needs_backfill:
        # Find the highest existing position
        max_pos = max((it.get("position") or 0) for it in items)
        try:
            with db_connect() as conn:
                for it in needs_backfill:
                    max_pos += 1
                    conn.execute(
                        "UPDATE draft_items SET position=?, updated_at=? WHERE id=? AND draft_id=?",
                        (max_pos, _now_iso(), it["id"], draft_id),
                    )
                    it["position"] = max_pos
                conn.commit()
        except Exception:
            pass  # non-critical

    # Compute per-item quality (reuse same logic as editor)
    for it in items:
        try:
            score, is_low = _compute_item_quality(it)
        except Exception:
            score, is_low = 70, False
        it["quality"] = score
        it["low_confidence"] = bool(is_low)

    # Initialize wizard category tracking (idempotent)
    drafts_store.init_wizard_categories(draft_id)
    progress = drafts_store.get_wizard_progress(draft_id)

    # Group items by category, and within each category by subcategory
    flat_groups = {}
    subcat_groups = {}  # {category: {subcategory: [items]}}
    for it in items:
        cat = (it.get("category") or "Uncategorized").strip()
        subcat = (it.get("subcategory") or "").strip()
        flat_groups.setdefault(cat, []).append(it)
        subcat_groups.setdefault(cat, {}).setdefault(subcat, []).append(it)

    # Determine current step: summary (first page), or a category name
    requested_step = request.args.get("step", "").strip()
    requested_cat = request.args.get("category", "").strip()
    requested_subcat = request.args.get("subcategory", "").strip() or None
    # Day 141: progress.categories now includes subcategory rows too.
    # category_list should remain a flat de-duplicated list of MAIN categories
    # only, in the order they were first seen.
    category_list = []
    _seen_cats = set()
    for c in progress["categories"]:
        if not c.get("subcategory") and c["name"] not in _seen_cats:
            _seen_cats.add(c["name"])
            category_list.append(c["name"])

    # Apply saved category order if it exists
    try:
        saved_order = drafts_store.get_category_order(draft_id)
        if saved_order:
            order_map = {name: i for i, name in enumerate(saved_order)}
            category_list.sort(key=lambda c: order_map.get(c, 9999))
            progress["categories"].sort(key=lambda c: order_map.get(c["name"], 9999))
    except Exception:
        pass

    # Default: summary page on first visit, otherwise first unreviewed
    wizard_step = "summary"  # "summary", "confirmation", or a category name
    current_category = None

    if requested_step == "confirmation":
        wizard_step = "confirmation"
    elif requested_cat and requested_cat in category_list:
        wizard_step = "category"
        current_category = requested_cat
    elif requested_step == "summary" or (not requested_step and not requested_cat):
        wizard_step = "summary"
    else:
        # First unreviewed category
        wizard_step = "category"
        for c in progress["categories"]:
            if not c["reviewed"]:
                current_category = c["name"]
                break
        # All reviewed — show the last one
        if current_category is None and category_list:
            current_category = category_list[-1]

    # Day 141.7: price intel no longer loaded in wizard — moved to
    # post-confirmation trigger, displayed only in the editor.

    # Restaurant info for summary page
    restaurant = None
    if draft.get("restaurant_id"):
        try:
            with db_connect() as conn:
                restaurant = conn.execute(
                    "SELECT id, name, cuisine_type, zip_code FROM restaurants WHERE id = ?",
                    (draft["restaurant_id"],),
                ).fetchone()
                if restaurant:
                    restaurant = dict(restaurant)
        except Exception:
            pass

    # Day 141.6: Load user's restaurants for confirmation screen
    # Day 141.7: include cuisine_type so user can verify it before price scrape
    user_restaurants = []
    if wizard_step == "confirmation":
        u = session.get("user") or {}
        uid = u.get("user_id")
        if uid and users_store:
            try:
                user_restaurants = users_store.get_user_restaurants(uid)
                for ur in user_restaurants:
                    try:
                        full = users_store.get_restaurant(ur["restaurant_id"])
                        if full:
                            ur["address"] = full.get("address") or ""
                            ur["phone"] = full.get("phone") or ""
                            ur["city"] = full.get("city") or ""
                            ur["state"] = full.get("state") or ""
                            ur["zip_code"] = full.get("zip_code") or ""
                            ur["cuisine_type"] = full.get("cuisine_type") or ""
                    except Exception:
                        pass
            except Exception:
                pass

    # Source job id for original file preview
    source_job_id = draft.get("source_job_id")

    # Day 141.6: Detect structured imports for file table view
    source_is_structured = False
    try:
        import json as _json
        src = _json.loads(draft.get("source") or "{}")
        if src.get("source_type", "").startswith("structured_"):
            source_is_structured = True
    except Exception:
        pass

    # Day 139: Load bounding box coordinates for item highlighting
    item_coordinates = {}
    try:
        if drafts_store and hasattr(drafts_store, "get_draft_coordinates"):
            raw_coords = drafts_store.get_draft_coordinates(draft_id)
            for c in raw_coords:
                item_coordinates[c["item_id"]] = {
                    "x_pct": c["x_pct"],
                    "y_pct": c["y_pct"],
                    "w_pct": c["w_pct"],
                    "h_pct": c["h_pct"],
                    "page": c["page"],
                }
    except Exception:
        pass

    return _safe_render(
        "wizard.html",
        draft=draft,
        items=items,
        flat_groups=flat_groups,
        subcat_groups=subcat_groups,
        progress=progress,
        category_list=category_list,
        current_category=current_category,
        current_subcategory=requested_subcat,
        wizard_step=wizard_step,
        restaurant=restaurant,
        source_job_id=source_job_id,
        source_is_structured=source_is_structured,
        item_coordinates=item_coordinates,
        user_restaurants=user_restaurants,
    )


@app.post("/drafts/<int:draft_id>/wizard/confirm")
@login_required
def wizard_confirm_category(draft_id: int):
    """Mark a category (or subcategory) as reviewed and advance to next.

    Day 141: subcategory-aware. The next view in line might be a subcategory
    of the just-confirmed category, not the next main category.
    """
    _require_drafts_storage()
    draft = drafts_store.get_draft(draft_id)
    if not draft:
        abort(404)

    category = (request.form.get("category") or "").strip()
    subcategory = (request.form.get("subcategory") or "").strip() or None
    if not category:
        flash("No category specified.", "error")
        return redirect(url_for("draft_wizard", draft_id=draft_id))

    drafts_store.mark_category_reviewed(draft_id, category, subcategory)
    progress = drafts_store.get_wizard_progress(draft_id)

    # Day 141: Confirm & Next NEVER leaves the wizard. Only the
    # "Complete Review" button in the sidebar can send the user to the editor.
    # If the user just confirmed the last unreviewed entry, stay put on the
    # current category so they can click Complete Review.
    if progress["complete"]:
        kwargs = {"draft_id": draft_id, "category": category}
        if subcategory:
            kwargs["subcategory"] = subcategory
        return redirect(url_for("draft_wizard", **kwargs))

    # Day 141: advance FORWARD only — find the next unreviewed entry that
    # comes AFTER the one we just confirmed. If nothing forward is unreviewed,
    # fall back to the first unreviewed at the top of the list (so the user
    # eventually finishes everything they skipped).
    cats = progress["categories"]
    current_idx = -1
    for i, c in enumerate(cats):
        if c["name"] == category and (c.get("subcategory") or None) == (subcategory or None):
            current_idx = i
            break

    next_entry = None
    # Search forward from the position right after the current entry
    if current_idx >= 0:
        for c in cats[current_idx + 1 :]:
            if not c["reviewed"]:
                next_entry = c
                break

    # No forward entry → fall back to the first unreviewed anywhere
    if next_entry is None:
        for c in cats:
            if not c["reviewed"]:
                next_entry = c
                break

    if next_entry is not None:
        kwargs = {"draft_id": draft_id, "category": next_entry["name"]}
        if next_entry.get("subcategory"):
            kwargs["subcategory"] = next_entry["subcategory"]
        return redirect(url_for("draft_wizard", **kwargs))

    return redirect(url_for("draft_wizard", draft_id=draft_id))


@app.post("/drafts/<int:draft_id>/wizard/complete_review")
@login_required
def wizard_complete_review(draft_id: int):
    """Day 141: 'Complete Review' button — only allowed if every category and
    subcategory has been confirmed. If anything is unreviewed, flash an error
    and bounce the user to the first unreviewed entry. The user has to walk
    through every step to actually finish.

    Day 141.6: Instead of going straight to the editor, redirect to the
    wizard confirmation screen where the user names the menu and picks a
    restaurant.
    """
    _require_drafts_storage()
    draft = drafts_store.get_draft(draft_id)
    if not draft:
        abort(404)

    progress = drafts_store.get_wizard_progress(draft_id)

    if progress["complete"]:
        drafts_store.mark_wizard_completed(draft_id)
        # Day 141.6: go to confirmation screen instead of editor
        return redirect(url_for("draft_wizard", draft_id=draft_id, step="confirmation"))

    # Not done — bounce to first unreviewed
    flash("Every category must be confirmed before completing review.", "error")
    for c in progress["categories"]:
        if not c["reviewed"]:
            kwargs = {"draft_id": draft_id, "category": c["name"]}
            if c.get("subcategory"):
                kwargs["subcategory"] = c["subcategory"]
            return redirect(url_for("draft_wizard", **kwargs))
    return redirect(url_for("draft_wizard", draft_id=draft_id))


@app.post("/drafts/<int:draft_id>/wizard/finalize")
@login_required
def wizard_finalize(draft_id: int):
    """Day 141.6 / 141.8: Confirmation screen submit — assign restaurant + menu
    title. Premium users with a restaurant get routed through the price-intel
    analyzer; everyone else lands directly in the editor."""
    _require_drafts_storage()
    draft = drafts_store.get_draft(draft_id)
    if not draft:
        abort(404)

    menu_title = (request.form.get("menu_title") or "").strip()
    restaurant_id = request.form.get("restaurant_id") or None

    if not menu_title:
        from datetime import datetime
        menu_title = f"Imported {datetime.utcnow().strftime('%b %d')}"

    drafts_store.save_draft_metadata(draft_id, title=menu_title)

    rid_int = None
    if restaurant_id:
        try:
            rid_int = int(restaurant_id)
            drafts_store.save_draft_metadata(draft_id, restaurant_id=rid_int)
        except (ValueError, TypeError):
            rid_int = None

    # Day 141.7: route premium users through the price-intel analyzer first.
    u = session.get("user") or {}
    tier = (u.get("account_tier") or "free").lower()
    if tier == "premium" and rid_int and ai_price_intel:
        return redirect(url_for("draft_analyzing_prices", draft_id=draft_id))

    flash("All categories reviewed! Welcome to the full editor.", "success")
    return redirect(url_for("draft_editor", draft_id=draft_id))


# Day 141.7: background status tracker for post-wizard price analysis.
# Simple in-memory dict keyed by draft_id. Each entry:
#   {"status": "running"|"done"|"error", "started_at": ts, "error": str|None}
_PRICE_ANALYSIS_JOBS: dict = {}
_PRICE_ANALYSIS_LOCK = threading.Lock()

# Editor competitor-list cache. Built lazily on first render of the editor,
# reused on every subsequent reload, blown away whenever a new price
# analysis run starts. Saves ~5-10s per reload (avoids the Gemini-restaurant
# Google Places enrichment). Keyed by draft_id.
_EDITOR_COMPETITORS_CACHE: dict = {}
_EDITOR_COMPETITORS_LOCK = threading.Lock()


def _invalidate_editor_cache(draft_id: int, restaurant_id: int | None = None) -> None:
    """Wipe all editor-side caches for a draft so the next render rebuilds
    from fresh pipeline output. Call this at the start of every new run —
    wizard finish, rerun button, or direct analyze trigger. Safe to call
    even if nothing is cached yet."""
    with _EDITOR_COMPETITORS_LOCK:
        _EDITOR_COMPETITORS_CACHE.pop(draft_id, None)
    # Nuke the per-restaurant Places temp file too — it's populated from
    # the same Gemini sources and would otherwise keep serving stale
    # coords/ratings alongside the freshly-computed results.
    if restaurant_id:
        try:
            import tempfile as _tf
            _gpc = os.path.join(_tf.gettempdir(), f"menuflow_places_cache_{restaurant_id}.json")
            if os.path.exists(_gpc):
                os.remove(_gpc)
        except Exception:
            pass


def _run_price_analysis_job(draft_id: int, restaurant_id: int) -> None:
    """Background worker: runs Claude Call 4 and updates the status dict.

    Status is set to "done" only after EVERYTHING the editor needs to
    render fully is in place: pricing pipeline, Places nearby search,
    editor competitor list (with shell + website backfills). The user
    waits a bit longer on the loading screen but lands on a fully
    populated editor — no shells, no missing websites, no race with
    background daemons."""
    _invalidate_editor_cache(draft_id, restaurant_id)
    try:
        result = ai_price_intel.analyze_menu_prices(
            draft_id=draft_id,
            restaurant_id=restaurant_id,
        )
        if result and not result.get("error"):
            # Pre-warm the editor competitor cache so the first render is
            # instant AND fully populated (shells/websites already filled
            # in by _build_editor_competitors's backfill passes).
            try:
                _draft = drafts_store.get_draft(draft_id)
                _restaurant = None
                if _draft and _draft.get("restaurant_id"):
                    with db_connect() as _conn:
                        _r = _conn.execute(
                            "SELECT id, name, address, city, state, zip_code, cuisine_type "
                            "FROM restaurants WHERE id = ?",
                            (_draft["restaurant_id"],),
                        ).fetchone()
                        if _r:
                            _restaurant = dict(_r)
                _price_intel = None
                try:
                    from storage.ai_price_intel import get_price_intelligence
                    _price_intel = get_price_intelligence(draft_id)
                except Exception:
                    pass
                if _draft:
                    _comps = _build_editor_competitors(
                        _draft, _price_intel, _restaurant,
                    )
                    with _EDITOR_COMPETITORS_LOCK:
                        _EDITOR_COMPETITORS_CACHE[draft_id] = _comps
            except Exception as _ce:
                log.warning(
                    "Editor competitor pre-warm failed for draft %d: %s",
                    draft_id, _ce,
                )

        with _PRICE_ANALYSIS_LOCK:
            if result and result.get("error"):
                _PRICE_ANALYSIS_JOBS[draft_id] = {
                    "status": "error",
                    "error": result["error"],
                }
            else:
                _PRICE_ANALYSIS_JOBS[draft_id] = {"status": "done", "error": None}
    except Exception as e:
        with _PRICE_ANALYSIS_LOCK:
            _PRICE_ANALYSIS_JOBS[draft_id] = {
                "status": "error",
                "error": f"{type(e).__name__}: {e}",
            }


@app.get("/drafts/<int:draft_id>/analyzing-prices")
@login_required
def draft_analyzing_prices(draft_id: int):
    """Loading screen that kicks off Call 4 in a background thread and
    polls for completion before redirecting to the editor."""
    _require_drafts_storage()
    draft = drafts_store.get_draft(draft_id)
    if not draft:
        abort(404)

    rid = draft.get("restaurant_id")
    if not rid or not ai_price_intel:
        return redirect(url_for("draft_editor", draft_id=draft_id))

    with _PRICE_ANALYSIS_LOCK:
        job = _PRICE_ANALYSIS_JOBS.get(draft_id)
        if not job or job.get("status") == "error":
            _PRICE_ANALYSIS_JOBS[draft_id] = {"status": "running", "error": None}
            threading.Thread(
                target=_run_price_analysis_job,
                args=(draft_id, int(rid)),
                daemon=True,
            ).start()

    return _safe_render("analyzing_prices.html", draft=draft)


@app.get("/api/drafts/<int:draft_id>/analyzing-prices/status")
@login_required
def api_analyzing_prices_status(draft_id: int):
    """Poll endpoint for the analyzing-prices page."""
    with _PRICE_ANALYSIS_LOCK:
        job = _PRICE_ANALYSIS_JOBS.get(draft_id) or {"status": "unknown"}
    return jsonify(job)


@app.post("/drafts/<int:draft_id>/rerun_price_analysis")
@login_required
def rerun_price_analysis(draft_id: int):
    """Admin/dev helper: wipe price-intel cache for this draft's restaurant
    and re-run the full analyzer. Short-circuits the wizard re-run loop."""
    u = session.get("user") or {}
    dev_tools = os.environ.get("DEV_TOOLS_ENABLED", "").strip().lower() in ("1", "true", "yes")
    if u.get("role") != "admin" and not dev_tools:
        abort(403)

    _require_drafts_storage()
    draft = drafts_store.get_draft(draft_id)
    if not draft:
        abort(404)
    rid = draft.get("restaurant_id")
    if not rid or not ai_price_intel:
        flash("Restaurant or price intel not available.", "error")
        return redirect(url_for("draft_editor", draft_id=draft_id))

    # Wipe the caches that would block a fresh full run:
    # - per-restaurant Google Places results
    # - every cached competitor menu scrape (shared across restaurants, but
    #   for testing we clear globally so stale poisoned data can't linger)
    # - this draft's Claude Call 4 results + per-competitor comparisons
    # Cache is keyed by (zip, cuisine) — resolve from the restaurant so we
    # also clear empty-result cache rows (which have no price_comparison_results
    # rows and would otherwise be invisible to a results-table subquery).
    from storage.users import get_restaurant
    _rest = get_restaurant(int(rid)) or {}
    _zip = (_rest.get("zip_code") or "").strip().split("-", 1)[0]
    _cuisine = (_rest.get("cuisine_type") or "other").strip().lower()
    with db_connect() as conn:
        if _zip:
            conn.execute(
                "DELETE FROM price_comparison_cache WHERE zip_code=? AND cuisine_type=?",
                (_zip, _cuisine),
            )
        conn.execute("DELETE FROM price_comparison_results WHERE restaurant_id=?", (rid,))
        conn.execute("DELETE FROM competitor_menus")
        conn.execute("DELETE FROM competitor_comparisons WHERE draft_id=?", (draft_id,))
        conn.execute("DELETE FROM price_intelligence_results WHERE draft_id=?", (draft_id,))
        conn.execute("DELETE FROM price_intelligence_summary WHERE draft_id=?", (draft_id,))
        conn.commit()

    # Temp Places file + in-memory competitor list cache both get wiped
    # by _run_price_analysis_job's call to _invalidate_editor_cache, so we
    # don't need to duplicate that here.

    with _PRICE_ANALYSIS_LOCK:
        _PRICE_ANALYSIS_JOBS[draft_id] = {"status": "running", "error": None}
        threading.Thread(
            target=_run_price_analysis_job,
            args=(draft_id, int(rid)),
            daemon=True,
        ).start()

    return redirect(url_for("draft_analyzing_prices", draft_id=draft_id))


# Day 141.8: Web proxy for competitor browser iframe embedding.
# Fetches external page server-side and strips X-Frame-Options / CSP
# so it can render inside the editor's competitor browser panel.
@app.get("/browse/pdf/<pdf_id>")
def browse_pdf_data(pdf_id):
    """Serve cached PDF data for the inline viewer."""
    import tempfile
    pdf_path = os.path.join(tempfile.gettempdir(), f"menuflow_pdf_{pdf_id}.pdf")
    if not os.path.exists(pdf_path):
        return "PDF not found", 404
    with open(pdf_path, "rb") as f:
        data = f.read()
    response = make_response(data)
    response.headers["Content-Type"] = "application/pdf"
    response.headers["Content-Disposition"] = "inline"
    return response


@app.post("/browse/pdf-upload")
def browse_pdf_upload():
    """Accept a PDF file upload and return zoomable image viewer HTML."""
    if 'pdf' not in request.files:
        return "No PDF file", 400
    pdf_file = request.files['pdf']
    content = pdf_file.read()
    if not content:
        return "Empty file", 400
    try:
        import base64, io
        from pdf2image import convert_from_bytes
        images = convert_from_bytes(content, dpi=200)
        img_tags = ""
        for img in images:
            buf = io.BytesIO()
            img.save(buf, format="PNG", optimize=True)
            b64 = base64.b64encode(buf.getvalue()).decode()
            img_tags += f'<img src="data:image/png;base64,{b64}" style="width:100%;display:block;margin-bottom:4px;">\n'
        return f"""<!DOCTYPE html><html><head><meta charset="utf-8">
<style>
*{{margin:0;padding:0;box-sizing:border-box}}
body{{background:#2c2c2c;overflow:hidden}}
.tb{{position:fixed;top:0;left:0;right:0;height:36px;background:#333;display:flex;align-items:center;gap:6px;padding:0 10px;z-index:100}}
.tb button{{background:#555;border:none;color:#fff;padding:4px 10px;border-radius:4px;cursor:pointer;font-size:.75rem}}
.tb button:hover{{background:#777}}
.tb span{{color:#aaa;font-size:.75rem}}
#v{{position:fixed;top:36px;left:0;right:0;bottom:0;overflow:hidden;cursor:grab}}
#v.d{{cursor:grabbing}}
#p{{transform-origin:0 0;position:absolute}}
#p img{{display:block;margin-bottom:4px;box-shadow:0 2px 8px rgba(0,0,0,.4)}}
</style></head><body>
<div class="tb">
<button onclick="z(1.25)">+</button>
<button onclick="z(0.8)">&minus;</button>
<button onclick="f()">Fit</button>
<span id="zl">100%</span>
</div>
<div id="v"><div id="p">{img_tags}</div></div>
<script>
var s=1,px=0,py=0,p=document.getElementById('p'),v=document.getElementById('v');
function a(){{p.style.transform='translate('+px+'px,'+py+'px) scale('+s+')';document.getElementById('zl').textContent=Math.round(s*100)+'%';}}
function f(){{s=1;px=0;py=0;p.style.width=v.clientWidth+'px';a();}}
function z(fc){{var vw=v.clientWidth,vh=v.clientHeight,cx=vw/2,cy=vh/2,os=s;s=Math.min(10,Math.max(0.1,s*fc));var r=s/os;px=cx-r*(cx-px);py=cy-r*(cy-py);a();}}
var dr=false,sx,sy,ox,oy;
v.onmousedown=function(e){{e.preventDefault();dr=true;v.classList.add('d');sx=e.clientX;sy=e.clientY;ox=px;oy=py;}};
document.onmousemove=function(e){{if(!dr)return;px=ox+(e.clientX-sx);py=oy+(e.clientY-sy);a();}};
document.onmouseup=function(){{dr=false;v.classList.remove('d');}};
v.onwheel=function(e){{e.preventDefault();var r=v.getBoundingClientRect();var mx=e.clientX-r.left,my=e.clientY-r.top,os=s;var fc=e.deltaY>0?0.9:1.1;s=Math.min(10,Math.max(0.1,s*fc));var rt=s/os;px=mx-rt*(mx-px);py=my-rt*(my-py);a();}};
var imgs=p.querySelectorAll('img');var ld2=0;
imgs.forEach(function(im){{im.onload=function(){{ld2++;if(ld2===1)f();}};if(im.complete&&im.naturalWidth){{ld2++;if(ld2===1)f();}}}});
if(imgs.length===0)setTimeout(f,200);
</script></body></html>"""
    except Exception as e:
        return f"<h3 style='padding:40px;color:#999;text-align:center;'>Could not render PDF: {type(e).__name__}</h3>", 502


@app.get("/browse/pdf-viewer")
def browse_pdf_viewer():
    """Fetch a PDF URL, convert to images, return zoomable viewer HTML."""
    import requests as _req
    url = request.args.get("url", "").strip()
    if not url:
        return "No URL", 400
    # Unwrap any proxy prefixes
    from urllib.parse import unquote
    while '/browse?url=' in url or '/browse%3Furl%3D' in url:
        url = url.split('/browse?url=')[-1] if '/browse?url=' in url else unquote(url.split('/browse%3Furl%3D')[-1])
    if url.startswith('http://localhost') or url.startswith('http://127.0.0.1'):
        from urllib.parse import urlparse as _up, parse_qs
        p = _up(url)
        qs = parse_qs(p.query)
        if 'url' in qs:
            url = qs['url'][0]
    try:
        resp = _req.get(url, timeout=15, headers={
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36",
        }, allow_redirects=True)
        if resp.status_code != 200:
            return f"<h3 style='padding:40px;color:#999;text-align:center;'>PDF fetch failed: {resp.status_code}</h3>", 502
        import base64, io
        from pdf2image import convert_from_bytes
        images = convert_from_bytes(resp.content, dpi=200)
        img_tags = ""
        for img in images:
            buf = io.BytesIO()
            img.save(buf, format="PNG", optimize=True)
            b64 = base64.b64encode(buf.getvalue()).decode()
            img_tags += f'<img src="data:image/png;base64,{b64}" style="width:100%;display:block;margin-bottom:4px;">\n'
        viewer_html = f"""<!DOCTYPE html><html><head><meta charset="utf-8">
<style>
*{{margin:0;padding:0;box-sizing:border-box}}
body{{background:#2c2c2c;overflow:hidden}}
.tb{{position:fixed;top:0;left:0;right:0;height:36px;background:#333;display:flex;align-items:center;gap:6px;padding:0 10px;z-index:100}}
.tb button{{background:#555;border:none;color:#fff;padding:4px 10px;border-radius:4px;cursor:pointer;font-size:.75rem}}
.tb button:hover{{background:#777}}
.tb span{{color:#aaa;font-size:.75rem}}
#v{{position:fixed;top:36px;left:0;right:0;bottom:0;overflow:hidden;cursor:grab}}
#v.d{{cursor:grabbing}}
#p{{transform-origin:0 0;position:absolute}}
#p img{{display:block;margin-bottom:4px;box-shadow:0 2px 8px rgba(0,0,0,.4)}}
</style></head><body>
<div class="tb">
<button onclick="z(1.25)">+</button>
<button onclick="z(0.8)">&minus;</button>
<button onclick="f()">Fit</button>
<span id="zl">100%</span>
</div>
<div id="v"><div id="p">{img_tags}</div></div>
<script>
var s=1,px=0,py=0,p=document.getElementById('p'),v=document.getElementById('v');
function a(){{p.style.transform='translate('+px+'px,'+py+'px) scale('+s+')';document.getElementById('zl').textContent=Math.round(s*100)+'%';}}
function f(){{s=1;px=0;py=0;p.style.width=v.clientWidth+'px';a();}}
function z(fc){{var vw=v.clientWidth,vh=v.clientHeight,cx=vw/2,cy=vh/2,os=s;s=Math.min(10,Math.max(0.1,s*fc));var r=s/os;px=cx-r*(cx-px);py=cy-r*(cy-py);a();}}
var dr=false,sx,sy,ox,oy;
v.onmousedown=function(e){{e.preventDefault();dr=true;v.classList.add('d');sx=e.clientX;sy=e.clientY;ox=px;oy=py;}};
document.onmousemove=function(e){{if(!dr)return;px=ox+(e.clientX-sx);py=oy+(e.clientY-sy);a();}};
document.onmouseup=function(){{dr=false;v.classList.remove('d');}};
v.onwheel=function(e){{e.preventDefault();var r=v.getBoundingClientRect();var mx=e.clientX-r.left,my=e.clientY-r.top,os=s;var fc=e.deltaY>0?0.9:1.1;s=Math.min(10,Math.max(0.1,s*fc));var rt=s/os;px=mx-rt*(mx-px);py=my-rt*(my-py);a();}};
var imgs=p.querySelectorAll('img');var ld2=0;
imgs.forEach(function(im){{im.onload=function(){{ld2++;if(ld2===1)f();}};if(im.complete&&im.naturalWidth){{ld2++;if(ld2===1)f();}}}});
if(imgs.length===0)setTimeout(f,200);
</script></body></html>"""
        return viewer_html
    except Exception as e:
        return f"<h3 style='padding:40px;color:#999;text-align:center;'>Could not render PDF: {e}</h3>", 502


@app.get("/browse")
def browse_proxy():
    """Proxy an external URL for iframe embedding."""
    import requests as _req
    from urllib.parse import urljoin, urlparse

    url = request.args.get("url", "").strip()
    if not url:
        return "No URL provided", 400

    # Unwrap double-proxied URLs (e.g. /browse?url=/browse?url=https://...)
    from urllib.parse import unquote
    while '/browse?url=' in url or '/browse%3Furl%3D' in url:
        if '/browse?url=' in url:
            url = url.split('/browse?url=', 1)[1]
        elif '/browse%3Furl%3D' in url:
            url = unquote(url.split('/browse%3Furl%3D', 1)[1])
    # Also strip localhost prefix if the href rewriter added it
    if url.startswith('http://localhost') or url.startswith('http://127.0.0.1'):
        from urllib.parse import urlparse as _up
        p = _up(url)
        if p.query and 'url=' in p.query:
            url = p.query.split('url=', 1)[1]
            url = unquote(url)

    # Basic allowlist — only http/https
    parsed = urlparse(url)
    if parsed.scheme not in ("http", "https"):
        return "Invalid URL scheme", 400

    # Block sites with aggressive bot detection that won't work through proxy
    blocked_hosts = ("yelp.com", "tripadvisor.com", "facebook.com", "instagram.com", "doordash.com", "grubhub.com", "ubereats.com")
    host = (parsed.hostname or "").lower()
    if any(host == b or host.endswith("." + b) for b in blocked_hosts):
        return (f'<div style="display:flex;align-items:center;justify-content:center;height:100%;font-family:sans-serif;color:#666;text-align:center;padding:40px;">'
                f'<div><h3 style="margin:0 0 16px;">{parsed.hostname} blocks embedded viewing</h3>'
                f'<a href="{url}" target="_blank" style="color:#B85C38;font-weight:600;">Open directly in new tab &rarr;</a></div></div>'), 200

    try:
        resp = _req.get(url, timeout=15, headers={
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
                          "(KHTML, like Gecko) Chrome/131.0.0.0 Safari/537.36",
            "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
            "Accept-Language": "en-US,en;q=0.9",
        }, allow_redirects=True, verify=True)

        content = resp.content
        ct = resp.headers.get("Content-Type", "text/html")

        # Detect failed pages (403, tiny SPA shells, error pages)
        if resp.status_code == 403:
            return (f'<div style="display:flex;align-items:center;justify-content:center;height:100vh;font-family:sans-serif;color:#666;text-align:center;padding:40px;">'
                    f'<div><h3>This website blocked our request</h3>'
                    f'<p style="margin:8px 0 16px;">Try viewing it directly:</p>'
                    f'<a href="{url}" target="_blank" style="color:#B85C38;font-weight:600;font-size:1.1rem;">Open in new tab &rarr;</a></div></div>'), 200

        # PDF/binary files: convert to images or show error
        # Check content type AND first bytes for PDF magic number
        is_pdf = "pdf" in ct.lower() or url.lower().endswith(".pdf") or content[:5] == b"%PDF-"
        if is_pdf:
            import hashlib, tempfile, base64
            try:
                from pdf2image import convert_from_bytes
                images = convert_from_bytes(content, dpi=200)
                img_tags = ""
                for i, img in enumerate(images):
                    import io
                    buf = io.BytesIO()
                    img.save(buf, format="PNG", optimize=True)
                    b64 = base64.b64encode(buf.getvalue()).decode()
                    img_tags += f'<img src="data:image/png;base64,{b64}" style="width:100%;display:block;margin-bottom:4px;">\n'
                viewer_html = f"""<!DOCTYPE html><html><head><meta charset="utf-8">
<style>
*{{margin:0;padding:0;box-sizing:border-box}}
body{{background:#2c2c2c;overflow:hidden}}
.tb{{position:fixed;top:0;left:0;right:0;height:36px;background:#333;display:flex;align-items:center;gap:6px;padding:0 10px;z-index:100}}
.tb button{{background:#555;border:none;color:#fff;padding:4px 10px;border-radius:4px;cursor:pointer;font-size:.75rem}}
.tb button:hover{{background:#777}}
.tb span{{color:#aaa;font-size:.75rem}}
#v{{position:fixed;top:36px;left:0;right:0;bottom:0;overflow:hidden;cursor:grab}}
#v.d{{cursor:grabbing}}
#p{{transform-origin:0 0;position:absolute}}
#p img{{display:block;margin-bottom:4px;box-shadow:0 2px 8px rgba(0,0,0,.4)}}
</style></head><body>
<div class="tb">
<button onclick="z(1.25)">+</button>
<button onclick="z(0.8)">&minus;</button>
<button onclick="f()">Fit</button>
<span id="zl">100%</span>
</div>
<div id="v"><div id="p">{img_tags}</div></div>
<script>
var s=1,px=0,py=0,p=document.getElementById('p'),v=document.getElementById('v');
function a(){{p.style.transform='translate('+px+'px,'+py+'px) scale('+s+')';document.getElementById('zl').textContent=Math.round(s*100)+'%';}}
function f(){{s=1;px=0;py=0;p.style.width=v.clientWidth+'px';a();}}
function z(fc){{var vw=v.clientWidth,vh=v.clientHeight,cx=vw/2,cy=vh/2,os=s;s=Math.min(10,Math.max(0.1,s*fc));var r=s/os;px=cx-r*(cx-px);py=cy-r*(cy-py);a();}}
var dr=false,sx,sy,ox,oy;
v.onmousedown=function(e){{e.preventDefault();dr=true;v.classList.add('d');sx=e.clientX;sy=e.clientY;ox=px;oy=py;}};
document.onmousemove=function(e){{if(!dr)return;px=ox+(e.clientX-sx);py=oy+(e.clientY-sy);a();}};
document.onmouseup=function(){{dr=false;v.classList.remove('d');}};
v.onwheel=function(e){{e.preventDefault();var r=v.getBoundingClientRect();var mx=e.clientX-r.left,my=e.clientY-r.top,os=s;var fc=e.deltaY>0?0.9:1.1;s=Math.min(10,Math.max(0.1,s*fc));var rt=s/os;px=mx-rt*(mx-px);py=my-rt*(my-py);a();}};
var imgs=p.querySelectorAll('img');var ld2=0;
imgs.forEach(function(im){{im.onload=function(){{ld2++;if(ld2===1)f();}};if(im.complete&&im.naturalWidth){{ld2++;if(ld2===1)f();}}}});
if(imgs.length===0)setTimeout(f,200);
</script></body></html>"""
            except Exception as e:
                viewer_html = f'<html><body style="padding:40px;color:#999;text-align:center;"><h3>Could not render PDF: {type(e).__name__}</h3></body></html>'
            response = make_response(viewer_html)
            response.headers["Content-Type"] = "text/html; charset=utf-8"
            return response

        # Catch binary content that's not PDF and not text
        if "text/html" not in ct and "text/" not in ct and not is_pdf:
            # Binary file — don't render as text. Serve inline or show message.
            response = make_response(content)
            response.headers["Content-Type"] = ct
            response.headers["Content-Disposition"] = "inline"
            return response

        # For HTML pages: inject <base> tag so relative URLs resolve correctly
        if "text/html" in ct:
            base_url = resp.url  # final URL after redirects
            # Anti-frame-busting: inject BEFORE any other scripts to block window.top access
            anti_bust = b'<script>try{Object.defineProperty(window,"top",{get:function(){return window.self}});}catch(e){} try{Object.defineProperty(window,"parent",{get:function(){return window.self}});}catch(e){}</script>'
            base_tag = f'<base href="{base_url}" target="_self">'.encode()
            # Inject anti-bust + base tag after <head> (before any site scripts)
            if b"<head>" in content:
                content = content.replace(b"<head>", b"<head>" + anti_bust + base_tag, 1)
            elif b"<head " in content:
                idx = content.find(b">", content.find(b"<head "))
                if idx > 0:
                    content = content[:idx+1] + anti_bust + base_tag + content[idx+1:]
            elif b"<HEAD>" in content:
                content = content.replace(b"<HEAD>", b"<HEAD>" + anti_bust + base_tag, 1)
            else:
                content = anti_bust + base_tag + content

            # Keep navigation inside iframe — prevent top-window hijacking
            nav_intercept = b"""<script>
// Strip target=_blank/_top from all links
document.querySelectorAll('a[target]').forEach(function(a){
  var t=a.getAttribute('target');
  if(t==='_blank'||t==='_top'||t==='_parent') a.setAttribute('target','_self');
});
new MutationObserver(function(muts){
  muts.forEach(function(m){
    m.addedNodes.forEach(function(n){
      if(n.querySelectorAll) n.querySelectorAll('a[target]').forEach(function(a){
        var t=a.getAttribute('target');
        if(t==='_blank'||t==='_top'||t==='_parent') a.setAttribute('target','_self');
      });
    });
  });
}).observe(document.body,{childList:true,subtree:true});
// Override window.open
var _wo=window.open;
window.open=function(url,target){
  if(url && url.indexOf('http')===0){
    window.location.href='/browse?url='+encodeURIComponent(url);
    return null;
  }
  return _wo.apply(window,arguments);
};
// Block top-navigation attempts (IHOP-style frame busting)
try{
  if(window.top!==window.self){
    Object.defineProperty(window,'top',{get:function(){return window.self;}});
  }
}catch(e){}
try{
  if(window.parent!==window.self){
    Object.defineProperty(window,'parent',{get:function(){return window.self;}});
  }
}catch(e){}
</script>"""
            if b"</body>" in content:
                content = content.replace(b"</body>", nav_intercept + b"</body>", 1)
            elif b"</BODY>" in content:
                content = content.replace(b"</BODY>", nav_intercept + b"</BODY>", 1)
            else:
                content += nav_intercept

            # Rewrite links that navigate the frame to also go through proxy
            import re
            def _rewrite_href(m):
                href = m.group(2)
                if href.startswith(("#", "javascript:", "mailto:", "tel:")):
                    return m.group(0)
                abs_url = urljoin(base_url, href)
                return f'{m.group(1)}/browse?url={abs_url}'
            content = re.sub(
                rb'(href=["\'])([^"\']*)',
                lambda m: _rewrite_href(
                    type('M', (), {'group': lambda s, i: m.group(i).decode()})()
                ).encode(),
                content,
            )
            # Strip target="_blank" / "_top" / "_parent" from all tags
            content = re.sub(rb'\s+target=["\'](_blank|_top|_parent)["\']', b'', content, flags=re.IGNORECASE)
            # Detect frame-busting code — if present, show "open in new tab" instead
            has_framebust = bool(re.search(rb'top\.location\s*=\s*self\.location|if\s*\(\s*top\.frames\.length\s*!=\s*0\s*\)', content))
            if has_framebust:
                return (f'<div style="display:flex;align-items:center;justify-content:center;height:100vh;font-family:sans-serif;color:#666;text-align:center;padding:40px;">'
                        f'<div><h3 style="margin:0 0 16px;">{parsed.hostname} blocks embedded viewing</h3>'
                        f'<a href="{url}" target="_blank" style="color:#B85C38;font-weight:600;font-size:1.1rem;">Open directly in new tab &rarr;</a></div></div>'), 200

        response = make_response(content)
        response.headers["Content-Type"] = ct
        # Remove headers that block iframe embedding
        response.headers.pop("X-Frame-Options", None)
        response.headers.pop("Content-Security-Policy", None)
        # Force PDFs and other files to display inline (not download)
        response.headers["Content-Disposition"] = "inline"
        return response

    except _req.exceptions.Timeout:
        return "<h3 style='padding:40px;text-align:center;color:#999;'>Site took too long to respond</h3>", 504
    except Exception as e:
        return f"<h3 style='padding:40px;text-align:center;color:#999;'>Could not load page: {type(e).__name__}</h3>", 502


# Day 141.8: Resolve a competitor's website URL from their Google Place ID
@app.get("/api/competitor_website")
def api_competitor_website():
    """Look up a competitor's website URL by name, return proxied URL."""
    name = request.args.get("name", "").strip()
    addr = request.args.get("addr", "").strip()
    if not name:
        return jsonify({"error": "No name"}), 400

    # Try to find the place_id from our cached data
    from storage.price_intel import get_place_details
    website_url = None
    with db_connect() as conn:
        # Try exact name match first, then fuzzy
        row = conn.execute(
            "SELECT place_id FROM price_comparison_results WHERE place_name = ? LIMIT 1",
            (name,),
        ).fetchone()
        if not row:
            row = conn.execute(
                "SELECT place_id FROM price_comparison_results WHERE place_name LIKE ? LIMIT 1",
                (f"%{name}%",),
            ).fetchone()

    if row and row["place_id"]:
        details = get_place_details(row["place_id"])
        if details:
            website_url = details.get("website")

    # If not in DB, try a live Google Places lookup
    if not website_url:
        api_key = os.environ.get("GOOGLE_PLACES_API_KEY", "")
        if api_key:
            try:
                import requests as _rq2
                resp = _rq2.get("https://maps.googleapis.com/maps/api/place/findplacefromtext/json", params={
                    "key": api_key,
                    "input": f"{name} restaurant {addr}",
                    "inputtype": "textquery",
                    "fields": "place_id,name,formatted_address",
                }, timeout=5)
                if resp.status_code == 200:
                    candidates = resp.json().get("candidates", [])
                    if candidates and candidates[0].get("place_id"):
                        details = get_place_details(candidates[0]["place_id"])
                        if details:
                            website_url = details.get("website")
            except Exception:
                pass

    if not website_url:
        website_url = f"https://www.google.com/search?q={name}+{addr}+menu"

    return jsonify({"url": website_url, "proxy_url": f"/browse?url={website_url}"})


@app.get("/drafts/<int:draft_id>/source_file")
@login_required
def draft_source_file(draft_id: int):
    """Day 141.6: Serve original structured import file content as JSON rows
    for the Original Menu viewer table."""
    _require_drafts_storage()
    draft = drafts_store.get_draft(draft_id)
    if not draft:
        abort(404)
    try:
        import json as _json, csv as _csv
        src = _json.loads(draft.get("source") or "{}")
        meta = src.get("meta", {})
        # Try raw_rows from summary first (already parsed)
        summary = meta.get("summary", {})
        raw_rows = summary.get("raw_rows") or summary.get("sample_rows")
        if raw_rows:
            return jsonify({"rows": raw_rows, "filename": meta.get("filename", "")})
        # Fallback: read the file from disk
        fpath = meta.get("csv_path") or meta.get("file_path")
        if fpath:
            from pathlib import Path
            p = Path(fpath)
            if p.exists() and p.suffix.lower() == ".csv":
                with open(p, "r", encoding="utf-8-sig") as f:
                    reader = _csv.DictReader(f)
                    rows = [dict(r) for r in reader]
                return jsonify({"rows": rows, "filename": meta.get("filename", p.name)})
    except Exception:
        pass
    return jsonify({"rows": [], "filename": ""})


def _cleanup_empty_drafts(user_id: int, exclude_draft_id: int = 0):
    """Day 141.6: Auto-delete empty drafts (0 items).
    Called on navigation to avoid accumulating abandoned drafts."""
    if not drafts_store:
        return
    try:
        with db_connect() as conn:
            # Delete any draft with 0 items (regardless of status)
            empty_drafts = conn.execute(
                """SELECT d.id FROM drafts d
                   LEFT JOIN draft_items di ON di.draft_id = d.id
                   WHERE d.id != ?
                   GROUP BY d.id
                   HAVING COUNT(di.id) = 0""",
                (exclude_draft_id,),
            ).fetchall()
            for row in empty_drafts:
                try:
                    drafts_store.delete_draft(row["id"])
                except Exception:
                    pass
    except Exception:
        pass


@app.post("/drafts/<int:draft_id>/cleanup_if_empty")
@login_required
def draft_cleanup_if_empty(draft_id: int):
    """Day 141.6: Called via beacon when leaving an empty draft editor."""
    _require_drafts_storage()
    try:
        with db_connect() as conn:
            count = conn.execute(
                "SELECT COUNT(*) FROM draft_items WHERE draft_id = ?", (draft_id,)
            ).fetchone()[0]
        if count == 0:
            drafts_store.delete_draft(draft_id)
            return jsonify({"deleted": True}), 200
    except Exception:
        pass
    return jsonify({"deleted": False}), 200


@app.post("/drafts/<int:draft_id>/delete")
@login_required
def delete_draft_route(draft_id: int):
    """Day 141.6: Hard-delete a draft."""
    _require_drafts_storage()
    draft = drafts_store.get_draft(draft_id)
    if not draft:
        abort(404)
    # Determine redirect target
    rest_id = draft.get("restaurant_id")
    try:
        drafts_store.delete_draft(draft_id)
        flash("Draft deleted.", "success")
    except Exception as e:
        flash(f"Failed to delete draft: {e}", "error")
    if rest_id:
        return redirect(url_for("menus_page", rest_id=rest_id))
    return redirect(url_for("dashboard"))


@app.post("/drafts/<int:source_draft_id>/copy_to_restaurant")
@login_required
def copy_draft_to_restaurant(source_draft_id: int):
    """Day 141.6: Clone a draft and assign the clone to a different restaurant."""
    _require_drafts_storage()
    target_restaurant_id = request.form.get("restaurant_id")
    if not target_restaurant_id:
        flash("Please choose a target restaurant.", "error")
        return redirect(url_for("draft_editor", draft_id=source_draft_id))
    try:
        target_restaurant_id = int(target_restaurant_id)
    except (ValueError, TypeError):
        flash("Invalid restaurant.", "error")
        return redirect(url_for("draft_editor", draft_id=source_draft_id))

    if not hasattr(drafts_store, "clone_draft"):
        flash("Clone not supported.", "error")
        return redirect(url_for("draft_editor", draft_id=source_draft_id))

    try:
        clone = drafts_store.clone_draft(source_draft_id)
        new_id = int(clone.get("id") or clone.get("draft_id"))
        drafts_store.save_draft_metadata(new_id, restaurant_id=target_restaurant_id)
        flash(f"Menu copied to restaurant.", "success")
        return redirect(url_for("draft_editor", draft_id=new_id))
    except Exception as e:
        flash(f"Copy failed: {e}", "error")
        return redirect(url_for("draft_editor", draft_id=source_draft_id))


@app.post("/drafts/<int:draft_id>/wizard/unconfirm")
@login_required
def wizard_unconfirm_category(draft_id: int):
    """Unmark a category (or subcategory) to allow re-review."""
    _require_drafts_storage()
    category = (request.form.get("category") or "").strip()
    subcategory = (request.form.get("subcategory") or "").strip() or None
    if category:
        drafts_store.unmark_category_reviewed(draft_id, category, subcategory)
    kwargs = {"draft_id": draft_id, "category": category}
    if subcategory:
        kwargs["subcategory"] = subcategory
    return redirect(url_for("draft_wizard", **kwargs))


@app.get("/api/drafts/<int:draft_id>/wizard/progress")
@login_required
def wizard_progress_api(draft_id: int):
    """JSON endpoint for wizard progress (for AJAX polling)."""
    _require_drafts_storage()
    progress = drafts_store.get_wizard_progress(draft_id)
    return jsonify({"ok": True, **progress})


@app.post("/api/drafts/<int:draft_id>/wizard/item/<int:item_id>")
@login_required
def wizard_save_item(draft_id: int, item_id: int):
    """Save a single item edit from the wizard (AJAX)."""
    _require_drafts_storage()
    if not request.is_json:
        return jsonify({"ok": False, "error": "Expected JSON"}), 400
    data = request.get_json(silent=True) or {}

    incoming_category = (data.get("category") or "").strip()
    # Defense in depth: reject the literal string "None" or empty values.
    # Cheese-pizza-into-None bug (Day 141.9): the editor's wizSave was
    # serializing Python's `current_category` (which is None for the All
    # Items view) via Jinja, producing a literal "None" string in the
    # JSON payload. JS fix is in draft_editor.html, but defending here
    # guards every other code path too.
    if incoming_category in ("", "None", "none", "null"):
        # Look up the item's existing category and preserve it, rather
        # than clobber to a junk value.
        existing = drafts_store.get_draft_item(draft_id, item_id) if hasattr(
            drafts_store, "get_draft_item"
        ) else None
        if existing and existing.get("category"):
            incoming_category = existing["category"]
        else:
            with db_connect() as _conn:
                row = _conn.execute(
                    "SELECT category FROM draft_items WHERE id=? AND draft_id=?",
                    (item_id, draft_id),
                ).fetchone()
                if row and row["category"]:
                    incoming_category = row["category"]

    item = {
        "id": item_id,
        "name": (data.get("name") or "").strip(),
        "description": (data.get("description") or "").strip(),
        "price_cents": data.get("price_cents", 0),
        "category": incoming_category,
    }
    if not item["name"]:
        return jsonify({"ok": False, "error": "Name is required"}), 400

    try:
        drafts_store.upsert_draft_items(draft_id, [item])
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500

    return jsonify({"ok": True, "item_id": item_id})


@app.post("/api/drafts/<int:draft_id>/wizard/item/<int:item_id>/delete")
@login_required
def wizard_delete_item(draft_id: int, item_id: int):
    """Delete a single item from the wizard (AJAX)."""
    _require_drafts_storage()
    try:
        deleted = drafts_store.delete_draft_items(draft_id, [item_id])
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500
    return jsonify({"ok": True, "deleted": deleted})


@app.post("/api/drafts/<int:draft_id>/wizard/verify-category")
@login_required
def wizard_verify_category(draft_id: int):
    """Visual diff verification for one category (AJAX).

    Day 139.5v2: Crops the menu image to the category's region and asks Claude
    to compare extracted items vs the original menu (zoomed in).
    Returns corrections + missing items.
    """
    _require_drafts_storage()
    if not request.is_json:
        return jsonify({"ok": False, "error": "Expected JSON"}), 400
    data = request.get_json(silent=True) or {}
    category = (data.get("category") or "").strip()
    if not category:
        return jsonify({"ok": False, "error": "Category required"}), 400

    # Load items for this category
    all_items = drafts_store.get_draft_items(draft_id, include_variants=True) or []
    cat_items = [it for it in all_items if (it.get("category") or "").strip() == category]
    if not cat_items:
        return jsonify({"ok": False, "error": f"No items in '{category}'"}), 404

    # Load coordinates
    coordinates = {}
    try:
        raw_coords = drafts_store.get_draft_coordinates(draft_id)
        for c in raw_coords:
            coordinates[c["item_id"]] = {
                "x_pct": c["x_pct"],
                "y_pct": c["y_pct"],
                "w_pct": c["w_pct"],
                "h_pct": c["h_pct"],
                "page": c["page"],
            }
    except Exception:
        pass

    # Find source image path
    draft = drafts_store.get_draft(draft_id)
    source_path = None

    # Try source_file_path on draft (may be just a filename)
    sfp = draft.get("source_file_path") if draft else None
    if sfp:
        # Could be absolute or just a filename
        p = Path(sfp)
        if p.exists():
            source_path = str(p)
        else:
            # Try in UPLOAD_FOLDER
            p2 = UPLOAD_FOLDER / sfp
            if p2.exists():
                source_path = str(p2)

    # Fallback: try import job source_path or filename
    if not source_path:
        job_id = draft.get("source_job_id") if draft else None
        if job_id:
            try:
                with db_connect() as conn:
                    job = conn.execute(
                        "SELECT source_path, filename FROM import_jobs WHERE id=?",
                        (int(job_id),),
                    ).fetchone()
                    if job:
                        if job["source_path"] and Path(job["source_path"]).exists():
                            source_path = job["source_path"]
                        elif job["filename"]:
                            p3 = UPLOAD_FOLDER / job["filename"]
                            if p3.exists():
                                source_path = str(p3)
            except Exception:
                pass

    if not source_path:
        return jsonify({"ok": False, "error": "Source image not found"}), 404

    # Run visual diff verification
    try:
        from storage.ai_vision_verify import verify_category_visual
        result = verify_category_visual(
            str(source_path), category, cat_items, coordinates,
        )
        if result.get("error"):
            return jsonify({"ok": False, "error": result["error"]}), 500

        # Apply corrections to the database
        applied = []
        corrections = result.get("corrections") or []
        for corr in corrections:
            if not isinstance(corr, dict):
                continue
            item_id = corr.get("item_id")
            if not item_id:
                continue
            fixes = corr.get("fixes") or {}
            reason = corr.get("reason", "")

            # Apply field fixes
            update_fields = {}
            if "price_cents" in fixes:
                try:
                    update_fields["price_cents"] = int(round(float(fixes["price_cents"])))
                except (ValueError, TypeError):
                    pass
            if "name" in fixes:
                update_fields["name"] = str(fixes["name"]).strip()
            if "description" in fixes:
                update_fields["description"] = (str(fixes["description"]).strip()
                                                 if fixes["description"] else None)

            if update_fields:
                try:
                    with db_connect() as conn:
                        sets = ", ".join(f"{k}=?" for k in update_fields)
                        vals = list(update_fields.values()) + [_now_iso(), int(item_id), draft_id]
                        conn.execute(
                            f"UPDATE draft_items SET {sets}, updated_at=? "
                            f"WHERE id=? AND draft_id=?",
                            vals,
                        )
                        conn.commit()
                    applied.append({
                        "item_id": item_id,
                        "fixes": update_fields,
                        "reason": reason,
                    })
                except Exception as e:
                    print(f"[VisualDiff] Failed to apply fix for item {item_id}: {e}")

            # Apply variant fixes
            variant_fixes = corr.get("variant_fixes")
            if variant_fixes and isinstance(variant_fixes, list):
                try:
                    # Delete existing variants and re-insert
                    drafts_store.delete_all_variants_for_item(item_id)
                    for vi, vf in enumerate(variant_fixes):
                        if not isinstance(vf, dict):
                            continue
                        lbl = (vf.get("label") or "").strip()
                        vp = 0
                        try:
                            vp = int(round(float(vf.get("price_cents", 0))))
                        except (ValueError, TypeError):
                            pass
                        if lbl or vp:
                            drafts_store.insert_variants(item_id, [{
                                "label": lbl or f"Size {vi + 1}",
                                "price_cents": vp,
                                "kind": vf.get("kind", "size"),
                                "position": vi,
                            }])
                    applied.append({
                        "item_id": item_id,
                        "variant_fixes": len(variant_fixes),
                        "reason": reason,
                    })
                except Exception as e:
                    print(f"[VisualDiff] Failed to apply variant fix for item {item_id}: {e}")

        # Add missing items
        added_items = []
        for mi in (result.get("missing_items") or []):
            if not isinstance(mi, dict):
                continue
            name = (mi.get("name") or "").strip()
            if not name:
                continue
            price_cents = 0
            try:
                price_cents = int(round(float(mi.get("price_cents", 0))))
            except (ValueError, TypeError):
                pass
            try:
                new_item = drafts_store.insert_draft_item(draft_id, {
                    "name": name,
                    "description": (mi.get("description") or "").strip() or None,
                    "price_cents": price_cents,
                    "category": category,
                    "confidence": 85,
                })
                added_items.append({"name": name, "price_cents": price_cents})
            except Exception as e:
                print(f"[VisualDiff] Failed to add missing item '{name}': {e}")

        return jsonify({
            "ok": True,
            "category": category,
            "corrections_applied": len(applied),
            "items_added": len(added_items),
            "details": applied,
            "added": added_items,
            "notes": result.get("notes", ""),
        })

    except Exception as e:
        print(f"[VisualDiff] Error: {e}")
        return jsonify({"ok": False, "error": str(e)}), 500


@app.post("/api/drafts/<int:draft_id>/wizard/variant/<int:variant_id>")
@login_required
def wizard_save_variant(draft_id: int, variant_id: int):
    """Save a single variant edit from the wizard (AJAX)."""
    _require_drafts_storage()
    if not request.is_json:
        return jsonify({"ok": False, "error": "Expected JSON"}), 400
    data = request.get_json(silent=True) or {}
    label = (data.get("label") or "").strip()
    price_cents = data.get("price_cents", 0)

    if not label:
        return jsonify({"ok": False, "error": "Label is required"}), 400

    try:
        with db_connect() as conn:
            conn.execute(
                "UPDATE draft_item_variants SET label=?, price_cents=?, updated_at=? WHERE id=?",
                (label, int(price_cents), _now_iso(), variant_id),
            )
            conn.commit()
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500

    return jsonify({"ok": True, "variant_id": variant_id})


@app.post("/api/drafts/<int:draft_id>/wizard/variant/<int:variant_id>/delete")
@login_required
def wizard_delete_variant(draft_id: int, variant_id: int):
    """Delete a single variant (AJAX)."""
    _require_drafts_storage()
    try:
        with db_connect() as conn:
            conn.execute("DELETE FROM draft_item_variants WHERE id=?", (variant_id,))
            conn.commit()
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500
    return jsonify({"ok": True})


@app.post("/api/drafts/<int:draft_id>/wizard/item/<int:item_id>/add_variant")
@login_required
def wizard_add_variant(draft_id: int, item_id: int):
    """Add a new blank variant to an item (AJAX)."""
    _require_drafts_storage()
    try:
        with db_connect() as conn:
            conn.execute(
                """INSERT INTO draft_item_variants (item_id, label, price_cents, kind, position, created_at, updated_at)
                   VALUES (?, '', 0, 'size', (SELECT COALESCE(MAX(position),0)+1 FROM draft_item_variants WHERE item_id=?), ?, ?)""",
                (item_id, item_id, _now_iso(), _now_iso()),
            )
            new_id = conn.execute("SELECT last_insert_rowid()").fetchone()[0]
            conn.commit()
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500
    return jsonify({"ok": True, "variant_id": new_id})


@app.post("/api/drafts/<int:draft_id>/wizard/reorder")
@login_required
def wizard_reorder(draft_id: int):
    """Save new item positions after drag reorder (AJAX)."""
    _require_drafts_storage()
    if not request.is_json:
        return jsonify({"ok": False, "error": "Expected JSON"}), 400
    data = request.get_json(silent=True) or {}
    order = data.get("order") or []  # list of item_id ints in new order

    if not order:
        return jsonify({"ok": False, "error": "No order provided"}), 400

    try:
        with db_connect() as conn:
            for pos, item_id in enumerate(order, start=1):
                conn.execute(
                    "UPDATE draft_items SET position=?, updated_at=? WHERE id=? AND draft_id=?",
                    (pos, _now_iso(), int(item_id), draft_id),
                )
            conn.commit()
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500

    return jsonify({"ok": True, "count": len(order)})


@app.post("/api/drafts/<int:draft_id>/wizard/apply_variant_labels")
@login_required
def wizard_apply_variant_labels(draft_id: int):
    """Apply checked variant labels to all items in a category.

    For items with existing variants at the target position: update the label.
    For items with fewer variants (or none): create new variant rows.
    """
    _require_drafts_storage()
    if not request.is_json:
        return jsonify({"ok": False, "error": "Expected JSON"}), 400
    data = request.get_json(silent=True) or {}
    source_item_id = data.get("source_item_id")
    category = (data.get("category") or "").strip()
    if not source_item_id or not category:
        return jsonify({"ok": False, "error": "source_item_id and category required"}), 400
    # Subcategory scoping: only apply to items in the same subcategory
    req_subcat = data.get("subcategory")  # null/missing = main items only

    # Only apply specific positions (0-indexed). If omitted, apply all.
    positions = data.get("positions")
    # mode: "label" = apply labels by position (default), "price" = apply price to matching label names
    mode = data.get("mode", "label")

    try:
        items = drafts_store.get_draft_items(draft_id, include_modifier_groups=False) or []
        cat_items = [it for it in items
                     if (it.get("category") or "").strip() == category
                     and (it.get("subcategory") or None) == (req_subcat or None)]

        # Get source item's variants (sorted by position)
        source = None
        for it in cat_items:
            if it["id"] == source_item_id:
                source = it
                break
        if not source or not source.get("variants"):
            return jsonify({"ok": False, "error": "Source item has no variants"}), 400

        src_variants = sorted(source["variants"], key=lambda v: v.get("position", 0))
        src_count = len(src_variants)
        apply_positions = sorted(set(positions) if positions is not None else set(range(src_count)))

        now = _now_iso()
        updated_items = 0
        old_labels = []   # for undo: existing variants that were changed
        created_ids = []  # for undo: new variant IDs that were created

        # renames: {old_label: new_label} for rename operations
        renames = data.get("renames") or {}
        # price_updates: {label: price_cents} for price-only updates by name
        price_updates = data.get("price_updates") or {}

        with db_connect() as conn:
            # Price updates: match by label name, update price
            if price_updates:
                pu_lower = {k.strip().lower(): v for k, v in price_updates.items()}
                for it in cat_items:
                    if it["id"] == source_item_id:
                        continue
                    changed = False
                    for tv in (it.get("variants") or []):
                        tv_label_lower = (tv.get("label") or "").strip().lower()
                        if tv_label_lower in pu_lower:
                            new_price = pu_lower[tv_label_lower]
                            if tv.get("price_cents", 0) != new_price:
                                old_labels.append({
                                    "variant_id": tv["id"],
                                    "old_label": tv.get("label", ""),
                                    "price_cents": tv.get("price_cents", 0),
                                })
                                conn.execute(
                                    "UPDATE draft_item_variants SET price_cents=?, updated_at=? WHERE id=?",
                                    (new_price, now, tv["id"]),
                                )
                                changed = True
                    if changed:
                        updated_items += 1

            if renames:
                # Rename mode: find variants by old label, update label + price
                # Build lookup: old_label_lower → {new_label, price_cents}
                rename_lower = {}
                for old_lbl, new_lbl in renames.items():
                    # Find the source variant's price for this rename
                    price = 0
                    for sv in src_variants:
                        if (sv.get("label") or "").strip().lower() == new_lbl.strip().lower():
                            price = sv.get("price_cents", 0)
                            break
                    rename_lower[old_lbl.strip().lower()] = {"label": new_lbl, "price_cents": price}

                for it in cat_items:
                    if it["id"] == source_item_id:
                        continue
                    changed = False
                    for tv in (it.get("variants") or []):
                        tv_label_lower = (tv.get("label") or "").strip().lower()
                        if tv_label_lower in rename_lower:
                            rl = rename_lower[tv_label_lower]
                            old_labels.append({
                                "variant_id": tv["id"],
                                "old_label": tv.get("label", ""),
                                "price_cents": tv.get("price_cents", 0),
                            })
                            conn.execute(
                                "UPDATE draft_item_variants SET label=?, price_cents=?, updated_at=? WHERE id=?",
                                (rl["label"], rl["price_cents"], now, tv["id"]),
                            )
                            changed = True
                    if changed:
                        updated_items += 1

            elif mode == "price":
                # Price mode: match by label name, update price
                # Build lookup of source labels → price for checked positions
                src_label_price = {}
                for pos in apply_positions:
                    if pos < src_count:
                        sv = src_variants[pos]
                        label_key = (sv.get("label") or "").strip().lower()
                        if label_key:
                            src_label_price[label_key] = sv.get("price_cents", 0)

                for it in cat_items:
                    if it["id"] == source_item_id:
                        continue
                    changed = False
                    for tv in (it.get("variants") or []):
                        tv_label = (tv.get("label") or "").strip().lower()
                        if tv_label in src_label_price:
                            new_price = src_label_price[tv_label]
                            old_labels.append({
                                "variant_id": tv["id"],
                                "old_label": tv.get("label", ""),
                                "price_cents": tv.get("price_cents", 0),
                            })
                            conn.execute(
                                "UPDATE draft_item_variants SET price_cents=?, updated_at=? WHERE id=?",
                                (new_price, now, tv["id"]),
                            )
                            changed = True
                    if changed:
                        updated_items += 1
            else:
                # Label mode: for each checked variant, create it on items that
                # don't already have a variant with that label. Never overwrite.
                for it in cat_items:
                    if it["id"] == source_item_id:
                        continue

                    tgt_variants = it.get("variants") or []
                    existing_labels = {
                        (v.get("label") or "").strip().lower()
                        for v in tgt_variants
                    }
                    max_pos = max((v.get("position", 0) for v in tgt_variants), default=-1)
                    changed = False

                    for pos in apply_positions:
                        if pos >= src_count:
                            continue
                        sv = src_variants[pos]
                        sv_label_lower = (sv.get("label") or "").strip().lower()

                        if sv_label_lower and sv_label_lower not in existing_labels:
                            # Create the variant — doesn't exist on this item
                            max_pos += 1
                            conn.execute(
                                """INSERT INTO draft_item_variants
                                   (item_id, label, price_cents, kind, position, created_at, updated_at)
                                   VALUES (?, ?, ?, 'size', ?, ?, ?)""",
                                (it["id"], sv["label"], sv.get("price_cents", 0), max_pos, now, now),
                            )
                            new_id = conn.execute("SELECT last_insert_rowid()").fetchone()[0]
                            created_ids.append({
                                "variant_id": new_id,
                                "item_id": it["id"],
                                "label": sv["label"],
                                "price_cents": sv.get("price_cents", 0),
                                "position": max_pos,
                            })
                            existing_labels.add(sv_label_lower)
                            changed = True

                    if changed:
                        updated_items += 1
            conn.commit()

    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500

    return jsonify({
        "ok": True,
        "updated_items": updated_items,
        "label_count": len(apply_positions),
        "old_labels": old_labels,
        "created_ids": created_ids,
    })


@app.post("/api/drafts/<int:draft_id>/wizard/delete_variants_by_label")
@login_required
def wizard_delete_variants_by_label(draft_id: int):
    """Delete all variants matching given labels from all items in a category."""
    _require_drafts_storage()
    if not request.is_json:
        return jsonify({"ok": False, "error": "Expected JSON"}), 400
    data = request.get_json(silent=True) or {}
    source_item_id = data.get("source_item_id")
    category = (data.get("category") or "").strip()
    labels = data.get("labels") or []  # list of label strings to delete
    if not category or not labels:
        return jsonify({"ok": False, "error": "category and labels required"}), 400
    req_subcat = data.get("subcategory")  # null/missing = main items only

    label_set = {l.strip().lower() for l in labels if l.strip()}
    if not label_set:
        return jsonify({"ok": False, "error": "No valid labels provided"}), 400

    try:
        items = drafts_store.get_draft_items(draft_id, include_modifier_groups=False) or []
        cat_items = [it for it in items
                     if (it.get("category") or "").strip() == category
                     and (it.get("subcategory") or None) == (req_subcat or None)]

        deleted_variants = []  # for undo: [{variant_id, item_id, label, price_cents, kind, position}]
        updated_items = 0

        with db_connect() as conn:
            for it in cat_items:
                # Include all items (source card too)
                changed = False
                for tv in (it.get("variants") or []):
                    if (tv.get("label") or "").strip().lower() in label_set:
                        deleted_variants.append({
                            "variant_id": tv["id"],
                            "item_id": it["id"],
                            "label": tv.get("label", ""),
                            "price_cents": tv.get("price_cents", 0),
                            "kind": tv.get("kind", "size"),
                            "position": tv.get("position", 0),
                        })
                        conn.execute("DELETE FROM draft_item_variants WHERE id=?", (tv["id"],))
                        changed = True
                if changed:
                    updated_items += 1
            conn.commit()

    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500

    return jsonify({
        "ok": True,
        "updated_items": updated_items,
        "deleted_count": len(deleted_variants),
        "deleted_variants": deleted_variants,
    })


@app.post("/api/drafts/<int:draft_id>/wizard/batch_undo")
@login_required
def wizard_batch_undo(draft_id: int):
    """Batch undo: delete created variants and restore old labels/prices in one call."""
    _require_drafts_storage()
    if not request.is_json:
        return jsonify({"ok": False, "error": "Expected JSON"}), 400
    data = request.get_json(silent=True) or {}
    delete_ids = data.get("delete_ids") or []       # variant IDs to delete
    restore = data.get("restore") or []             # [{variant_id, label, price_cents}]

    try:
        now = _now_iso()
        with db_connect() as conn:
            for vid in delete_ids:
                conn.execute("DELETE FROM draft_item_variants WHERE id=?", (int(vid),))
            for r in restore:
                conn.execute(
                    "UPDATE draft_item_variants SET label=?, price_cents=?, updated_at=? WHERE id=?",
                    (r.get("label", ""), int(r.get("price_cents", 0)), now, int(r["variant_id"])),
                )
            conn.commit()
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500

    return jsonify({"ok": True, "deleted": len(delete_ids), "restored": len(restore)})


@app.post("/api/drafts/<int:draft_id>/wizard/batch_create_variants")
@login_required
def wizard_batch_create_variants(draft_id: int):
    """Batch create variants: insert multiple variants across items in one call.

    Day 141: returns the new variant IDs (in the same order as the input list)
    so clients can update the DOM in-place without a full page reload.
    """
    _require_drafts_storage()
    if not request.is_json:
        return jsonify({"ok": False, "error": "Expected JSON"}), 400
    data = request.get_json(silent=True) or {}
    variants = data.get("variants") or []  # [{item_id, label, price_cents, kind, position}]

    new_ids = []
    try:
        now = _now_iso()
        with db_connect() as conn:
            for v in variants:
                cur = conn.execute(
                    """INSERT INTO draft_item_variants
                       (item_id, label, price_cents, kind, position, created_at, updated_at)
                       VALUES (?, ?, ?, ?, ?, ?, ?)""",
                    (int(v["item_id"]), v.get("label", ""), int(v.get("price_cents", 0)),
                     v.get("kind", "size"), int(v.get("position", 0)), now, now),
                )
                new_ids.append(cur.lastrowid)
            conn.commit()
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500

    return jsonify({"ok": True, "created": len(variants), "variant_ids": new_ids})


@app.post("/api/drafts/<int:draft_id>/wizard/bulk_move")
@login_required
def wizard_bulk_move(draft_id: int):
    """Move items to a different category (or a new one)."""
    _require_drafts_storage()
    if not request.is_json:
        return jsonify({"ok": False, "error": "Expected JSON"}), 400
    data = request.get_json(silent=True) or {}
    item_ids = data.get("item_ids") or []
    target_category = (data.get("target_category") or "").strip()
    target_subcategory = (data.get("target_subcategory") or "").strip() or None
    if not item_ids or not target_category:
        return jsonify({"ok": False, "error": "item_ids and target_category required"}), 400

    try:
        now = _now_iso()
        with db_connect() as conn:
            for iid in item_ids:
                conn.execute(
                    "UPDATE draft_items SET category=?, subcategory=?, updated_at=? WHERE id=? AND draft_id=?",
                    (target_category, target_subcategory, now, int(iid), draft_id),
                )
            conn.commit()
        # Re-initialize wizard categories to pick up the new category
        drafts_store.init_wizard_categories(draft_id)
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500

    return jsonify({"ok": True, "moved": len(item_ids), "target_category": target_category})


@app.post("/api/drafts/<int:draft_id>/wizard/bulk_delete")
@login_required
def wizard_bulk_delete(draft_id: int):
    """Delete multiple items at once."""
    _require_drafts_storage()
    if not request.is_json:
        return jsonify({"ok": False, "error": "Expected JSON"}), 400
    data = request.get_json(silent=True) or {}
    item_ids = data.get("item_ids") or []
    if not item_ids:
        return jsonify({"ok": False, "error": "item_ids required"}), 400

    try:
        deleted = drafts_store.delete_draft_items(draft_id, [int(i) for i in item_ids])
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500

    return jsonify({"ok": True, "deleted": deleted})


@app.post("/api/drafts/<int:draft_id>/wizard/rename_category")
@login_required
def wizard_rename_category(draft_id: int):
    """Rename a category across all items and wizard tracking."""
    _require_drafts_storage()
    if not request.is_json:
        return jsonify({"ok": False, "error": "Expected JSON"}), 400
    data = request.get_json(silent=True) or {}
    old_name = (data.get("old_name") or "").strip()
    new_name = (data.get("new_name") or "").strip()
    if not old_name or not new_name:
        return jsonify({"ok": False, "error": "old_name and new_name required"}), 400

    try:
        now = _now_iso()
        with db_connect() as conn:
            # Update all items in this category
            conn.execute(
                "UPDATE draft_items SET category=?, updated_at=? WHERE draft_id=? AND category=?",
                (new_name, now, draft_id, old_name),
            )
            # Update wizard category tracking
            conn.execute(
                "UPDATE draft_category_reviews SET category=? WHERE draft_id=? AND category=?",
                (new_name, draft_id, old_name),
            )
            # Update category order if stored
            conn.commit()
        # Update category_order JSON on the draft
        try:
            order = drafts_store.get_category_order(draft_id)
            if order and old_name in order:
                order = [new_name if c == old_name else c for c in order]
                drafts_store.save_category_order(draft_id, order)
        except Exception:
            pass
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500

    return jsonify({"ok": True, "old_name": old_name, "new_name": new_name})


@app.post("/api/drafts/<int:draft_id>/wizard/add_item")
@login_required
def wizard_add_item(draft_id: int):
    """Add a new blank item to a category (AJAX)."""
    _require_drafts_storage()
    if not request.is_json:
        return jsonify({"ok": False, "error": "Expected JSON"}), 400
    data = request.get_json(silent=True) or {}
    category = (data.get("category") or "").strip()
    if not category:
        return jsonify({"ok": False, "error": "Category required"}), 400
    subcategory = (data.get("subcategory") or "").strip() or None

    # Insert at the end of the category
    result = drafts_store.upsert_draft_items(draft_id, [{
        "name": "",
        "description": "",
        "price_cents": 0,
        "category": category,
        "subcategory": subcategory,
    }])

    new_id = result.get("inserted_ids", [None])[0]
    return jsonify({"ok": True, "item_id": new_id})


@app.post("/api/drafts/<int:draft_id>/wizard/move_subcategory")
@login_required
def wizard_move_subcategory(draft_id: int):
    """Move all items in a subcategory to a different parent category."""
    _require_drafts_storage()
    if not request.is_json:
        return jsonify({"ok": False, "error": "Expected JSON"}), 400
    data = request.get_json(silent=True) or {}
    subcategory = (data.get("subcategory") or "").strip()
    old_category = (data.get("old_category") or "").strip()
    new_category = (data.get("new_category") or "").strip()
    if not subcategory or not old_category or not new_category:
        return jsonify({"ok": False, "error": "subcategory, old_category, and new_category required"}), 400

    try:
        now = _now_iso()
        with db_connect() as conn:
            result = conn.execute(
                "UPDATE draft_items SET category=?, updated_at=? "
                "WHERE draft_id=? AND category=? AND subcategory=?",
                (new_category, now, draft_id, old_category, subcategory),
            )
            conn.commit()
        drafts_store.init_wizard_categories(draft_id)
        return jsonify({"ok": True, "moved": result.rowcount})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500


@app.post("/api/drafts/<int:draft_id>/wizard/delete_category")
@login_required
def wizard_delete_category(draft_id: int):
    """Delete an empty category or subcategory (removes any placeholder items)."""
    _require_drafts_storage()
    if not request.is_json:
        return jsonify({"ok": False, "error": "Expected JSON"}), 400
    data = request.get_json(silent=True) or {}
    category = (data.get("category") or "").strip()
    subcategory = data.get("subcategory")
    force = bool(data.get("force"))
    if not category:
        return jsonify({"ok": False, "error": "category required"}), 400

    try:
        with db_connect() as conn:
            if subcategory:
                if force:
                    # Delete ALL items in this subcategory
                    conn.execute(
                        "DELETE FROM draft_items WHERE draft_id=? AND category=? AND subcategory=?",
                        (draft_id, category, subcategory),
                    )
                else:
                    # Only delete if empty or just placeholders
                    conn.execute(
                        "DELETE FROM draft_items WHERE draft_id=? AND category=? AND subcategory=? AND (name='' OR name='New Item') AND price_cents=0",
                        (draft_id, category, subcategory),
                    )
                    remaining = conn.execute(
                        "SELECT COUNT(*) FROM draft_items WHERE draft_id=? AND category=? AND subcategory=?",
                        (draft_id, category, subcategory),
                    ).fetchone()[0]
                    if remaining > 0:
                        return jsonify({"ok": False, "error": "Subcategory is not empty"}), 400
            else:
                if force:
                    # Delete ALL items in this category (including all subcategories)
                    conn.execute(
                        "DELETE FROM draft_items WHERE draft_id=? AND category=?",
                        (draft_id, category),
                    )
                else:
                    conn.execute(
                        "DELETE FROM draft_items WHERE draft_id=? AND category=? AND (name='' OR name='New Item') AND price_cents=0",
                        (draft_id, category),
                    )
                    remaining = conn.execute(
                        "SELECT COUNT(*) FROM draft_items WHERE draft_id=? AND category=?",
                        (draft_id, category),
                    ).fetchone()[0]
                    if remaining > 0:
                        return jsonify({"ok": False, "error": "Category is not empty"}), 400
            conn.commit()
        drafts_store.init_wizard_categories(draft_id)
        return jsonify({"ok": True})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500


# --- NEW: Draft status probe for polling ---
@app.get("/drafts/<int:draft_id>/status")
@login_required
def draft_status(draft_id: int):
    _require_drafts_storage()
    d = drafts_store.get_draft(draft_id) or {}
    status = (d.get("status") or "editing")
    return jsonify({"ok": True, "status": status, "draft_id": draft_id})

# --- AI Cleanup route (now supports AJAX JSON or redirect, with status flips) ---
@app.post("/drafts/<int:draft_id>/cleanup")
@login_required
def cleanup_draft(draft_id: int):
    from storage.ai_cleanup import apply_ai_cleanup
    _require_drafts_storage()

    # Detect redirect vs JSON
    ct = (request.headers.get("Content-Type") or "").lower()
    is_form_post = ct.startswith("application/x-www-form-urlencoded") or ct.startswith("multipart/form-data")
    wants_redirect = (
        request.args.get("redirect") == "1"
        or (request.form.get("redirect") == "1" if is_form_post else False)
        or is_form_post
        or (request.headers.get("X-Requested-With") == "fetch" and "text/html" in (request.headers.get("Accept") or ""))
    )

    # Flip to processing for polling UIs
    try:
        drafts_store.save_draft_metadata(int(draft_id), status="processing")
    except Exception:
        pass

    try:
        updated = apply_ai_cleanup(int(draft_id))
        try:
            drafts_store.save_draft_metadata(int(draft_id), status="finalized")
        except Exception:
            pass

        if wants_redirect:
            flash(f"AI cleanup complete: {updated} item(s) updated.", "success")
            return redirect(url_for("draft_editor", draft_id=int(draft_id)))
        return jsonify({"ok": True, "updated": int(updated), "status": "finalized"}), 200

    except Exception as e:
        app.logger.exception("AI cleanup failed")
        try:
            drafts_store.save_draft_metadata(int(draft_id), status="editing")
        except Exception:
            pass
        if wants_redirect:
            flash(f"AI cleanup failed: {e}", "error")
            return redirect(url_for("draft_editor", draft_id=int(draft_id)))
        return jsonify({"ok": False, "error": str(e), "status": "editing"}), 500

@app.post("/drafts/<int:draft_id>/save")
@login_required
def draft_save(draft_id: int):
    """Save title and bulk upsert/delete items. Expects JSON payload."""
    _require_drafts_storage()
    if not request.is_json:
        return jsonify({"ok": False, "error": "Expected JSON payload"}), 400
    payload = request.get_json(silent=True) or {}

    if payload.get("autosave_ping"):
        return jsonify({"ok": True, "saved_at": _now_iso(), "ping": True}), 200

    # Day 83: Block saves on approved/published drafts
    draft = drafts_store.get_draft(draft_id)
    if draft and draft.get("status") in ("approved", "published"):
        return jsonify({"ok": False, "error": "Draft is approved and cannot be edited"}), 403

    # 🔒 Validate payload contract (prevents UI/AI drift)
    probe = {
        "draft_id": draft_id,
        "items": payload.get("items") or [],
        # extra fields tolerated by the validator (ignored if present)
        "title": payload.get("title"),
        "restaurant_id": payload.get("restaurant_id"),
        "status": payload.get("status"),
        "deleted_modifier_group_ids": payload.get("deleted_modifier_group_ids") or [],
    }
    ok, err = validate_draft_payload(probe)
    if not ok:
        return jsonify({"ok": False, "error": f"schema: {err}"}), 400

    title = (payload.get("title") or "").strip() or None
    items = payload.get("items") or []
    deleted_ids = payload.get("deleted_item_ids") or []
    deleted_variant_ids = payload.get("deleted_variant_ids") or []
    deleted_modifier_group_ids = payload.get("deleted_modifier_group_ids") or []
    # Day 119: per-item modifier_groups[] for full save lifecycle
    modifier_groups_by_item = payload.get("modifier_groups_by_item") or {}

    # Day 119: server-side validation warnings for modifier groups
    save_warnings = []
    if isinstance(modifier_groups_by_item, dict):
        for _item_key, groups in modifier_groups_by_item.items():
            if not isinstance(groups, list):
                continue
            for grp in groups:
                if not isinstance(grp, dict):
                    continue
                if grp.get("required") and not (grp.get("modifiers") or []):
                    gname = grp.get("name") or "Unnamed group"
                    save_warnings.append(
                        f"Required modifier group \"{gname}\" has no modifiers defined."
                    )

    try:
        if title is not None:
            drafts_store.save_draft_metadata(draft_id, title=title)
        upsert_result = drafts_store.upsert_draft_items(draft_id, items)
        deleted_count = 0
        if deleted_ids:
            del_ints = []
            for x in deleted_ids:
                try:
                    del_ints.append(int(x))
                except Exception:
                    continue
            if del_ints:
                deleted_count = drafts_store.delete_draft_items(draft_id, del_ints)
        # Delete orphaned variant rows
        if deleted_variant_ids:
            for vid in deleted_variant_ids:
                try:
                    vid_int = int(vid)
                    drafts_store.delete_variants_by_id([vid_int])
                except Exception:
                    continue
        # Delete orphaned modifier group rows (Day 113)
        deleted_mg_count = 0
        if deleted_modifier_group_ids:
            for mgid in deleted_modifier_group_ids:
                try:
                    if drafts_store.delete_modifier_group(int(mgid)):
                        deleted_mg_count += 1
                except Exception:
                    continue
        # Day 119: sync modifier group metadata + modifiers from payload
        mg_synced = 0
        if isinstance(modifier_groups_by_item, dict):
            for _item_key, groups in modifier_groups_by_item.items():
                if not isinstance(groups, list):
                    continue
                for grp in groups:
                    if not isinstance(grp, dict):
                        continue
                    try:
                        gid = int(grp.get("id") or 0)
                    except (ValueError, TypeError):
                        continue
                    if not gid:
                        continue
                    # Update group metadata
                    update_kwargs = {}
                    if "name" in grp:
                        update_kwargs["name"] = str(grp["name"]).strip()
                    if "required" in grp:
                        update_kwargs["required"] = bool(grp["required"])
                    if "min_select" in grp:
                        try:
                            update_kwargs["min_select"] = int(grp["min_select"])
                        except (ValueError, TypeError):
                            pass
                    if "max_select" in grp:
                        try:
                            update_kwargs["max_select"] = int(grp["max_select"])
                        except (ValueError, TypeError):
                            pass
                    # Confirm group exists before counting as synced
                    found = bool(drafts_store.get_modifier_group(gid))
                    if not found:
                        continue
                    if update_kwargs:
                        drafts_store.update_modifier_group(gid, **update_kwargs)
                    # Full-replace modifiers for this group
                    if "modifiers" in grp:
                        drafts_store.upsert_group_modifiers(gid, grp["modifiers"] or [])
                    mg_synced += 1
        # Day 141.6: Mark draft as saved
        try:
            drafts_store.save_draft_metadata(draft_id, status="saved")
        except Exception:
            pass
        saved = {
            "ok": True,
            "saved_at": _now_iso(),
            "inserted_ids": upsert_result.get("inserted_ids", []),
            "updated_ids": upsert_result.get("updated_ids", []),
            "deleted_count": deleted_count,
            "deleted_mg_count": deleted_mg_count,
            "mg_synced": mg_synced,
            "warnings": save_warnings,
        }
        return jsonify(saved), 200
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 400

# ---------------------------------------------------------------------------
# Day 114 — Modifier Group Reorder, Template Apply, Bulk Migration
# ---------------------------------------------------------------------------


def _parse_ordered_ids(payload: dict):
    """
    Extract and coerce 'ordered_ids' from a JSON payload.

    Returns (ids: List[int], error_response) where exactly one value is None.
    On success: (list_of_ints, None).
    On failure: (None, (jsonify_response, http_status)).
    """
    raw = payload.get("ordered_ids") or []
    if not isinstance(raw, list):
        return None, (jsonify({"ok": False, "error": "ordered_ids must be a list"}), 400)
    try:
        return [int(x) for x in raw], None
    except (TypeError, ValueError):
        return None, (jsonify({"ok": False, "error": "ordered_ids must contain integers"}), 400)


@app.post("/drafts/<int:draft_id>/items/reorder")
@login_required
def items_reorder(draft_id: int):
    """
    Bulk-update item positions within a draft.

    Body: {"ordered_ids": [<item_id>, ...]}
    The array index becomes the new position value.
    IDs not belonging to draft_id are silently ignored.

    Returns: {"ok": true, "updated": <int>}
    """
    _require_drafts_storage()
    payload = request.get_json(silent=True) or {}
    ordered_ids, err = _parse_ordered_ids(payload)
    if err:
        return err
    updated = drafts_store.reorder_items(draft_id, ordered_ids)
    return jsonify({"ok": True, "updated": updated}), 200


@app.post("/drafts/<int:draft_id>/items/<int:item_id>/modifier_groups/reorder")
@login_required
def modifier_groups_reorder(draft_id: int, item_id: int):
    """
    Bulk-update modifier group positions for *item_id*.

    Body: {"ordered_ids": [<group_id>, ...]}
    The array index becomes the new position value.
    IDs not belonging to item_id are silently ignored.

    Returns: {"ok": true, "updated": <int>}
    """
    _require_drafts_storage()
    payload = request.get_json(silent=True) or {}
    ordered_ids, err = _parse_ordered_ids(payload)
    if err:
        return err
    updated = drafts_store.reorder_modifier_groups(item_id, ordered_ids)
    return jsonify({"ok": True, "updated": updated}), 200


@app.post("/drafts/<int:draft_id>/modifier_groups/<int:group_id>/modifiers/reorder")
@login_required
def modifiers_reorder(draft_id: int, group_id: int):
    """
    Bulk-update modifier (variant) positions within *group_id*.

    Body: {"ordered_ids": [<variant_id>, ...]}
    The array index becomes the new position value.
    IDs not belonging to group_id are silently ignored.

    Returns: {"ok": true, "updated": <int>}
    """
    _require_drafts_storage()
    payload = request.get_json(silent=True) or {}
    ordered_ids, err = _parse_ordered_ids(payload)
    if err:
        return err
    updated = drafts_store.reorder_modifiers(group_id, ordered_ids)
    return jsonify({"ok": True, "updated": updated}), 200


@app.post("/drafts/<int:draft_id>/reorder_categories")
@login_required
def reorder_categories(draft_id: int):
    """
    Persist the user-defined category display order for a draft.

    Body: {"categories": ["Cat A", "Cat B", ...]}
    Order is stored as a JSON array; unknown categories are ignored at render time.

    Returns: {"ok": true, "count": <int>}
    """
    _require_drafts_storage()
    draft = drafts_store.get_draft(draft_id)
    if not draft:
        return jsonify({"error": "Draft not found"}), 404
    payload = request.get_json(silent=True) or {}
    if "categories" not in payload:
        return jsonify({"error": "Missing 'categories' key"}), 400
    cats = payload["categories"]
    if not isinstance(cats, list):
        return jsonify({"error": "'categories' must be a list"}), 400
    if not all(isinstance(c, str) for c in cats):
        return jsonify({"error": "All categories must be strings"}), 400
    drafts_store.save_category_order(draft_id, cats)
    return jsonify({"ok": True, "count": len(cats)}), 200


@app.get("/restaurants/<int:restaurant_id>/modifier_templates")
@login_required
def list_modifier_templates(restaurant_id: int):
    """
    List modifier group templates for *restaurant_id* (+ global templates).

    Returns: {"ok": true, "templates": [...], "count": <int>}
    """
    _require_drafts_storage()
    templates = drafts_store.list_modifier_templates(restaurant_id)
    return jsonify({"ok": True, "templates": templates, "count": len(templates)}), 200


@app.post("/drafts/<int:draft_id>/items/<int:item_id>/apply_template")
@login_required
def apply_modifier_template(draft_id: int, item_id: int):
    """
    Apply a modifier group template to *item_id*, creating a new group + modifiers.

    Body: {"template_id": <int>}
    Non-idempotent: calling twice creates two independent groups.

    Returns: {"ok": true, "group_id": <int>, "modifier_ids": [...]}
    """
    _require_drafts_storage()
    payload = request.get_json(silent=True) or {}
    template_id = payload.get("template_id")
    if template_id is None:
        return jsonify({"ok": False, "error": "template_id required"}), 400
    try:
        template_id = int(template_id)
    except (TypeError, ValueError):
        return jsonify({"ok": False, "error": "template_id must be an integer"}), 400
    try:
        result = drafts_store.apply_modifier_template(item_id, template_id)
        return jsonify({"ok": True, **result}), 200
    except ValueError as exc:
        return jsonify({"ok": False, "error": str(exc)}), 404


@app.post("/drafts/<int:draft_id>/migrate_modifier_groups")
@login_required
def draft_migrate_modifier_groups(draft_id: int):
    """
    Batch-migrate all items in *draft_id* that have ungrouped variants.

    Groups existing variants by kind → creates named modifier groups.
    Items that already have groups are skipped (idempotent per item).

    Returns: {"ok": true, "item_count": <int>, "migrated_count": <int>}
    """
    _require_drafts_storage()
    result = drafts_store.migrate_draft_modifier_groups(draft_id)
    return jsonify({"ok": True, **result}), 200


@app.post("/drafts/<int:draft_id>/items/<int:item_id>/modifier_groups")
@login_required
def add_modifier_group(draft_id: int, item_id: int):
    """
    Add a single modifier group to *item_id*.

    Body (JSON):
      { "name": <str>, "required": <bool>,
        "min_select": <int>, "max_select": <int>, "position": <int> }

    Returns: { "ok": true, "group_id": <int> }
    """
    _require_drafts_storage()
    payload = request.get_json(silent=True) or {}
    name = (payload.get("name") or "").strip()
    if not name:
        return jsonify({"ok": False, "error": "name is required"}), 400

    required = bool(payload.get("required", False))
    try:
        min_select = int(payload.get("min_select") or 0)
        max_select = int(payload.get("max_select") or 0)
        position = int(payload.get("position") or 0)
    except (TypeError, ValueError):
        return jsonify({"ok": False, "error": "min_select, max_select, position must be integers"}), 400

    group_id = drafts_store.insert_modifier_group(
        item_id,
        name,
        required=required,
        min_select=min_select,
        max_select=max_select,
        position=position,
    )
    return jsonify({"ok": True, "group_id": group_id}), 201


@app.patch("/drafts/<int:draft_id>/modifier_groups/<int:group_id>")
@login_required
def update_modifier_group(draft_id: int, group_id: int):
    """
    Update fields on a modifier group.

    Allowed body fields: name, required, min_select, max_select, position.
    Returns: { "ok": true } on success, 404 if group not found.
    """
    _require_drafts_storage()
    payload = request.get_json(silent=True) or {}

    allowed = {"name", "required", "min_select", "max_select", "position"}
    updates: dict = {}

    if "name" in payload:
        name = (payload["name"] or "").strip()
        if not name:
            return jsonify({"ok": False, "error": "name cannot be empty"}), 400
        updates["name"] = name

    for int_field in ("min_select", "max_select", "position"):
        if int_field in payload:
            try:
                updates[int_field] = int(payload[int_field])
            except (TypeError, ValueError):
                return jsonify({"ok": False, "error": f"{int_field} must be an integer"}), 400

    if "required" in payload:
        updates["required"] = 1 if payload["required"] else 0

    if not updates:
        return jsonify({"ok": False, "error": "no valid fields provided"}), 400

    updated = drafts_store.update_modifier_group(group_id, **updates)
    if not updated:
        return jsonify({"ok": False, "error": "modifier group not found"}), 404
    return jsonify({"ok": True}), 200


@app.post("/drafts/<int:draft_id>/submit")
@login_required
def draft_submit(draft_id: int):
    """Mark draft as submitted."""
    _require_drafts_storage()
    try:
        drafts_store.submit_draft(draft_id)
        flash(f"Draft #{draft_id} submitted for review.", "success")
        return redirect(url_for("drafts_list"))
    except Exception as e:
        flash(f"Submit failed: {e}", "error")
        return redirect(url_for("draft_editor", draft_id=draft_id))

@app.post("/drafts/<int:draft_id>/publish_now")
@login_required
def draft_publish_now(draft_id: int):
    """
    Approve & publish from the Draft Editor.
    Requires restaurant_id to be assigned (in metadata) or provided in form/json.

    Day 88: If draft has menu_id assigned, creates a versioned snapshot
    via menus_store.create_menu_version() instead of legacy menu_items insert.
    """
    _require_drafts_storage()
    try:
        draft = drafts_store.get_draft(draft_id)
        if not draft:
            flash("Draft not found.", "error")
            return redirect(url_for("drafts_list"))

        rid = draft.get("restaurant_id")
        if request.is_json:
            payload = request.get_json(silent=True) or {}
            rid = payload.get("restaurant_id", rid)
        else:
            rid = request.form.get("restaurant_id", rid)

        try:
            restaurant_id = int(rid) if rid is not None else None
        except Exception:
            restaurant_id = None

        if not restaurant_id:
            flash("Assign a restaurant before publishing.", "error")
            return redirect(url_for("draft_editor", draft_id=draft_id))

        # Day 88: Versioned publish path — if draft has menu_id, create version
        assigned_menu_id = draft.get("menu_id")
        if assigned_menu_id and menus_store:
            menu = menus_store.get_menu(int(assigned_menu_id))
            if not menu:
                flash("Assigned menu not found. Please reassign.", "error")
                return redirect(url_for("draft_editor", draft_id=draft_id))
            # Day 91: capture session user in created_by
            _publish_user = None
            try:
                _publish_user = session.get("user", {}).get("email") or session.get("user", {}).get("name")
            except Exception:
                pass
            version = menus_store.create_menu_version(
                int(assigned_menu_id),
                source_draft_id=draft_id,
                notes=f"Published from draft #{draft_id}",
                created_by=_publish_user,
            )
            # Mark draft as published
            try:
                if hasattr(drafts_store, "approve_publish"):
                    drafts_store.approve_publish(draft_id)
                else:
                    drafts_store.save_draft_metadata(draft_id, status="published")
            except Exception:
                pass

            # Day 92: record publish activity
            try:
                menus_store.record_menu_activity(
                    int(assigned_menu_id), "version_published",
                    version_id=version["id"],
                    detail=f"{version['label']} ({version['item_count']} items, {version['variant_count']} variants)",
                    actor=_publish_user,
                )
            except Exception:
                pass

            flash(
                f"Published draft #{draft_id} → {menu['name']} "
                f"{version['label']} ({version['item_count']} items, "
                f"{version['variant_count']} variants).",
                "success",
            )
            return redirect(url_for("menu_detail", menu_id=int(assigned_menu_id)))

        # Legacy path: no menu_id assigned — flat publish to menu_items
        publish_rows = drafts_store.get_publish_rows(draft_id)
        with db_connect() as conn:
            menu_id = _find_or_create_menu_for_restaurant(conn, int(restaurant_id))
            cur = conn.cursor()
            inserted = 0
            for row in publish_rows:
                name = row.get("name", "")
                desc = row.get("description", "")
                price_cents = row.get("price_cents", 0)
                if not _dedupe_exists(conn, menu_id, name, int(price_cents)):
                    cur.execute(
                        "INSERT INTO menu_items (menu_id, name, description, price_cents, is_available) VALUES (?, ?, ?, ?, 1)",
                        (menu_id, name, desc, int(price_cents)),
                    )
                    inserted += 1
            conn.commit()

        try:
            if hasattr(drafts_store, "approve_publish"):
                drafts_store.approve_publish(draft_id)
            else:
                drafts_store.save_draft_metadata(draft_id, status="published")
        except Exception:
            pass

        flash(f"Published draft #{draft_id} to menu #{menu_id} ({inserted} items).", "success")
        return redirect(url_for("items_page", menu_id=menu_id))
    except Exception as e:
        flash(f"Publish failed: {e}", "error")
        return redirect(url_for("draft_editor", draft_id=draft_id))

@app.post("/drafts/<int:draft_id>/approve_export")
@login_required
@_require_tier_chosen
def draft_approve_export(draft_id: int):
    """Approve draft and export as Generic POS JSON.

    Day 83: One-click approve & export workflow.
    1. Validates draft items for export
    2. Builds Generic POS JSON payload
    3. Sets draft status to 'approved'
    4. Records export in history
    5. Returns POS JSON for client-side download
    """
    _require_drafts_storage()
    try:
        draft = drafts_store.get_draft(draft_id)
        if not draft:
            return jsonify({"ok": False, "error": "Draft not found"}), 404

        items = drafts_store.get_draft_items(draft_id, include_modifier_groups=True) or []
        warnings = _validate_draft_for_export(items)
        # Count all modifiers: grouped (in modifier_groups[].modifiers[]) + ungrouped
        variant_count = sum(
            sum(len(grp.get("modifiers") or []) for grp in (it.get("modifier_groups") or []))
            + len(it.get("ungrouped_variants") or it.get("variants") or [])
            for it in items
        )

        pos_json = _build_generic_pos_json(items, draft)

        drafts_store.approve_draft(draft_id)

        drafts_store.record_export(
            draft_id, "generic_pos",
            item_count=len(items),
            variant_count=variant_count,
            warning_count=len(warnings),
        )

        # Day 85: Fire webhook notifications
        rid = (draft or {}).get("restaurant_id")
        approved_ts = _now_iso()
        try:
            drafts_store.fire_webhooks(rid, "draft.approved", {
                "event": "draft.approved",
                "draft_id": draft_id,
                "title": (draft or {}).get("title", ""),
                "item_count": len(items),
                "variant_count": variant_count,
                "warning_count": len(warnings),
                "approved_at": approved_ts,
            })
        except Exception:
            pass  # webhook failure must not block approve flow
        try:
            drafts_store.fire_webhooks(rid, "draft.exported", {
                "event": "draft.exported",
                "draft_id": draft_id,
                "format": "generic_pos",
                "item_count": len(items),
                "variant_count": variant_count,
                "exported_at": approved_ts,
            })
        except Exception:
            pass

        return jsonify({
            "ok": True,
            "pos_json": pos_json,
            "item_count": len(items),
            "variant_count": variant_count,
            "warnings": warnings,
            "warning_count": len(warnings),
            "approved_at": approved_ts,
        })
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500


@app.get("/drafts/<int:draft_id>/export_history")
@login_required
def draft_export_history(draft_id: int):
    """Return export history records for a draft."""
    _require_drafts_storage()
    draft = drafts_store.get_draft(draft_id)
    if not draft:
        return jsonify({"ok": False, "error": "Draft not found"}), 404
    history = drafts_store.get_export_history(draft_id)
    return jsonify({"ok": True, "draft_id": draft_id, "history": history})


# ============================================================
# Day 84: REST API Endpoints for External POS Integrations
# ============================================================

@app.get("/api/drafts/<int:draft_id>/items")
@api_key_required
def api_get_draft_items(draft_id: int):
    """REST API: Retrieve draft items with variants."""
    draft = drafts_store.get_draft(draft_id)
    if not draft:
        return jsonify({"ok": False, "error": "Draft not found"}), 404
    items = drafts_store.get_draft_items(draft_id, include_variants=True) or []
    return jsonify({
        "ok": True,
        "draft_id": draft_id,
        "items": items,
        "count": len(items),
    })


@app.post("/api/drafts/<int:draft_id>/items")
@api_key_required
def api_create_draft_items(draft_id: int):
    """REST API: Create new items (with optional variants) on a draft."""
    if not request.is_json:
        return jsonify({"ok": False, "error": "Expected JSON payload"}), 400

    draft = drafts_store.get_draft(draft_id)
    if not draft:
        return jsonify({"ok": False, "error": "Draft not found"}), 404

    if draft.get("status") != "editing":
        return jsonify({
            "ok": False,
            "error": f"Draft is '{draft.get('status')}' and cannot be modified",
        }), 403

    payload = request.get_json(silent=True) or {}
    items = payload.get("items")
    if not isinstance(items, list):
        return jsonify({"ok": False, "error": "'items' must be a list"}), 400

    probe = {"draft_id": draft_id, "items": items}
    ok, err = validate_draft_payload(probe)
    if not ok:
        return jsonify({"ok": False, "error": f"Validation: {err}"}), 400

    try:
        result = drafts_store.upsert_draft_items(draft_id, items)
        return jsonify({
            "ok": True,
            "inserted_ids": result.get("inserted_ids", []),
            "updated_ids": result.get("updated_ids", []),
        }), 201
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500


@app.put("/api/drafts/<int:draft_id>/items/<int:item_id>")
@api_key_required
def api_update_draft_item(draft_id: int, item_id: int):
    """REST API: Update a single item (with optional variants) on a draft."""
    if not request.is_json:
        return jsonify({"ok": False, "error": "Expected JSON payload"}), 400

    draft = drafts_store.get_draft(draft_id)
    if not draft:
        return jsonify({"ok": False, "error": "Draft not found"}), 404

    if draft.get("status") != "editing":
        return jsonify({
            "ok": False,
            "error": f"Draft is '{draft.get('status')}' and cannot be modified",
        }), 403

    # Verify item belongs to this draft
    existing = drafts_store.get_draft_items(draft_id, include_variants=False)
    item_ids = {it["id"] for it in existing}
    if item_id not in item_ids:
        return jsonify({"ok": False, "error": "Item not found in this draft"}), 404

    payload = request.get_json(silent=True) or {}
    payload["id"] = item_id

    probe = {"draft_id": draft_id, "items": [payload]}
    ok, err = validate_draft_payload(probe)
    if not ok:
        return jsonify({"ok": False, "error": f"Validation: {err}"}), 400

    try:
        result = drafts_store.upsert_draft_items(draft_id, [payload])
        return jsonify({
            "ok": True,
            "updated_ids": result.get("updated_ids", []),
        })
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500


# ============================================================
# Day 85: Webhook Management API
# ============================================================

@app.post("/api/webhooks")
@api_key_required
def api_register_webhook():
    """REST API: Register a new webhook for notifications."""
    if not request.is_json:
        return jsonify({"ok": False, "error": "Expected JSON payload"}), 400

    data = request.get_json(silent=True) or {}
    url = (data.get("url") or "").strip()
    if not url:
        return jsonify({"ok": False, "error": "url is required"}), 400

    if not (url.startswith("http://") or url.startswith("https://")):
        return jsonify({"ok": False, "error": "url must start with http:// or https://"}), 400

    event_types = data.get("event_types")
    if not event_types or not isinstance(event_types, list):
        return jsonify({"ok": False, "error": "event_types must be a non-empty list"}), 400

    restaurant_id = data.get("restaurant_id")
    key_record = g.api_key
    if key_record.get("restaurant_id"):
        restaurant_id = key_record["restaurant_id"]

    try:
        result = drafts_store.register_webhook(
            url=url,
            event_types=event_types,
            restaurant_id=restaurant_id,
        )
        return jsonify({"ok": True, "webhook": result}), 201
    except ValueError as e:
        return jsonify({"ok": False, "error": str(e)}), 400
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500


@app.get("/api/webhooks")
@api_key_required
def api_list_webhooks():
    """REST API: List active webhooks."""
    key_record = g.api_key
    restaurant_id = key_record.get("restaurant_id")
    hooks = drafts_store.list_webhooks(restaurant_id=restaurant_id)
    return jsonify({"ok": True, "webhooks": hooks, "count": len(hooks)})


@app.delete("/api/webhooks/<int:webhook_id>")
@api_key_required
def api_delete_webhook(webhook_id: int):
    """REST API: Delete a webhook."""
    hook = drafts_store.get_webhook(webhook_id)
    if not hook:
        return jsonify({"ok": False, "error": "Webhook not found"}), 404

    key_record = g.api_key
    if key_record.get("restaurant_id"):
        if hook.get("restaurant_id") != key_record["restaurant_id"]:
            return jsonify({"ok": False, "error": "Webhook not found"}), 404

    deleted = drafts_store.delete_webhook(webhook_id)
    return jsonify({"ok": True, "deleted": deleted})


# ============================================================
# Day 85: API Documentation (public, no auth required)
# ============================================================

@app.get("/api/docs")
def api_docs_page():
    """Public API documentation page."""
    return _safe_render("api_docs.html")


@app.post("/drafts/<int:draft_id>/backfill_variants")
@login_required
def draft_backfill_variants(draft_id: int):
    """Merge legacy 'Name (Size)' items into structured parent + variant rows."""
    _require_drafts_storage()
    try:
        draft = drafts_store.get_draft(draft_id)
        if not draft:
            return jsonify({"ok": False, "error": "Draft not found"}), 404
        if draft.get("status") != "editing":
            return jsonify({"ok": False, "error": "Draft is not in editing state"}), 400
        result = drafts_store.backfill_variants_from_names(draft_id)
        return jsonify({
            "ok": True,
            "groups_found": result.get("groups_found", 0),
            "variants_created": result.get("variants_created", 0),
            "items_deleted": result.get("items_deleted", 0),
        }), 200
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500

@app.post("/drafts/<int:draft_id>/assign_restaurant")
@login_required
def draft_assign_restaurant(draft_id: int):
    _require_drafts_storage()
    rid = request.form.get("restaurant_id")
    if not rid:
        flash("Please choose a restaurant.", "error")
        return redirect(url_for("draft_editor", draft_id=draft_id))
    try:
        restaurant_id = int(rid)
    except Exception:
        flash("Invalid restaurant id.", "error")
        return redirect(url_for("draft_editor", draft_id=draft_id))
    try:
        drafts_store.save_draft_metadata(draft_id, restaurant_id=restaurant_id)
        flash("Restaurant assigned to draft.", "success")
    except Exception as e:
        flash(f"Failed to assign restaurant: {e}", "error")
    return redirect(url_for("draft_editor", draft_id=draft_id))

@app.post("/drafts/<int:draft_id>/assign_menu")
@login_required
def draft_assign_menu(draft_id: int):
    """Assign a menu to a draft (Phase 10 Day 87)."""
    _require_drafts_storage()
    mid = request.form.get("menu_id")
    if not mid:
        flash("Please choose a menu.", "error")
        return redirect(url_for("draft_editor", draft_id=draft_id))
    try:
        menu_id = int(mid)
    except Exception:
        flash("Invalid menu id.", "error")
        return redirect(url_for("draft_editor", draft_id=draft_id))
    try:
        drafts_store.save_draft_metadata(draft_id, menu_id=menu_id)
        flash("Menu assigned to draft.", "success")
    except Exception as e:
        flash(f"Failed to assign menu: {e}", "error")
    return redirect(url_for("draft_editor", draft_id=draft_id))

# OCR Inspector debug endpoints
@app.get("/drafts/<int:draft_id>/ocr-debug.json")
@login_required
def draft_ocr_debug_json(draft_id: int):
    _require_drafts_storage()
    if not hasattr(drafts_store, "load_ocr_debug"):
        return jsonify({"error": "OCR debug storage not available"}), 404
    dbg = drafts_store.load_ocr_debug(draft_id)
    if not dbg:
        return jsonify({"error": "No OCR debug payload found for this draft"}), 404
    return jsonify(dbg)


@app.get("/drafts/<int:draft_id>/ocr-debug.csv")
@login_required
def draft_ocr_debug_csv(draft_id: int):
    _require_drafts_storage()
    load_fn = getattr(drafts_store, "load_ocr_debug", None)
    if load_fn is None:
        return make_response("OCR debug storage not available", 404)
    dbg = load_fn(draft_id)
    if not dbg:
        return make_response("No OCR debug payload found for this draft", 404)

    rows = []
    items = dbg.get("items") or []
    if items:
        for it in items:
            src = it.get("source") or {}
            rows.append({
                "id": it.get("id"),
                "name": it.get("name"),
                "description": it.get("desc") or it.get("description"),
                "price": it.get("price"),
                "category": it.get("category"),
                "confidence": it.get("confidence"),
                "page": src.get("page"),
                "line_idx": src.get("line_idx"),
                "bbox": json.dumps(src.get("bbox")) if src.get("bbox") is not None else "",
                "matched_rule": src.get("matched_rule"),
            })
    else:
        for a in (dbg.get("assignments") or [])[:200]:
            rows.append({
                "id": "",
                "name": a.get("name"),
                "description": "",
                "price": "",
                "category": a.get("category"),
                "confidence": a.get("score"),
                "page": "",
                "line_idx": "",
                "bbox": "",
                "matched_rule": a.get("reason"),
            })

    buf = io.StringIO()
    fieldnames = ["id", "name", "description", "price", "category", "confidence", "page", "line_idx", "bbox", "matched_rule"]
    writer = csv.DictWriter(buf, fieldnames=fieldnames)
    writer.writeheader()
    for r in rows:
        writer.writerow({k: ("" if r.get(k) is None else r.get(k)) for k in fieldnames})
    data = buf.getvalue().encode("utf-8-sig")
    resp = make_response(data)
    resp.headers["Content-Type"] = "text/csv; charset=utf-8"
    resp.headers["Content-Disposition"] = f'attachment; filename="draft_{draft_id}_ocr_debug.csv"'
    return resp

# ------------------------------------------------------------
# NEW: Layout Debug JSON (job -> draft -> stored payload)
# ------------------------------------------------------------

@app.get("/drafts/<int:draft_id>/layout-debug.json")
@login_required
def draft_layout_debug_json(draft_id: int):
    _require_drafts_storage()

    layout = _load_layout_debug_for_draft(draft_id)

    if not layout:
        # Back-compat fallback: older keys may have been stored at root
        dbg = _load_debug_for_draft(draft_id) or {}
        if not isinstance(dbg, dict):
            dbg = {}

        payload = (
            dbg.get("layout_debug")
            or dbg.get("layout")
            or dbg.get("debug_layout")
            or dbg.get("blocks_layout")
            or None
        )
        if isinstance(payload, dict):
            layout = payload


    if not layout:
        return jsonify({
            "ok": True,
            "draft_id": draft_id,
            "note": "No layout_debug payload present yet",
        }), 200

    # If present, meta.orientation should carry your per-page rotation audit trail
    meta = layout.get("meta") if isinstance(layout, dict) else None
    if not isinstance(meta, dict):
        meta = {}

    # Always provide the orientation key so clients can rely on it
    if "orientation" not in meta or not isinstance(meta.get("orientation"), dict):
        meta["orientation"] = {}

    return jsonify({
        "ok": True,
        "draft_id": draft_id,
        "layout_debug": layout,
        "meta": meta,
    }), 200




@app.get("/imports/<int:job_id>/layout-debug.json")
@login_required
def imports_layout_debug_json(job_id: int):
    _require_drafts_storage()
    draft_id = _get_or_create_draft_for_job(job_id)
    if not draft_id:
        return jsonify({"ok": False, "job_id": job_id, "error": "No draft found for job"}), 404
    return draft_layout_debug_json(int(draft_id))


@app.get("/ocr/layout-debug/<int:job_id>.json")
@login_required
def ocr_layout_debug_alias(job_id: int):
    return imports_layout_debug_json(job_id)


# Exporters
@app.get("/drafts/<int:draft_id>/export.csv")
@login_required
def draft_export_csv(draft_id: int):
    _require_drafts_storage()
    draft = drafts_store.get_draft(draft_id) or {}
    items = drafts_store.get_draft_items(draft_id) or []
    buf = io.StringIO()
    writer = csv.DictWriter(buf, fieldnames=["id", "name", "description", "price_cents", "category", "position"])
    writer.writeheader()
    for it in items:
        writer.writerow({
            "id": it.get("id"),
            "name": it.get("name", ""),
            "description": it.get("description", ""),
            "price_cents": it.get("price_cents", 0),
            "category": it.get("category") or "",
            "position": it.get("position") if it.get("position") is not None else ""
        })
    csv_data = buf.getvalue().encode("utf-8-sig")
    resp = make_response(csv_data)
    resp.headers["Content-Type"] = "text/csv; charset=utf-8"
    resp.headers["Content-Disposition"] = f'attachment; filename="draft_{draft_id}.csv"'
    return resp

@app.get("/drafts/<int:draft_id>/export_variants.csv")
@login_required
def draft_export_csv_variants(draft_id: int):
    """CSV export with variant sub-rows under each parent item.

    When modifier groups are present, emits modifier_group header rows
    followed by modifier child rows.  Ungrouped variants continue as
    plain variant rows for backward compatibility.
    """
    _require_drafts_storage()
    items = drafts_store.get_draft_items(draft_id, include_variants=True,
                                          include_modifier_groups=True) or []
    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(["type", "id", "name", "description", "price_cents",
                      "category", "kind", "label", "group_name", "required"])
    for it in items:
        writer.writerow([
            "item",
            it.get("id", ""),
            it.get("name", ""),
            it.get("description", ""),
            it.get("price_cents", 0),
            it.get("category") or "",
            "", "", "", "",
        ])

        # POS-native modifier groups → group header + modifier rows
        for grp in (it.get("modifier_groups") or []):
            grp_name = grp.get("name", "Option")
            required = "Y" if grp.get("required") else "N"
            writer.writerow([
                "modifier_group", "", "", "", "",
                "", "", "", grp_name, required,
            ])
            for mod in (grp.get("modifiers") or []):
                writer.writerow([
                    "modifier", "", "", "",
                    mod.get("price_cents", 0),
                    "", mod.get("kind", "size"), mod.get("label", ""),
                    grp_name, "",
                ])

        # Ungrouped variants (backward compat)
        for v in (it.get("ungrouped_variants") or it.get("variants") or []):
            writer.writerow([
                "variant", "", "", "",
                v.get("price_cents", 0),
                "", v.get("kind", "size"), v.get("label", ""),
                "", "",
            ])
    csv_data = buf.getvalue().encode("utf-8-sig")
    resp = make_response(csv_data)
    resp.headers["Content-Type"] = "text/csv; charset=utf-8"
    resp.headers["Content-Disposition"] = f'attachment; filename="draft_{draft_id}_variants.csv"'
    return resp


@app.get("/drafts/<int:draft_id>/export_wide.csv")
@login_required
def draft_export_csv_wide(draft_id: int):
    """CSV export with variant prices as extra columns (one row per item).

    Modifier group modifiers become columns prefixed with the group name
    (e.g. ``Size:Small``, ``Size:Large``).  Ungrouped variants use their
    label directly (backward compat).
    """
    _require_drafts_storage()
    items = drafts_store.get_draft_items(draft_id, include_variants=True,
                                          include_modifier_groups=True) or []

    # Collect all unique column labels across items (in first-appearance order)
    seen_labels: dict = {}  # label -> insertion order

    for it in items:
        # Modifier group modifiers → "GroupName:Label"
        for grp in (it.get("modifier_groups") or []):
            grp_name = grp.get("name", "Option")
            for mod in (grp.get("modifiers") or []):
                col = f"{grp_name}:{mod.get('label', '')}"
                if col not in seen_labels:
                    seen_labels[col] = len(seen_labels)
        # Ungrouped variants → plain label (backward compat)
        for v in (it.get("ungrouped_variants") or it.get("variants") or []):
            lbl = (v.get("label") or "").strip()
            if lbl and lbl not in seen_labels:
                seen_labels[lbl] = len(seen_labels)

    label_order = sorted(seen_labels.keys(), key=lambda x: seen_labels[x])

    buf = io.StringIO()
    writer = csv.writer(buf)
    base_headers = ["id", "name", "description", "price_cents", "category"]
    writer.writerow(base_headers + [f"price_{lbl}" for lbl in label_order])

    for it in items:
        row = [
            it.get("id", ""),
            it.get("name", ""),
            it.get("description", ""),
            it.get("price_cents", 0),
            it.get("category") or "",
        ]
        # Build label -> price mapping
        vpmap: dict = {}
        for grp in (it.get("modifier_groups") or []):
            grp_name = grp.get("name", "Option")
            for mod in (grp.get("modifiers") or []):
                col = f"{grp_name}:{mod.get('label', '')}"
                vpmap[col] = mod.get("price_cents", 0)
        for v in (it.get("ungrouped_variants") or it.get("variants") or []):
            lbl = (v.get("label") or "").strip()
            if lbl:
                vpmap[lbl] = v.get("price_cents", 0)
        # Append price columns
        for lbl in label_order:
            row.append(vpmap.get(lbl, ""))
        writer.writerow(row)

    csv_data = buf.getvalue().encode("utf-8-sig")
    resp = make_response(csv_data)
    resp.headers["Content-Type"] = "text/csv; charset=utf-8"
    resp.headers["Content-Disposition"] = f'attachment; filename="draft_{draft_id}_wide.csv"'
    return resp


@app.get("/drafts/<int:draft_id>/export.json")
@login_required
def draft_export_json(draft_id: int):
    _require_drafts_storage()
    draft = drafts_store.get_draft(draft_id) or {}
    items = drafts_store.get_draft_items(draft_id, include_variants=True,
                                          include_modifier_groups=True) or []

    # Clean items for export: include nested variants + modifier_groups
    export_items = []
    for it in items:
        eitem = {
            "id": it.get("id"),
            "name": it.get("name", ""),
            "description": it.get("description", ""),
            "price_cents": it.get("price_cents", 0),
            "category": it.get("category") or "",
            "position": it.get("position"),
        }

        # Modifier groups (POS-native)
        modifier_groups = it.get("modifier_groups") or []
        if modifier_groups:
            eitem["modifier_groups"] = [
                {
                    "name": grp.get("name", ""),
                    "required": bool(grp.get("required")),
                    "min_select": grp.get("min_select") or 0,
                    "max_select": grp.get("max_select") or 0,
                    "modifiers": [
                        {
                            "label": mod.get("label", ""),
                            "price_cents": mod.get("price_cents", 0),
                            "kind": mod.get("kind", "size"),
                        }
                        for mod in (grp.get("modifiers") or [])
                    ],
                }
                for grp in modifier_groups
            ]
        else:
            eitem["modifier_groups"] = []

        # Ungrouped variants (backward compat)
        ungrouped = it.get("ungrouped_variants") or it.get("variants") or []
        if ungrouped:
            eitem["variants"] = [
                {
                    "label": v.get("label", ""),
                    "price_cents": v.get("price_cents", 0),
                    "kind": v.get("kind", "size"),
                }
                for v in ungrouped
            ]
        else:
            eitem["variants"] = []
        export_items.append(eitem)

    payload = {
        "draft_id": draft_id,
        "title": draft.get("title"),
        "restaurant_id": draft.get("restaurant_id"),
        "status": draft.get("status"),
        "items": export_items,
        "exported_at": _now_iso(),
    }

    resp = make_response(json.dumps(payload, indent=2))
    resp.headers["Content-Type"] = "application/json; charset=utf-8"
    resp.headers["Content-Disposition"] = f'attachment; filename="draft_{draft_id}.json"'
    return resp


# ---------------------------------------------------------------------------
# XLSX shared helper (Day 125)
# ---------------------------------------------------------------------------

def _xlsx_write_sheet(ws, items, xl, *, include_category: bool = True):
    """Write items with modifier group headers + modifier/variant sub-rows to a worksheet.

    Row types and styles:
      - Header row: bold white text on dark bg (#1a2236)
      - Item row: bold text
      - Modifier group row: bold text on light blue bg (#D6EAF8), shows group name + required
      - Modifier row: gray text on light gray bg (#F2F2F2), indented with "    "
      - Ungrouped variant row: gray text on light gray bg (#F2F2F2), indented with "  "
    """
    from openpyxl.styles import Font, PatternFill  # type: ignore[import]

    # -- Styles --
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1a2236", end_color="1a2236", fill_type="solid")
    parent_font = Font(bold=True)
    group_font = Font(bold=True, color="1a5276")
    group_fill = PatternFill(start_color="D6EAF8", end_color="D6EAF8", fill_type="solid")
    modifier_font = Font(color="666666")
    modifier_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

    # -- Headers --
    if include_category:
        headers = ["name", "description", "price_cents", "category", "group_name", "required"]
    else:
        headers = ["name", "description", "price_cents", "group_name", "required"]

    ws.append(headers)
    for ci in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=ci)
        cell.font = header_font
        cell.fill = header_fill

    # -- Data rows --
    for it in items:
        mod_groups = it.get("modifier_groups") or []
        ungrouped = it.get("ungrouped_variants") or []
        # Fallback: if fetched with include_variants=True (no modifier_groups key)
        if not mod_groups and not ungrouped and "variants" in it:
            ungrouped = it.get("variants") or []

        # Item row
        if include_category:
            row_data = [
                it.get("name", ""),
                it.get("description", ""),
                it.get("price_cents", 0),
                it.get("category") or "",
                "",  # group_name
                "",  # required
            ]
        else:
            row_data = [
                it.get("name", ""),
                it.get("description", ""),
                it.get("price_cents", 0),
                "",  # group_name
                "",  # required
            ]
        ws.append(row_data)
        row_num = ws.max_row
        for ci in range(1, len(headers) + 1):
            ws.cell(row=row_num, column=ci).font = parent_font

        # Modifier group header + modifier rows
        for grp in mod_groups:
            grp_name = grp.get("name") or "(unnamed)"
            required = "Y" if grp.get("required") else "N"
            if include_category:
                grp_row = ["", "", "", "", grp_name, required]
            else:
                grp_row = ["", "", "", grp_name, required]
            ws.append(grp_row)
            grow_num = ws.max_row
            for ci in range(1, len(headers) + 1):
                cell = ws.cell(row=grow_num, column=ci)
                cell.font = group_font
                cell.fill = group_fill

            for mod in (grp.get("modifiers") or []):
                if include_category:
                    mrow = [
                        "    " + (mod.get("label") or ""),
                        mod.get("kind", "size"),
                        mod.get("price_cents", 0),
                        "",  # category
                        grp_name,
                        "",  # required
                    ]
                else:
                    mrow = [
                        "    " + (mod.get("label") or ""),
                        mod.get("kind", "size"),
                        mod.get("price_cents", 0),
                        grp_name,
                        "",  # required
                    ]
                ws.append(mrow)
                mrow_num = ws.max_row
                for ci in range(1, len(headers) + 1):
                    cell = ws.cell(row=mrow_num, column=ci)
                    cell.font = modifier_font
                    cell.fill = modifier_fill

        # Ungrouped variant sub-rows
        for v in ungrouped:
            if include_category:
                vrow = [
                    "  " + (v.get("label") or ""),
                    v.get("kind", "size"),
                    v.get("price_cents", 0),
                    "",  # category
                    "",  # group_name
                    "",  # required
                ]
            else:
                vrow = [
                    "  " + (v.get("label") or ""),
                    v.get("kind", "size"),
                    v.get("price_cents", 0),
                    "",  # group_name
                    "",  # required
                ]
            ws.append(vrow)
            vrow_num = ws.max_row
            for ci in range(1, len(headers) + 1):
                cell = ws.cell(row=vrow_num, column=ci)
                cell.font = modifier_font
                cell.fill = modifier_fill

    # Auto-width columns
    for ci in range(1, len(headers) + 1):
        ws.column_dimensions[xl.utils.get_column_letter(ci)].width = 18


@app.get("/drafts/<int:draft_id>/export.xlsx")
@login_required
def draft_export_xlsx(draft_id: int):
    """Excel export with modifier group headers, modifier sub-rows, and ungrouped variant sub-rows."""
    _require_drafts_storage()

    try:
        import openpyxl as xl  # type: ignore[import]
        from openpyxl.styles import Font, PatternFill, Alignment  # type: ignore[import]
    except Exception:
        xl = None

    if xl is None:
        return make_response("openpyxl not installed. pip install openpyxl", 500)

    draft = drafts_store.get_draft(draft_id) or {}
    items = drafts_store.get_draft_items(draft_id, include_modifier_groups=True) or []

    wb = xl.Workbook()
    ws = wb.active
    ws.title = (draft.get("title") or f"Draft {draft_id}")[:31]

    _xlsx_write_sheet(ws, items, xl, include_category=True)

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)

    resp = make_response(out.read())
    resp.headers["Content-Type"] = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    resp.headers["Content-Disposition"] = f'attachment; filename="draft_{draft_id}.xlsx"'
    return resp


@app.get("/drafts/<int:draft_id>/export_by_category.xlsx")
@login_required
def draft_export_xlsx_by_category(draft_id: int):
    """Excel export with one sheet per category.  Each sheet has modifier group + variant sub-rows."""
    _require_drafts_storage()

    try:
        import openpyxl as xl  # type: ignore[import]
        from openpyxl.styles import Font, PatternFill  # type: ignore[import]
    except Exception:
        xl = None

    if xl is None:
        return make_response("openpyxl not installed. pip install openpyxl", 500)

    draft = drafts_store.get_draft(draft_id) or {}
    items = drafts_store.get_draft_items(draft_id, include_modifier_groups=True) or []

    # Group items by category
    cat_map: dict = {}  # category -> list of items
    for it in items:
        cat = (it.get("category") or "Uncategorized").strip()
        cat_map.setdefault(cat, []).append(it)

    wb = xl.Workbook()
    wb.remove(wb.active)  # remove default empty sheet

    for cat_name in sorted(cat_map.keys()):
        cat_items = cat_map[cat_name]
        sheet_title = cat_name[:31] or "Uncategorized"
        ws = wb.create_sheet(title=sheet_title)
        _xlsx_write_sheet(ws, cat_items, xl, include_category=False)

    # If no categories at all, create a placeholder sheet
    if not cat_map:
        ws = wb.create_sheet(title="Empty")
        ws.append(["No items"])

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)

    resp = make_response(out.read())
    resp.headers["Content-Type"] = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    resp.headers["Content-Disposition"] = f'attachment; filename="draft_{draft_id}_by_category.xlsx"'
    return resp


# ---------------------------------------------------------------------------
# POS Export Templates (Day 80)
# ---------------------------------------------------------------------------

def _validate_draft_for_export(items):
    """Pre-export validation: returns list of warning dicts."""
    warnings = []
    for it in items:
        item_id = it.get("id")
        name = it.get("name", "")
        price = it.get("price_cents", 0)
        cat = it.get("category") or ""
        # Collect all modifiers regardless of source format:
        #   - flat variants (include_variants=True path)
        #   - ungrouped_variants + modifier_groups[].modifiers[] (include_modifier_groups=True path)
        variants = list(it.get("variants") or [])
        variants += list(it.get("ungrouped_variants") or [])
        for grp in (it.get("modifier_groups") or []):
            variants += list(grp.get("modifiers") or [])

        if not price and not variants:
            warnings.append({
                "item_id": item_id,
                "name": name,
                "type": "missing_price",
                "message": f"Item '{name}' has no price and no variants",
            })
        if not cat:
            warnings.append({
                "item_id": item_id,
                "name": name,
                "type": "missing_category",
                "message": f"Item '{name}' has no category",
            })
        if not name.strip():
            warnings.append({
                "item_id": item_id,
                "name": name,
                "type": "missing_name",
                "message": "Item has no name",
            })
        for v in variants:
            if not v.get("price_cents"):
                warnings.append({
                    "item_id": item_id,
                    "name": name,
                    "type": "variant_missing_price",
                    "message": f"Variant '{v.get('label','')}' on '{name}' has no price",
                })

        # --- Day 81: variant_missing_label ---
        for v in variants:
            lbl = (v.get("label") or "").strip()
            if not lbl:
                warnings.append({
                    "item_id": item_id,
                    "name": name,
                    "type": "variant_missing_label",
                    "message": f"A variant on '{name}' has no label",
                })

        # --- Day 81: duplicate_variant_label (case-insensitive, one per item) ---
        seen_labels = set()
        for v in variants:
            lbl = (v.get("label") or "").strip().lower()
            if lbl:
                if lbl in seen_labels:
                    warnings.append({
                        "item_id": item_id,
                        "name": name,
                        "type": "duplicate_variant_label",
                        "message": f"Duplicate variant label '{v.get('label', '')}' on '{name}'",
                    })
                    break
                seen_labels.add(lbl)

        # --- Day 81: price_inversion (size-kind only) ---
        size_variants = [v for v in variants
                         if v.get("kind") == "size" and v.get("price_cents")]
        if len(size_variants) >= 2:
            try:
                from storage.parsers.size_vocab import normalize_size_token, size_ordinal
                ordinal_prices = []
                for v in size_variants:
                    raw_label = (v.get("label") or "").strip()
                    normalized = normalize_size_token(raw_label)
                    ordinal = size_ordinal(normalized)
                    if ordinal is not None:
                        ordinal_prices.append((ordinal, v.get("price_cents", 0)))
                if len(ordinal_prices) >= 2:
                    ordinal_prices.sort(key=lambda x: x[0])
                    for i in range(1, len(ordinal_prices)):
                        if ordinal_prices[i][1] < ordinal_prices[i - 1][1]:
                            warnings.append({
                                "item_id": item_id,
                                "name": name,
                                "type": "price_inversion",
                                "message": f"Size variant prices on '{name}' are not in ascending order",
                            })
                            break
            except ImportError:
                pass  # size_vocab not available — skip price inversion check

        # --- Day 123: modifier group warnings ---
        for grp in (it.get("modifier_groups") or []):
            grp_name = grp.get("name") or "(unnamed)"
            mods = grp.get("modifiers") or []
            mod_count = len(mods)

            # Empty modifier group
            if mod_count == 0:
                warnings.append({
                    "item_id": item_id,
                    "name": name,
                    "type": "modifier_group_empty",
                    "message": f"Modifier group '{grp_name}' on '{name}' has no modifiers",
                })

            # Required group with no modifiers (can't satisfy requirement)
            if grp.get("required") and mod_count == 0:
                warnings.append({
                    "item_id": item_id,
                    "name": name,
                    "type": "required_group_empty",
                    "message": f"Required modifier group '{grp_name}' on '{name}' has no modifiers",
                })

            # min/max consistency
            min_sel = grp.get("min_select") or 0
            max_sel = grp.get("max_select") or 0
            if min_sel and max_sel and min_sel > max_sel:
                warnings.append({
                    "item_id": item_id,
                    "name": name,
                    "type": "group_min_exceeds_max",
                    "message": f"Modifier group '{grp_name}' on '{name}': min_select ({min_sel}) > max_select ({max_sel})",
                })
            if max_sel and mod_count and max_sel > mod_count:
                warnings.append({
                    "item_id": item_id,
                    "name": name,
                    "type": "group_max_exceeds_count",
                    "message": f"Modifier group '{grp_name}' on '{name}': max_select ({max_sel}) > modifier count ({mod_count})",
                })

    # --- Day 125: cross-item modifier group consistency ---
    # Within each category, collect which group names each item uses.
    # If most items in a category share a group name but some don't, flag outliers.
    cat_groups: dict = {}  # category -> {group_name: [item_name, ...]}
    for it in items:
        cat = (it.get("category") or "").strip()
        if not cat:
            continue
        for grp in (it.get("modifier_groups") or []):
            gn = (grp.get("name") or "").strip()
            if gn:
                cat_groups.setdefault(cat, {}).setdefault(gn, []).append(it.get("name", ""))

    for cat, group_map in cat_groups.items():
        for gn, item_names in group_map.items():
            # Count total items in this category (that have at least one modifier group)
            cat_items_with_groups = set()
            for it in items:
                it_cat = (it.get("category") or "").strip()
                if it_cat == cat and (it.get("modifier_groups") or []):
                    cat_items_with_groups.add(it.get("name", ""))
            total_with_groups = len(cat_items_with_groups)
            using_this_group = set(item_names)
            # If >=50% of items in category have this group but some don't, flag missing ones
            if total_with_groups >= 3 and len(using_this_group) >= total_with_groups * 0.5:
                missing = cat_items_with_groups - using_this_group
                for miss_name in missing:
                    warnings.append({
                        "item_id": None,
                        "name": miss_name,
                        "type": "modifier_group_inconsistent",
                        "message": (
                            f"Most items in '{cat}' have modifier group '{gn}', "
                            f"but '{miss_name}' does not"
                        ),
                    })

    return warnings


def _format_price_dollars(cents):
    """Convert cents to dollar string: 1299 -> '12.99'."""
    if not cents:
        return "0.00"
    return f"{int(cents) / 100:.2f}"


# ---------------------------------------------------------------------------
# Export Metrics & Round-Trip Verification (Day 81)
# ---------------------------------------------------------------------------

def _compute_export_metrics(items):
    """Compute export metrics: counts, breakdowns, price stats."""
    total_items = len(items)
    items_with_variants = 0
    total_variants = 0
    variants_by_kind = {}
    category_breakdown = {}
    all_prices = []

    for it in items:
        variants = it.get("variants") or []
        cat = it.get("category") or "Uncategorized"

        if cat not in category_breakdown:
            category_breakdown[cat] = {"item_count": 0, "variant_count": 0}
        category_breakdown[cat]["item_count"] += 1

        if variants:
            items_with_variants += 1
            total_variants += len(variants)
            category_breakdown[cat]["variant_count"] += len(variants)
            for v in variants:
                k = v.get("kind", "size")
                variants_by_kind[k] = variants_by_kind.get(k, 0) + 1
                vp = v.get("price_cents", 0)
                if vp:
                    all_prices.append(vp)

        base_price = it.get("price_cents", 0)
        if base_price:
            all_prices.append(base_price)

    price_stats = {
        "min_cents": min(all_prices) if all_prices else 0,
        "max_cents": max(all_prices) if all_prices else 0,
        "avg_cents": round(statistics.mean(all_prices)) if all_prices else 0,
        "price_count": len(all_prices),
    }

    return {
        "total_items": total_items,
        "items_with_variants": items_with_variants,
        "items_without_variants": total_items - items_with_variants,
        "total_variants": total_variants,
        "variants_by_kind": variants_by_kind,
        "category_breakdown": category_breakdown,
        "price_stats": price_stats,
    }


def _verify_csv_round_trip(items):
    """Export to CSV variants format, parse back, verify counts match."""
    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(["type", "id", "name", "description", "price_cents",
                      "category", "kind", "label"])
    for it in items:
        writer.writerow(["item", it.get("id", ""), it.get("name", ""),
                          it.get("description", ""), it.get("price_cents", 0),
                          it.get("category") or "", "", ""])
        for v in (it.get("variants") or []):
            writer.writerow(["variant", "", "", "", v.get("price_cents", 0),
                              "", v.get("kind", "size"), v.get("label", "")])
    text = buf.getvalue()

    reader = csv.DictReader(io.StringIO(text))
    rows = list(reader)
    item_rows = [r for r in rows if r.get("type") == "item"]
    variant_rows = [r for r in rows if r.get("type") == "variant"]

    expected_items = len(items)
    expected_variants = sum(len(it.get("variants") or []) for it in items)

    return {
        "ok": len(item_rows) == expected_items and len(variant_rows) == expected_variants,
        "expected_items": expected_items,
        "actual_items": len(item_rows),
        "expected_variants": expected_variants,
        "actual_variants": len(variant_rows),
    }


def _verify_json_round_trip(items):
    """Export to JSON format, parse back, verify structure matches."""
    export_items = []
    for it in items:
        eitem = {
            "id": it.get("id"),
            "name": it.get("name", ""),
            "description": it.get("description", ""),
            "price_cents": it.get("price_cents", 0),
            "category": it.get("category") or "",
            "position": it.get("position"),
        }
        variants = it.get("variants") or []
        eitem["variants"] = [
            {"label": v.get("label", ""), "price_cents": v.get("price_cents", 0),
             "kind": v.get("kind", "size")}
            for v in variants
        ] if variants else []
        export_items.append(eitem)

    payload = {"items": export_items}
    text = json.dumps(payload, indent=2)
    parsed = json.loads(text)

    actual_items = parsed.get("items", [])
    actual_variants = sum(len(i.get("variants", [])) for i in actual_items)
    expected_variants = sum(len(it.get("variants") or []) for it in items)

    return {
        "ok": len(actual_items) == len(items) and actual_variants == expected_variants,
        "expected_items": len(items),
        "actual_items": len(actual_items),
        "expected_variants": expected_variants,
        "actual_variants": actual_variants,
    }


def _verify_pos_json_round_trip(items, draft=None):
    """Export to Generic POS JSON, parse back, verify structure."""
    payload = _build_generic_pos_json(items, draft)
    text = json.dumps(payload, indent=2)
    parsed = json.loads(text)

    actual_items = 0
    actual_modifiers = 0
    for cat in parsed.get("menu", {}).get("categories", []):
        for item in cat.get("items", []):
            actual_items += 1
            actual_modifiers += len(item.get("modifiers", []))

    expected_variants = sum(len(it.get("variants") or []) for it in items)

    return {
        "ok": actual_items == len(items) and actual_modifiers == expected_variants,
        "expected_items": len(items),
        "actual_items": actual_items,
        "expected_modifiers": expected_variants,
        "actual_modifiers": actual_modifiers,
        "metadata": parsed.get("metadata", {}),
    }


@app.get("/drafts/<int:draft_id>/export/validate")
@login_required
def draft_export_validate(draft_id: int):
    """Pre-export validation: returns warnings for items missing data."""
    _require_drafts_storage()
    items = drafts_store.get_draft_items(draft_id, include_variants=True) or []
    warnings = _validate_draft_for_export(items)
    return jsonify({
        "draft_id": draft_id,
        "item_count": len(items),
        "variant_count": sum(len(it.get("variants") or []) for it in items),
        "warnings": warnings,
        "warning_count": len(warnings),
    })


@app.get("/drafts/<int:draft_id>/export/metrics")
@login_required
def draft_export_metrics(draft_id: int):
    """Export metrics: counts, breakdowns, price statistics."""
    _require_drafts_storage()
    items = drafts_store.get_draft_items(draft_id, include_variants=True) or []
    metrics = _compute_export_metrics(items)
    metrics["draft_id"] = draft_id
    return jsonify(metrics)


@app.get("/drafts/<int:draft_id>/export/preview")
@login_required
def draft_export_preview(draft_id: int):
    """Export preview: returns formatted output as JSON for pre-download review."""
    _require_drafts_storage()
    fmt = request.args.get("format", "generic_pos")
    items = drafts_store.get_draft_items(draft_id, include_variants=True) or []
    draft = drafts_store.get_draft(draft_id) or {}
    warnings = _validate_draft_for_export(items)

    if fmt == "square":
        rows = _build_square_rows(items)
        # Return first 50 rows as preview
        preview_lines = []
        headers = ["Token", "Item Name", "Description", "Category",
                    "Price", "Modifier Set Name", "Modifier Name", "Modifier Price"]
        preview_lines.append(",".join(headers))
        for r in rows[:50]:
            preview_lines.append(",".join(str(c) for c in r))
        content = "\n".join(preview_lines)
    elif fmt == "toast":
        rows = _build_toast_rows(items)
        headers = ["Menu Group", "Menu Item", "Base Price",
                    "Option Group", "Option", "Option Price"]
        preview_lines = [",".join(headers)]
        for r in rows[:50]:
            preview_lines.append(",".join(str(c) for c in r))
        content = "\n".join(preview_lines)
    else:
        payload = _build_generic_pos_json(items, draft)
        content = json.dumps(payload, indent=2)

    return jsonify({
        "format": fmt,
        "content": content,
        "item_count": len(items),
        "warnings": warnings,
        "truncated": fmt != "generic_pos" and len(items) > 50,
    })


def _build_square_rows(items):
    """Build Square CSV rows: items + modifier sets from modifier groups.

    Square import format:
      Token, Item Name, Description, Category, Price,
      Modifier Set Name, Modifier Name, Modifier Price,
      Required, Min Select, Max Select

    Items without modifiers: single row with base price.
    Items with modifier_groups: parent row, then modifier rows under each
    group (mapped 1:1 to Square Modifier Sets with selection rules).
    Items with only ungrouped_variants: fall back to kind-based grouping.
    """
    _KIND_LABELS = {
        "size": "Size",
        "combo": "Combo Add-on",
        "flavor": "Flavor",
        "style": "Style",
        "other": "Option",
    }

    rows = []
    for it in items:
        name = it.get("name", "")
        desc = it.get("description") or ""
        cat = it.get("category") or ""
        price = _format_price_dollars(it.get("price_cents", 0))

        modifier_groups = it.get("modifier_groups") or []
        ungrouped = it.get("ungrouped_variants") or it.get("variants") or []

        # Parent row (always emitted)
        rows.append([
            "item", name, desc, cat, price, "", "", "", "", "", "",
        ])

        if modifier_groups:
            # POS-native modifier groups → Square Modifier Sets (1:1)
            for grp in modifier_groups:
                set_name = grp.get("name", "Option")
                required = "Y" if grp.get("required") else "N"
                min_sel = grp.get("min_select") or 0
                max_sel = grp.get("max_select") or 0
                for mod in (grp.get("modifiers") or []):
                    mod_price = _format_price_dollars(mod.get("price_cents", 0))
                    rows.append([
                        "modifier", name, "", "", "",
                        set_name, mod.get("label", ""), mod_price,
                        required, str(min_sel), str(max_sel),
                    ])

        if ungrouped:
            # Ungrouped variants → kind-based modifier sets (backward compat)
            kind_groups: dict = {}
            for v in ungrouped:
                k = v.get("kind", "size")
                kind_groups.setdefault(k, []).append(v)

            for kind, vlist in kind_groups.items():
                set_name = _KIND_LABELS.get(kind, "Option")
                for v in vlist:
                    mod_price = _format_price_dollars(v.get("price_cents", 0))
                    rows.append([
                        "modifier", name, "", "", "",
                        set_name, v.get("label", ""), mod_price,
                        "", "", "",
                    ])
    return rows


def _build_toast_rows(items):
    """Build Toast CSV rows: menu group/item/option hierarchy.

    Toast import format:
      Menu Group, Menu Item, Base Price,
      Option Group, Option, Option Price, Required

    Items map to Menu Items under their category (Menu Group).
    Modifier groups map 1:1 to Toast Option Groups.
    Ungrouped variants fall back to kind-based grouping.
    """
    _KIND_LABELS = {
        "size": "Size",
        "combo": "Combo Add-on",
        "flavor": "Flavor",
        "style": "Style",
        "other": "Option",
    }

    rows = []
    for it in items:
        name = it.get("name", "")
        cat = it.get("category") or "Uncategorized"
        price = _format_price_dollars(it.get("price_cents", 0))

        modifier_groups = it.get("modifier_groups") or []
        ungrouped = it.get("ungrouped_variants") or it.get("variants") or []

        # Parent row (always emitted)
        rows.append([cat, name, price, "", "", "", ""])

        if modifier_groups:
            # POS-native modifier groups → Toast Option Groups (1:1)
            for grp in modifier_groups:
                group_name = grp.get("name", "Option")
                required = "Y" if grp.get("required") else "N"
                for mod in (grp.get("modifiers") or []):
                    opt_price = _format_price_dollars(mod.get("price_cents", 0))
                    rows.append([
                        "", "", "", group_name,
                        mod.get("label", ""), opt_price, required,
                    ])

        if ungrouped:
            # Ungrouped variants → kind-based option groups (backward compat)
            kind_groups: dict = {}
            for v in ungrouped:
                k = v.get("kind", "size")
                kind_groups.setdefault(k, []).append(v)

            for kind, vlist in kind_groups.items():
                group_name = _KIND_LABELS.get(kind, "Option")
                for v in vlist:
                    opt_price = _format_price_dollars(v.get("price_cents", 0))
                    rows.append([
                        "", "", "", group_name,
                        v.get("label", ""), opt_price, "",
                    ])
    return rows


def _build_generic_pos_json(items, draft=None):
    """Build Generic POS JSON: universal item/variant/modifier schema.

    Structure:
      { menu: { id, title, categories: [
          { name, items: [
              { name, kitchen_name, description, base_price,
                modifier_groups: [               # POS-native (when present)
                  { name, required, min_select, max_select,
                    modifiers: [{ name, price }] }
                ],
                modifiers: [                     # legacy flat fallback
                  { group, name, price }
                ]
              }
          ]}
      ]}, metadata: { ... } }

    When an item carries modifier_groups[], they are emitted as POS-native
    nested groups (with required/min_select/max_select metadata).  The legacy
    flat `modifiers` list is still included for backward compatibility, built
    by flattening all modifier groups.  Items that only have ungrouped
    variants use the old flat-only format (modifier_groups=[]).
    """
    _KIND_LABELS = {
        "size": "Size",
        "combo": "Combo Add-on",
        "flavor": "Flavor",
        "style": "Style",
        "other": "Option",
    }

    draft = draft or {}
    cat_map: dict = {}
    for it in items:
        cat = it.get("category") or "Uncategorized"
        cat_map.setdefault(cat, []).append(it)

    categories = []
    for cat_name in sorted(cat_map.keys()):
        cat_items = []
        for it in cat_map[cat_name]:
            kitchen = (it.get("kitchen_name") or "").strip() or None
            item_entry = {
                "name": it.get("name", ""),
                "description": it.get("description") or "",
                "base_price": _format_price_dollars(it.get("price_cents", 0)),
            }
            if kitchen:
                item_entry["kitchen_name"] = kitchen

            modifier_groups = it.get("modifier_groups") or []
            ungrouped = it.get("ungrouped_variants") or it.get("variants") or []

            if modifier_groups:
                # POS-native nested groups
                pos_groups = []
                flat_modifiers = []
                for grp in modifier_groups:
                    grp_modifiers = []
                    for mod in (grp.get("modifiers") or []):
                        mod_entry = {
                            "name": mod.get("label", ""),
                            "price": _format_price_dollars(mod.get("price_cents", 0)),
                        }
                        grp_modifiers.append(mod_entry)
                        flat_modifiers.append({
                            "group": grp.get("name", ""),
                            "name": mod.get("label", ""),
                            "price": _format_price_dollars(mod.get("price_cents", 0)),
                        })
                    pos_groups.append({
                        "name": grp.get("name", ""),
                        "required": bool(grp.get("required")),
                        "min_select": grp.get("min_select") or 0,
                        "max_select": grp.get("max_select") or 0,
                        "modifiers": grp_modifiers,
                    })
                # also flatten any remaining ungrouped variants
                for v in ungrouped:
                    flat_modifiers.append({
                        "group": _KIND_LABELS.get(v.get("kind", "other"), "Option"),
                        "name": v.get("label", ""),
                        "price": _format_price_dollars(v.get("price_cents", 0)),
                    })
                item_entry["modifier_groups"] = pos_groups
                item_entry["modifiers"] = flat_modifiers
            else:
                # legacy flat-only format (no named groups)
                item_entry["modifier_groups"] = []
                if ungrouped:
                    item_entry["modifiers"] = [
                        {
                            "group": _KIND_LABELS.get(v.get("kind", "other"), "Option"),
                            "name": v.get("label", ""),
                            "price": _format_price_dollars(v.get("price_cents", 0)),
                        }
                        for v in ungrouped
                    ]
                else:
                    item_entry["modifiers"] = []

            cat_items.append(item_entry)
        categories.append({"name": cat_name, "items": cat_items})

    return {
        "menu": {
            "id": draft.get("id"),
            "title": draft.get("title") or "",
            "categories": categories,
        },
        "metadata": {
            "exported_at": _now_iso(),
            "format": "generic_pos",
            "version": "1.1",
            "item_count": len(items),
            "category_count": len(categories),
        },
    }


@app.get("/drafts/<int:draft_id>/export_square.csv")
@login_required
@_require_tier_chosen
def draft_export_square_csv(draft_id: int):
    """Square POS CSV export: items + modifier groups."""
    _require_drafts_storage()
    items = drafts_store.get_draft_items(draft_id, include_modifier_groups=True) or []
    rows = _build_square_rows(items)

    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(["Token", "Item Name", "Description", "Category",
                      "Price", "Modifier Set Name", "Modifier Name", "Modifier Price",
                      "Required", "Min Select", "Max Select"])
    for r in rows:
        writer.writerow(r)

    csv_data = buf.getvalue().encode("utf-8-sig")
    resp = make_response(csv_data)
    resp.headers["Content-Type"] = "text/csv; charset=utf-8"
    resp.headers["Content-Disposition"] = f'attachment; filename="draft_{draft_id}_square.csv"'
    return resp


@app.get("/drafts/<int:draft_id>/export_toast.csv")
@login_required
@_require_tier_chosen
def draft_export_toast_csv(draft_id: int):
    """Toast POS CSV export: menu group / item / option hierarchy."""
    _require_drafts_storage()
    items = drafts_store.get_draft_items(draft_id, include_modifier_groups=True) or []
    rows = _build_toast_rows(items)

    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(["Menu Group", "Menu Item", "Base Price",
                      "Option Group", "Option", "Option Price", "Required"])
    for r in rows:
        writer.writerow(r)

    csv_data = buf.getvalue().encode("utf-8-sig")
    resp = make_response(csv_data)
    resp.headers["Content-Type"] = "text/csv; charset=utf-8"
    resp.headers["Content-Disposition"] = f'attachment; filename="draft_{draft_id}_toast.csv"'
    return resp


@app.get("/drafts/<int:draft_id>/export_pos.json")
@login_required
@_require_tier_chosen
def draft_export_pos_json(draft_id: int):
    """Generic POS JSON export: universal item/variant/modifier schema."""
    _require_drafts_storage()
    draft = drafts_store.get_draft(draft_id) or {}
    items = drafts_store.get_draft_items(draft_id, include_modifier_groups=True) or []
    payload = _build_generic_pos_json(items, draft)

    resp = make_response(json.dumps(payload, indent=2))
    resp.headers["Content-Type"] = "application/json; charset=utf-8"
    resp.headers["Content-Disposition"] = f'attachment; filename="draft_{draft_id}_pos.json"'
    return resp


# ------------------------
# NEW: Fix descriptions endpoint used by the editor button
# ------------------------
def _clean_desc(name: str, desc: Optional[str]) -> Optional[str]:
    """
    Heuristics:
      - Trim, collapse whitespace.
      - Drop leading repeats of the name: "Pepperoni: pepperoni slices..." -> "pepperoni slices..."
      - Remove surrounding quotes and stray punctuation at ends.
      - If desc becomes empty, return None.
    """
    nm = (name or "").strip()
    d = (desc or "").strip()

    if not d:
        return None

    # If description begins with the name (case-insensitive), remove that prefix + common separators
    lowered = d.lower()
    nm_low = nm.lower()
    prefixes = [f"{nm_low} - ", f"{nm_low} — ", f"{nm_low} – ", f"{nm_low}: ", f"{nm_low} —", f"{nm_low} –", f"{nm_low}:"]
    for pre in prefixes:
        if lowered.startswith(pre):
            d = d[len(pre):].lstrip()
            lowered = d.lower()
            break

    # Strip wrapping quotes/parens
    d = d.strip(" '\"\t\r\n")
    # Fix bad spacing around punctuation
    d = re.sub(r"\s+([,.;:!?])", r"\1", d)
    d = re.sub(r"\s{2,}", " ", d).strip()

    return d or None

def _split_name_into_desc_if_needed(name: str, desc: Optional[str]) -> Tuple[str, Optional[str]]:
    """
    If description is empty but name contains " - " or " — ", split once and move the tail to description.
    Keep conservative: only when the left part isn't too long (title-like) and right part has letters.
    """
    nm = (name or "").strip()
    if (desc or "").strip():
        return nm, (desc or None)

    for sep in [" - ", " — ", " – ", ":", " · "]:
        if sep in nm:
            left, right = nm.split(sep, 1)
            left, right = left.strip(), right.strip()
            if left and right and any(c.isalpha() for c in right) and len(left) <= 80:
                return left, right
    return nm, (desc or None)

@app.route("/drafts/<int:draft_id>/fix-descriptions", methods=["POST", "GET"])
@login_required
def fix_descriptions_for_draft(draft_id: int):
    """
    Clean up funky descriptions in-place. Safe, idempotent heuristics.

    Behavior:
      - GET  → redirect back to editor (no mutation).
      - POST (AJAX/JSON) → returns JSON {ok, updated_count} (editor JS can reload).
      - POST (regular form OR ?redirect=1) → flash + redirect back to editor (no JSON blank page).
    """
    _require_drafts_storage()

    if request.method == "GET":
        return redirect(url_for("draft_editor", draft_id=draft_id))

    # --- detect if caller wants redirect instead of JSON ---
    ct = (request.headers.get("Content-Type") or "").lower()
    is_form_post = ct.startswith("application/x-www-form-urlencoded") or ct.startswith("multipart/form-data")
    wants_redirect = (
        request.args.get("redirect") == "1"
        or (request.form.get("redirect") == "1" if is_form_post else False)
        or is_form_post
    )

    items = drafts_store.get_draft_items(draft_id) or []
    updates = []

    for it in items:
        _id = it.get("id")
        if _id is None:
            continue
        name = (it.get("name") or "").strip()
        desc = it.get("description")
        # Optionally split long names into name+desc if desc is empty
        new_name, maybe_desc = _split_name_into_desc_if_needed(name, desc)
        # Always run the desc cleaner
        new_desc = _clean_desc(new_name, maybe_desc)

        # Only push updates when something has changed
        changed = False
        upd = {"id": int(_id)}

        if new_name != name:
            upd["name"] = new_name
            changed = True
        # Normalize empty -> None; DB layer should treat None as NULL
        norm_desc = new_desc if (new_desc and new_desc.strip()) else None
        # Only update if different from current (normalize current too)
        cur_norm = (desc or None)
        if (norm_desc or None) != (cur_norm or None):
            upd["description"] = norm_desc
            changed = True

        if changed:
            updates.append(upd)

    updated_count = 0
    if updates:
        try:
            res = drafts_store.upsert_draft_items(draft_id, updates)
            updated_count = len(res.get("updated_ids", [])) + len(res.get("inserted_ids", []))
            # bump updated_at
            try:
                ds = drafts_store.get_draft(draft_id) or {}
                drafts_store.save_draft_metadata(draft_id, title=ds.get("title"))
            except Exception:
                pass
        except Exception as e:
            if wants_redirect:
                flash(f"Description cleanup failed: {e}", "error")
                return redirect(url_for("draft_editor", draft_id=draft_id))
            return jsonify({"ok": False, "error": f"update failed: {e}"}), 500

    if wants_redirect:
        flash(f"Cleaned descriptions — {int(updated_count)} item(s) updated.", "success")
        return redirect(url_for("draft_editor", draft_id=draft_id))

    return jsonify({"ok": True, "updated_count": int(updated_count)}), 200

def _ensure_draft_for_job(job_id: int, row=None) -> Optional[int]:
    """
    Ensure there is a DB-backed draft for an import job.

    Strategy:
      1) Use existing helper _get_or_create_draft_for_job(job_id)
      2) If legacy JSON exists (import_jobs.draft_path), upgrade it into a DB draft
         using drafts_store.create_draft_from_structured_items(...)
      3) Otherwise attempt to create via drafts_store if it exposes
         create_draft_from_import / create_draft_from_import_job.

    This function does NOT run OCR.
    """
    try:
        _require_drafts_storage()
    except Exception:
        return None

    # First: try the canonical helper
    try:
        did = _get_or_create_draft_for_job(job_id)
        if did:
            did_int = int(did)
            try:
                update_import_job(job_id, draft_id=did_int)
            except Exception:
                pass
            return did_int
    except Exception:
        pass

    filename = ""
    source_type = ""
    restaurant_id = None
    legacy_draft_path = None

    try:
        if row is not None:
            filename = (row["filename"] or "").strip()
            try:
                source_type = (row.get("source_type") or "").strip()
            except Exception:
                source_type = ""
            try:
                legacy_draft_path = (row.get("draft_path") or "").strip() or None
            except Exception:
                legacy_draft_path = None
            rid = None
            try:
                rid = row.get("restaurant_id")
            except Exception:
                rid = None
            try:
                restaurant_id = int(rid) if rid is not None else None
            except Exception:
                restaurant_id = None
    except Exception:
        filename = ""
        source_type = ""
        restaurant_id = None
        legacy_draft_path = None

    title = (Path(filename).stem if filename else f"Import {job_id}").strip() or f"Import {job_id}"

    # ---------------------------------------------------------------------
    # NEW: If the job only has a legacy JSON draft_path, upgrade it to DB.
    # ---------------------------------------------------------------------
    try:
        if legacy_draft_path:
            abs_legacy = _abs_from_rel(legacy_draft_path)
            if abs_legacy and abs_legacy.exists():
                legacy = json.loads(abs_legacy.read_text(encoding="utf-8"))

                items: List[Dict[str, Any]] = []
                cats = legacy.get("categories") or []
                for cat_obj in cats:
                    cat_name = (cat_obj.get("name") or "").strip() or None
                    for it in (cat_obj.get("items") or []):
                        base_name = (it.get("name") or "").strip()
                        if not base_name:
                            continue
                        desc = (it.get("description") or "").strip() or None

                        sizes = it.get("sizes") or []
                        if sizes:
                            for s in sizes:
                                size_name = (s.get("name") or "").strip()
                                price_val = s.get("price", 0)
                                try:
                                    price_cents = int(round(float(price_val) * 100))
                                except Exception:
                                    price_cents = 0
                                display_name = f"{base_name} ({size_name})" if size_name else base_name
                                items.append(
                                    {
                                        "name": display_name,
                                        "description": desc,
                                        "category": cat_name,
                                        "subcategory": None,
                                        "price_cents": max(int(price_cents), 0),
                                    }
                                )
                        else:
                            price_val = it.get("price", 0)
                            try:
                                price_cents = int(round(float(price_val) * 100))
                            except Exception:
                                price_cents = 0
                            items.append(
                                {
                                    "name": base_name,
                                    "description": desc,
                                    "category": cat_name,
                                    "subcategory": None,
                                    "price_cents": max(int(price_cents), 0),
                                }
                            )

                create_structured = getattr(drafts_store, "create_draft_from_structured_items", None)
                if callable(create_structured) and items:
                    source_meta = {
                        "job_id": int(job_id),
                        "filename": filename,
                        "source_type": source_type or "legacy_json",
                        "legacy_draft_path": legacy_draft_path,
                    }
                    draft = create_structured(
                        title=title,
                        restaurant_id=restaurant_id,
                        items=items,
                        source_type="legacy_json",
                        source_job_id=int(job_id),
                        source_meta=source_meta,
                    )
                    did_int = int(draft.get("id") or draft.get("draft_id") or 0)

                    if did_int > 0:
                        try:
                            update_import_job(job_id, draft_id=did_int)
                        except Exception:
                            pass
                        return did_int
    except Exception:
        pass

    # Second: attempt explicit creation via drafts_store (if supported)
    create_fn = getattr(drafts_store, "create_draft_from_import", None)
    if not callable(create_fn):
        create_fn = getattr(drafts_store, "create_draft_from_import_job", None)

    if not callable(create_fn):
        return None

    source_file_path = None
    try:
        if filename:
            p = (UPLOAD_FOLDER / filename).resolve()
            if p.exists():
                source_file_path = str(p)
    except Exception:
        source_file_path = None

    source_meta = {
        "job_id": int(job_id),
        "filename": filename,
        "source_type": source_type,
    }

    draft = None

    # Try rich kwargs first
    try:
        draft = create_fn(
            title=title,
            restaurant_id=restaurant_id,
            source_type=source_type or "ocr",
            source_job_id=int(job_id),
            source_file_path=source_file_path,
            source_meta=source_meta,
        )
    except TypeError:
        draft = None
    except Exception:
        draft = None

    # Try minimal kwargs
    if draft is None:
        try:
            draft = create_fn(
                title=title,
                restaurant_id=restaurant_id,
                source_job_id=int(job_id),
            )
        except TypeError:
            draft = None
        except Exception:
            draft = None

    # Try positional job_id only
    if draft is None:
        try:
            draft = create_fn(int(job_id))
        except Exception:
            draft = None

    did_int = 0
    try:
        if isinstance(draft, dict):
            did_int = int(draft.get("id") or draft.get("draft_id") or 0)
        elif draft is not None:
            did_int = int(draft)
    except Exception:
        did_int = 0

    if did_int > 0:
        try:
            update_import_job(job_id, draft_id=did_int)
        except Exception:
            pass
        return did_int

    return None


# (Day 100.5: AI Heuristics routes removed — imports_ai_preview, imports_ai_commit, imports_ai_finalize)
# Heuristic fallback produced low-quality garble; free tier now gets empty draft for manual input.
# Pipeline Debug view replaces the old heuristic preview.

# ------------------------
# Pipeline Debug View (Day 100.5)
# ------------------------
@app.get("/drafts/<int:draft_id>/pipeline-debug")
@login_required
def draft_pipeline_debug(draft_id: int):
    """
    Day 100.5: Pipeline Debug view — shows OCR text, Claude extraction results,
    vision verification, semantic pipeline, and pipeline metrics for a draft.
    All data comes from the stored debug payload (saved during run_ocr_and_make_draft).
    """
    _require_drafts_storage()
    draft = drafts_store.get_draft(draft_id)
    if not draft:
        abort(404)

    dbg = {}
    if hasattr(drafts_store, "load_ocr_debug"):
        dbg = drafts_store.load_ocr_debug(draft_id) or {}

    has_debug = bool(dbg and dbg.get("extraction_strategy"))

    return render_template(
        "pipeline_debug.html",
        draft=draft,
        debug_payload=dbg,
        has_debug=has_debug,
        extraction_strategy=dbg.get("extraction_strategy", "unknown"),
        clean_ocr_chars=dbg.get("clean_ocr_chars", 0),
        raw_ocr_text=dbg.get("raw_ocr_text", ""),
        vision=dbg.get("vision_verification"),
        semantic=dbg.get("semantic_pipeline"),
        metrics=dbg.get("pipeline_metrics"),
    )


# ------------------------
# Diagnostics
# ------------------------
@app.get("/__ping")
def __ping():

    return jsonify({"ok": True, "time": _now_iso()})

@app.get("/__routes")
def __routes():
    """
    Debug helper: list all registered routes.

    Made defensive so that if any weird rule object blows up during
    stringification or sorting, we surface the error instead of a bare 500.
    """
    try:
        routes = []
        for r in app.url_map.iter_rules():
            try:
                routes.append(str(r.rule))
            except Exception as inner:
                # Fallback so a single bad rule doesn't kill the whole endpoint
                routes.append(f"<unprintable rule: {inner.__class__.__name__}>")

        routes.sort()
        return jsonify({"ok": True, "count": len(routes), "routes": routes})
    except Exception as e:
        # With FLASK_DEBUG=1 and the dev errorhandler, this will also log the traceback.
        app.logger.exception("__routes failed")
        return jsonify({"ok": False, "error": str(e)}), 500


@app.get("/__boom")
def __boom():
    raise RuntimeError("Intentional test error")

@app.get("/__ocrtxt/<int:job_id>")
def __ocrtxt(job_id: int):
    """
    Debug: return AI helper items + structured categories + superimport bundle
    for the given import job.
    """
    from storage.ocr_facade import extract_menu_from_pdf
    from storage.drafts import find_draft_by_source_job

    draft = find_draft_by_source_job(job_id)
    if not draft:
        return jsonify({"ok": False, "error": "draft not found"}), 404

    # Original upload path
    path = draft.get("source_file_path")
    if not path:
        return jsonify({"ok": False, "error": "no source_file_path"}), 400

    structured, debug = extract_menu_from_pdf(path)

    return jsonify({
        "ok": True,
        "job_id": job_id,
        "structured": structured,
        "superimport": debug.get("superimport"),
        "ai_items": debug.get("ai_preview", {}).get("items"),
        "hierarchy": debug.get("ai_preview", {}).get("hierarchy"),
    })

# ------------------------
# Blueprint registration (core)
# ------------------------
try:
    from .routes.core import core_bp  # type: ignore
except Exception:
    from routes.core import core_bp  # fallback if relative import fails

app.register_blueprint(core_bp)


# --------------------------------------------------------
# TEMPORARY TEST ROUTE — bypass browser validation
# --------------------------------------------------------
@app.get("/test_csv_form")
def test_csv_form():
    return """
    <form action="/api/drafts/import_structured" method="post" enctype="multipart/form-data">
        <input type="file" name="file">

        <button type="submit">Test CSV Upload</button>
    </form>
    """


# ------------------------
# Run
# ------------------------
if __name__ == "__main__":
    app.run(host="0.0.0.0", port=5000, debug=True, threaded=True)

# === DEBUG APPEND (auto-added) ===
import os as _os
if _os.environ.get("FLASK_DEBUG") == "1":
    @app.errorhandler(500)
    def _dev_rethrow(e):
        import traceback
        traceback.print_exc()
        raise e

    @app.get("/__debug/imports/<int:job_id>/draft")
    @login_required
    def imports_draft_debug(job_id: int):
        try:
            draft_id = _get_or_create_draft_for_job(job_id)
            if draft_id:
                return redirect(url_for("draft_editor", draft_id=draft_id))
            flash("Draft not ready for editor yet. Showing legacy import view.", "info")
            return redirect(url_for("imports_detail", job_id=job_id))
        except Exception:
            import traceback, html
            tb = traceback.format_exc()
            return f"<pre>{html.escape(tb)}</pre>", 500
# === Day 122: Editor Stats + Bulk Card Actions ===


@app.get("/drafts/<int:draft_id>/stats")
@login_required
def draft_stats(draft_id: int):
    """Return live editor stats as JSON."""
    _require_drafts_storage()
    draft = drafts_store.get_draft(draft_id)
    if not draft:
        return jsonify({"error": "Draft not found"}), 404
    items = drafts_store.get_draft_items(draft_id, include_modifier_groups=True) or []
    stats = _compute_editor_stats(items)
    return jsonify({"ok": True, "stats": stats}), 200


@app.post("/drafts/<int:draft_id>/bulk_delete")
@login_required
def bulk_delete_items(draft_id: int):
    """
    Delete multiple items by ID from a draft (card-view bulk action).
    Body: {"item_ids": [1, 2, 3]}
    """
    _require_drafts_storage()
    draft = drafts_store.get_draft(draft_id)
    if not draft:
        return jsonify({"error": "Draft not found"}), 404
    if draft.get("status") != "editing":
        return jsonify({"ok": False, "error": "Draft is not editable"}), 403
    payload = request.get_json(silent=True) or {}
    raw_ids = payload.get("item_ids")
    if not isinstance(raw_ids, list):
        return jsonify({"ok": False, "error": "'item_ids' must be a list"}), 400
    int_ids = []
    for x in raw_ids:
        try:
            int_ids.append(int(x))
        except (ValueError, TypeError):
            return jsonify({"ok": False, "error": f"Invalid item id: {x}"}), 400
    deleted = drafts_store.delete_draft_items(draft_id, int_ids) if int_ids else 0
    return jsonify({"ok": True, "deleted": deleted}), 200


@app.post("/drafts/<int:draft_id>/bulk_move_category")
@login_required
def bulk_move_category(draft_id: int):
    """
    Move multiple items to a new category.
    Body: {"item_ids": [1, 2, 3], "category": "Appetizers"}
    """
    _require_drafts_storage()
    draft = drafts_store.get_draft(draft_id)
    if not draft:
        return jsonify({"error": "Draft not found"}), 404
    if draft.get("status") != "editing":
        return jsonify({"ok": False, "error": "Draft is not editable"}), 403
    payload = request.get_json(silent=True) or {}
    raw_ids = payload.get("item_ids")
    category = payload.get("category")
    if not isinstance(raw_ids, list):
        return jsonify({"ok": False, "error": "'item_ids' must be a list"}), 400
    if not isinstance(category, str) or not category.strip():
        return jsonify({"ok": False, "error": "'category' must be a non-empty string"}), 400
    category = category.strip()
    int_ids = []
    for x in raw_ids:
        try:
            int_ids.append(int(x))
        except (ValueError, TypeError):
            return jsonify({"ok": False, "error": f"Invalid item id: {x}"}), 400
    if not int_ids:
        return jsonify({"ok": True, "updated": 0}), 200
    with db_connect() as conn:
        qmarks = ",".join(["?"] * len(int_ids))
        conn.execute(
            f"UPDATE draft_items SET category=? WHERE draft_id=? AND id IN ({qmarks})",
            (category, int(draft_id), *int_ids),
        )
        conn.commit()
    return jsonify({"ok": True, "updated": len(int_ids), "category": category}), 200


# === /DEBUG APPEND ===
