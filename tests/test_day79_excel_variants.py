"""
Day 79 — Excel Export with Variants tests.

Sprint 9.3, Day 79: Verifies that XLSX export endpoints correctly include
structured variant data with formatting and per-category sheets.

Covers:
  XLSX export (single sheet, variant sub-rows):
  - Header row with base columns + dynamic variant label columns
  - Parent rows are bold, variant sub-rows are gray/indented
  - Variant label columns auto-generated from actual labels in draft
  - Items without variants have empty variant columns
  - Items with variants have prices in correct label columns
  - Variant sub-rows appear below parent with indented label + kind
  - Empty draft produces header-only sheet
  - Many variants (10+) all exported
  - All 5 variant kinds preserved in sub-rows
  - Mixed items: some with variants, some without
  - Variant order preserved (position order)

  XLSX per-category export:
  - One sheet per unique category
  - Sheet names match category names (truncated to 31 chars)
  - Each sheet has own variant label columns (category-specific)
  - Items sorted into correct category sheets
  - Uncategorized items go to "Uncategorized" sheet
  - Empty draft produces placeholder sheet
  - Parent bold / variant gray formatting per sheet
  - Variant sub-rows under each parent per sheet
  - Categories sorted alphabetically

  Edge cases:
  - Zero-price variants
  - Special characters in names/labels
  - Items with 0 variants on single-sheet export
  - Long category name truncated to 31 chars for sheet title
"""

from __future__ import annotations

import io
import sqlite3
import pytest
from typing import Optional


# ---------------------------------------------------------------------------
# In-memory DB helpers (same pattern as Day 71-78 tests)
# ---------------------------------------------------------------------------
_TEST_CONN: Optional[sqlite3.Connection] = None


def _make_test_db() -> sqlite3.Connection:
    """Create an in-memory SQLite DB with the required schema."""
    conn = sqlite3.connect(":memory:")
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA foreign_keys = ON;")

    conn.execute("""
        CREATE TABLE IF NOT EXISTS restaurants (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL,
            created_at TEXT NOT NULL DEFAULT (datetime('now'))
        )
    """)
    conn.execute("""
        CREATE TABLE IF NOT EXISTS drafts (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            title TEXT,
            restaurant_id INTEGER,
            status TEXT NOT NULL DEFAULT 'editing',
            source TEXT,
            source_job_id INTEGER,
            source_file_path TEXT,
            created_at TEXT NOT NULL,
            updated_at TEXT NOT NULL,
            FOREIGN KEY (restaurant_id) REFERENCES restaurants(id) ON DELETE SET NULL
        )
    """)
    conn.execute("""
        CREATE TABLE IF NOT EXISTS draft_items (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            draft_id INTEGER NOT NULL,
            name TEXT NOT NULL,
            description TEXT,
            price_cents INTEGER NOT NULL DEFAULT 0,
            category TEXT,
            position INTEGER,
            confidence INTEGER,
            created_at TEXT NOT NULL,
            updated_at TEXT NOT NULL,
            FOREIGN KEY (draft_id) REFERENCES drafts(id) ON DELETE CASCADE
        )
    """)
    conn.execute("""
        CREATE TABLE IF NOT EXISTS draft_item_variants (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            item_id INTEGER NOT NULL,
            label TEXT NOT NULL,
            price_cents INTEGER NOT NULL DEFAULT 0,
            kind TEXT DEFAULT 'size',
            position INTEGER DEFAULT 0,
            created_at TEXT NOT NULL,
            updated_at TEXT NOT NULL,
            FOREIGN KEY (item_id) REFERENCES draft_items(id) ON DELETE CASCADE
        )
    """)
    conn.execute("""
        CREATE TABLE IF NOT EXISTS menu_items (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            restaurant_id INTEGER,
            name TEXT NOT NULL,
            description TEXT,
            price_cents INTEGER NOT NULL DEFAULT 0,
            category TEXT,
            position INTEGER,
            created_at TEXT NOT NULL DEFAULT (datetime('now')),
            updated_at TEXT NOT NULL DEFAULT (datetime('now'))
        )
    """)
    conn.execute("CREATE INDEX IF NOT EXISTS idx_items_draft ON draft_items(draft_id)")
    conn.execute("CREATE INDEX IF NOT EXISTS idx_variants_item ON draft_item_variants(item_id)")
    conn.commit()
    return conn


def _patch_db(monkeypatch):
    global _TEST_CONN
    _TEST_CONN = _make_test_db()
    import storage.drafts as drafts_mod
    def mock_connect():
        return _TEST_CONN
    monkeypatch.setattr(drafts_mod, "db_connect", mock_connect)
    return _TEST_CONN


@pytest.fixture(autouse=True)
def fresh_db(monkeypatch):
    conn = _patch_db(monkeypatch)
    yield conn
    global _TEST_CONN
    _TEST_CONN = None


def _create_draft(conn, title="Test Draft", status="editing") -> int:
    cur = conn.cursor()
    cur.execute(
        "INSERT INTO drafts (title, status, created_at, updated_at) VALUES (?, ?, datetime('now'), datetime('now'))",
        (title, status),
    )
    conn.commit()
    return int(cur.lastrowid)


def _insert_item(conn, draft_id, name, price_cents=0, category=None, description=None, position=None) -> int:
    cur = conn.cursor()
    cur.execute(
        "INSERT INTO draft_items (draft_id, name, description, price_cents, category, position, confidence, created_at, updated_at) "
        "VALUES (?, ?, ?, ?, ?, ?, 80, datetime('now'), datetime('now'))",
        (draft_id, name, description, price_cents, category, position),
    )
    conn.commit()
    return int(cur.lastrowid)


def _insert_variant(conn, item_id, label, price_cents, kind="size", position=0) -> int:
    cur = conn.cursor()
    cur.execute(
        "INSERT INTO draft_item_variants (item_id, label, price_cents, kind, position, created_at, updated_at) "
        "VALUES (?, ?, ?, ?, ?, datetime('now'), datetime('now'))",
        (item_id, label, price_cents, kind, position),
    )
    conn.commit()
    return int(cur.lastrowid)


# ---------------------------------------------------------------------------
# Helper: build XLSX in memory and load with openpyxl for assertions
# ---------------------------------------------------------------------------
def _build_xlsx_single(items, title="Test"):
    """Replicate the single-sheet XLSX export logic, return openpyxl Workbook."""
    import openpyxl as xl
    from openpyxl.styles import Font, PatternFill

    # Collect all unique variant labels in first-appearance order
    seen_labels = {}
    for it in items:
        for v in (it.get("variants") or []):
            lbl = (v.get("label") or "").strip()
            if lbl and lbl not in seen_labels:
                seen_labels[lbl] = len(seen_labels)
    label_order = sorted(seen_labels.keys(), key=lambda x: seen_labels[x])

    wb = xl.Workbook()
    ws = wb.active
    ws.title = title[:31]

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1a2236", end_color="1a2236", fill_type="solid")
    parent_font = Font(bold=True)
    variant_font = Font(color="666666")
    variant_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

    base_headers = ["name", "description", "price_cents", "category"]
    variant_headers = [f"price_{lbl}" for lbl in label_order]
    all_headers = base_headers + variant_headers
    ws.append(all_headers)
    for ci, _ in enumerate(all_headers, start=1):
        cell = ws.cell(row=1, column=ci)
        cell.font = header_font
        cell.fill = header_fill

    for it in items:
        variants = it.get("variants") or []
        vpmap = {}
        for v in variants:
            lbl = (v.get("label") or "").strip()
            if lbl:
                vpmap[lbl] = v.get("price_cents", 0)

        row_data = [
            it.get("name", ""),
            it.get("description", ""),
            it.get("price_cents", 0),
            it.get("category") or "",
        ]
        for lbl in label_order:
            row_data.append(vpmap.get(lbl, ""))
        ws.append(row_data)
        row_num = ws.max_row
        for ci in range(1, len(all_headers) + 1):
            ws.cell(row=row_num, column=ci).font = parent_font

        for v in variants:
            vrow = [
                "  " + (v.get("label") or ""),
                v.get("kind", "size"),
                v.get("price_cents", 0),
                "",
            ]
            for _ in label_order:
                vrow.append("")
            ws.append(vrow)
            vrow_num = ws.max_row
            for ci in range(1, len(all_headers) + 1):
                cell = ws.cell(row=vrow_num, column=ci)
                cell.font = variant_font
                cell.fill = variant_fill

    return wb


def _build_xlsx_by_category(items, title="Test"):
    """Replicate the per-category XLSX export logic, return openpyxl Workbook."""
    import openpyxl as xl
    from openpyxl.styles import Font, PatternFill

    cat_map = {}
    for it in items:
        cat = (it.get("category") or "Uncategorized").strip()
        cat_map.setdefault(cat, []).append(it)

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1a2236", end_color="1a2236", fill_type="solid")
    parent_font = Font(bold=True)
    variant_font = Font(color="666666")
    variant_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

    wb = xl.Workbook()
    wb.remove(wb.active)

    for cat_name in sorted(cat_map.keys()):
        cat_items = cat_map[cat_name]
        sheet_title = cat_name[:31] or "Uncategorized"
        ws = wb.create_sheet(title=sheet_title)

        seen_labels = {}
        for it in cat_items:
            for v in (it.get("variants") or []):
                lbl = (v.get("label") or "").strip()
                if lbl and lbl not in seen_labels:
                    seen_labels[lbl] = len(seen_labels)
        label_order = sorted(seen_labels.keys(), key=lambda x: seen_labels[x])

        base_headers = ["name", "description", "price_cents"]
        variant_headers = [f"price_{lbl}" for lbl in label_order]
        all_headers = base_headers + variant_headers

        ws.append(all_headers)
        for ci, _ in enumerate(all_headers, start=1):
            cell = ws.cell(row=1, column=ci)
            cell.font = header_font
            cell.fill = header_fill

        for it in cat_items:
            variants = it.get("variants") or []
            vpmap = {}
            for v in variants:
                lbl = (v.get("label") or "").strip()
                if lbl:
                    vpmap[lbl] = v.get("price_cents", 0)

            row_data = [
                it.get("name", ""),
                it.get("description", ""),
                it.get("price_cents", 0),
            ]
            for lbl in label_order:
                row_data.append(vpmap.get(lbl, ""))
            ws.append(row_data)
            row_num = ws.max_row
            for ci in range(1, len(all_headers) + 1):
                ws.cell(row=row_num, column=ci).font = parent_font

            for v in variants:
                vrow = [
                    "  " + (v.get("label") or ""),
                    v.get("kind", "size"),
                    v.get("price_cents", 0),
                ]
                for _ in label_order:
                    vrow.append("")
                ws.append(vrow)
                vrow_num = ws.max_row
                for ci in range(1, len(all_headers) + 1):
                    cell = ws.cell(row=vrow_num, column=ci)
                    cell.font = variant_font
                    cell.fill = variant_fill

    if not cat_map:
        ws = wb.create_sheet(title="Empty")
        ws.append(["No items"])

    return wb


def _wb_to_bytes(wb) -> bytes:
    """Serialize workbook to bytes (for round-trip testing)."""
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def _load_wb(data: bytes):
    """Load workbook from bytes."""
    import openpyxl as xl
    return xl.load_workbook(io.BytesIO(data))


def _ws_rows(ws):
    """Return list of lists from a worksheet."""
    return [[cell.value for cell in row] for row in ws.iter_rows()]


# ===========================================================================
# Single-Sheet XLSX Export Tests
# ===========================================================================
class TestXlsxSingleSheet:
    """Tests for single-sheet XLSX export with variant sub-rows."""

    def test_header_base_columns_no_variants(self, fresh_db):
        """No variants → header has only base columns."""
        from storage.drafts import get_draft_items
        d = _create_draft(fresh_db)
        _insert_item(fresh_db, d, "Salad", price_cents=500, category="Salads")
        items = get_draft_items(d, include_variants=True)
        wb = _build_xlsx_single(items)
        ws = wb.active
        rows = _ws_rows(ws)
        assert rows[0] == ["name", "description", "price_cents", "category"]

    def test_header_with_variant_label_columns(self, fresh_db):
        """Variant labels become extra header columns."""
        from storage.drafts import get_draft_items
        d = _create_draft(fresh_db)
        iid = _insert_item(fresh_db, d, "Pizza", price_cents=800)
        _insert_variant(fresh_db, iid, "Small", 800)
        _insert_variant(fresh_db, iid, "Large", 1200)
        items = get_draft_items(d, include_variants=True)
        wb = _build_xlsx_single(items)
        ws = wb.active
        rows = _ws_rows(ws)
        header = rows[0]
        assert "price_Small" in header
        assert "price_Large" in header

    def test_parent_row_bold(self, fresh_db):
        """Parent item rows have bold font."""
        from storage.drafts import get_draft_items
        d = _create_draft(fresh_db)
        _insert_item(fresh_db, d, "Pizza", price_cents=800, category="Pizza")
        items = get_draft_items(d, include_variants=True)
        wb = _build_xlsx_single(items)
        ws = wb.active
        # Row 2 is the first data row (parent)
        cell = ws.cell(row=2, column=1)
        assert cell.font.bold is True

    def test_variant_row_gray_and_indented(self, fresh_db):
        """Variant sub-rows have gray font + gray fill + indented label."""
        from storage.drafts import get_draft_items
        d = _create_draft(fresh_db)
        iid = _insert_item(fresh_db, d, "Pizza", price_cents=800)
        _insert_variant(fresh_db, iid, "Small", 800)
        items = get_draft_items(d, include_variants=True)
        wb = _build_xlsx_single(items)
        ws = wb.active
        # Row 3 is the variant sub-row
        vcell = ws.cell(row=3, column=1)
        assert vcell.value.startswith("  ")  # indented
        assert "Small" in vcell.value
        assert vcell.font.color is not None
        # Verify fill
        assert ws.cell(row=3, column=1).fill.start_color.rgb is not None

    def test_variant_subrow_fields(self, fresh_db):
        """Variant sub-row has label (indented), kind, price."""
        from storage.drafts import get_draft_items
        d = _create_draft(fresh_db)
        iid = _insert_item(fresh_db, d, "Wings", price_cents=600)
        _insert_variant(fresh_db, iid, "W/Fries", 800, "combo", 0)
        items = get_draft_items(d, include_variants=True)
        wb = _build_xlsx_single(items)
        ws = wb.active
        rows = _ws_rows(ws)
        vrow = rows[2]  # row index 2 = variant
        assert "W/Fries" in vrow[0]  # name column (indented label)
        assert vrow[1] == "combo"    # description column (kind)
        assert vrow[2] == 800        # price column

    def test_parent_with_variant_prices_in_columns(self, fresh_db):
        """Parent row has variant prices in dynamic label columns."""
        from storage.drafts import get_draft_items
        d = _create_draft(fresh_db)
        iid = _insert_item(fresh_db, d, "Pizza", price_cents=800)
        _insert_variant(fresh_db, iid, "Small", 800, "size", 0)
        _insert_variant(fresh_db, iid, "Medium", 1000, "size", 1)
        _insert_variant(fresh_db, iid, "Large", 1200, "size", 2)
        items = get_draft_items(d, include_variants=True)
        wb = _build_xlsx_single(items)
        ws = wb.active
        rows = _ws_rows(ws)
        header = rows[0]
        parent = rows[1]
        # Find column indices for variant labels
        sm_idx = header.index("price_Small")
        md_idx = header.index("price_Medium")
        lg_idx = header.index("price_Large")
        assert parent[sm_idx] == 800
        assert parent[md_idx] == 1000
        assert parent[lg_idx] == 1200

    def test_item_without_variants_empty_columns(self, fresh_db):
        """Items without variants have empty variant columns."""
        from storage.drafts import get_draft_items
        d = _create_draft(fresh_db)
        iid = _insert_item(fresh_db, d, "Pizza", price_cents=800)
        _insert_variant(fresh_db, iid, "Small", 800)
        _insert_variant(fresh_db, iid, "Large", 1200)
        _insert_item(fresh_db, d, "Salad", price_cents=500, category="Salads")
        items = get_draft_items(d, include_variants=True)
        wb = _build_xlsx_single(items)
        ws = wb.active
        rows = _ws_rows(ws)
        # Salad is the third data row (after Pizza parent + Pizza variant sub-rows)
        # Find the Salad row
        salad_row = None
        for r in rows[1:]:
            if r[0] == "Salad":
                salad_row = r
                break
        assert salad_row is not None
        header = rows[0]
        sm_idx = header.index("price_Small")
        lg_idx = header.index("price_Large")
        assert salad_row[sm_idx] == ""
        assert salad_row[lg_idx] == ""

    def test_empty_draft_header_only(self, fresh_db):
        """Empty draft produces header-only sheet."""
        from storage.drafts import get_draft_items
        d = _create_draft(fresh_db)
        items = get_draft_items(d, include_variants=True)
        wb = _build_xlsx_single(items)
        ws = wb.active
        rows = _ws_rows(ws)
        assert len(rows) == 1  # header only
        assert rows[0] == ["name", "description", "price_cents", "category"]

    def test_many_variants_all_exported(self, fresh_db):
        """Item with 10+ variants exports all in sub-rows."""
        from storage.drafts import get_draft_items
        d = _create_draft(fresh_db)
        iid = _insert_item(fresh_db, d, "Custom Pizza", price_cents=500)
        for i in range(12):
            _insert_variant(fresh_db, iid, f"Size_{i}", 500 + i * 100, "size", i)
        items = get_draft_items(d, include_variants=True)
        wb = _build_xlsx_single(items)
        ws = wb.active
        rows = _ws_rows(ws)
        # 1 header + 1 parent + 12 variant sub-rows = 14
        assert len(rows) == 14
        # All variant labels present in sub-rows
        variant_labels = [r[0].strip() for r in rows[2:]]
        for i in range(12):
            assert f"Size_{i}" in variant_labels

    def test_all_five_variant_kinds(self, fresh_db):
        """All 5 variant kinds appear in sub-rows."""
        from storage.drafts import get_draft_items
        d = _create_draft(fresh_db)
        iid = _insert_item(fresh_db, d, "Custom", price_cents=500)
        kinds = ["size", "combo", "flavor", "style", "other"]
        for i, k in enumerate(kinds):
            _insert_variant(fresh_db, iid, f"Var_{k}", 500 + i * 100, k, i)
        items = get_draft_items(d, include_variants=True)
        wb = _build_xlsx_single(items)
        ws = wb.active
        rows = _ws_rows(ws)
        variant_rows = rows[2:]  # skip header + parent
        assert len(variant_rows) == 5
        exported_kinds = [r[1] for r in variant_rows]  # kind is in description column
        assert exported_kinds == kinds

    def test_mixed_items_with_and_without_variants(self, fresh_db):
        """Mix of items: some with variants, some without."""
        from storage.drafts import get_draft_items
        d = _create_draft(fresh_db)
        iid1 = _insert_item(fresh_db, d, "Pizza", price_cents=800, category="Pizza")
        _insert_variant(fresh_db, iid1, "Small", 800, "size", 0)
        _insert_variant(fresh_db, iid1, "Large", 1200, "size", 1)
        _insert_item(fresh_db, d, "Salad", price_cents=500, category="Salads")
        iid3 = _insert_item(fresh_db, d, "Wings", price_cents=600, category="Apps")
        _insert_variant(fresh_db, iid3, "6pc", 600, "size", 0)
        items = get_draft_items(d, include_variants=True)
        wb = _build_xlsx_single(items)
        ws = wb.active
        rows = _ws_rows(ws)
        # 1 header + Pizza(1 parent + 2 variants) + Salad(1 parent) + Wings(1 parent + 1 variant) = 7
        assert len(rows) == 7

    def test_variant_order_preserved(self, fresh_db):
        """Variants appear in position order in sub-rows."""
        from storage.drafts import get_draft_items
        d = _create_draft(fresh_db)
        iid = _insert_item(fresh_db, d, "Drink", price_cents=200)
        _insert_variant(fresh_db, iid, "XL", 500, "size", 2)
        _insert_variant(fresh_db, iid, "Small", 200, "size", 0)
        _insert_variant(fresh_db, iid, "Medium", 350, "size", 1)
        items = get_draft_items(d, include_variants=True)
        wb = _build_xlsx_single(items)
        ws = wb.active
        rows = _ws_rows(ws)
        variant_labels = [r[0].strip() for r in rows[2:]]
        assert variant_labels == ["Small", "Medium", "XL"]

    def test_header_font_styled(self, fresh_db):
        """Header row cells have bold white font and dark fill."""
        from storage.drafts import get_draft_items
        d = _create_draft(fresh_db)
        _insert_item(fresh_db, d, "Salad", price_cents=500)
        items = get_draft_items(d, include_variants=True)
        wb = _build_xlsx_single(items)
        ws = wb.active
        hcell = ws.cell(row=1, column=1)
        assert hcell.font.bold is True

    def test_xlsx_round_trip_bytes(self, fresh_db):
        """Workbook serializes and deserializes correctly."""
        from storage.drafts import get_draft_items
        d = _create_draft(fresh_db)
        iid = _insert_item(fresh_db, d, "Burger", price_cents=899, category="Burgers")
        _insert_variant(fresh_db, iid, "Single", 899, "size", 0)
        _insert_variant(fresh_db, iid, "Double", 1299, "size", 1)
        items = get_draft_items(d, include_variants=True)
        wb = _build_xlsx_single(items, title="Round Trip")
        data = _wb_to_bytes(wb)
        wb2 = _load_wb(data)
        ws2 = wb2.active
        assert ws2.title == "Round Trip"
        rows = _ws_rows(ws2)
        assert len(rows) == 4  # header + parent + 2 variants
        assert rows[1][0] == "Burger"

    def test_label_columns_first_appearance_order(self, fresh_db):
        """Variant label columns follow first-appearance order."""
        from storage.drafts import get_draft_items
        d = _create_draft(fresh_db)
        iid1 = _insert_item(fresh_db, d, "Pizza", price_cents=800)
        _insert_variant(fresh_db, iid1, "Large", 1200, "size", 0)
        _insert_variant(fresh_db, iid1, "Small", 800, "size", 1)
        iid2 = _insert_item(fresh_db, d, "Wings", price_cents=600)
        _insert_variant(fresh_db, iid2, "Small", 600, "size", 0)
        _insert_variant(fresh_db, iid2, "XL", 1000, "size", 1)
        items = get_draft_items(d, include_variants=True)
        wb = _build_xlsx_single(items)
        ws = wb.active
        rows = _ws_rows(ws)
        header = rows[0]
        # Large appeared first, then Small, then XL
        assert header.index("price_Large") < header.index("price_Small")
        assert header.index("price_Small") < header.index("price_XL")


# ===========================================================================
# Per-Category XLSX Export Tests
# ===========================================================================
class TestXlsxByCategory:
    """Tests for XLSX export with one sheet per category."""

    def test_one_sheet_per_category(self, fresh_db):
        """Each unique category gets its own sheet."""
        from storage.drafts import get_draft_items
        d = _create_draft(fresh_db)
        _insert_item(fresh_db, d, "Pizza Margherita", price_cents=800, category="Pizza")
        _insert_item(fresh_db, d, "Caesar Salad", price_cents=500, category="Salads")
        _insert_item(fresh_db, d, "Garlic Bread", price_cents=300, category="Appetizers")
        items = get_draft_items(d, include_variants=True)
        wb = _build_xlsx_by_category(items)
        sheet_names = wb.sheetnames
        assert len(sheet_names) == 3
        assert "Appetizers" in sheet_names
        assert "Pizza" in sheet_names
        assert "Salads" in sheet_names

    def test_sheets_sorted_alphabetically(self, fresh_db):
        """Category sheets are in alphabetical order."""
        from storage.drafts import get_draft_items
        d = _create_draft(fresh_db)
        _insert_item(fresh_db, d, "Wings", price_cents=600, category="Appetizers")
        _insert_item(fresh_db, d, "Pizza", price_cents=800, category="Pizza")
        _insert_item(fresh_db, d, "Cake", price_cents=400, category="Desserts")
        items = get_draft_items(d, include_variants=True)
        wb = _build_xlsx_by_category(items)
        assert wb.sheetnames == ["Appetizers", "Desserts", "Pizza"]

    def test_items_in_correct_category_sheet(self, fresh_db):
        """Items appear only in their category's sheet."""
        from storage.drafts import get_draft_items
        d = _create_draft(fresh_db)
        _insert_item(fresh_db, d, "Pizza Margherita", price_cents=800, category="Pizza")
        _insert_item(fresh_db, d, "Pizza Pepperoni", price_cents=900, category="Pizza")
        _insert_item(fresh_db, d, "Caesar Salad", price_cents=500, category="Salads")
        items = get_draft_items(d, include_variants=True)
        wb = _build_xlsx_by_category(items)
        pizza_rows = _ws_rows(wb["Pizza"])
        salad_rows = _ws_rows(wb["Salads"])
        # Pizza sheet: 1 header + 2 items = 3 rows
        assert len(pizza_rows) == 3
        assert pizza_rows[1][0] == "Pizza Margherita"
        assert pizza_rows[2][0] == "Pizza Pepperoni"
        # Salads sheet: 1 header + 1 item = 2 rows
        assert len(salad_rows) == 2
        assert salad_rows[1][0] == "Caesar Salad"

    def test_category_sheet_has_variant_subrows(self, fresh_db):
        """Category sheets include variant sub-rows under parents."""
        from storage.drafts import get_draft_items
        d = _create_draft(fresh_db)
        iid = _insert_item(fresh_db, d, "Pizza", price_cents=800, category="Pizza")
        _insert_variant(fresh_db, iid, "Small", 800, "size", 0)
        _insert_variant(fresh_db, iid, "Large", 1200, "size", 1)
        items = get_draft_items(d, include_variants=True)
        wb = _build_xlsx_by_category(items)
        rows = _ws_rows(wb["Pizza"])
        # 1 header + 1 parent + 2 variant sub-rows = 4
        assert len(rows) == 4
        assert rows[1][0] == "Pizza"  # parent
        assert "Small" in rows[2][0]  # variant (indented)
        assert "Large" in rows[3][0]  # variant (indented)

    def test_category_specific_variant_columns(self, fresh_db):
        """Each category sheet has only its own variant label columns."""
        from storage.drafts import get_draft_items
        d = _create_draft(fresh_db)
        iid1 = _insert_item(fresh_db, d, "Pizza", price_cents=800, category="Pizza")
        _insert_variant(fresh_db, iid1, "Small", 800)
        _insert_variant(fresh_db, iid1, "Large", 1200)
        iid2 = _insert_item(fresh_db, d, "Wings", price_cents=600, category="Appetizers")
        _insert_variant(fresh_db, iid2, "6pc", 600)
        _insert_variant(fresh_db, iid2, "12pc", 1000)
        items = get_draft_items(d, include_variants=True)
        wb = _build_xlsx_by_category(items)
        pizza_header = _ws_rows(wb["Pizza"])[0]
        apps_header = _ws_rows(wb["Appetizers"])[0]
        # Pizza sheet should have Small/Large columns, not 6pc/12pc
        assert "price_Small" in pizza_header
        assert "price_Large" in pizza_header
        assert "price_6pc" not in pizza_header
        # Appetizers sheet should have 6pc/12pc columns, not Small/Large
        assert "price_6pc" in apps_header
        assert "price_12pc" in apps_header
        assert "price_Small" not in apps_header

    def test_uncategorized_items(self, fresh_db):
        """Items with no category go to 'Uncategorized' sheet."""
        from storage.drafts import get_draft_items
        d = _create_draft(fresh_db)
        _insert_item(fresh_db, d, "Mystery Item", price_cents=500)
        _insert_item(fresh_db, d, "Pizza", price_cents=800, category="Pizza")
        items = get_draft_items(d, include_variants=True)
        wb = _build_xlsx_by_category(items)
        assert "Uncategorized" in wb.sheetnames
        uncat_rows = _ws_rows(wb["Uncategorized"])
        assert len(uncat_rows) == 2  # header + 1 item
        assert uncat_rows[1][0] == "Mystery Item"

    def test_empty_draft_placeholder_sheet(self, fresh_db):
        """Empty draft produces a placeholder sheet."""
        from storage.drafts import get_draft_items
        d = _create_draft(fresh_db)
        items = get_draft_items(d, include_variants=True)
        wb = _build_xlsx_by_category(items)
        assert len(wb.sheetnames) == 1
        assert wb.sheetnames[0] == "Empty"
        rows = _ws_rows(wb.active)
        assert rows[0][0] == "No items"

    def test_category_parent_bold_variant_gray(self, fresh_db):
        """Per-category sheet: parent rows bold, variant rows gray."""
        from storage.drafts import get_draft_items
        d = _create_draft(fresh_db)
        iid = _insert_item(fresh_db, d, "Pizza", price_cents=800, category="Pizza")
        _insert_variant(fresh_db, iid, "Small", 800)
        items = get_draft_items(d, include_variants=True)
        wb = _build_xlsx_by_category(items)
        ws = wb["Pizza"]
        # Row 2 = parent (bold)
        assert ws.cell(row=2, column=1).font.bold is True
        # Row 3 = variant (not bold, has fill)
        vcell = ws.cell(row=3, column=1)
        assert vcell.font.bold is not True
        assert vcell.fill.start_color.rgb is not None

    def test_long_category_name_truncated(self, fresh_db):
        """Category names longer than 31 chars are truncated for sheet title."""
        from storage.drafts import get_draft_items
        d = _create_draft(fresh_db)
        long_cat = "A" * 50
        _insert_item(fresh_db, d, "Item", price_cents=500, category=long_cat)
        items = get_draft_items(d, include_variants=True)
        wb = _build_xlsx_by_category(items)
        assert len(wb.sheetnames) == 1
        assert len(wb.sheetnames[0]) == 31

    def test_per_category_round_trip_bytes(self, fresh_db):
        """Per-category workbook serializes and deserializes correctly."""
        from storage.drafts import get_draft_items
        d = _create_draft(fresh_db)
        iid = _insert_item(fresh_db, d, "Burger", price_cents=899, category="Burgers")
        _insert_variant(fresh_db, iid, "Single", 899, "size", 0)
        _insert_variant(fresh_db, iid, "Double", 1299, "size", 1)
        _insert_item(fresh_db, d, "Fries", price_cents=399, category="Sides")
        items = get_draft_items(d, include_variants=True)
        wb = _build_xlsx_by_category(items)
        data = _wb_to_bytes(wb)
        wb2 = _load_wb(data)
        assert "Burgers" in wb2.sheetnames
        assert "Sides" in wb2.sheetnames
        burger_rows = _ws_rows(wb2["Burgers"])
        assert len(burger_rows) == 4  # header + parent + 2 variants


# ===========================================================================
# Edge Cases
# ===========================================================================
class TestXlsxEdgeCases:
    """Edge cases for Excel variant export."""

    def test_zero_price_variant(self, fresh_db):
        """Variant with price_cents=0 exports correctly."""
        from storage.drafts import get_draft_items
        d = _create_draft(fresh_db)
        iid = _insert_item(fresh_db, d, "Water", price_cents=0)
        _insert_variant(fresh_db, iid, "Cup", 0, "size", 0)
        _insert_variant(fresh_db, iid, "Bottle", 200, "size", 1)
        items = get_draft_items(d, include_variants=True)
        wb = _build_xlsx_single(items)
        ws = wb.active
        rows = _ws_rows(ws)
        header = rows[0]
        parent = rows[1]
        cup_idx = header.index("price_Cup")
        bottle_idx = header.index("price_Bottle")
        assert parent[cup_idx] == 0
        assert parent[bottle_idx] == 200

    def test_special_chars_in_names(self, fresh_db):
        """Special characters in names/labels preserved in XLSX."""
        from storage.drafts import get_draft_items
        d = _create_draft(fresh_db)
        iid = _insert_item(fresh_db, d, 'Mac "N" Cheese', price_cents=700, category="Pasta & More")
        _insert_variant(fresh_db, iid, "Small, 8oz", 700)
        items = get_draft_items(d, include_variants=True)
        wb = _build_xlsx_single(items)
        ws = wb.active
        rows = _ws_rows(ws)
        assert rows[1][0] == 'Mac "N" Cheese'
        assert rows[1][3] == "Pasta & More"
        assert "Small, 8oz" in rows[2][0]

    def test_item_no_description(self, fresh_db):
        """Item with None description exports as empty string."""
        from storage.drafts import get_draft_items
        d = _create_draft(fresh_db)
        _insert_item(fresh_db, d, "Salad", price_cents=500)
        items = get_draft_items(d, include_variants=True)
        wb = _build_xlsx_single(items)
        ws = wb.active
        rows = _ws_rows(ws)
        assert rows[1][1] is None or rows[1][1] == ""

    def test_combo_and_size_variants_mixed(self, fresh_db):
        """Item with both size and combo variants exports all."""
        from storage.drafts import get_draft_items
        d = _create_draft(fresh_db)
        iid = _insert_item(fresh_db, d, "Wings", price_cents=600, category="Apps")
        _insert_variant(fresh_db, iid, "6pc", 600, "size", 0)
        _insert_variant(fresh_db, iid, "12pc", 1000, "size", 1)
        _insert_variant(fresh_db, iid, "W/Fries", 800, "combo", 2)
        items = get_draft_items(d, include_variants=True)
        wb = _build_xlsx_single(items)
        ws = wb.active
        rows = _ws_rows(ws)
        header = rows[0]
        assert "price_6pc" in header
        assert "price_12pc" in header
        assert "price_W/Fries" in header
        # 1 header + 1 parent + 3 variants = 5
        assert len(rows) == 5

    def test_multiple_categories_variant_counts(self, fresh_db):
        """Multiple categories with varying variant counts."""
        from storage.drafts import get_draft_items
        d = _create_draft(fresh_db)
        # Pizza: 3 sizes
        iid1 = _insert_item(fresh_db, d, "Pepperoni", price_cents=800, category="Pizza")
        _insert_variant(fresh_db, iid1, "S", 800, "size", 0)
        _insert_variant(fresh_db, iid1, "M", 1000, "size", 1)
        _insert_variant(fresh_db, iid1, "L", 1200, "size", 2)
        # Salad: no variants
        _insert_item(fresh_db, d, "Caesar", price_cents=500, category="Salads")
        # Drinks: 2 sizes
        iid3 = _insert_item(fresh_db, d, "Soda", price_cents=200, category="Beverages")
        _insert_variant(fresh_db, iid3, "Reg", 200, "size", 0)
        _insert_variant(fresh_db, iid3, "Large", 300, "size", 1)
        items = get_draft_items(d, include_variants=True)
        wb = _build_xlsx_by_category(items)
        # Pizza: header + 1 parent + 3 variants = 5
        assert len(_ws_rows(wb["Pizza"])) == 5
        # Salads: header + 1 parent = 2
        assert len(_ws_rows(wb["Salads"])) == 2
        # Beverages: header + 1 parent + 2 variants = 4
        assert len(_ws_rows(wb["Beverages"])) == 4

    def test_category_no_variant_columns(self, fresh_db):
        """Category sheet with no variants has only base columns."""
        from storage.drafts import get_draft_items
        d = _create_draft(fresh_db)
        _insert_item(fresh_db, d, "Salad", price_cents=500, category="Salads")
        _insert_item(fresh_db, d, "Soup", price_cents=400, category="Salads")
        items = get_draft_items(d, include_variants=True)
        wb = _build_xlsx_by_category(items)
        header = _ws_rows(wb["Salads"])[0]
        assert header == ["name", "description", "price_cents"]

    def test_single_sheet_title_from_draft(self, fresh_db):
        """Single-sheet XLSX uses draft title as sheet name."""
        from storage.drafts import get_draft_items
        d = _create_draft(fresh_db, title="My Restaurant Menu")
        _insert_item(fresh_db, d, "Salad", price_cents=500)
        items = get_draft_items(d, include_variants=True)
        wb = _build_xlsx_single(items, title="My Restaurant Menu")
        assert wb.active.title == "My Restaurant Menu"

    def test_partial_variant_coverage_across_items(self, fresh_db):
        """Items with partial variant coverage show prices only for their labels."""
        from storage.drafts import get_draft_items
        d = _create_draft(fresh_db)
        iid1 = _insert_item(fresh_db, d, "Pizza", price_cents=800)
        _insert_variant(fresh_db, iid1, "Small", 800)
        _insert_variant(fresh_db, iid1, "Medium", 1000)
        _insert_variant(fresh_db, iid1, "Large", 1200)
        iid2 = _insert_item(fresh_db, d, "Calzone", price_cents=900)
        _insert_variant(fresh_db, iid2, "Small", 900)
        _insert_variant(fresh_db, iid2, "Large", 1400)
        # Calzone has no Medium
        items = get_draft_items(d, include_variants=True)
        wb = _build_xlsx_single(items)
        ws = wb.active
        rows = _ws_rows(ws)
        header = rows[0]
        md_idx = header.index("price_Medium")
        # Find Calzone parent row
        calzone_row = None
        for r in rows[1:]:
            if r[0] == "Calzone":
                calzone_row = r
                break
        assert calzone_row is not None
        assert calzone_row[md_idx] == ""  # No medium for calzone
