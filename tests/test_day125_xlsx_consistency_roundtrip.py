"""
Day 125 — XLSX Modifier Groups, Cross-Item Consistency & Round-Trip Tests
=========================================================================
Tests for:
  - XLSX export: modifier group header rows (blue) + modifier sub-rows + ungrouped variants
  - XLSX by-category: same treatment per category sheet
  - Cross-item modifier group consistency validation warnings
  - End-to-end round-trip tests for all export formats with modifier group data
  - Backward compat: items without modifier groups export cleanly

Test plan (~32 tests):
  Class 1: XLSX export with modifier groups (8 tests)
    1.  Item with no modifiers → item row only
    2.  Item with modifier group → group header + modifier rows
    3.  Group header row has group_name + required columns
    4.  Modifier rows indented with 4 spaces + carry group_name
    5.  Ungrouped variants → indented with 2 spaces (backward compat)
    6.  Mixed: modifier_groups + ungrouped_variants both emitted
    7.  Multiple groups → multiple group header rows in order
    8.  Route GET → 200 XLSX content type

  Class 2: XLSX by-category with modifier groups (8 tests)
    9.  Single category sheet with modifier group rows
    10. Multiple categories → one sheet per category
    11. Group header rows styled differently from modifier rows
    12. Empty draft → placeholder "Empty" sheet
    13. Items without groups → plain item rows (backward compat)
    14. Category sheet has no category column (already implicit)
    15. Route GET → 200 XLSX content type
    16. Unauthenticated → 302

  Class 3: Cross-item modifier group consistency checks (8 tests)
    17. No modifier groups → no consistency warnings
    18. All items in category have same group → no warnings
    19. Most items have group, one missing → modifier_group_inconsistent warning
    20. Fewer than 3 items with groups → no consistency check (too few)
    21. Multiple categories checked independently
    22. Group used by exactly 50% → flagged (>=50% threshold)
    23. Group used by <50% → not flagged
    24. Items without any groups not counted in denominator

  Class 4: End-to-end round-trip tests (8 tests)
    25. XLSX round-trip: items with modifier groups survive export
    26. XLSX by-category round-trip: modifier group rows per sheet
    27. CSV variants round-trip with modifier groups
    28. JSON round-trip with modifier groups
    29. Square CSV backward compat: items without groups export cleanly
    30. Toast CSV backward compat: items without groups export cleanly
    31. XLSX backward compat: items without groups → item rows only
    32. Validation with modifier groups: all warning types fire correctly
"""
from __future__ import annotations

import csv
import io
import json
import sqlite3

import pytest

import storage.drafts as drafts_mod


# ---------------------------------------------------------------------------
# Schema (matches Day 124)
# ---------------------------------------------------------------------------

_SCHEMA_SQL = """
CREATE TABLE IF NOT EXISTS restaurants (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    name TEXT NOT NULL,
    active INTEGER DEFAULT 1
);
CREATE TABLE IF NOT EXISTS drafts (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    title TEXT,
    restaurant_id INTEGER,
    status TEXT NOT NULL DEFAULT 'editing',
    source TEXT,
    source_job_id INTEGER,
    source_file_path TEXT,
    menu_id INTEGER,
    category_order TEXT,
    created_at TEXT NOT NULL,
    updated_at TEXT NOT NULL,
    FOREIGN KEY (restaurant_id) REFERENCES restaurants(id) ON DELETE SET NULL
);
CREATE TABLE IF NOT EXISTS draft_items (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    draft_id INTEGER NOT NULL,
    name TEXT NOT NULL,
    description TEXT,
    price_cents INTEGER NOT NULL DEFAULT 0,
    category TEXT,
    position INTEGER,
    confidence INTEGER,
    kitchen_name TEXT,
    created_at TEXT NOT NULL,
    updated_at TEXT NOT NULL,
    FOREIGN KEY (draft_id) REFERENCES drafts(id) ON DELETE CASCADE
);
CREATE TABLE IF NOT EXISTS draft_item_variants (
    id                INTEGER PRIMARY KEY AUTOINCREMENT,
    item_id           INTEGER NOT NULL,
    label             TEXT NOT NULL,
    price_cents       INTEGER NOT NULL DEFAULT 0,
    kind              TEXT DEFAULT 'size',
    position          INTEGER DEFAULT 0,
    modifier_group_id INTEGER,
    created_at        TEXT NOT NULL,
    updated_at        TEXT NOT NULL,
    FOREIGN KEY (item_id) REFERENCES draft_items(id) ON DELETE CASCADE
);
CREATE TABLE IF NOT EXISTS draft_modifier_groups (
    id          INTEGER PRIMARY KEY AUTOINCREMENT,
    item_id     INTEGER NOT NULL,
    name        TEXT NOT NULL,
    required    INTEGER DEFAULT 0,
    min_select  INTEGER DEFAULT 0,
    max_select  INTEGER DEFAULT 0,
    position    INTEGER DEFAULT 0,
    created_at  TEXT NOT NULL,
    updated_at  TEXT NOT NULL,
    FOREIGN KEY (item_id) REFERENCES draft_items(id) ON DELETE CASCADE
);
CREATE TABLE IF NOT EXISTS draft_modifier_group_templates (
    id            INTEGER PRIMARY KEY AUTOINCREMENT,
    restaurant_id INTEGER,
    name          TEXT NOT NULL,
    required      INTEGER DEFAULT 0,
    min_select    INTEGER DEFAULT 0,
    max_select    INTEGER DEFAULT 0,
    position      INTEGER DEFAULT 0,
    modifiers     TEXT NOT NULL DEFAULT '[]',
    created_at    TEXT NOT NULL,
    updated_at    TEXT NOT NULL
);
CREATE TABLE IF NOT EXISTS draft_export_history (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    draft_id INTEGER NOT NULL,
    format TEXT NOT NULL,
    item_count INTEGER NOT NULL DEFAULT 0,
    variant_count INTEGER NOT NULL DEFAULT 0,
    warning_count INTEGER NOT NULL DEFAULT 0,
    exported_at TEXT NOT NULL,
    FOREIGN KEY (draft_id) REFERENCES drafts(id) ON DELETE CASCADE
);
CREATE TABLE IF NOT EXISTS menu_items (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    restaurant_id INTEGER,
    name TEXT NOT NULL,
    price_cents INTEGER NOT NULL DEFAULT 0,
    category TEXT,
    created_at TEXT NOT NULL DEFAULT (datetime('now')),
    updated_at TEXT NOT NULL DEFAULT (datetime('now'))
);
CREATE TABLE IF NOT EXISTS api_keys (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    key_hash TEXT UNIQUE NOT NULL,
    restaurant_id INTEGER,
    label TEXT,
    active INTEGER DEFAULT 1,
    rate_limit_rpm INTEGER DEFAULT 60,
    created_at TEXT NOT NULL
);
CREATE TABLE IF NOT EXISTS pipeline_rejections (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    job_id INTEGER,
    gate_score REAL,
    gate_threshold REAL,
    reason TEXT,
    customer_message TEXT,
    rejected_at TEXT NOT NULL
);
"""

_NOW = "2026-03-18T10:00:00"


def _make_conn() -> sqlite3.Connection:
    conn = sqlite3.connect(":memory:")
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA foreign_keys = ON;")
    conn.executescript(_SCHEMA_SQL)
    return conn


def _create_draft(conn, status="editing", title="Menu") -> tuple[int, int]:
    rid = conn.execute(
        "INSERT INTO restaurants (name, active) VALUES ('R', 1)"
    ).lastrowid
    did = conn.execute(
        "INSERT INTO drafts (title, restaurant_id, status, created_at, updated_at) "
        "VALUES (?, ?, ?, ?, ?)",
        (title, rid, status, _NOW, _NOW),
    ).lastrowid
    conn.commit()
    return rid, did


def _create_item(conn, draft_id, name="Burger", price_cents=999, category="Entrees") -> int:
    iid = conn.execute(
        "INSERT INTO draft_items (draft_id, name, price_cents, category, position, created_at, updated_at) "
        "VALUES (?, ?, ?, ?, 0, ?, ?)",
        (draft_id, name, price_cents, category, _NOW, _NOW),
    ).lastrowid
    conn.commit()
    return iid


def _create_modifier_group(conn, item_id, name="Size", required=0,
                            min_select=0, max_select=0, position=0) -> int:
    gid = conn.execute(
        "INSERT INTO draft_modifier_groups "
        "(item_id, name, required, min_select, max_select, position, created_at, updated_at) "
        "VALUES (?, ?, ?, ?, ?, ?, ?, ?)",
        (item_id, name, required, min_select, max_select, position, _NOW, _NOW),
    ).lastrowid
    conn.commit()
    return gid


def _create_variant(conn, item_id, label="Small", price_cents=0, kind="size",
                     modifier_group_id=None, position=0) -> int:
    vid = conn.execute(
        "INSERT INTO draft_item_variants "
        "(item_id, label, price_cents, kind, position, modifier_group_id, created_at, updated_at) "
        "VALUES (?, ?, ?, ?, ?, ?, ?, ?)",
        (item_id, label, price_cents, kind, position, modifier_group_id, _NOW, _NOW),
    ).lastrowid
    conn.commit()
    return vid


# ---------------------------------------------------------------------------
# Fixtures
# ---------------------------------------------------------------------------

@pytest.fixture()
def app_ctx(monkeypatch):
    """Patch db_connect for both storage + portal; yield (app, client, conn)."""
    import portal.app as _app_mod

    conn = _make_conn()

    def _fake_connect():
        return conn

    monkeypatch.setattr(drafts_mod, "db_connect", _fake_connect)
    monkeypatch.setattr(_app_mod, "db_connect", _fake_connect)

    _app_mod.app.config["TESTING"] = True
    _app_mod.app.config["WTF_CSRF_ENABLED"] = False
    _app_mod.app.secret_key = "test"

    with _app_mod.app.test_client() as client:
        with client.session_transaction() as sess:
            sess["user"] = "tester"
        yield _app_mod.app, client, conn


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _item(name="Burger", price_cents=999, category="Entrees",
          modifier_groups=None, ungrouped_variants=None):
    d = {
        "id": 1, "name": name, "description": "",
        "price_cents": price_cents, "category": category,
    }
    if modifier_groups is not None:
        d["modifier_groups"] = modifier_groups
    if ungrouped_variants is not None:
        d["ungrouped_variants"] = ungrouped_variants
    return d


def _mg(name="Size", required=False, min_select=0, max_select=0, modifiers=None):
    return {
        "id": 1, "name": name,
        "required": 1 if required else 0,
        "min_select": min_select, "max_select": max_select,
        "modifiers": modifiers or [],
    }


def _mod(label="Small", price_cents=0, kind="size"):
    return {"id": 1, "label": label, "price_cents": price_cents, "kind": kind}


def _parse_csv(csv_text: str) -> list[list[str]]:
    reader = csv.reader(io.StringIO(csv_text))
    return [row for row in reader]


def _xlsx_rows(resp_data: bytes) -> list[list]:
    """Parse XLSX response bytes into list of row lists."""
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(resp_data))
    ws = wb.active
    return [[cell.value for cell in row] for row in ws.iter_rows()]


def _xlsx_sheet_rows(resp_data: bytes, sheet_name: str) -> list[list]:
    """Parse a specific sheet from XLSX response bytes."""
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(resp_data))
    ws = wb[sheet_name]
    return [[cell.value for cell in row] for row in ws.iter_rows()]


def _xlsx_cell_fill(resp_data: bytes, row: int, col: int) -> str:
    """Get fill color of a cell (1-indexed row/col)."""
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(resp_data))
    ws = wb.active
    return ws.cell(row=row, column=col).fill.start_color.rgb or ""


def _xlsx_sheet_cell_fill(resp_data: bytes, sheet_name: str, row: int, col: int) -> str:
    """Get fill color of a cell in a specific sheet."""
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(resp_data))
    ws = wb[sheet_name]
    return ws.cell(row=row, column=col).fill.start_color.rgb or ""


# ---------------------------------------------------------------------------
# Class 1: XLSX export with modifier groups (tests 1–8)
# ---------------------------------------------------------------------------

class TestXLSXModifierGroups:
    """XLSX export upgraded for POS-native modifier groups."""

    def test_no_modifiers_item_row_only(self, app_ctx):
        """1. Item with no modifiers → header + item row only."""
        app, client, conn = app_ctx
        _, did = _create_draft(conn)
        _create_item(conn, did, "Burger", 999, "Entrees")

        resp = client.get(f"/drafts/{did}/export.xlsx")
        assert resp.status_code == 200
        rows = _xlsx_rows(resp.data)
        # Header + 1 item row
        assert len(rows) == 2
        assert rows[0][0] == "name"  # header
        assert rows[1][0] == "Burger"

    def test_modifier_group_emits_group_and_modifier_rows(self, app_ctx):
        """2. Item with modifier group → group header + modifier sub-rows."""
        app, client, conn = app_ctx
        _, did = _create_draft(conn)
        iid = _create_item(conn, did, "Burger", 999, "Entrees")
        gid = _create_modifier_group(conn, iid, "Size", required=1)
        _create_variant(conn, iid, "Small", 599, "size", modifier_group_id=gid)
        _create_variant(conn, iid, "Large", 899, "size", modifier_group_id=gid)

        resp = client.get(f"/drafts/{did}/export.xlsx")
        rows = _xlsx_rows(resp.data)
        # Header + item + group header + 2 modifiers = 5 rows
        assert len(rows) == 5
        assert rows[1][0] == "Burger"  # item row

    def test_group_header_has_name_and_required(self, app_ctx):
        """3. Group header row shows group_name and required flag."""
        app, client, conn = app_ctx
        _, did = _create_draft(conn)
        iid = _create_item(conn, did, "Burger", 999, "Entrees")
        gid = _create_modifier_group(conn, iid, "Size", required=1)
        _create_variant(conn, iid, "Small", 599, "size", modifier_group_id=gid)

        resp = client.get(f"/drafts/{did}/export.xlsx")
        rows = _xlsx_rows(resp.data)
        # Row 3 is group header (1=header, 2=item, 3=group)
        group_row = rows[2]
        # group_name column (index 4 with category)
        assert group_row[4] == "Size"
        # required column (index 5)
        assert group_row[5] == "Y"

    def test_modifier_rows_indented_with_group_name(self, app_ctx):
        """4. Modifier rows indented with 4 spaces and carry group_name."""
        app, client, conn = app_ctx
        _, did = _create_draft(conn)
        iid = _create_item(conn, did, "Burger", 999, "Entrees")
        gid = _create_modifier_group(conn, iid, "Size", required=0)
        _create_variant(conn, iid, "Small", 599, "size", modifier_group_id=gid)

        resp = client.get(f"/drafts/{did}/export.xlsx")
        rows = _xlsx_rows(resp.data)
        # Row 4 is modifier (1=header, 2=item, 3=group, 4=modifier)
        mod_row = rows[3]
        assert mod_row[0].startswith("    ")  # 4-space indent
        assert "Small" in mod_row[0]
        assert mod_row[4] == "Size"  # group_name column

    def test_ungrouped_variants_indented_two_spaces(self, app_ctx):
        """5. Ungrouped variants → 2-space indent (backward compat)."""
        app, client, conn = app_ctx
        _, did = _create_draft(conn)
        iid = _create_item(conn, did, "Burger", 999, "Entrees")
        _create_variant(conn, iid, "Extra Cheese", 150, "combo")  # no group

        resp = client.get(f"/drafts/{did}/export.xlsx")
        rows = _xlsx_rows(resp.data)
        # Header + item + ungrouped variant = 3 rows
        assert len(rows) == 3
        variant_row = rows[2]
        assert variant_row[0].startswith("  ")  # 2-space indent
        assert "Extra Cheese" in variant_row[0]

    def test_mixed_groups_and_ungrouped(self, app_ctx):
        """6. Item with both modifier_groups and ungrouped_variants."""
        app, client, conn = app_ctx
        _, did = _create_draft(conn)
        iid = _create_item(conn, did, "Burger", 999, "Entrees")
        gid = _create_modifier_group(conn, iid, "Size", required=1)
        _create_variant(conn, iid, "Small", 599, "size", modifier_group_id=gid)
        _create_variant(conn, iid, "Extra Cheese", 150, "combo")  # ungrouped

        resp = client.get(f"/drafts/{did}/export.xlsx")
        rows = _xlsx_rows(resp.data)
        # Header + item + group header + modifier + ungrouped variant = 5 rows
        assert len(rows) == 5
        # Group header row
        assert rows[2][4] == "Size"
        # Modifier row (4-space)
        assert rows[3][0].startswith("    ")
        # Ungrouped variant (2-space)
        assert rows[4][0].startswith("  ") and not rows[4][0].startswith("    ")

    def test_multiple_groups_multiple_headers(self, app_ctx):
        """7. Multiple modifier groups → multiple group header rows."""
        app, client, conn = app_ctx
        _, did = _create_draft(conn)
        iid = _create_item(conn, did, "Burger", 999, "Entrees")
        g1 = _create_modifier_group(conn, iid, "Size", required=1, position=0)
        g2 = _create_modifier_group(conn, iid, "Toppings", required=0, position=1)
        _create_variant(conn, iid, "Small", 599, "size", modifier_group_id=g1)
        _create_variant(conn, iid, "Bacon", 150, "combo", modifier_group_id=g2)

        resp = client.get(f"/drafts/{did}/export.xlsx")
        rows = _xlsx_rows(resp.data)
        # Header + item + group1 + mod1 + group2 + mod2 = 6 rows
        assert len(rows) == 6
        group_names = [r[4] for r in rows[1:] if r[4]]
        assert "Size" in group_names
        assert "Toppings" in group_names

    def test_route_returns_xlsx_content_type(self, app_ctx):
        """8. Route GET → 200 with XLSX content type."""
        app, client, conn = app_ctx
        _, did = _create_draft(conn)
        _create_item(conn, did, "Burger", 999, "Entrees")

        resp = client.get(f"/drafts/{did}/export.xlsx")
        assert resp.status_code == 200
        assert "spreadsheetml" in resp.headers["Content-Type"]
        assert f"draft_{did}.xlsx" in resp.headers["Content-Disposition"]


# ---------------------------------------------------------------------------
# Class 2: XLSX by-category with modifier groups (tests 9–16)
# ---------------------------------------------------------------------------

class TestXLSXByCategoryModifierGroups:
    """XLSX by-category export with modifier group support."""

    def test_single_category_with_modifier_groups(self, app_ctx):
        """9. Single category sheet has modifier group rows."""
        app, client, conn = app_ctx
        _, did = _create_draft(conn)
        iid = _create_item(conn, did, "Burger", 999, "Entrees")
        gid = _create_modifier_group(conn, iid, "Size", required=1)
        _create_variant(conn, iid, "Small", 599, "size", modifier_group_id=gid)

        resp = client.get(f"/drafts/{did}/export_by_category.xlsx")
        assert resp.status_code == 200
        rows = _xlsx_sheet_rows(resp.data, "Entrees")
        # Header + item + group + modifier = 4 rows
        assert len(rows) == 4
        # No category column in by-category sheets
        assert "category" not in rows[0]

    def test_multiple_categories_separate_sheets(self, app_ctx):
        """10. Multiple categories → one sheet per category."""
        app, client, conn = app_ctx
        _, did = _create_draft(conn)
        _create_item(conn, did, "Burger", 999, "Entrees")
        _create_item(conn, did, "Fries", 499, "Sides")

        resp = client.get(f"/drafts/{did}/export_by_category.xlsx")
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(resp.data))
        sheet_names = wb.sheetnames
        assert "Entrees" in sheet_names
        assert "Sides" in sheet_names

    def test_group_header_styled_differently(self, app_ctx):
        """11. Group header rows have distinct fill color (light blue D6EAF8)."""
        app, client, conn = app_ctx
        _, did = _create_draft(conn)
        iid = _create_item(conn, did, "Burger", 999, "Entrees")
        gid = _create_modifier_group(conn, iid, "Size", required=1)
        _create_variant(conn, iid, "Small", 599, "size", modifier_group_id=gid)

        resp = client.get(f"/drafts/{did}/export_by_category.xlsx")
        # Group header is row 3 (1=header, 2=item, 3=group)
        fill = _xlsx_sheet_cell_fill(resp.data, "Entrees", 3, 1)
        assert "D6EAF8" in fill.upper()

    def test_empty_draft_placeholder_sheet(self, app_ctx):
        """12. Empty draft → placeholder 'Empty' sheet."""
        app, client, conn = app_ctx
        _, did = _create_draft(conn)

        resp = client.get(f"/drafts/{did}/export_by_category.xlsx")
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(resp.data))
        assert "Empty" in wb.sheetnames

    def test_items_without_groups_plain_rows(self, app_ctx):
        """13. Items without modifier groups → plain item rows."""
        app, client, conn = app_ctx
        _, did = _create_draft(conn)
        _create_item(conn, did, "Burger", 999, "Entrees")

        resp = client.get(f"/drafts/{did}/export_by_category.xlsx")
        rows = _xlsx_sheet_rows(resp.data, "Entrees")
        # Header + 1 item row
        assert len(rows) == 2
        assert rows[1][0] == "Burger"

    def test_category_sheet_no_category_column(self, app_ctx):
        """14. Category sheets don't include category column (already implicit)."""
        app, client, conn = app_ctx
        _, did = _create_draft(conn)
        _create_item(conn, did, "Burger", 999, "Entrees")

        resp = client.get(f"/drafts/{did}/export_by_category.xlsx")
        rows = _xlsx_sheet_rows(resp.data, "Entrees")
        headers = [str(c) for c in rows[0]]
        assert "category" not in headers

    def test_route_returns_xlsx_content_type(self, app_ctx):
        """15. Route GET → 200 with XLSX content type."""
        app, client, conn = app_ctx
        _, did = _create_draft(conn)
        _create_item(conn, did, "Burger", 999, "Entrees")

        resp = client.get(f"/drafts/{did}/export_by_category.xlsx")
        assert resp.status_code == 200
        assert "spreadsheetml" in resp.headers["Content-Type"]

    def test_unauthenticated_redirect(self, app_ctx):
        """16. Unauthenticated → 302 redirect."""
        app, client, conn = app_ctx
        _, did = _create_draft(conn)

        # Clear session
        with client.session_transaction() as sess:
            sess.pop("user", None)

        resp = client.get(f"/drafts/{did}/export_by_category.xlsx")
        assert resp.status_code == 302


# ---------------------------------------------------------------------------
# Class 3: Cross-item modifier group consistency checks (tests 17–24)
# ---------------------------------------------------------------------------

class TestCrossItemModifierGroupConsistency:
    """Validation: cross-item modifier group consistency within categories."""

    def _validate(self, items):
        import portal.app as _app_mod
        return _app_mod._validate_draft_for_export(items)

    def test_no_modifier_groups_no_warnings(self):
        """17. Items without modifier groups → no consistency warnings."""
        items = [_item("A"), _item("B"), _item("C")]
        warnings = self._validate(items)
        types = [w["type"] for w in warnings]
        assert "modifier_group_inconsistent" not in types

    def test_all_items_same_group_no_warnings(self):
        """18. All items in category have same group → no warnings."""
        mods = [_mod("Small", 599), _mod("Large", 899)]
        items = [
            _item("A", modifier_groups=[_mg("Size", modifiers=mods)]),
            _item("B", modifier_groups=[_mg("Size", modifiers=mods)]),
            _item("C", modifier_groups=[_mg("Size", modifiers=mods)]),
        ]
        warnings = self._validate(items)
        types = [w["type"] for w in warnings]
        assert "modifier_group_inconsistent" not in types

    def test_most_items_have_group_one_missing_flagged(self):
        """19. 3 items with groups, one missing 'Size' → flagged."""
        mods = [_mod("Small", 599), _mod("Large", 899)]
        items = [
            _item("A", modifier_groups=[_mg("Size", modifiers=mods)]),
            _item("B", modifier_groups=[_mg("Size", modifiers=mods)]),
            _item("C", modifier_groups=[_mg("Toppings", modifiers=[_mod("Bacon", 150, "combo")])]),
        ]
        warnings = self._validate(items)
        inconsistent = [w for w in warnings if w["type"] == "modifier_group_inconsistent"]
        # "C" doesn't have "Size" but A and B do (2/3 = 66.7% >= 50%)
        flagged_names = [w["name"] for w in inconsistent]
        assert "C" in flagged_names

    def test_fewer_than_3_items_no_check(self):
        """20. Fewer than 3 items with groups → no consistency check."""
        mods = [_mod("Small", 599)]
        items = [
            _item("A", modifier_groups=[_mg("Size", modifiers=mods)]),
            _item("B", modifier_groups=[_mg("Toppings", modifiers=[_mod("Bacon")])]),
        ]
        warnings = self._validate(items)
        types = [w["type"] for w in warnings]
        assert "modifier_group_inconsistent" not in types

    def test_multiple_categories_independent(self):
        """21. Consistency checks run per-category independently."""
        mods = [_mod("Small", 599), _mod("Large", 899)]
        items = [
            _item("A", category="Entrees", modifier_groups=[_mg("Size", modifiers=mods)]),
            _item("B", category="Entrees", modifier_groups=[_mg("Size", modifiers=mods)]),
            _item("C", category="Entrees", modifier_groups=[_mg("Size", modifiers=mods)]),
            _item("D", category="Sides", modifier_groups=[_mg("Temp", modifiers=[_mod("Hot")])]),
        ]
        warnings = self._validate(items)
        inconsistent = [w for w in warnings if w["type"] == "modifier_group_inconsistent"]
        # "Sides" only has 1 item → no check
        assert len(inconsistent) == 0

    def test_group_at_50_percent_flagged(self):
        """22. Group used by exactly 50% of items → threshold met, outliers flagged."""
        mods = [_mod("Small", 599)]
        items = [
            _item("A", modifier_groups=[_mg("Size", modifiers=mods)]),
            _item("B", modifier_groups=[_mg("Size", modifiers=mods)]),
            _item("C", modifier_groups=[_mg("Toppings", modifiers=[_mod("Bacon")])]),
            _item("D", modifier_groups=[_mg("Toppings", modifiers=[_mod("Cheese")])]),
        ]
        warnings = self._validate(items)
        inconsistent = [w for w in warnings if w["type"] == "modifier_group_inconsistent"]
        # Size: 2/4 = 50% → flagged; C and D don't have Size
        size_flagged = [w for w in inconsistent if "Size" in w["message"]]
        assert len(size_flagged) >= 1

    def test_group_below_50_percent_not_flagged(self):
        """23. Group used by <50% → not flagged."""
        mods = [_mod("Small", 599)]
        items = [
            _item("A", modifier_groups=[_mg("Size", modifiers=mods)]),
            _item("B", modifier_groups=[_mg("Toppings", modifiers=[_mod("Bacon")])]),
            _item("C", modifier_groups=[_mg("Toppings", modifiers=[_mod("Cheese")])]),
            _item("D", modifier_groups=[_mg("Toppings", modifiers=[_mod("Lettuce")])]),
        ]
        warnings = self._validate(items)
        inconsistent = [w for w in warnings if w["type"] == "modifier_group_inconsistent"]
        # Size: 1/4 = 25% < 50% → not flagged
        size_flagged = [w for w in inconsistent if "Size" in w["message"]]
        assert len(size_flagged) == 0

    def test_items_without_groups_not_in_denominator(self):
        """24. Items without any modifier groups not counted in denominator."""
        mods = [_mod("Small", 599)]
        items = [
            _item("A", modifier_groups=[_mg("Size", modifiers=mods)]),
            _item("B", modifier_groups=[_mg("Size", modifiers=mods)]),
            _item("C", modifier_groups=[_mg("Size", modifiers=mods)]),
            _item("D"),  # no modifier groups at all
            _item("E"),  # no modifier groups at all
        ]
        warnings = self._validate(items)
        inconsistent = [w for w in warnings if w["type"] == "modifier_group_inconsistent"]
        # Only 3 items have groups, all 3 have "Size" → no inconsistency
        assert len(inconsistent) == 0


# ---------------------------------------------------------------------------
# Class 4: End-to-end round-trip tests (tests 25–32)
# ---------------------------------------------------------------------------

class TestEndToEndRoundTrip:
    """Round-trip and backward compat tests across all export formats."""

    def test_xlsx_roundtrip_modifier_groups(self, app_ctx):
        """25. XLSX export with modifier groups → rows survive round-trip."""
        app, client, conn = app_ctx
        _, did = _create_draft(conn)
        iid = _create_item(conn, did, "Burger", 999, "Entrees")
        gid = _create_modifier_group(conn, iid, "Size", required=1, min_select=1, max_select=1)
        _create_variant(conn, iid, "Small", 599, "size", modifier_group_id=gid)
        _create_variant(conn, iid, "Large", 899, "size", modifier_group_id=gid)

        resp = client.get(f"/drafts/{did}/export.xlsx")
        rows = _xlsx_rows(resp.data)
        # Verify structure: header + item + group + 2 mods = 5
        assert len(rows) == 5
        # Item row
        assert rows[1][0] == "Burger"
        assert rows[1][2] == 999
        # Group row
        assert rows[2][4] == "Size"
        assert rows[2][5] == "Y"
        # Modifier rows
        assert "Small" in rows[3][0]
        assert rows[3][2] == 599
        assert "Large" in rows[4][0]
        assert rows[4][2] == 899

    def test_xlsx_by_category_roundtrip(self, app_ctx):
        """26. XLSX by-category: modifier group rows present per sheet."""
        app, client, conn = app_ctx
        _, did = _create_draft(conn)
        iid = _create_item(conn, did, "Pizza", 1299, "Pizza")
        gid = _create_modifier_group(conn, iid, "Size", required=1)
        _create_variant(conn, iid, "10 inch", 1299, "size", modifier_group_id=gid)
        _create_variant(conn, iid, "14 inch", 1599, "size", modifier_group_id=gid)

        resp = client.get(f"/drafts/{did}/export_by_category.xlsx")
        rows = _xlsx_sheet_rows(resp.data, "Pizza")
        assert len(rows) == 5  # header + item + group + 2 mods
        assert rows[2][3] == "Size"  # group_name col (no category col)
        assert rows[2][4] == "Y"  # required

    def test_csv_variants_roundtrip_with_groups(self, app_ctx):
        """27. CSV variants export with modifier groups → correct row types."""
        app, client, conn = app_ctx
        _, did = _create_draft(conn)
        iid = _create_item(conn, did, "Burger", 999, "Entrees")
        gid = _create_modifier_group(conn, iid, "Size", required=1)
        _create_variant(conn, iid, "Small", 599, "size", modifier_group_id=gid)
        _create_variant(conn, iid, "Extra Cheese", 150, "combo")  # ungrouped

        resp = client.get(f"/drafts/{did}/export_variants.csv")
        rows = _parse_csv(resp.data.decode("utf-8-sig"))
        types = [r[0] for r in rows[1:]]
        assert "item" in types
        assert "modifier_group" in types
        assert "modifier" in types
        assert "variant" in types  # ungrouped

    def test_json_roundtrip_with_groups(self, app_ctx):
        """28. JSON export with modifier groups → structured output."""
        app, client, conn = app_ctx
        _, did = _create_draft(conn)
        iid = _create_item(conn, did, "Burger", 999, "Entrees")
        gid = _create_modifier_group(conn, iid, "Size", required=1)
        _create_variant(conn, iid, "Small", 599, "size", modifier_group_id=gid)

        resp = client.get(f"/drafts/{did}/export.json")
        data = json.loads(resp.data)
        items = data["items"]
        assert len(items) == 1
        assert len(items[0]["modifier_groups"]) == 1
        mg = items[0]["modifier_groups"][0]
        assert mg["name"] == "Size"
        assert mg["required"] is True
        assert len(mg["modifiers"]) == 1
        assert mg["modifiers"][0]["label"] == "Small"

    def test_square_csv_backward_compat_no_groups(self, app_ctx):
        """29. Square CSV: items without modifier groups export cleanly."""
        app, client, conn = app_ctx
        _, did = _create_draft(conn)
        _create_item(conn, did, "Burger", 999, "Entrees")
        _create_variant(conn, did, "Small", 599, "size")  # ungrouped
        # The ungrouped variant is on the item itself
        # Actually need to attach variant to item, not draft
        conn.execute("DELETE FROM draft_item_variants")
        conn.commit()
        iid = conn.execute(
            "SELECT id FROM draft_items WHERE draft_id=?", (did,)
        ).fetchone()[0]
        _create_variant(conn, iid, "Small", 599, "size")

        resp = client.get(f"/drafts/{did}/export_square.csv")
        assert resp.status_code == 200
        rows = _parse_csv(resp.data.decode("utf-8-sig"))
        tokens = [r[0] for r in rows[1:]]
        assert "item" in tokens

    def test_toast_csv_backward_compat_no_groups(self, app_ctx):
        """30. Toast CSV: items without modifier groups export cleanly."""
        app, client, conn = app_ctx
        _, did = _create_draft(conn)
        _create_item(conn, did, "Burger", 999, "Entrees")

        resp = client.get(f"/drafts/{did}/export_toast.csv")
        assert resp.status_code == 200
        rows = _parse_csv(resp.data.decode("utf-8-sig"))
        assert len(rows) >= 2  # header + at least 1 item

    def test_xlsx_backward_compat_no_groups(self, app_ctx):
        """31. XLSX: items without modifier groups → simple item rows only."""
        app, client, conn = app_ctx
        _, did = _create_draft(conn)
        _create_item(conn, did, "Burger", 999, "Entrees")
        _create_item(conn, did, "Fries", 499, "Sides")

        resp = client.get(f"/drafts/{did}/export.xlsx")
        rows = _xlsx_rows(resp.data)
        # Header + 2 item rows
        assert len(rows) == 3
        names = [r[0] for r in rows[1:]]
        assert "Burger" in names
        assert "Fries" in names

    def test_validation_all_modifier_group_warnings(self, app_ctx):
        """32. Validation fires all modifier group warning types correctly."""
        import portal.app as _app_mod

        items = [
            _item("A", modifier_groups=[
                _mg("Empty Group", modifiers=[]),  # empty
                _mg("Bad Limits", required=True, min_select=5, max_select=2,
                     modifiers=[_mod("X", 100)]),  # required_empty=no (has mods), min>max, max>count
            ]),
        ]
        warnings = _app_mod._validate_draft_for_export(items)
        types = {w["type"] for w in warnings}
        assert "modifier_group_empty" in types
        assert "group_min_exceeds_max" in types
        assert "group_max_exceeds_count" in types
